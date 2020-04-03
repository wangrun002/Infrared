#!/usr/bin/python3
# -*- coding: utf-8 -*-

from datetime import datetime
from random import choice,sample
from threading import Timer
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment
from openpyxl.styles.colors import RED
from openpyxl.utils import get_column_letter,column_index_from_string
import serial
import serial.tools.list_ports
import logging
import re
import time
import copy
import sys
import os
import random
import threading
import shelve
import platform

class MyGlobal():
    def __init__(self):
        self.main_loop_state = True                         # 主程序循环状态变量
        self.control_stage_delay_state = True               # 控制延时的状态变量
        self.get_group_channel_total_info_state = True      # 控制获取组别等信息的进程与用例执行代码切换的状态变量
        self.current_stage = 0                              # 控制执行用例的各个阶段
        self.sub_stage = 0                                  # 控制信息获取的各个阶段
        self.get_ch_attr_sub_stage = 0                      # 控制获取节目属性的各个阶段
        self.read_ch_attr_sub_stage = 0                     # 控制读取节目属性后的处理的各个阶段
        self.commd_global_pos = 0                           # 发送指令在list中的位置
        self.commd_global_length = 0                        # 发送指令的总长度
        self.specify_group_prog_total = 0                   # 所选case指定分组的节目总数
        self.TV_channel_groups = {}                         # 存放电视节目的组别和节目数信息
        self.Radio_channel_groups = {}                      # 存放广播节目的组别和节目数信息

        # [频道号,频道名称,tp,ttx,sub,lock,scramble,频道类型,视频编码,音频编码,视频高度,视频宽度,组别,切台时间]
        self.channel_info = ['','','','','','','','','','','','','','']
        self.prog_group_name = ''                           # 组别名称
        self.prog_group_total = ''                          # 组别下的节目总数
        self.numb_key_switch_commd = []                     # 数字键切台指令集
        self.screen_switch_commd = []                       # 大画面下切台指令集
        self.channel_list_switch_commd = []                 # 节目列表界面切台指令集
        self.epg_switch_commd = []                          # EPG界面切台指令集
        self.channel_edit_switch_commd = []                 # 节目编辑界面切台指令集
        self.tv_and_radio_switch_commd = []                 # 广电切换指令集
        self.recall_switch_commd = []                       # 回看切换指令集
        self.all_test_case = []                             # 存放所有播放控制测试用例的数据集合
        self.scene_commd_length = []                        # 存放不同界面的指令长度
        self.report_data = [0,0,0,0,0,0,[],[]]
        self.choice_channel = ['','']                       # 存放显示回看或广电切换时,所选两个节目的频道信息
        self.TV_ch_attribute = [[], [], []]                 # 用于存放TV节目属性的列表(免费\加密\加锁)
        self.Radio_ch_attribute = [[], [], []]              # 用于存放Radio节目属性的列表(免费\加密\加锁)
        self.free_tv_tp_ch_dict = {}                        # 用于存放免费TV节目按TP归类的字典
        self.free_tv_codec_ch_dict = {}                     # 用于存放免费TV节目按codec归类的字典
        self.free_tv_resolution_ch_dict = {}                # 用于存放免费TV节目按resolution归类的字典


def check_ports():
    global send_com,receive_com
    ports_info = []
    if platform.system() == "Windows":
        ser_cable_num = 5
        serial_ser = {
            "1": "FTDVKA2HA",
            "2": "FTGDWJ64A",
            "3": "FT9SP964A",
            "4": "FTHB6SSTA",
            "5": "FTDVKPRSA",
            "6": "FTHI8UIHA",
             }
        send_port_desc = "USB-SERIAL CH340"
        receive_port_desc = serial_ser[str(ser_cable_num)]
    elif platform.system() == "Linux":
        ser_cable_num = 5
        serial_ser = {
            "1": "FTDVKA2H",
            "2": "FTGDWJ64",
            "3": "FT9SP964",
            "4": "FTHB6SST",
            "5": "FTDVKPRS",
            "6": "FTHI8UIH",
            }
        send_port_desc = "USB2.0-Serial"
        receive_port_desc = serial_ser[str(ser_cable_num)]
    ports = list(serial.tools.list_ports.comports())
    for i in range(len(ports)):
        logging.info("可用端口:名称:{} + 描述:{} + 硬件id:{}".format(ports[i].device, ports[i].description, ports[i].hwid))
        print("可用端口:名称:{} + 描述:{} + 硬件id:{}".format(ports[i].device, ports[i].description, ports[i].hwid))
        ports_info.append("{}~{}~{}".format(ports[i].device, ports[i].description, ports[i].hwid))
    if len(ports) <= 0:
        logging.info("无可用端口")
    elif len(ports) == 1:
        logging.info("只有一个可用端口:{}".format(ports[0].device))
    elif len(ports) >= 2:
        for i in range(len(ports_info)):
            if send_port_desc in ports_info[i]:
                send_com = ports_info[i].split("~")[0]
            elif receive_port_desc in ports_info[i]:
                receive_com = ports_info[i].split("~")[0]
    return send_com,receive_com

def serial_set(ser, ser_name, ser_baudrate):
    ser.port = ser_name
    ser.baudrate = ser_baudrate
    ser.bytesize = 8
    ser.parity = "N"
    ser.stopbits = 1
    ser.timeout = 1
    ser.write_timeout = 0
    ser.open()

def hex_strs_to_bytes(strings):
    return bytes.fromhex(strings)

def send_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    time.sleep(1)

def send_numb_key_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    time.sleep(0.5)

def write_logs_to_txt(file_path,logs):
    with open(file_path, "a+", encoding="utf-8") as fo:
        fo.write(logs)

def unlock_channel():
    for i in range(4):
        send_numb_key_commd(KEY["0"])

def check_if_log_and_report_file_path_exists():
    global case_log_file_directory, full_log_file_directory
    global report_file_directory, transfer_data_file_directory, logging_file_directory
    parent_path = os.path.dirname(os.getcwd())
    test_file_folder_name = "test_data"
    test_file_directory = os.path.join(parent_path, test_file_folder_name)
    case_log_folder_name = "print_log"
    case_log_file_directory = os.path.join(parent_path, test_file_folder_name, case_log_folder_name)
    full_log_folder_name = "print_log_full"
    full_log_file_directory = os.path.join(parent_path, test_file_folder_name, full_log_folder_name)
    report_folder_name = "report"
    report_file_directory = os.path.join(parent_path, test_file_folder_name, report_folder_name)
    transfer_data_folder_name = "transfer_data"
    transfer_data_file_directory = os.path.join(parent_path, test_file_folder_name, transfer_data_folder_name)
    logging_file_folder_name = "logging_file"
    logging_file_directory = os.path.join(parent_path, test_file_folder_name, logging_file_folder_name)

    if not os.path.exists(test_file_directory):
        os.mkdir(test_file_directory)
    if not os.path.exists(case_log_file_directory):
        os.mkdir(case_log_file_directory)
    if not os.path.exists(full_log_file_directory):
        os.mkdir(full_log_file_directory)
    if not os.path.exists(report_file_directory):
        os.mkdir(report_file_directory)
    if not os.path.exists(transfer_data_file_directory):
        os.mkdir(transfer_data_file_directory)
    if not os.path.exists(logging_file_directory):
        os.mkdir(logging_file_directory)

def build_print_log_and_report_file_path():
    global case_log_txt_path, full_log_txt_path, report_file_path, sheet_name, time_info, fmt_name
    global transfer_tv_data_file_path, transfer_radio_data_file_path, logging_file_path
    case_info = ALL_TEST_CASE[choice_switch_case]
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    fmt_name = "{}_{}_{}_{}_{}".format(case_info[4],case_info[3], case_info[0], case_info[1], case_info[2])

    case_log_file_name = "{}_{}_{}.txt".format(choice_switch_case,fmt_name, time_info)
    case_log_txt_path = os.path.join(case_log_file_directory, case_log_file_name)
    full_log_file_name = "full_{}_{}_{}.txt".format(choice_switch_case,fmt_name, time_info)
    full_log_txt_path = os.path.join(full_log_file_directory, full_log_file_name)
    report_file_name = "{}_{}_{}.xlsx".format(choice_switch_case,fmt_name, time_info)
    report_file_path = os.path.join(report_file_directory, report_file_name)
    transfer_tv_data_file_name = "tv_ch_attr_info.db"
    transfer_tv_data_file_path = os.path.join(transfer_data_file_directory, transfer_tv_data_file_name)
    transfer_radio_data_file_name = "radio_ch_attr_info.db"
    transfer_radio_data_file_path = os.path.join(transfer_data_file_directory, transfer_radio_data_file_name)
    logging_file_name = "{}_{}_{}_test.log".format(choice_switch_case,fmt_name, time_info)
    logging_file_path = os.path.join(logging_file_directory, logging_file_name)
    sheet_name = "{}".format(case_info[1])

    GL.report_data[0] = "{}_{}_{}".format(case_info[0], case_info[1], case_info[2])     # 报告名称
    GL.report_data[1] = "{}".format(case_info[3])       # 分组信息
    GL.report_data[3] = "{}".format(case_info[4])       # 节目类型
    GL.report_data[4] = "{} + {}".format(case_info[1],case_info[2])     # 切台模式

def check_if_report_exists_and_write_data_to_report():
    report_title = [
        "报告名称",
        "分组名称",
        "分组节目总数",
        "节目类别",
        "切台模式",
        "切台次数",
        {"命令": EXCEL_CH_INFO_KWS},
    ]
    font = Font(color=RED)
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    if not os.path.exists(report_file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 16
        for i in range(len(report_title)):
            ws.row_dimensions[(i + 1)].height = 13.5
            if i < len(report_title) - 1:
                ws.cell(i + 1, 1).value = report_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(report_title) - 1:
                ws.cell(i + 1, 1).value = list(report_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
                for j in range(len(list(report_title[i].values())[0])):
                    all_column_numb = column_index_from_string("A") + (j + 1)
                    all_column_char = get_column_letter(all_column_numb)
                    if j == 1 or j == 2:
                        ws.column_dimensions[all_column_char].width = 16
                    else:
                        ws.column_dimensions[all_column_char].width = 9
                    ws.cell((i + 1), (1 + j + 1)).value = list(report_title[i].values())[0][j]
                    ws.cell((i + 1), (1 + j + 1)).alignment = alignment
        if GL.all_test_case[choice_switch_case][1] == READ_CH_ATTRIBUTE:
            ws.cell(5,8).value = ALL_TEST_CASE[choice_switch_case][1]
            ws.cell(5,9).value = "".join(GL.choice_channel[0])
            ws.cell(5,10).value = ALL_TEST_CASE[choice_switch_case][2]
            ws.cell(5,11).value = "".join(GL.choice_channel[1])
            for i in range(8,12):
                ws.cell(5,i).alignment = alignment

    elif os.path.exists(report_file_path):
        wb = load_workbook(report_file_path)
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 16
        for i in range(len(report_title)):
            ws.row_dimensions[(i + 1)].height = 13.5
            if i < len(report_title) - 1:
                ws.cell(i + 1, 1).value = report_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(report_title) - 1:
                ws.cell(i + 1, 1).value = list(report_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
                for j in range(len(list(report_title[i].values())[0])):
                    all_column_numb = column_index_from_string("A") + j
                    all_column_char = get_column_letter(all_column_numb)
                    if j == 1 or j == 2:
                        ws.column_dimensions[all_column_char].width = 16
                    else:
                        ws.column_dimensions[all_column_char].width = 9
                    ws.cell((i + 1), (1 + j + 1)).value = list(report_title[i].values())[0][j]
                    ws.cell((i + 1), (1 + j + 1)).alignment = alignment
        if GL.all_test_case[choice_switch_case][1] == READ_CH_ATTRIBUTE:
            ws.cell(5,8).value = ALL_TEST_CASE[choice_switch_case][1]
            ws.cell(5,9).value = "".join(GL.choice_channel[0])
            ws.cell(5,10).value = ALL_TEST_CASE[choice_switch_case][2]
            ws.cell(5,11).value = "".join(GL.choice_channel[1])
            for i in range(8,12):
                ws.cell(5,i).alignment = alignment

    for m in range(len(GL.report_data)):
        if m < len(GL.report_data) - 2:     # 写title项对应的数据
            ws.cell(m + 1, 2).value = GL.report_data[m]
            ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=7)
            ws.cell(m + 1, 2).alignment = alignment
        elif m == len(GL.report_data) - 2:  # 写发送的命令列表
            for n in range(len(GL.report_data[m])):
                ws.cell(m + 2 + n, 1).value = GL.report_data[m][n]
                ws.cell(m + 2 + n, 1).alignment = alignment
        elif m == len(GL.report_data) - 1:  # 写发送命令后记录的数据
            for k in range(len(GL.report_data[m])):
                prog_info_list = GL.report_data[m][k]
                for l in range(len(prog_info_list)):
                    if l == len(prog_info_list) - 1:
                        if prog_info_list[l] != '' and float(prog_info_list[l]) > float(3.0):
                            ws.cell((m + 1 + k), (2 + l)).value = GL.report_data[m][k][l]
                            ws.cell((m + 1 + k), (2 + l)).alignment = alignment
                            ws.cell((m + 1 + k), (2 + l)).font = font
                        else:
                            ws.cell((m + 1 + k), (2 + l)).value = GL.report_data[m][k][l]
                            ws.cell((m + 1 + k), (2 + l)).alignment = alignment
                    else:
                        ws.cell((m + 1 + k),(2 + l)).value = GL.report_data[m][k][l]
                        ws.cell((m + 1 + k),(2 + l)).alignment = alignment
                ws.row_dimensions[(m + 1 + k)].height = 13.5
    wb.save(report_file_path)

def delay_time(interval_time,expect_delay_time):
    global t
    expect_delay_time -= interval_time
    print(expect_delay_time)
    if expect_delay_time == 0:
        GL.control_stage_delay_state = True
    if expect_delay_time > 0:
        t = Timer(interval_time, delay_time, args=(interval_time, expect_delay_time)).start()

def build_ch_numbs_list(chs_total):
    # 新建指定个数的数值列表
    # ch_numbs = GL.channel_numbers
    ch_numbs_list = []
    for i in range(chs_total):
        ch_numbs_list.append(str(i + 1))
    return ch_numbs_list

def build_random_ch_commds_list(commds_list,number_of_time):
    # 在某指令集中随机抽取指定个数的指令集
    random_channel_commds_list = []
    for i in range(number_of_time):
        random_channel_commds_list.append(choice(commds_list))
    return random_channel_commds_list

def change_numbs_to_commds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_commds_list = []
    for i in range(len(numbs_list)):
        channel_commds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_commds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_commds_list[i].append(KEY[numbs_list[i][j]])
    return channel_commds_list

def change_commds_to_numbs_list(commds_list):
    # 将指令集列表转化为数值列表
    channel_numbs_list = []
    for i in range(len(commds_list)):
        channel_numbs_list.append([])
        for j in range(len(commds_list[i])):
            channel_numbs_list[i].append(REVERSE_KEY[commds_list[i][j]])
    return channel_numbs_list

def change_numbs_to_str_list(numbs_list):
    # 将数值列表转化为字符串
    channel_str_list = []
    for m in range(len(numbs_list)):
        channel_str_list.append("")
        for n in range(len(numbs_list[m])):
            try:
                int(numbs_list[m][n])
            except:
                if len(numbs_list[m]) > 1:
                    if n > 0:
                        channel_str_list[m] = channel_str_list[m] + "_" + numbs_list[m][n]
                    else:
                        channel_str_list[m] = channel_str_list[m] + numbs_list[m][n]
                else:
                    channel_str_list[m] = channel_str_list[m] + numbs_list[m][n]
            else:
                channel_str_list[m] = channel_str_list[m] + numbs_list[m][n]
    return channel_str_list

def commds_add_key_list(old_commds_list,key_name):
    # 在每个指令集后增加指定的单一指令
    new_commds_list = copy.deepcopy(old_commds_list)
    for i in range(len(new_commds_list)):
        new_commds_list[i].append(key_name)
    return new_commds_list

def build_single_commds_list(single_commd,number_of_time):
    # 创建指定次数的单一指令集
    single_commds_list = []
    for i in range(number_of_time):
        single_commds_list.append([])
        single_commds_list[i].append(single_commd)
    return single_commds_list

def build_random_move_focus_list(single_commd,number_of_time,group_total_numb):
    # 创建指定次数的随机个数单一指令集
    random_max_value = 20
    if group_total_numb > random_max_value:
        random_max_value = 20
    elif 1 < group_total_numb <= random_max_value:
        random_max_value = group_total_numb - 1
    elif group_total_numb == 1:
        logging.info("节目数量太少,不能进行上下键随机切台")
    random_mv_focus_list = []
    single_commd_list = [single_commd]
    for i in range(number_of_time):
        random_mv_focus_list.append(single_commd_list * random.randint(1,random_max_value))
    return random_mv_focus_list

def commds_add_random_move_focus_list(old_commds_list,single_commd):
    # 在已有的commds_list追加随机个数的单一指令集
    random_max_value = 5
    if ALL_TEST_CASE[choice_switch_case][4] == "TV":
        min_group_total = min([int(i) for i in list(GL.TV_channel_groups.values())])
    elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
        min_group_total = min([int(i) for i in list(GL.Radio_channel_groups.values())])

    if min_group_total > random_max_value:
        random_max_value = 5
    elif 1 < min_group_total <= random_max_value:
        random_max_value = min_group_total - 1
    elif min_group_total == 1:
        logging.info("节目数量太少,不能分组切换后进行上下键随机切台")
    new_commds_list = copy.deepcopy(old_commds_list)
    # single_commd_list = [single_commd]
    for i in range(len(new_commds_list)):
        for j in range(random.randint(1,random_max_value)):
            new_commds_list[i].append(single_commd)
    return new_commds_list

def build_numb_key_switch_list():
    # 电视下数字键超时切台、即时切台、随机切台（tv_numb、radio_numb、numb+ok、numb+random）
    # scene_list = ["numb", "numb+ok", "numb+random"]
    random_switch_time = 100
    numb_key_total_commd = []
    if ALL_TEST_CASE[choice_switch_case][4] == "TV":
        GL.specify_group_prog_total = int(GL.TV_channel_groups[ALL_TEST_CASE[choice_switch_case][3]])
    elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
        GL.specify_group_prog_total = int(GL.Radio_channel_groups[ALL_TEST_CASE[choice_switch_case][3]])
    chs_numb_list = build_ch_numbs_list(GL.specify_group_prog_total)
    # 生成各个case的commd list
    tv_numb_key_chs_commd_list = change_numbs_to_commds_list(chs_numb_list)
    radio_numb_key_chs_commd_list = change_numbs_to_commds_list(chs_numb_list)
    numb_key_chs_commd_add_ok_list = commds_add_key_list(tv_numb_key_chs_commd_list, KEY["OK"])
    numb_key_random_chs_commd_list = build_random_ch_commds_list(tv_numb_key_chs_commd_list, random_switch_time)
    # 将case_commd_list添加到numb_key_total_commd中
    numb_key_total_commd.append(tv_numb_key_chs_commd_list)
    numb_key_total_commd.append(radio_numb_key_chs_commd_list)
    numb_key_total_commd.append(numb_key_chs_commd_add_ok_list)
    numb_key_total_commd.append(numb_key_random_chs_commd_list)
    return numb_key_total_commd

def build_screen_switch_list():
    # 大画面上键切台、下键切台、随机切台（up、down、random、continuous）
    # scene_list = ["up", "down", "random", "continuous"]
    up_down_list = [[KEY["UP"]],[KEY["DOWN"]]]
    screen_total_commd = []
    # 根据获取的分组节目总数的奇偶来设定随机切换的次数
    if GL.specify_group_prog_total % 2 == 0:
        screen_random_switch_times = GL.specify_group_prog_total // 2
    elif GL.specify_group_prog_total % 2 != 0:
        screen_random_switch_times = (GL.specify_group_prog_total + 1) // 2
    # 生成各个case的commd list
    screen_up_commd_list = build_single_commds_list(KEY["UP"],GL.specify_group_prog_total)
    screen_down_commd_list = build_single_commds_list(KEY["DOWN"],GL.specify_group_prog_total)
    screen_random_commd_list = build_random_ch_commds_list(up_down_list,screen_random_switch_times)
    screen_cont_commd_list = build_single_commds_list(KEY["UP"],GL.specify_group_prog_total)
    # 将case_commd_list添加到screen_total_commd中
    screen_total_commd.append(screen_up_commd_list)
    screen_total_commd.append(screen_down_commd_list)
    screen_total_commd.append(screen_random_commd_list)
    screen_total_commd.append(screen_cont_commd_list)
    return screen_total_commd

def build_channel_list_switch_list():
    # 频道列表界面，上下键、翻页键、分组切台
    # （up+ok、down+ok、up_random+ok、down_random+ok、page_up+ok、page_down+ok、left、right、left+down+ok、right+down+ok）
    left_right_list = [[KEY["LEFT"]],[KEY["RIGHT"]]]
    random_up_or_down_time = 50
    group_switch_time = 50
    page_switch_time = 100
    channel_list_total_commd = []
    # 生成各个case的commd list
    ch_list_up_commd_list = build_single_commds_list(KEY["UP"],GL.specify_group_prog_total)
    ch_list_down_commd_list = build_single_commds_list(KEY["DOWN"],GL.specify_group_prog_total)
    ch_list_random_up_commd_list = build_random_move_focus_list(KEY["UP"],random_up_or_down_time,GL.specify_group_prog_total)
    ch_list_random_down_commd_list = build_random_move_focus_list(KEY["DOWN"],random_up_or_down_time,GL.specify_group_prog_total)
    ch_list_page_up_commd_list = build_single_commds_list(KEY["PAGE_UP"],page_switch_time)
    ch_list_page_down_commd_list = build_single_commds_list(KEY["PAGE_DOWN"],page_switch_time)
    ch_list_left_commd_list = build_single_commds_list(KEY["LEFT"],group_switch_time)
    ch_list_right_commd_list = build_single_commds_list(KEY["RIGHT"],group_switch_time)
    ch_list_left_random_down_commd_list = commds_add_random_move_focus_list(ch_list_left_commd_list,KEY["DOWN"])
    ch_list_left_random_down_ok_commd_list = commds_add_key_list(ch_list_left_random_down_commd_list,KEY["OK"])
    ch_list_right_random_down_commd_list = commds_add_random_move_focus_list(ch_list_right_commd_list,KEY["DOWN"])
    ch_list_right_random_down_ok_commd_list = commds_add_key_list(ch_list_right_random_down_commd_list,KEY["OK"])
    # 将case_commd_list添加到channel_list_total_commd中
    channel_list_total_commd.append(commds_add_key_list(ch_list_up_commd_list,KEY["OK"]))
    channel_list_total_commd.append(commds_add_key_list(ch_list_down_commd_list, KEY["OK"]))
    channel_list_total_commd.append(commds_add_key_list(ch_list_random_up_commd_list, KEY["OK"]))
    channel_list_total_commd.append(commds_add_key_list(ch_list_random_down_commd_list, KEY["OK"]))
    channel_list_total_commd.append(commds_add_key_list(ch_list_page_up_commd_list, KEY["OK"]))
    channel_list_total_commd.append(commds_add_key_list(ch_list_page_down_commd_list, KEY["OK"]))
    channel_list_total_commd.append(ch_list_left_commd_list)
    channel_list_total_commd.append(ch_list_right_commd_list)
    channel_list_total_commd.append(ch_list_left_random_down_ok_commd_list)
    channel_list_total_commd.append(ch_list_right_random_down_ok_commd_list)
    return channel_list_total_commd

def build_epg_switch_list():
    # EPG界面上下键、翻页键、随机、连续切台(up、down、page_up、page_down、random、continuous)
    up_down_list = [[KEY["UP"]], [KEY["DOWN"]]]
    page_switch_time = 50
    epg_total_commd = []
    # 根据获取的分组节目总数的奇偶来设定随机切换的次数
    if GL.specify_group_prog_total % 2 == 0:
        epg_random_switch_times = GL.specify_group_prog_total // 2
    elif GL.specify_group_prog_total % 2 != 0:
        epg_random_switch_times = (GL.specify_group_prog_total + 1) // 2
    # 生成各个case的commd list
    epg_up_commd_list = build_single_commds_list(KEY["UP"],GL.specify_group_prog_total)
    epg_down_commd_list = build_single_commds_list(KEY["DOWN"],GL.specify_group_prog_total)
    epg_page_up_commd_list = build_single_commds_list(KEY["PAGE_UP"],page_switch_time)
    epg_page_down_commd_list = build_single_commds_list(KEY["PAGE_DOWN"],page_switch_time)
    epg_random_commd_list = build_random_ch_commds_list(up_down_list,epg_random_switch_times)
    epg_continuous_commd_list = build_single_commds_list(KEY["UP"],GL.specify_group_prog_total)
    # 将case_commd_list添加到epg_total_commd中
    epg_total_commd.append(epg_up_commd_list)
    epg_total_commd.append(epg_down_commd_list)
    epg_total_commd.append(epg_page_up_commd_list)
    epg_total_commd.append(epg_page_down_commd_list)
    epg_total_commd.append(epg_random_commd_list)
    epg_total_commd.append(epg_continuous_commd_list)
    return epg_total_commd

def build_channel_edit_switch_list():
    # 频道编辑界面，上下键、翻页键、分组切台
    # （up+ok、down+ok、up_random+ok、down_random+ok、page_up+ok、page_down+ok、left、right、left+down+ok、right+down+ok）
    left_right_list = [[KEY["LEFT"]], [KEY["RIGHT"]]]
    random_up_or_down_time = 50
    group_switch_time = 50
    page_switch_time = 100
    channel_edit_total_commd = []
    # 生成各个case的commd list
    ch_edit_up_commd_list = build_single_commds_list(KEY["UP"], GL.specify_group_prog_total)
    ch_edit_down_commd_list = build_single_commds_list(KEY["DOWN"], GL.specify_group_prog_total)
    ch_edit_random_up_commd_list = build_random_move_focus_list(KEY["UP"], random_up_or_down_time,GL.specify_group_prog_total)
    ch_edit_random_down_commd_list = build_random_move_focus_list(KEY["DOWN"], random_up_or_down_time,GL.specify_group_prog_total)
    ch_edit_page_up_commd_list = build_single_commds_list(KEY["PAGE_UP"], page_switch_time)
    ch_edit_page_down_commd_list = build_single_commds_list(KEY["PAGE_DOWN"], page_switch_time)
    ch_edit_left_commd_list = build_single_commds_list(KEY["LEFT"], group_switch_time)
    ch_edit_right_commd_list = build_single_commds_list(KEY["RIGHT"], group_switch_time)
    ch_edit_left_random_down_commd_list = commds_add_random_move_focus_list(ch_edit_left_commd_list, KEY["DOWN"])
    ch_edit_left_random_down_ok_commd_list = commds_add_key_list(ch_edit_left_random_down_commd_list, KEY["OK"])
    ch_edit_right_random_down_commd_list = commds_add_random_move_focus_list(ch_edit_right_commd_list, KEY["DOWN"])
    ch_edit_right_random_down_ok_commd_list = commds_add_key_list(ch_edit_right_random_down_commd_list, KEY["OK"])
    # 将case_commd_list添加到channel_list_total_commd中
    channel_edit_total_commd.append(commds_add_key_list(ch_edit_up_commd_list, KEY["OK"]))
    channel_edit_total_commd.append(commds_add_key_list(ch_edit_down_commd_list, KEY["OK"]))
    channel_edit_total_commd.append(commds_add_key_list(ch_edit_random_up_commd_list, KEY["OK"]))
    channel_edit_total_commd.append(commds_add_key_list(ch_edit_random_down_commd_list, KEY["OK"]))
    channel_edit_total_commd.append(commds_add_key_list(ch_edit_page_up_commd_list, KEY["OK"]))
    channel_edit_total_commd.append(commds_add_key_list(ch_edit_page_down_commd_list, KEY["OK"]))
    channel_edit_total_commd.append(ch_edit_left_commd_list)
    channel_edit_total_commd.append(ch_edit_right_commd_list)
    channel_edit_total_commd.append(ch_edit_left_random_down_ok_commd_list)
    channel_edit_total_commd.append(ch_edit_right_random_down_ok_commd_list)
    return channel_edit_total_commd

def build_tv_and_radio_switch_list():
    # 广电切换
    # free_tv/free_radio,free_tv/scr_radio,free_tv/lock_radio,
    # scr_tv/free_radio,scr_tv/scr_radio,scr_tv/lock_radio,
    # lock_tv/free_radio,lock_tv/scr_radio,lock_tv/lock_radio,
    tv_radio_total_commd = []
    tv_radio_switch_time = 50
    # 生成各个case的commd list
    free_tv_and_free_radio_list = build_single_commds_list(KEY["TV/R"],tv_radio_switch_time)
    free_tv_and_scr_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    free_tv_and_lock_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    scr_tv_and_free_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    scr_tv_and_scr_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    scr_tv_and_lock_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    lock_tv_and_free_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    lock_tv_and_scr_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    lock_tv_and_lock_radio_list = build_single_commds_list(KEY["TV/R"], tv_radio_switch_time)
    # 将case_commd_list添加到tv_radio_total_commd中
    tv_radio_total_commd.append(free_tv_and_free_radio_list)
    tv_radio_total_commd.append(free_tv_and_scr_radio_list)
    tv_radio_total_commd.append(free_tv_and_lock_radio_list)
    tv_radio_total_commd.append(scr_tv_and_free_radio_list)
    tv_radio_total_commd.append(scr_tv_and_scr_radio_list)
    tv_radio_total_commd.append(scr_tv_and_lock_radio_list)
    tv_radio_total_commd.append(lock_tv_and_free_radio_list)
    tv_radio_total_commd.append(lock_tv_and_scr_radio_list)
    tv_radio_total_commd.append(lock_tv_and_lock_radio_list)
    return tv_radio_total_commd

def build_recall_switch_list():
    # 回看
    # tv_to_tv
    # radio_to_radio
    # tv_to_radio
    # same_or_diff_tp_tv_ch
    # same_or_diff_codec_tv_ch
    recall_total_commd = []
    recall_switch_time = 50
    # 各个case的commd list
    # tv_to_tv
    free_tv_and_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    free_tv_and_scr_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    free_tv_and_lock_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_tv_and_scr_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_tv_and_lock_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    lock_tv_and_lock_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    # radio_to_radio
    free_radio_and_free_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    free_radio_and_scr_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    free_radio_and_lock_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_radio_and_scr_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_radio_and_lock_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    lock_radio_and_lock_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    # tv_to_radio
    free_tv_and_free_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    free_tv_and_scr_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    free_tv_and_lock_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_tv_and_free_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_tv_and_scr_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    scr_tv_and_lock_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    lock_tv_and_free_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    lock_tv_and_scr_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    lock_tv_and_lock_radio_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    # same_or_diff_tp_tv_ch
    same_tp_free_tv_and_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    diff_tp_free_tv_and_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    # same_or_diff_codec_tv_ch
    same_codec_free_tv_and_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    diff_codec_free_tv_and_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    # hd_or_sd_tv_ch
    hd_free_tv_and_hd_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    sd_free_tv_and_sd_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    hd_free_tv_and_sd_free_tv_list = build_single_commds_list(KEY["RECALL"], recall_switch_time)
    # 将case_commd_list添加到racall_total_commd中
    recall_total_commd.append(free_tv_and_free_tv_list)
    recall_total_commd.append(free_tv_and_scr_tv_list)
    recall_total_commd.append(free_tv_and_lock_tv_list)
    recall_total_commd.append(scr_tv_and_scr_tv_list)
    recall_total_commd.append(scr_tv_and_lock_tv_list)
    recall_total_commd.append(lock_tv_and_lock_tv_list)
    recall_total_commd.append(free_radio_and_free_radio_list)
    recall_total_commd.append(free_radio_and_scr_radio_list)
    recall_total_commd.append(free_radio_and_lock_radio_list)
    recall_total_commd.append(scr_radio_and_scr_radio_list)
    recall_total_commd.append(scr_radio_and_lock_radio_list)
    recall_total_commd.append(lock_radio_and_lock_radio_list)
    recall_total_commd.append(free_tv_and_free_radio_list)
    recall_total_commd.append(free_tv_and_scr_radio_list)
    recall_total_commd.append(free_tv_and_lock_radio_list)
    recall_total_commd.append(scr_tv_and_free_radio_list)
    recall_total_commd.append(scr_tv_and_scr_radio_list)
    recall_total_commd.append(scr_tv_and_lock_radio_list)
    recall_total_commd.append(lock_tv_and_free_radio_list)
    recall_total_commd.append(lock_tv_and_scr_radio_list)
    recall_total_commd.append(lock_tv_and_lock_radio_list)
    recall_total_commd.append(same_tp_free_tv_and_free_tv_list)
    recall_total_commd.append(diff_tp_free_tv_and_free_tv_list)
    recall_total_commd.append(same_codec_free_tv_and_free_tv_list)
    recall_total_commd.append(diff_codec_free_tv_and_free_tv_list)
    recall_total_commd.append(hd_free_tv_and_hd_free_tv_list)
    recall_total_commd.append(sd_free_tv_and_sd_free_tv_list)
    recall_total_commd.append(hd_free_tv_and_sd_free_tv_list)
    return recall_total_commd

def build_all_scene_commd_list():
    # 数字键切台
    GL.numb_key_switch_commd = build_numb_key_switch_list()
    for i in range(len(GL.numb_key_switch_commd)):
        if i == 0:
            GL.all_test_case.append(
                [ALL_TEST_CASE[i + sum(GL.scene_commd_length)], GET_CH_ATTRIBUTE, GL.numb_key_switch_commd[i],
                 EXIT_TO_SCREEN])
        elif i == 1:
            GL.all_test_case.append(
                [ALL_TEST_CASE[i + sum(GL.scene_commd_length)], GET_CH_ATTRIBUTE, GL.numb_key_switch_commd[i],
                 EXIT_TO_SCREEN])
        else:
            GL.all_test_case.append(
                [ALL_TEST_CASE[i + sum(GL.scene_commd_length)], NOT_PREPARATORY_WORK, GL.numb_key_switch_commd[i],
                 EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.numb_key_switch_commd))
    # 大画面切台
    GL.screen_switch_commd = build_screen_switch_list()
    for j in range(len(GL.screen_switch_commd)):
        GL.all_test_case.append(
            [ALL_TEST_CASE[j + sum(GL.scene_commd_length)], NOT_PREPARATORY_WORK, GL.screen_switch_commd[j],
             EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.screen_switch_commd))
    # 频道列表界面切台
    GL.channel_list_switch_commd = build_channel_list_switch_list()
    for k in range(len(GL.channel_list_switch_commd)):
        GL.all_test_case.append(
            [ALL_TEST_CASE[k + sum(GL.scene_commd_length)], CH_LIST_PREPARATORY_WORK, GL.channel_list_switch_commd[k],
             EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.channel_list_switch_commd))
    # EPG界面切台
    GL.epg_switch_commd = build_epg_switch_list()
    for l in range(len(GL.epg_switch_commd)):
        GL.all_test_case.append(
            [ALL_TEST_CASE[l + sum(GL.scene_commd_length)], EPG_PREPARATORY_WORK, GL.epg_switch_commd[l],
             EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.epg_switch_commd))
    # 节目编辑界面切台
    GL.channel_edit_switch_commd = build_channel_edit_switch_list()
    for m in range(len(GL.channel_edit_switch_commd)):
        GL.all_test_case.append(
            [ALL_TEST_CASE[m + sum(GL.scene_commd_length)], CH_EDIT_PREPARATORY_WORK, GL.channel_edit_switch_commd[m],
             EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.channel_edit_switch_commd))

    # 广电切换
    GL.tv_and_radio_switch_commd = build_tv_and_radio_switch_list()
    for n in range(len(GL.tv_and_radio_switch_commd)):
        GL.all_test_case.append(
            [ALL_TEST_CASE[n + sum(GL.scene_commd_length)], READ_CH_ATTRIBUTE, GL.tv_and_radio_switch_commd[n],
             EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.tv_and_radio_switch_commd))

    # 回看
    GL.recall_switch_commd = build_recall_switch_list()
    for p in range(len(GL.recall_switch_commd)):
        GL.all_test_case.append(
            [ALL_TEST_CASE[p + sum(GL.scene_commd_length)], READ_CH_ATTRIBUTE, GL.recall_switch_commd[p],
             EXIT_TO_SCREEN])
    GL.scene_commd_length.append(len(GL.recall_switch_commd))

def from_file_read_ch_attribute():
    f1 = shelve.open(transfer_tv_data_file_path)
    GL.TV_ch_attribute = f1["tv_ch_attr"]
    GL.free_tv_tp_ch_dict = f1["free_tv_tp_ch_dict"]
    GL.free_tv_codec_ch_dict = f1["free_tv_codec_ch_dict"]
    GL.free_tv_resolution_ch_dict = f1["free_tv_resolution_ch_dict"]
    f1.close()
    f2 = shelve.open(transfer_radio_data_file_path)
    GL.Radio_ch_attribute = f2["radio_ch_attr"]
    f2.close()
    logging.info("免费电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[0]), GL.TV_ch_attribute[0]))
    logging.info("加密电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[1]), GL.TV_ch_attribute[1]))
    logging.info("加锁电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[2]), GL.TV_ch_attribute[2]))
    logging.info("免费广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[0]), GL.Radio_ch_attribute[0]))
    logging.info("加密广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[1]), GL.Radio_ch_attribute[1]))
    logging.info("加锁广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[2]), GL.Radio_ch_attribute[2]))
    logging.info("免费电视按TP归类字典:{}--{}".format(len(GL.free_tv_tp_ch_dict), GL.free_tv_tp_ch_dict))
    logging.info("免费电视按编码归类字典:{}--{}".format(len(GL.free_tv_codec_ch_dict), GL.free_tv_codec_ch_dict))
    logging.info("免费电视按分辨率归类字典:{}--{}".format(len(GL.free_tv_resolution_ch_dict), GL.free_tv_resolution_ch_dict))

def choice_same_attribute_ch1_and_ch2():
    global ch1,ch2
    if GL.read_ch_attr_sub_stage == 0:
        logging.debug("GL.read_ch_attr_sub_stage == 0")
        ch1 = ''
        ch2 = ''
        from_file_read_ch_attribute()
        GL.read_ch_attr_sub_stage += 1
    elif GL.read_ch_attr_sub_stage == 1:
        logging.debug("GL.read_ch_attr_sub_stage == 1")
        if "tv" in ALL_TEST_CASE[choice_switch_case][1]:
            if GL.channel_info[7] == "TV":
                if "free_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.TV_ch_attribute[0]) < 2:
                        logging.info("没有足够的免费电视节目用于回看")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[0]) >= 2:
                        free_tv_numb = sample(GL.TV_ch_attribute[0], 2)
                        logging.debug("所选免费电视节目为:{}".format(free_tv_numb))
                        GL.choice_channel[0] = free_tv_numb[0]
                        GL.choice_channel[1] = free_tv_numb[1]
                        free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                        free_tv_commd_add_ok = commds_add_key_list(free_tv_commd, KEY["OK"])
                        ch1 = free_tv_commd_add_ok[0]
                        ch2 = free_tv_commd_add_ok[1]

                elif "scr_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.TV_ch_attribute[1]) < 2:
                        logging.info("没有足够的加密电视节目用于回看")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[1]) >= 2:
                        scr_tv_numb = sample(GL.TV_ch_attribute[1], 2)
                        logging.debug("所选加密电视节目为:{}".format(scr_tv_numb))
                        GL.choice_channel[0] = scr_tv_numb[0]
                        GL.choice_channel[1] = scr_tv_numb[1]
                        scr_tv_commd = change_numbs_to_commds_list(scr_tv_numb)
                        scr_tv_commd_add_ok = commds_add_key_list(scr_tv_commd, KEY["OK"])
                        ch1 = scr_tv_commd_add_ok[0]
                        ch2 = scr_tv_commd_add_ok[1]

                elif "lock_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.TV_ch_attribute[2]) < 2:
                        logging.info("没有足够的加锁电视节目用于回看")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[2]) >= 2:
                        lock_tv_numb = sample(GL.TV_ch_attribute[2], 2)
                        logging.debug("所选加锁电视节目为:{}".format(lock_tv_numb))
                        GL.choice_channel[0] = lock_tv_numb[0]
                        GL.choice_channel[1] = lock_tv_numb[1]
                        lock_tv_commd = change_numbs_to_commds_list(lock_tv_numb)
                        lock_tv_commd_add_ok = commds_add_key_list(lock_tv_commd, KEY["OK"])
                        ch1 = lock_tv_commd_add_ok[0]
                        ch2 = lock_tv_commd_add_ok[1]

                elif "same_tp_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    prog_greater_than_or_rqual_to_2_tp_list = []  # 节目个数大于等于2的TP列表
                    if len(GL.free_tv_tp_ch_dict) <= 0:
                        logging.info("没有TP来进行异频点免费电视回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_tp_ch_dict) > 0:
                        key_tp_list = list(GL.free_tv_tp_ch_dict.keys())  # 字典中的key值列表(TP信息列表)
                        for i in range(len(GL.free_tv_tp_ch_dict)):  # 筛选免费电视节目大于等于2的TP
                            if len(GL.free_tv_tp_ch_dict[key_tp_list[i]]) >= 2:
                                prog_greater_than_or_rqual_to_2_tp_list.append(key_tp_list[i])
                        if len(prog_greater_than_or_rqual_to_2_tp_list) <= 0:
                            logging.info("没有TP有两个及两个以上的免费电视来进行同频点免费电视回看操作")
                            GL.main_loop_state = False
                        elif len(prog_greater_than_or_rqual_to_2_tp_list) > 0:
                            random_choice_tp = choice(prog_greater_than_or_rqual_to_2_tp_list)  # 随机选择某个TP
                            random_choice_tp_ch = sample(GL.free_tv_tp_ch_dict[random_choice_tp], 2)  # 所选TP下随机选择两个节目
                            logging.debug("所选同TP下的免费电视节目为:{}".format(random_choice_tp_ch))
                            GL.choice_channel[0] = random_choice_tp_ch[0]
                            GL.choice_channel[1] = random_choice_tp_ch[1]
                            random_choice_tp_ch_commd = change_numbs_to_commds_list(random_choice_tp_ch)
                            random_choice_tp_ch_commd_add_ok = commds_add_key_list(random_choice_tp_ch_commd, KEY["OK"])
                            ch1 = random_choice_tp_ch_commd_add_ok[0]
                            ch2 = random_choice_tp_ch_commd_add_ok[1]

                elif "diff_tp_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    random_choice_diff_tp_ch = []
                    if len(GL.free_tv_tp_ch_dict) <= 0:
                        logging.info("没有TP来进行异频点免费电视回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_tp_ch_dict) <= 1:
                        logging.info("没有足够的TP来进行异频点免费电视回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_tp_ch_dict) >= 2:
                        key_tp_list = list(GL.free_tv_tp_ch_dict.keys())  # 字典中的key值列表(TP信息列表)
                        random_choice_diff_tp = sample(key_tp_list, 2)  # 随机选择两个频点
                        random_choice_diff_tp_ch.append(choice(GL.free_tv_tp_ch_dict[random_choice_diff_tp[0]]))
                        random_choice_diff_tp_ch.append(choice(GL.free_tv_tp_ch_dict[random_choice_diff_tp[1]]))
                        logging.debug("异TP下的所选两个免费电视节目为:{}".format(random_choice_diff_tp_ch))
                        GL.choice_channel[0] = random_choice_diff_tp_ch[0]
                        GL.choice_channel[1] = random_choice_diff_tp_ch[1]
                        random_choice_diff_tp_ch_commd = change_numbs_to_commds_list(random_choice_diff_tp_ch)
                        random_choice_diff_tp_ch_commd_add_ok = commds_add_key_list(random_choice_diff_tp_ch_commd, KEY["OK"])
                        ch1 = random_choice_diff_tp_ch_commd_add_ok[0]
                        ch2 = random_choice_diff_tp_ch_commd_add_ok[1]

                elif "same_codec_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    prog_greater_than_or_rqual_to_2_codec_list = []  # 节目个数大于等于2的codec列表
                    if len(GL.free_tv_codec_ch_dict) <= 0:
                        logging.info("没有codec来进行相同编码的免费电视回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_codec_ch_dict) > 0:
                        key_codec_list = list(GL.free_tv_codec_ch_dict.keys())  # 字典中的key值列表(编码信息列表)
                        for i in range(len(GL.free_tv_codec_ch_dict)):  # 筛选免费电视节目大于等于2的编码
                            if len(GL.free_tv_codec_ch_dict[key_codec_list[i]]) >= 2:
                                prog_greater_than_or_rqual_to_2_codec_list.append(key_codec_list[i])
                        if len(prog_greater_than_or_rqual_to_2_codec_list) <= 0:
                            logging.info("没有codec有两个及两个以上的免费电视来进行同编码免费电视回看操作")
                            GL.main_loop_state = False
                        elif len(prog_greater_than_or_rqual_to_2_codec_list) > 0:
                            random_choice_codec = choice(prog_greater_than_or_rqual_to_2_codec_list)  # 随机选择某个编码
                            random_choice_codec_ch = sample(GL.free_tv_codec_ch_dict[random_choice_codec], 2)  # 所选编码下随机选择两个节目
                            logging.debug("所选codec下的两个同编码的免费电视节目为:{}".format(random_choice_codec_ch))
                            GL.choice_channel[0] = random_choice_codec_ch[0]
                            GL.choice_channel[1] = random_choice_codec_ch[1]
                            random_choice_codec_ch_commd = change_numbs_to_commds_list(random_choice_codec_ch)
                            random_choice_codec_ch_commd_add_ok = commds_add_key_list(random_choice_codec_ch_commd, KEY["OK"])
                            ch1 = random_choice_codec_ch_commd_add_ok[0]
                            ch2 = random_choice_codec_ch_commd_add_ok[1]

                elif "diff_codec_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    random_choice_diff_codec_ch = []
                    if len(GL.free_tv_codec_ch_dict) <= 0:
                        logging.info("没有codec来进行不同编码的免费电视回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_codec_ch_dict) <= 1:
                        logging.info("没有足够的codec来进行不同编码的免费电视回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_codec_ch_dict) >= 2:
                        key_codec_list = list(GL.free_tv_codec_ch_dict.keys())  # 字典中的key值列表(TP信息列表)
                        random_choice_diff_codec = sample(key_codec_list, 2)  # 随机选择两个频点
                        random_choice_diff_codec_ch.append(choice(GL.free_tv_codec_ch_dict[random_choice_diff_codec[0]]))
                        random_choice_diff_codec_ch.append(choice(GL.free_tv_codec_ch_dict[random_choice_diff_codec[1]]))
                        logging.debug("不同编码下的所选两个免费电视节目为:{}".format(random_choice_diff_codec_ch))
                        GL.choice_channel[0] = random_choice_diff_codec_ch[0]
                        GL.choice_channel[1] = random_choice_diff_codec_ch[1]
                        random_choice_diff_codec_ch_commd = change_numbs_to_commds_list(random_choice_diff_codec_ch)
                        random_choice_diff_codec_ch_commd_add_ok = commds_add_key_list(random_choice_diff_codec_ch_commd,
                                                                                       KEY["OK"])
                        ch1 = random_choice_diff_codec_ch_commd_add_ok[0]
                        ch2 = random_choice_diff_codec_ch_commd_add_ok[1]

                elif "hd_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    prog_greater_than_or_rqual_to_2_resolution_list = []  # 节目个数大于等于2的resolution列表
                    if len(GL.free_tv_resolution_ch_dict) <= 0:
                        logging.info("没有分辨率来进行高清免费电视之间回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_resolution_ch_dict) > 0:
                        key_resolution_list = list(GL.free_tv_resolution_ch_dict.keys())  # 字典中的key值列表(分辨率信息列表)
                        for i in range(len(GL.free_tv_resolution_ch_dict)):  # 筛选免费电视节目个数大于等于2的分辨率
                            if len(GL.free_tv_resolution_ch_dict[key_resolution_list[i]]) >= 2:
                                prog_greater_than_or_rqual_to_2_resolution_list.append(key_resolution_list[i])
                        if len(prog_greater_than_or_rqual_to_2_resolution_list) <= 0:
                            logging.info("没有分辨率有两个及两个以上的免费电视来进行HD免费电视之间回看操作")
                            GL.main_loop_state = False
                        elif len(prog_greater_than_or_rqual_to_2_resolution_list) > 0:
                            hd_res = []
                            for j in range(len(prog_greater_than_or_rqual_to_2_resolution_list)):
                                res_height = re.split(r"x", prog_greater_than_or_rqual_to_2_resolution_list[j])[0]
                                res_width = re.split(r"x", prog_greater_than_or_rqual_to_2_resolution_list[j])[1]
                                if int(res_height) > 576 and int(res_width) > 720:
                                    hd_res.append(prog_greater_than_or_rqual_to_2_resolution_list[j])
                            if len(hd_res) <= 0:
                                logging.info("没有两个及两个以上的高清节目来进行回看操作")
                                GL.main_loop_state = False
                            elif len(hd_res) > 0:
                                random_choice_hd_res = choice(hd_res)  # 随机选择某个hd_res
                                random_choice_hd_ch = sample(GL.free_tv_resolution_ch_dict[random_choice_hd_res],
                                                             2)  # 所选res下随机选择两个节目
                                logging.debug("所选HD分辨率下的两个高清免费电视节目为:{}".format(random_choice_hd_ch))
                                GL.choice_channel[0] = random_choice_hd_ch[0]
                                GL.choice_channel[1] = random_choice_hd_ch[1]
                                random_choice_hd_ch_commd = change_numbs_to_commds_list(random_choice_hd_ch)
                                random_choice_hd_ch_commd_add_ok = commds_add_key_list(random_choice_hd_ch_commd, KEY["OK"])
                                ch1 = random_choice_hd_ch_commd_add_ok[0]
                                ch2 = random_choice_hd_ch_commd_add_ok[1]

                elif "sd_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    prog_greater_than_or_rqual_to_2_resolution_list = []  # 节目个数大于等于2的resolution列表
                    if len(GL.free_tv_resolution_ch_dict) <= 0:
                        logging.info("没有分辨率来进行标清免费电视之间回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_resolution_ch_dict) > 0:
                        key_resolution_list = list(GL.free_tv_resolution_ch_dict.keys())  # 字典中的key值列表(分辨率信息列表)
                        for i in range(len(GL.free_tv_resolution_ch_dict)):  # 筛选免费电视节目个数大于等于2的分辨率
                            if len(GL.free_tv_resolution_ch_dict[key_resolution_list[i]]) >= 2:
                                prog_greater_than_or_rqual_to_2_resolution_list.append(key_resolution_list[i])
                        if len(prog_greater_than_or_rqual_to_2_resolution_list) <= 0:
                            logging.info("没有分辨率有两个及两个以上的免费电视来进行标清免费电视之间回看操作")
                            GL.main_loop_state = False
                        elif len(prog_greater_than_or_rqual_to_2_resolution_list) > 0:
                            sd_res = []
                            for j in range(len(prog_greater_than_or_rqual_to_2_resolution_list)):
                                res_height = re.split(r"x", prog_greater_than_or_rqual_to_2_resolution_list[j])[0]
                                res_width = re.split(r"x", prog_greater_than_or_rqual_to_2_resolution_list[j])[1]
                                if int(res_height) <= 576 and int(res_width) <= 720:
                                    sd_res.append(prog_greater_than_or_rqual_to_2_resolution_list[j])
                            if len(sd_res) <= 0:
                                logging.info("没有两个及两个以上的标清节目来进行回看操作")
                                GL.main_loop_state = False
                            elif len(sd_res) > 0:
                                random_choice_sd_res = choice(sd_res)  # 随机选择某个sd_res
                                random_choice_sd_ch = sample(GL.free_tv_resolution_ch_dict[random_choice_sd_res],
                                                             2)  # 所选res下随机选择两个节目
                                logging.debug("所选SD分辨率下的两个标清免费电视节目为:{}".format(random_choice_sd_ch))
                                GL.choice_channel[0] = random_choice_sd_ch[0]
                                GL.choice_channel[1] = random_choice_sd_ch[1]
                                random_choice_sd_ch_commd = change_numbs_to_commds_list(random_choice_sd_ch)
                                random_choice_sd_ch_commd_add_ok = commds_add_key_list(random_choice_sd_ch_commd, KEY["OK"])
                                ch1 = random_choice_sd_ch_commd_add_ok[0]
                                ch2 = random_choice_sd_ch_commd_add_ok[1]

                elif "hd_sd_tv" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    hd_res = []
                    sd_res = []
                    if len(GL.free_tv_resolution_ch_dict) <= 0:
                        logging.info("没有分辨率来进行高清和标清的免费电视之间回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_resolution_ch_dict) <= 1:
                        logging.info("没有足够多的分辨率来进行高清和标清的免费电视之间回看操作")
                        GL.main_loop_state = False
                    elif len(GL.free_tv_resolution_ch_dict) >= 2:
                        key_resolution_list = list(GL.free_tv_resolution_ch_dict.keys())  # 字典中的key值列表(分辨率信息列表)
                        for j in range(len(key_resolution_list)):
                            res_height = re.split(r"x", key_resolution_list[j])[0]
                            res_width = re.split(r"x", key_resolution_list[j])[1]
                            if int(res_height) <= 576 and int(res_width) <= 720:
                                sd_res.append(key_resolution_list[j])
                            elif int(res_height) > 576 and int(res_width) > 720:
                                hd_res.append(key_resolution_list[j])
                        if len(sd_res) <= 0 or len(hd_res) <= 0:
                            logging.info("没有足够多的HD或SD分辨率来进行高清和标清的免费电视之间回看操作")
                            GL.main_loop_state = False
                        elif len(sd_res) > 0 and len(hd_res) > 0:
                            random_choice_sd_res = choice(sd_res)  # 随机选择某个sd_res
                            random_choice_sd_ch = sample(GL.free_tv_resolution_ch_dict[random_choice_sd_res],
                                                         1)  # 所选res下随机选择sd节目
                            random_choice_hd_res = choice(hd_res)  # 随机选择某个hd_res
                            random_choice_hd_ch = sample(GL.free_tv_resolution_ch_dict[random_choice_hd_res],
                                                         1)  # 所选res下随机选择hd节目
                            logging.debug("所选HD分辨率下的高清免费电视节目为:{}".format(random_choice_hd_ch))
                            logging.debug("所选SD分辨率下的标清免费电视节目为:{}".format(random_choice_sd_ch))
                            GL.choice_channel[0] = random_choice_hd_ch[0]
                            GL.choice_channel[1] = random_choice_sd_ch[0]
                            random_choice_hd_ch_commd = change_numbs_to_commds_list(random_choice_hd_ch)
                            random_choice_hd_ch_commd_add_ok = commds_add_key_list(random_choice_hd_ch_commd, KEY["OK"])
                            random_choice_sd_ch_commd = change_numbs_to_commds_list(random_choice_sd_ch)
                            random_choice_sd_ch_commd_add_ok = commds_add_key_list(random_choice_sd_ch_commd, KEY["OK"])
                            ch1 = random_choice_hd_ch_commd_add_ok[0]
                            ch2 = random_choice_sd_ch_commd_add_ok[0]

                time.sleep(2)
                send_commd(KEY["EXIT"])
                GL.read_ch_attr_sub_stage += 1
            elif GL.channel_info[7] != "TV":
                send_commd(KEY["TV/R"])
                if GL.channel_info[5] == "1":
                    unlock_channel()
        elif "radio" in ALL_TEST_CASE[choice_switch_case][1]:
            if GL.channel_info[7] == "Radio":
                if "free_radio" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.Radio_ch_attribute[0]) < 2:
                        logging.info("没有足够的免费广播节目用于回看")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[0]) >= 2:
                        free_radio_numb = sample(GL.Radio_ch_attribute[0], 2)
                        logging.debug("所选免费广播节目为:{}".format(free_radio_numb))
                        GL.choice_channel[0] = free_radio_numb[0]
                        GL.choice_channel[1] = free_radio_numb[1]
                        free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                        free_radio_commd_add_ok = commds_add_key_list(free_radio_commd, KEY["OK"])
                        ch1 = free_radio_commd_add_ok[0]
                        ch2 = free_radio_commd_add_ok[1]

                elif "scr_radio" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.Radio_ch_attribute[1]) < 2:
                        logging.info("没有足够的加密广播节目用于回看")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[1]) >= 2:
                        scr_radio_numb = sample(GL.Radio_ch_attribute[1], 2)
                        logging.debug("所选加密广播节目为:{}".format(scr_radio_numb))
                        GL.choice_channel[0] = scr_radio_numb[0]
                        GL.choice_channel[1] = scr_radio_numb[1]
                        scr_radio_commd = change_numbs_to_commds_list(scr_radio_numb)
                        scr_radio_commd_add_ok = commds_add_key_list(scr_radio_commd, KEY["OK"])
                        ch1 = scr_radio_commd_add_ok[0]
                        ch2 = scr_radio_commd_add_ok[1]

                elif "lock_radio" == ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.Radio_ch_attribute[2]) < 2:
                        logging.info("没有足够的加锁广播节目用于回看")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[2]) >= 2:
                        lock_radio_numb = sample(GL.Radio_ch_attribute[2], 2)
                        logging.debug("所选加锁广播节目为:{}".format(lock_radio_numb))
                        GL.choice_channel[0] = lock_radio_numb[0]
                        GL.choice_channel[1] = lock_radio_numb[1]
                        lock_radio_commd = change_numbs_to_commds_list(lock_radio_numb)
                        lock_radio_commd_add_ok = commds_add_key_list(lock_radio_commd, KEY["OK"])
                        ch1 = lock_radio_commd_add_ok[0]
                        ch2 = lock_radio_commd_add_ok[1]

                time.sleep(2)
                send_commd(KEY["EXIT"])
                GL.read_ch_attr_sub_stage += 1
            elif GL.channel_info[7] != "Radio":
                send_commd(KEY["TV/R"])
                if GL.channel_info[5] == "1":
                    unlock_channel()
    elif GL.read_ch_attr_sub_stage == 2:
        time.sleep(2)
        logging.debug("GL.read_ch_attr_sub_stage == 2")
        for i in range(len(ch1)):
            send_numb_key_commd(ch1[i])
        send_commd(KEY["EXIT"])
        GL.read_ch_attr_sub_stage += 1
    elif GL.read_ch_attr_sub_stage == 3:
        time.sleep(2)
        logging.debug("GL.read_ch_attr_sub_stage == 3")
        for i in range(len(ch2)):
            send_numb_key_commd(ch2[i])
        send_commd(KEY["EXIT"])
        GL.read_ch_attr_sub_stage += 1
    elif GL.read_ch_attr_sub_stage == 4:
        time.sleep(2)
        logging.debug("GL.read_ch_attr_sub_stage == 4")
        send_commd(KEY["RECALL"])
        GL.read_ch_attr_sub_stage += 1
    elif GL.read_ch_attr_sub_stage == 5:
        logging.debug("GL.read_ch_attr_sub_stage == 5")
        GL.get_group_channel_total_info_state = False
        GL.sub_stage == 0

def choice_diff_attribute_ch1_and_ch2():
    global ch1, ch2
    if GL.read_ch_attr_sub_stage == 0:
        logging.debug("GL.read_ch_attr_sub_stage == 0")
        ch1 = ''
        ch2 = ''
        from_file_read_ch_attribute()
        GL.read_ch_attr_sub_stage += 1
    elif GL.read_ch_attr_sub_stage == 1:
        logging.debug("GL.read_ch_attr_sub_stage == 1")
        if "tv" in ALL_TEST_CASE[choice_switch_case][1]:
            if GL.channel_info[7] == "TV":
                if "free" in ALL_TEST_CASE[choice_switch_case][1]:
                    if len(GL.TV_ch_attribute[0]) == 0:
                        logging.info("没有免费电视节目")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[0]) > 0:
                        free_tv_numb = sample(GL.TV_ch_attribute[0], 1)
                        logging.debug("所选免费电视节目为:{}".format(free_tv_numb))
                        GL.choice_channel[0] = free_tv_numb
                        free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                        for i in range(len(free_tv_commd)):
                            for j in range(len(free_tv_commd[i])):
                                send_numb_key_commd(free_tv_commd[i][j])
                        send_commd(KEY["OK"])

                elif "scr" in ALL_TEST_CASE[choice_switch_case][1]:
                    if len(GL.TV_ch_attribute[1]) == 0:
                        logging.info("没有加密电视节目")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[1]) > 0:
                        scr_tv_numb = sample(GL.TV_ch_attribute[1], 1)
                        logging.debug("所选加密电视节目为:{}".format(scr_tv_numb))
                        GL.choice_channel[0] = scr_tv_numb
                        scr_tv_commd = change_numbs_to_commds_list(scr_tv_numb)
                        for i in range(len(scr_tv_commd)):
                            for j in range(len(scr_tv_commd[i])):
                                send_numb_key_commd(scr_tv_commd[i][j])
                        send_commd(KEY["OK"])

                elif "lock" in ALL_TEST_CASE[choice_switch_case][1]:
                    if len(GL.TV_ch_attribute[2]) == 0:
                        logging.info("没有加锁电视节目")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[2]) > 0:
                        lock_tv_numb = sample(GL.TV_ch_attribute[2], 1)
                        logging.debug("所选加锁电视节目为:{}".format(lock_tv_numb))
                        GL.choice_channel[0] = lock_tv_numb
                        lock_tv_commd = change_numbs_to_commds_list(lock_tv_numb)
                        for i in range(len(lock_tv_commd)):
                            for j in range(len(lock_tv_commd[i])):
                                send_numb_key_commd(lock_tv_commd[i][j])
                        send_commd(KEY["OK"])

                time.sleep(2)
                send_commd(KEY["EXIT"])
                GL.read_ch_attr_sub_stage += 1
            elif GL.channel_info[7] != "TV":
                send_commd(KEY["TV/R"])
                if GL.channel_info[5] == "1":
                    unlock_channel()
        elif "radio" in ALL_TEST_CASE[choice_switch_case][1]:
            if GL.channel_info[7] == "Radio":
                if "free" in ALL_TEST_CASE[choice_switch_case][1]:
                    if len(GL.Radio_ch_attribute[0]) == 0:
                        logging.info("没有免费广播节目")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[0]) > 0:
                        free_radio_numb = sample(GL.Radio_ch_attribute[0], 1)
                        logging.debug("所选免费广播节目为:{}".format(free_radio_numb))
                        GL.choice_channel[0] = free_radio_numb
                        free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                        for i in range(len(free_radio_commd)):
                            for j in range(len(free_radio_commd[i])):
                                send_numb_key_commd(free_radio_commd[i][j])
                        send_commd(KEY["OK"])

                elif "scr" in ALL_TEST_CASE[choice_switch_case][1]:
                    if len(GL.Radio_ch_attribute[1]) == 0:
                        logging.info("没有加密广播节目")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[1]) > 0:
                        scr_radio_numb = sample(GL.Radio_ch_attribute[1], 1)
                        logging.debug("所选加密广播节目为:{}".format(scr_radio_numb))
                        GL.choice_channel[0] = scr_radio_numb
                        scr_radio_commd = change_numbs_to_commds_list(scr_radio_numb)
                        for i in range(len(scr_radio_commd)):
                            for j in range(len(scr_radio_commd[i])):
                                send_numb_key_commd(scr_radio_commd[i][j])
                        send_commd(KEY["OK"])

                elif "lock" in ALL_TEST_CASE[choice_switch_case][1]:
                    if len(GL.Radio_ch_attribute[2]) == 0:
                        logging.info("没有加锁广播节目")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[2]) > 0:
                        lock_radio_numb = sample(GL.Radio_ch_attribute[2], 1)
                        logging.debug("所选加锁广播节目为:{}".format(lock_radio_numb))
                        GL.choice_channel[0] = lock_radio_numb
                        lock_radio_commd = change_numbs_to_commds_list(lock_radio_numb)
                        for i in range(len(lock_radio_commd)):
                            for j in range(len(lock_radio_commd[i])):
                                send_numb_key_commd(lock_radio_commd[i][j])
                        send_commd(KEY["OK"])
                time.sleep(2)
                send_commd(KEY["EXIT"])
                GL.read_ch_attr_sub_stage += 1
            elif GL.channel_info[7] != "Radio":
                send_commd(KEY["TV/R"])
                if GL.channel_info[5] == "1":
                    unlock_channel()
    elif GL.read_ch_attr_sub_stage == 2:
        logging.debug("GL.read_ch_attr_sub_stage == 2")
        if "tv" in ALL_TEST_CASE[choice_switch_case][2]:
            if GL.channel_info[7] == "TV":
                if "free" in ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.TV_ch_attribute[0]) == 0:
                        logging.info("没有免费电视节目")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[0]) > 0:
                        free_tv_numb = sample(GL.TV_ch_attribute[0], 1)
                        logging.debug("所选免费电视节目为:{}".format(free_tv_numb))
                        GL.choice_channel[1] = free_tv_numb
                        free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                        for i in range(len(free_tv_commd)):
                            for j in range(len(free_tv_commd[i])):
                                send_numb_key_commd(free_tv_commd[i][j])
                        send_commd(KEY["OK"])

                elif "scr" in ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.TV_ch_attribute[1]) == 0:
                        logging.info("没有加密电视节目")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[1]) > 0:
                        scr_tv_numb = sample(GL.TV_ch_attribute[1], 1)
                        logging.debug("所选加密电视节目为:{}".format(scr_tv_numb))
                        GL.choice_channel[1] = scr_tv_numb
                        scr_tv_commd = change_numbs_to_commds_list(scr_tv_numb)
                        for i in range(len(scr_tv_commd)):
                            for j in range(len(scr_tv_commd[i])):
                                send_numb_key_commd(scr_tv_commd[i][j])
                        send_commd(KEY["OK"])

                elif "lock" in ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.TV_ch_attribute[2]) == 0:
                        logging.info("没有加锁电视节目")
                        GL.main_loop_state = False
                    elif len(GL.TV_ch_attribute[2]) > 0:
                        lock_tv_numb = sample(GL.TV_ch_attribute[2], 1)
                        logging.debug("所选加锁电视节目为:{}".format(lock_tv_numb))
                        GL.choice_channel[1] = lock_tv_numb
                        lock_tv_commd = change_numbs_to_commds_list(lock_tv_numb)
                        for i in range(len(lock_tv_commd)):
                            for j in range(len(lock_tv_commd[i])):
                                send_numb_key_commd(lock_tv_commd[i][j])
                        send_commd(KEY["OK"])
                time.sleep(2)
                GL.read_ch_attr_sub_stage += 1
            elif GL.channel_info[7] != "TV":
                send_commd(KEY["TV/R"])
                if GL.channel_info[5] == "1":
                    unlock_channel()
        elif "radio" in ALL_TEST_CASE[choice_switch_case][2]:
            if GL.channel_info[7] == "Radio":
                if "free" in ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.Radio_ch_attribute[0]) == 0:
                        logging.info("没有免费广播节目")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[0]) > 0:
                        free_radio_numb = sample(GL.Radio_ch_attribute[0], 1)
                        logging.debug("所选免费广播节目为:{}".format(free_radio_numb))
                        GL.choice_channel[1] = free_radio_numb
                        free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                        for i in range(len(free_radio_commd)):
                            for j in range(len(free_radio_commd[i])):
                                send_numb_key_commd(free_radio_commd[i][j])
                        send_commd(KEY["OK"])

                elif "scr" in ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.Radio_ch_attribute[1]) == 0:
                        logging.info("没有加密广播节目")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[1]) > 0:
                        scr_radio_numb = sample(GL.Radio_ch_attribute[1], 1)
                        logging.debug("所选加密广播节目为:{}".format(scr_radio_numb))
                        GL.choice_channel[1] = scr_radio_numb
                        scr_radio_commd = change_numbs_to_commds_list(scr_radio_numb)
                        for i in range(len(scr_radio_commd)):
                            for j in range(len(scr_radio_commd[i])):
                                send_numb_key_commd(scr_radio_commd[i][j])
                        send_commd(KEY["OK"])

                elif "lock" in ALL_TEST_CASE[choice_switch_case][2]:
                    if len(GL.Radio_ch_attribute[2]) == 0:
                        logging.info("没有加锁广播节目")
                        GL.main_loop_state = False
                    elif len(GL.Radio_ch_attribute[2]) > 0:
                        lock_radio_numb = sample(GL.Radio_ch_attribute[2], 1)
                        logging.debug("所选加锁广播节目为:{}".format(lock_radio_numb))
                        GL.choice_channel[1] = lock_radio_numb
                        lock_radio_commd = change_numbs_to_commds_list(lock_radio_numb)
                        for i in range(len(lock_radio_commd)):
                            for j in range(len(lock_radio_commd[i])):
                                send_numb_key_commd(lock_radio_commd[i][j])
                        send_commd(KEY["OK"])
                time.sleep(2)
                send_commd(KEY["EXIT"])
                GL.read_ch_attr_sub_stage += 1
            elif GL.channel_info[7] != "Radio":
                send_commd(KEY["TV/R"])
                if GL.channel_info[5] == "1":
                    unlock_channel()
    elif GL.read_ch_attr_sub_stage == 3:
        logging.debug("GL.read_ch_attr_sub_stage == 3")
        if ALL_TEST_CASE[choice_switch_case][0] == "tv_radio":
            send_commd(KEY["TV/R"])
            send_commd(KEY["EXIT"])
        elif ALL_TEST_CASE[choice_switch_case][0] == "recall":
            if "tv" in ALL_TEST_CASE[choice_switch_case][1] and "radio" in ALL_TEST_CASE[choice_switch_case][2]:
                send_commd(KEY["TV/R"])
                send_commd(KEY["EXIT"])
            else:
                send_commd(KEY["RECALL"])
                send_commd(KEY["EXIT"])
        GL.read_ch_attr_sub_stage += 1
    elif GL.read_ch_attr_sub_stage == 4:
        logging.debug("GL.read_ch_attr_sub_stage == 4")
        GL.get_group_channel_total_info_state = False
        GL.sub_stage == 0

def get_group_channel_total_info():
    # 搜索前获取case节目类别,分组,分组节目数量,以及获取节目属性前的去除加锁的判断
    if GL.sub_stage == 0:  # 切台用于判断当前节目类别属性(TV/Radio)
        # 根据所选case切换到对应类型节目的界面
        logging.debug("GL.sub_stage == 0")
        while GL.channel_info[7] != ALL_TEST_CASE[choice_switch_case][4]:
            send_commd(KEY["TV/R"])
            if GL.channel_info[5] == "1":
                send_commd(KEY["EXIT"])
        GL.sub_stage += 1
    elif GL.sub_stage == 1:  # 调出频道列表,用于判断组别信息
        logging.debug("GL.sub_stage == 1")
        send_commd(KEY["OK"])
        GL.sub_stage += 1
    elif GL.sub_stage == 2: # 采集所有分组的名称和分组下节目总数信息
        logging.debug("GL.sub_stage == 2")
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            while GL.prog_group_name not in GL.TV_channel_groups.keys():
                print(GL.prog_group_name)
                GL.TV_channel_groups[GL.prog_group_name] = GL.prog_group_total
                send_commd(KEY["RIGHT"])
                if GL.channel_info[5] == "1":
                    send_commd(KEY["EXIT"])
            GL.sub_stage += 1
        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            while GL.prog_group_name not in GL.Radio_channel_groups.keys():
                GL.Radio_channel_groups[GL.prog_group_name] = GL.prog_group_total
                send_commd(KEY["RIGHT"])
                if GL.channel_info[5] == "1":
                    send_commd(KEY["EXIT"])
            GL.sub_stage += 1
    elif GL.sub_stage == 3:  # 根据所选case切换到对应的分组
        logging.debug("GL.sub_stage == 3")
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            while GL.prog_group_name != ALL_TEST_CASE[choice_switch_case][3]:
                send_commd(KEY["RIGHT"])
                if GL.channel_info[5] == "1":
                    send_commd(KEY["EXIT"])
            #     if GL.prog_group_name not in GL.TV_channel_groups.keys():
            #         GL.TV_channel_groups[GL.prog_group_name] = GL.prog_group_total
            # if GL.prog_group_name not in GL.TV_channel_groups.keys():
            #     GL.TV_channel_groups[GL.prog_group_name] = GL.prog_group_total
            GL.sub_stage += 1
        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            while GL.prog_group_name != ALL_TEST_CASE[choice_switch_case][3]:
                send_commd(KEY["RIGHT"])
                if GL.channel_info[5] == "1":
                    send_commd(KEY["EXIT"])
            #     if GL.prog_group_name not in GL.Radio_channel_groups.keys():
            #         GL.Radio_channel_groups[GL.prog_group_name] = GL.prog_group_total
            # if GL.prog_group_name not in GL.Radio_channel_groups.keys():
            #     GL.Radio_channel_groups[GL.prog_group_name] = GL.prog_group_total
            GL.sub_stage += 1
    elif GL.sub_stage == 4:  # 退出频道列表,回到大画面界面
        logging.debug("GL.sub_stage == 3")
        send_commd(KEY["EXIT"])
        logging.debug(GL.channel_info)
        GL.sub_stage += 1
    elif GL.sub_stage == 5:  # 创建所有场景的测试用例,测试数据,文件log保存目录
        logging.debug("GL.sub_stage == 4")
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            logging.info("电视节目分组信息:{}".format(GL.TV_channel_groups.keys()))
            logging.debug("电视分组及其分组节目个数:{}".format(GL.TV_channel_groups))
        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            logging.info("广播节目分组信息:{}".format(GL.Radio_channel_groups.keys()))
            logging.debug("广播分组及其分组节目个数:{}".format(GL.Radio_channel_groups))
        build_all_scene_commd_list()
        logging.info(GL.all_test_case[choice_switch_case][2])
        GL.sub_stage += 1
    elif GL.sub_stage == 6:
        logging.debug("GL.sub_stage == 5")
        if GL.all_test_case[choice_switch_case][1] == GET_CH_ATTRIBUTE:
            logging.debug("对所有节目进行加锁")
            for i in range(len(LOCK_ALL_CH)):
                send_commd(LOCK_ALL_CH[i])
            send_commd(KEY["TV/R"])
            send_commd(KEY["TV/R"])
            send_commd(KEY["EXIT"])
            logging.debug("对所有节目进行解锁")
            for j in range(len(UNLOCK_ALL_CH)):
                send_commd(UNLOCK_ALL_CH[j])
            GL.get_group_channel_total_info_state = False
            GL.sub_stage == 0

        if GL.all_test_case[choice_switch_case][1] == READ_CH_ATTRIBUTE:
            if ALL_TEST_CASE[choice_switch_case][1] == ALL_TEST_CASE[choice_switch_case][2]:
                choice_same_attribute_ch1_and_ch2()
            elif ALL_TEST_CASE[choice_switch_case][1] != ALL_TEST_CASE[choice_switch_case][2]:
                choice_diff_attribute_ch1_and_ch2()
        else:
            GL.get_group_channel_total_info_state = False
            GL.sub_stage += 1

def get_ch_attribute():
    global prepare_lock_free_tv_ch,prepare_lock_scramble_tv_ch,prepare_lock_free_radio_ch,prepare_lock_scramble_radio_ch
    if GL.get_ch_attr_sub_stage == 0:
        prepare_lock_free_tv_ch = []
        prepare_lock_scramble_tv_ch = []
        prepare_lock_free_radio_ch = []
        prepare_lock_scramble_radio_ch = []
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            logging.info("免费电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[0]), GL.TV_ch_attribute[0]))
            logging.info("加密电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[1]), GL.TV_ch_attribute[1]))
            logging.info("加锁电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[2]), GL.TV_ch_attribute[2]))
            logging.info("免费电视按TP归类字典:{}--{}".format(len(GL.free_tv_tp_ch_dict), GL.free_tv_tp_ch_dict))
            logging.info("免费电视按编码归类字典:{}--{}".format(len(GL.free_tv_codec_ch_dict), GL.free_tv_codec_ch_dict))
            logging.info("免费电视按分辨率归类字典:{}--{}".format(len(GL.free_tv_resolution_ch_dict), GL.free_tv_resolution_ch_dict))
            if len(GL.TV_ch_attribute[0]) > 20:
                prepare_lock_free_tv_ch = sample(GL.TV_ch_attribute[0], len(GL.TV_ch_attribute[0]) * 5 // 100)
            elif len(GL.TV_ch_attribute[0]) <= 20:
                prepare_lock_free_tv_ch = sample(GL.TV_ch_attribute[0], len(GL.TV_ch_attribute[0]) // 2)
            if len(GL.TV_ch_attribute[1]) > 20:
                prepare_lock_scramble_tv_ch = sample(GL.TV_ch_attribute[1], len(GL.TV_ch_attribute[1]) * 5 // 100)
            elif len(GL.TV_ch_attribute[1]) <= 20:
                prepare_lock_scramble_tv_ch = sample(GL.TV_ch_attribute[1], len(GL.TV_ch_attribute[1]) // 2)
            GL.get_ch_attr_sub_stage += 1
        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            logging.info("免费广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[0]), GL.Radio_ch_attribute[0]))
            logging.info("加密广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[1]), GL.Radio_ch_attribute[1]))
            logging.info("加锁广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[2]), GL.Radio_ch_attribute[2]))
            if len(GL.Radio_ch_attribute[0]) > 20:
                prepare_lock_free_radio_ch = sample(GL.Radio_ch_attribute[0], len(GL.Radio_ch_attribute[0]) * 5 // 100)
            elif len(GL.Radio_ch_attribute[0]) <= 20:
                prepare_lock_free_radio_ch = sample(GL.Radio_ch_attribute[0], len(GL.Radio_ch_attribute[0]) // 2)
            if len(GL.Radio_ch_attribute[1]) > 20:
                prepare_lock_scramble_radio_ch = sample(GL.Radio_ch_attribute[1], len(GL.Radio_ch_attribute[1]) * 5 // 100)
            elif len(GL.Radio_ch_attribute[1]) <= 20:
                prepare_lock_scramble_radio_ch = sample(GL.Radio_ch_attribute[1], len(GL.Radio_ch_attribute[1]) // 2)
            GL.get_ch_attr_sub_stage += 1

    elif GL.get_ch_attr_sub_stage == 1:
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            if len(prepare_lock_free_tv_ch) > 0:
                tv_tp_ch_key_list = list(GL.free_tv_tp_ch_dict.keys())
                tv_codec_ch_key_list = list(GL.free_tv_codec_ch_dict.keys())
                tv_resolution_ch_key_list = list(GL.free_tv_resolution_ch_dict.keys())
                prepare_lock_free_tv_ch_commd = change_numbs_to_commds_list(prepare_lock_free_tv_ch)
                prepare_lock_free_tv_ch_commd_add_ok = commds_add_key_list(prepare_lock_free_tv_ch_commd, KEY["OK"])
                for i in range(len(prepare_lock_free_tv_ch_commd_add_ok)):  # 对准备加锁的免费电视节目进行加锁
                    for j in range(len(prepare_lock_free_tv_ch_commd_add_ok[i])):
                        send_numb_key_commd(prepare_lock_free_tv_ch_commd_add_ok[i][j])
                        if j == (len(prepare_lock_free_tv_ch_commd_add_ok[i]) - 1):
                            for k in range(len(LOCK_CH)):
                                send_commd(LOCK_CH[k])
                                if k == (len(LOCK_CH) - 1):
                                    send_commd(KEY["EXIT"])
                    GL.TV_ch_attribute[0].remove(prepare_lock_free_tv_ch[i])  # 对加锁的免费电视节目在免费电视列表中去除
                    GL.TV_ch_attribute[2].append(prepare_lock_free_tv_ch[i])  # 对加锁的免费电视节目加入到加锁电视节目列表
                    time.sleep(2)
                # 将加锁的免费电视节目在tp,codec,resolution类别的字典中去除
                for m in range(len(prepare_lock_free_tv_ch)):
                    for h in range(len(tv_tp_ch_key_list)):
                        if prepare_lock_free_tv_ch[m] in GL.free_tv_tp_ch_dict[tv_tp_ch_key_list[h]]:
                            GL.free_tv_tp_ch_dict[tv_tp_ch_key_list[h]].remove(prepare_lock_free_tv_ch[m])
                    for k in range(len(tv_codec_ch_key_list)):
                        if prepare_lock_free_tv_ch[m] in GL.free_tv_codec_ch_dict[tv_codec_ch_key_list[k]]:
                            GL.free_tv_codec_ch_dict[tv_codec_ch_key_list[k]].remove(prepare_lock_free_tv_ch[m])
                    for l in range(len(tv_resolution_ch_key_list)):
                        if prepare_lock_free_tv_ch[m] in GL.free_tv_resolution_ch_dict[tv_resolution_ch_key_list[l]]:
                            GL.free_tv_resolution_ch_dict[tv_resolution_ch_key_list[l]].remove(prepare_lock_free_tv_ch[m])
                # 将tp,codec,resolution类别的字典中去除加锁免费后为空的元素删除
                for h in range(len(tv_tp_ch_key_list)):
                    if len(GL.free_tv_tp_ch_dict[tv_tp_ch_key_list[h]]) == 0:
                        del GL.free_tv_tp_ch_dict[tv_tp_ch_key_list[h]]
                for k in range(len(tv_codec_ch_key_list)):
                    if len(GL.free_tv_codec_ch_dict[tv_codec_ch_key_list[k]]) == 0:
                        del GL.free_tv_codec_ch_dict[tv_codec_ch_key_list[k]]
                logging.debug(GL.free_tv_resolution_ch_dict)
                for l in range(len(tv_resolution_ch_key_list)):
                    if len(GL.free_tv_resolution_ch_dict[tv_resolution_ch_key_list[l]]) == 0:
                        del GL.free_tv_resolution_ch_dict[tv_resolution_ch_key_list[l]]
                    if tv_resolution_ch_key_list[l] == "x":  # 去除因为信号不好或者没有检测到分辨率信息的节目
                        del GL.free_tv_resolution_ch_dict[tv_resolution_ch_key_list[l]]
                logging.debug(GL.free_tv_resolution_ch_dict)

                GL.get_ch_attr_sub_stage += 1
            elif len(prepare_lock_free_tv_ch) <= 0:
                logging.info("没有免费电视节目能被加锁")
                GL.get_ch_attr_sub_stage += 1
        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            if len(prepare_lock_free_radio_ch) > 0:
                prepare_lock_free_radio_ch_commd = change_numbs_to_commds_list(prepare_lock_free_radio_ch)
                prepare_lock_free_radio_ch_commd_add_ok = commds_add_key_list(prepare_lock_free_radio_ch_commd, KEY["OK"])
                for i in range(len(prepare_lock_free_radio_ch_commd_add_ok)):  # 对准备加锁的免费广播节目进行加锁
                    for j in range(len(prepare_lock_free_radio_ch_commd_add_ok[i])):
                        send_numb_key_commd(prepare_lock_free_radio_ch_commd_add_ok[i][j])
                        if j == (len(prepare_lock_free_radio_ch_commd_add_ok[i]) - 1):
                            for k in range(len(LOCK_CH)):
                                send_commd(LOCK_CH[k])
                                if k == (len(LOCK_CH) - 1):
                                    send_commd(KEY["EXIT"])
                    GL.Radio_ch_attribute[0].remove(prepare_lock_free_radio_ch[i])  # 对加锁的免费广播节目在免费广播列表中去除
                    GL.Radio_ch_attribute[2].append(prepare_lock_free_radio_ch[i])  # 对加锁的免费广播节目加入到加锁广播节目列表
                    time.sleep(2)
                GL.get_ch_attr_sub_stage += 1
            elif len(prepare_lock_free_radio_ch) <= 0:
                logging.info("没有免费广播节目能被加锁")
                GL.get_ch_attr_sub_stage += 1

    elif GL.get_ch_attr_sub_stage == 2:
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            if len(prepare_lock_scramble_tv_ch) > 0:  # 对准备加锁的加密电视节目进行加锁
                prepare_lock_scramble_tv_ch_commd = change_numbs_to_commds_list(prepare_lock_scramble_tv_ch)
                prepare_lock_scramble_tv_ch_commd_add_ok = commds_add_key_list(prepare_lock_scramble_tv_ch_commd, KEY["OK"])
                for l in range(len(prepare_lock_scramble_tv_ch_commd_add_ok)):
                    for m in range(len(prepare_lock_scramble_tv_ch_commd_add_ok[l])):
                        send_numb_key_commd(prepare_lock_scramble_tv_ch_commd_add_ok[l][m])
                        if m == (len(prepare_lock_scramble_tv_ch_commd_add_ok[l]) - 1):
                            for n in range(len(LOCK_CH)):
                                send_commd(LOCK_CH[n])
                                if n == (len(LOCK_CH) - 1):
                                    send_commd(KEY["EXIT"])
                    GL.TV_ch_attribute[1].remove(prepare_lock_scramble_tv_ch[l])  # 对加锁的免费电视节目在免费电视列表中去除
                    GL.TV_ch_attribute[2].append(prepare_lock_scramble_tv_ch[l])  # 对加锁的免费电视节目加入到加锁电视节目列表
                    time.sleep(2)
                GL.get_ch_attr_sub_stage += 1
            elif len(prepare_lock_scramble_tv_ch) <= 0:
                logging.info("没有加密电视节目能被加锁")
                GL.get_ch_attr_sub_stage += 1

        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            if len(prepare_lock_scramble_radio_ch) > 0:  # 对准备加锁的加密广播节目进行加锁
                prepare_lock_scramble_radio_ch_commd = change_numbs_to_commds_list(prepare_lock_scramble_radio_ch)
                prepare_lock_scramble_radio_ch_commd_add_ok = commds_add_key_list(prepare_lock_scramble_radio_ch_commd, KEY["OK"])
                for l in range(len(prepare_lock_scramble_radio_ch_commd_add_ok)):
                    for m in range(len(prepare_lock_scramble_radio_ch_commd_add_ok[l])):
                        send_numb_key_commd(prepare_lock_scramble_radio_ch_commd_add_ok[l][m])
                        if m == (len(prepare_lock_scramble_radio_ch_commd_add_ok[l]) - 1):
                            for n in range(len(LOCK_CH)):
                                send_commd(LOCK_CH[n])
                                if n == (len(LOCK_CH) - 1):
                                    send_commd(KEY["EXIT"])
                    GL.Radio_ch_attribute[1].remove(prepare_lock_scramble_radio_ch[l])  # 对加锁的免费广播节目在免费广播列表中去除
                    GL.Radio_ch_attribute[2].append(prepare_lock_scramble_radio_ch[l])  # 对加锁的免费广播节目加入到加锁广播节目列表
                    time.sleep(2)
                GL.get_ch_attr_sub_stage += 1
            elif len(prepare_lock_scramble_radio_ch) <= 0:
                logging.info("没有加密广播节目能被加锁")
                GL.get_ch_attr_sub_stage += 1

    elif GL.get_ch_attr_sub_stage == 3:
        if ALL_TEST_CASE[choice_switch_case][4] == "TV":
            logging.info("免费电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[0]), GL.TV_ch_attribute[0]))
            logging.info("加密电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[1]), GL.TV_ch_attribute[1]))
            logging.info("加锁电视节目列表:{}--{}".format(len(GL.TV_ch_attribute[2]), GL.TV_ch_attribute[2]))
            logging.info("免费电视按TP归类字典:{}--{}".format(len(GL.free_tv_tp_ch_dict), GL.free_tv_tp_ch_dict))
            logging.info("免费电视按编码归类字典:{}--{}".format(len(GL.free_tv_codec_ch_dict), GL.free_tv_codec_ch_dict))
            logging.info("免费电视按分辨率归类字典:{}--{}".format(len(GL.free_tv_resolution_ch_dict), GL.free_tv_resolution_ch_dict))
            f1 = shelve.open(transfer_tv_data_file_path)
            f1["tv_ch_attr"] = GL.TV_ch_attribute
            f1["free_tv_tp_ch_dict"] = GL.free_tv_tp_ch_dict
            f1["free_tv_codec_ch_dict"] = GL.free_tv_codec_ch_dict
            f1["free_tv_resolution_ch_dict"] = GL.free_tv_resolution_ch_dict
            f1.close()
            GL.get_ch_attr_sub_stage += 1
            GL.current_stage += 1
        elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
            logging.info("免费广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[0]), GL.Radio_ch_attribute[0]))
            logging.info("加密广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[1]), GL.Radio_ch_attribute[1]))
            logging.info("加锁广播节目列表:{}--{}".format(len(GL.Radio_ch_attribute[2]), GL.Radio_ch_attribute[2]))
            f2 = shelve.open(transfer_radio_data_file_path)
            f2["radio_ch_attr"] = GL.Radio_ch_attribute
            f2.close()
            GL.get_ch_attr_sub_stage += 1
            GL.current_stage += 1

def check_ch_type():
    while GL.channel_info[7] != GL.all_test_case[choice_switch_case][0][4]:
        send_commd(KEY["TV/R"])
        if GL.channel_info[5] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])
    GL.current_stage += 1

def check_preparatory_work():
    if isinstance(GL.all_test_case[choice_switch_case][1], str):
        GL.current_stage += 1
    elif isinstance(GL.all_test_case[choice_switch_case][1], list):
        send_data = GL.all_test_case[choice_switch_case][1]
        for i in range(len(send_data)):
            send_commd(send_data[i])
        if GL.channel_info[5] == "1":
            unlock_channel()
        GL.current_stage += 1

def data_send_thread():
    while GL.main_loop_state:
        if GL.get_group_channel_total_info_state:
            get_group_channel_total_info()

        elif not GL.get_group_channel_total_info_state:
            if GL.current_stage == 0:   # 判断节目类型:TV/Radio
                check_ch_type()

            elif GL.current_stage == 1: # 判断是否有准备工作,比如进入某界面
                check_preparatory_work()

            elif GL.current_stage == 2: # 发送所选用例的切台指令
                send_data = GL.all_test_case[choice_switch_case][2]
                for i in range(len(send_data)):
                    for j in range(len(send_data[i])):
                        if j == 0:
                            GL.channel_info = ['', '', '', '', '', '', '', '', '', '', '', '', GL.prog_group_name, '']
                        send_numb_key_commd(send_data[i][j])
                        if j == len(send_data[i]) - 1:
                            if GL.all_test_case[choice_switch_case][0][-1] == INTERVAL_TIME[0]:
                                time.sleep(1)
                                if GL.channel_info[5] == "1":
                                    unlock_channel()
                                print(GL.channel_info)
                                logging.info(GL.channel_info)
                                GL.report_data[7].append(GL.channel_info)

                            elif GL.all_test_case[choice_switch_case][0][-1] == INTERVAL_TIME[1]:
                                time.sleep(2)   # 加入没有这个暂停,有时会出现不能即时获取到加锁信息,可能线程没有即时切换
                                if GL.channel_info[5] == "1":
                                    unlock_channel()
                                time.sleep(3)
                                if GL.all_test_case[choice_switch_case][1] == GET_CH_ATTRIBUTE:
                                    if ALL_TEST_CASE[choice_switch_case][4] == "TV":
                                        if GL.channel_info[5] == "1":  # 加锁电视节目
                                            GL.TV_ch_attribute[2].append(GL.channel_info[0])
                                        elif GL.channel_info[6] == "0":  # 免费电视节目
                                            GL.TV_ch_attribute[0].append(GL.channel_info[0])
                                            if GL.channel_info[2] not in GL.free_tv_tp_ch_dict.keys():  # 检测TP信息
                                                GL.free_tv_tp_ch_dict[GL.channel_info[2]] = []
                                                GL.free_tv_tp_ch_dict[GL.channel_info[2]].append(GL.channel_info[0])
                                            elif GL.channel_info[2] in GL.free_tv_tp_ch_dict.keys():
                                                GL.free_tv_tp_ch_dict[GL.channel_info[2]].append(GL.channel_info[0])
                                            if GL.channel_info[8] not in GL.free_tv_codec_ch_dict.keys():  # 检测视频编码类型
                                                GL.free_tv_codec_ch_dict[GL.channel_info[8]] = []
                                                GL.free_tv_codec_ch_dict[GL.channel_info[8]].append(GL.channel_info[0])
                                            elif GL.channel_info[8] in GL.free_tv_codec_ch_dict.keys():
                                                GL.free_tv_codec_ch_dict[GL.channel_info[8]].append(GL.channel_info[0])
                                            free_tv_resolution = "{}x{}".format(GL.channel_info[10],GL.channel_info[11])
                                            if free_tv_resolution not in GL.free_tv_resolution_ch_dict.keys():
                                                GL.free_tv_resolution_ch_dict[free_tv_resolution] = []
                                                GL.free_tv_resolution_ch_dict[free_tv_resolution].append(GL.channel_info[0])
                                            elif free_tv_resolution in GL.free_tv_resolution_ch_dict.keys():
                                                GL.free_tv_resolution_ch_dict[free_tv_resolution].append(GL.channel_info[0])
                                        elif GL.channel_info[6] == "1":  # 加密电视节目
                                            GL.TV_ch_attribute[1].append(GL.channel_info[0])
                                    elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
                                        if GL.channel_info[5] == "1":  # 加锁广播节目
                                            GL.Radio_ch_attribute[2].append(GL.channel_info[0])
                                        elif GL.channel_info[6] == "0":  # 免费广播节目
                                            GL.Radio_ch_attribute[0].append(GL.channel_info[0])
                                        elif GL.channel_info[6] == "1":  # 加密广播节目
                                            GL.Radio_ch_attribute[1].append(GL.channel_info[0])
                                print(GL.channel_info)
                                logging.info(GL.channel_info)
                                GL.report_data[7].append(GL.channel_info)
                GL.current_stage += 1

            elif GL.current_stage == 3: # 发送切台后的退回大画面的指令
                GL.commd_global_length = len(GL.all_test_case[choice_switch_case][3])
                if GL.commd_global_pos < GL.commd_global_length and GL.control_stage_delay_state:
                    GL.control_stage_delay_state = False
                    exit_commd = GL.all_test_case[choice_switch_case][3][GL.commd_global_pos]
                    send_commd(exit_commd)
                    GL.commd_global_pos += 1
                    delay_time(0.25, INTERVAL_TIME[0])
                elif GL.commd_global_pos == GL.commd_global_length:
                    GL.current_stage += 1
                    GL.commd_global_length = 0
                    GL.commd_global_pos = 0

            elif GL.current_stage == 4: # 写报告
                if ALL_TEST_CASE[choice_switch_case][4] == "TV":
                    GL.report_data[2] = int(GL.TV_channel_groups["All"])        # 分组节目总数
                elif ALL_TEST_CASE[choice_switch_case][4] == "Radio":
                    GL.report_data[2] = int(GL.Radio_channel_groups["All"])     # 分组节目总数

                scene_commd_list = GL.all_test_case[choice_switch_case][2]
                GL.report_data[5] = len(scene_commd_list)           # 切台次数
                GL.report_data[6] = change_numbs_to_str_list(change_commds_to_numbs_list(scene_commd_list))     # 将命令转化为字符串
                logging.debug(GL.report_data[7])
                check_if_report_exists_and_write_data_to_report()
                GL.current_stage += 1

            elif GL.current_stage == 5:
                if GL.all_test_case[choice_switch_case][1] == GET_CH_ATTRIBUTE:
                    get_ch_attribute()
                else:
                    GL.current_stage += 1
                    GL.control_stage_delay_state = False
                    delay_time(0.5, INTERVAL_TIME[1])

            elif GL.current_stage == 6 and GL.control_stage_delay_state: # 结束程序
                GL.main_loop_state = False

def data_receiver_thread():
    while GL.main_loop_state:
        data = receive_ser.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("ISO-8859-1", "replace")
            # data2 = data.decode('ISO-8859-1', 'replace').replace('\ufffd', '').strip()
            # data1 = data.decode("GB18030", "replace")
            # data2 = data.decode('GB18030', 'replace').replace('\ufffd', '').strip()
            data1 = data.decode("GB18030", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            # data3 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data2).strip()
            data4 = "[{}]     {}\n".format(str(tt), data2)
            # print(data2)
            write_logs_to_txt(full_log_txt_path, data4)
            if not GL.get_group_channel_total_info_state:
                write_logs_to_txt(case_log_txt_path, data4)

            if SWITCH_CHANNEL_KWS[0] in data2:
                ch_info_split = re.split(r"[\],]", data2)
                for i in range(len(ch_info_split)):
                    if CH_INFO_KWS[0] in ch_info_split[i]:  # 提取频道号
                        GL.channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if CH_INFO_KWS[1] in ch_info_split[i]:  # 提取频道名称
                        GL.channel_info[1] = re.split("=", ch_info_split[i])[-1]
                # print(GL.channel_info[0],GL.channel_info[1])

            if SWITCH_CHANNEL_KWS[1] in data2:
                flag_info_split = re.split(r"[\],]", data2)
                for i in range(len(flag_info_split)):
                    if CH_INFO_KWS[2] in flag_info_split[i]:  # 提取频道所属TP
                        GL.channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if CH_INFO_KWS[3] in flag_info_split[i]:  # 提取频道TTX_flag
                        GL.channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[4] in flag_info_split[i]:  # 提取频道SUB_flag
                        GL.channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[5] in flag_info_split[i]:  # 提取频道Lock_flag
                        GL.channel_info[5] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[6] in flag_info_split[i]:  # 提取频道Scramble_flag
                        GL.channel_info[6] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[7] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        GL.channel_info[7] = re.split(r"=", flag_info_split[i])[-1]
                    if GL.channel_info[7] == "TV":
                        if CH_INFO_KWS[8] in flag_info_split[i]:  # 提取TV频道视频编码
                            GL.channel_info[8] = re.split(r"=", flag_info_split[i])[-1]
                        if CH_INFO_KWS[9] in flag_info_split[i]:  # 提取TV频道音频编码
                            GL.channel_info[9] = re.split(r"=", flag_info_split[i])[-1]
                    if GL.channel_info[7] == "Radio":
                        GL.channel_info[8] = "0"  # 指定Radio频道视频编码为空
                        if CH_INFO_KWS[9] in flag_info_split[i]:  # 提取Radio频道音频编码
                            GL.channel_info[9] = re.split(r"=", flag_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[2] in data2:
                wide_high_info_split = re.split(r"[\],]", data2)
                for i in range(len(wide_high_info_split)):
                    if CH_INFO_KWS[10] in wide_high_info_split[i]:  # 提取TV频道画面高度
                        GL.channel_info[10] = re.split(r"=", wide_high_info_split[i])[-1]
                    if CH_INFO_KWS[11] in wide_high_info_split[i]:  # 提取TV频道画面宽度
                        GL.channel_info[11] = re.split(r"=", wide_high_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[3] in data2:
                group_info_split = re.split(r"[\],]", data2)
                for i in range(len(group_info_split)):
                    if GROUP_INFO_KWS[0] in group_info_split[i]:  # 提取频道所属组别
                        GL.prog_group_name = re.split(r"=", group_info_split[i])[-1]
                        GL.channel_info[12] = GL.prog_group_name
                    if GROUP_INFO_KWS[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        GL.prog_group_total = re.split(r"=", group_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[4] in data2:      # 提取切台时间
                GL.channel_info[13] = re.split(r"[:\s]", data2)[-2].strip()

if __name__ == "__main__":


    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "ZOOM": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "R/L": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "FIND": "A1 F1 22 DD 46", "PAUSE": "A1 F1 22 DD 45", "SUB": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIMESHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D",

        "CH+": "A1 F1 22 DD 11", "CH-": "A1 F1 22 DD 14", "VOL-": "A1 F1 22 DD 12", "VOL+": "A1 F1 22 DD 13",
        "MULTIFEED": "A1 F1 22 DD 0F", "SAT": "A1 F1 22 DD 17", "AUDIO": "A1 F1 22 DD 19", "TTX": "A1 F1 22 DD 1C",
    }

    REVERSE_KEY = dict([val,key] for key,val in KEY.items())

    SWITCH_CHANNEL_KWS = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    CH_INFO_KWS = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "TTX_flag",
        "SUB_flag",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "V_codec",
        "A_codec",
        "video_height",
        "video_width",
        "Group_name",]

    EXCEL_CH_INFO_KWS = [
        "频道号",
        "频道名称",
        "TP",
        "TTX标志",
        "SUB标志",
        "加锁标志",
        "加密标志",
        "节目类型",
        "视频编码",
        "音频编码",
        "视频高度",
        "视频宽度",
        "组别",
        "切台时间"]

    GROUP_INFO_KWS = [
        "Group_name",
        "Prog_total"
    ]

    NOT_PREPARATORY_WORK = "not_preparatory_work"
    GET_CH_ATTRIBUTE = "get_ch_attribute"
    READ_CH_ATTRIBUTE = "read_ch_attribute"
    CH_LIST_PREPARATORY_WORK = [KEY["OK"]]
    EPG_PREPARATORY_WORK = [KEY["EPG"]]
    CH_EDIT_PREPARATORY_WORK = [KEY["OK"],KEY["YELLOW"]]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    LOCK_CH = [KEY["OK"],KEY["YELLOW"],KEY["4"],KEY["0"],KEY["0"],KEY["0"],KEY["0"],KEY["OK"],KEY["EXIT"],KEY["OK"]]
    LOCK_ALL_CH = [KEY["OK"], KEY["YELLOW"],KEY["0"], KEY["0"], KEY["0"], KEY["0"], KEY["4"], KEY["0"], KEY["0"], KEY["0"], KEY["0"], KEY["RED"], KEY["EXIT"], KEY["OK"], KEY["EXIT"]]
    UNLOCK_ALL_CH = [KEY["OK"], KEY["YELLOW"], KEY["0"], KEY["0"], KEY["0"], KEY["0"], KEY["4"], KEY["0"], KEY["0"], KEY["0"], KEY["0"], KEY["RED"], KEY["EXIT"], KEY["OK"]]
    INTERVAL_TIME = [1.0, 5.0]

    ALL_TEST_CASE = [
        ["numb_key", "one_by_one", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["numb_key", "one_by_one", "timeout", "All", "Radio", INTERVAL_TIME[1]],
        ["numb_key", "one_by_one", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["numb_key", "random", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["screen", "up", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["screen", "down", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["screen", "random", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["screen", "up", "continuous", "All", "TV", INTERVAL_TIME[0]],
        ["ch_list", "up", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "random_up", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "random_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "page_up", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "page_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "left", "group", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "right", "group", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "left_group_random_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_list", "right_group_random_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["epg", "up", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["epg", "down", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["epg", "page_up", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["epg", "page_down", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["epg", "random", "timeout", "All", "TV", INTERVAL_TIME[1]],
        ["epg", "up", "continuous", "All", "TV", INTERVAL_TIME[0]],
        ["ch_edit", "up", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "random_up", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "random_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "page_up", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "page_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "left", "group", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "right", "group", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "left_group_random_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["ch_edit", "right_group_random_down", "add_ok", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "free_tv", "free_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "free_tv", "scr_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "free_tv", "lock_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "scr_tv", "free_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "scr_tv", "scr_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "scr_tv", "lock_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "lock_tv", "free_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "lock_tv", "scr_radio", "All", "TV", INTERVAL_TIME[1]],
        ["tv_radio", "lock_tv", "lock_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "free_tv", "free_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "free_tv", "scr_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "free_tv", "lock_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "scr_tv", "scr_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "scr_tv", "lock_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "lock_tv", "lock_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "free_radio", "free_radio", "All", "Radio", INTERVAL_TIME[1]],
        ["recall", "free_radio", "scr_radio", "All", "Radio", INTERVAL_TIME[1]],
        ["recall", "free_radio", "lock_radio", "All", "Radio", INTERVAL_TIME[1]],
        ["recall", "scr_radio", "scr_radio", "All", "Radio", INTERVAL_TIME[1]],
        ["recall", "scr_radio", "lock_radio", "All", "Radio", INTERVAL_TIME[1]],
        ["recall", "lock_radio", "lock_radio", "All", "Radio", INTERVAL_TIME[1]],
        ["recall", "free_tv", "free_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "free_tv", "scr_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "free_tv", "lock_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "scr_tv", "free_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "scr_tv", "scr_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "scr_tv", "lock_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "lock_tv", "free_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "lock_tv", "scr_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "lock_tv", "lock_radio", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "same_tp_tv", "same_tp_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "diff_tp_tv", "diff_tp_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "same_codec_tv", "same_codec_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "diff_codec_tv", "diff_codec_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "hd_tv", "hd_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "sd_tv", "sd_tv", "All", "TV", INTERVAL_TIME[1]],
        ["recall", "hd_sd_tv", "hd_sd_tv", "All", "TV", INTERVAL_TIME[1]],
    ]

    GL = MyGlobal()
    # 指定切台用例
    choice_switch_case = int(sys.argv[1])
    # choice_switch_case = 1
    # 检查打印日志和报告的目录,以及创建文件名称
    check_if_log_and_report_file_path_exists()
    build_print_log_and_report_file_path()

    # 配置日志信息
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT, filename=logging_file_path)
    # logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)
    # filename=r"d:\test\test.log" #有了filename参数就不会直接输出显示到控制台，而是直接写入文件

    # 获取串口并配置串口信息
    send_com = ''
    receive_com = ''
    send_ser_name,receive_ser_name = check_ports()
    send_ser = serial.Serial(send_ser_name, 9600)
    receive_ser = serial.Serial(receive_ser_name, 115200, timeout=1)
    # serial_set(send_ser,send_ser_name,9600)
    # serial_set(receive_ser,receive_ser_name,115200)

    msg = "现在开始执行:{}-{}-switch-channel".format(choice_switch_case,fmt_name)
    logging.critical(format(msg, '*^150'))
    print(format(msg, '*^150'))

    thread_send = threading.Thread(target= data_send_thread)
    thread_receive = threading.Thread(target= data_receiver_thread)
    thread_receive.start()
    thread_send.start()
