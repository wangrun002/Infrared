#!/usr/bin/python
# -*- coding: utf-8 -*-

from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import serial
import serial.tools.list_ports
import time
import re
import os,sys

def check_ports():
	ports_list = list(serial.tools.list_ports.comports())
	list_port = []
	if len(ports_list) <= 0:
		print("无可用端口")
	elif len(ports_list) <= 1:
		list_port.append(ports_list[0][0])
		print("可用端口:{}".format(list_port[0]))
	elif len(ports_list) >= 2:
#		print(ports_list)
		for i in range(len(ports_list)):
			if CheckPort["PortDescriptor"][0] in str(ports_list[i]):
				CheckPort["SendPort"] = ports_list[i][0]
			if CheckPort["PortDescriptor"][1] in str(ports_list[i]) and Serial_SER_Dict[str(sys.argv[3])] in str(ports_list[i][2]):
				print(ports_list[i][2])
				CheckPort["ReceivePort"] = ports_list[i][0]
			list_port.append(str(ports_list[i]))
		print("可用端口:{}".format(list_port))
	return list_port

def serial_set(ser,ser_name,ser_baudrate):
	ser.port = ser_name
	ser.baudrate = ser_baudrate
	ser.bytesize = 8
	ser.parity = "N"
	ser.stopbits = 1
	ser.timeout = 2

	ser.open()

def hexStringTobytes(str):
	str = str.replace(" ","")
	return bytes.fromhex(str)

def ReadDataFromFile(file_name,send_data):
	#将文本中的指令变成list
	with open(file_name,'r',encoding='utf-8') as fo:
		send_data = fo.read().split("\n")
	return send_data

def WriteDataToFile(file_name,write_data):
	with open(file_name,"a+",encoding='utf-8') as fo:
		fo.write(write_data)

def WriteDataToXlsx():
	alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	column_number = column_index_from_string("A")+button["Data_interval"]
	column_number_1 = column_index_from_string("A")+button["Data_interval"] + 1
	column_number_2 = column_index_from_string("A")+button["Data_interval"] + 2
	column_number_3 = column_index_from_string("A")+button["Data_interval"] + 3
	column_char = get_column_letter(column_number)
	column_char_1 = get_column_letter(column_number_1)
	column_char_2 = get_column_letter(column_number_2)
	column_char_3 = get_column_letter(column_number_3)
	ws.column_dimensions[column_char].width = 12
	ws.column_dimensions[column_char_1].width = 3
	ws.column_dimensions[column_char_2].width = 3
	ws.column_dimensions[column_char_3].width = 3
#	print(column_number,column_char)

	for i in range(len(Search_data)):
		if i < len(Search_data) - 2:		#按顺序写入Search_data[0]~[6]位元素的数据
			ws.cell((i+1),(1+button["Data_interval"])).value = Search_data[i]
			ws.merge_cells(start_row=(i+1),start_column=(1+button["Data_interval"]),end_row=(i+1),end_column=(1+button["Data_interval"]+4))
			ws.cell((i+1),(1+button["Data_interval"])).alignment = alignment
		elif i == len(Search_data) - 2:		#写入Search_data[7]元素字典的value列表的各个元素（表头）
			for n in range(len(xlxs_title[7]["数据类别"])):
				ws.cell((i+1),(1+button["Data_interval"]+n)).value = list(xlxs_title[i].values())[0][n]
				ws.cell((i+1),(1+button["Data_interval"]+n)).alignment = alignment
				ws.row_dimensions[(i+1)].height = 13.5
		elif i == len(Search_data) - 1:		#写入Search_data[8]元素（按照Search_data[7]的表头写入数据）
			for j in range(len(TP_LIST)):
				ws.cell((i+1+j),(1+button["Data_interval"])).value = Search_data[i][j]
				ws.cell((i+1+j),(1+button["Data_interval"]+1)).value = len(ChannelInfo[str(j+1)][0]) + len(ChannelInfo[str(j+1)][1]) #单个TP下的电视和广播总数
				ws.cell((i+1+j),(1+button["Data_interval"]+2)).value = len(ChannelInfo[str(j+1)][0])  #单个TP下的电视总数
				ws.cell((i+1+j),(1+button["Data_interval"]+3)).value = len(ChannelInfo[str(j+1)][1])  #单个TP下的广播总数
				ws.cell((i+1+j),(1+button["Data_interval"]+4)).value = ','.join(ChannelInfo[str(j+1)][0] + ChannelInfo[str(j+1)][1])
				for m in range(len(xlxs_title[7]["数据类别"])):
					ws.cell((i+1+j),(1+button["Data_interval"]+m)).alignment = alignment
				ws.row_dimensions[(i+1+j)].height = 13.5
	wb.save(WriteFileName["TotalExcel"])

kws_list = [
			"[PTD]SearchStart",		#0
			"[PTD]TV------",		#1
			"[PTD]Radio-----",		#2
			"[PTD]SearchFinish",	#3
			"[PTD]get :  fre",		#4
			"[PTD]TP_save=",		#5
			"[PTD]TV_save=",		#6
			"TV_save=",				#7
			"Radio_save=",			#8
			"[PTD]maximum_tp",		#9
			"[PTD]maximum_channel",	#10
			]

xlxs_title = [
				"搜索模式",
				"搜索次数",
				"搜索TP数",
				"搜索节目数",
				"保存TP数",
				"保存节目数",
				"搜索时间",
				{"数据类别":["TP","All","TV","Radio","CH_Name"]},
				"TP"
			]

ReadFileName = [
                'UpperLimitTPSearchCommand(Z6).txt',        #0
                'UpperLimitChannelSearchCommand(Y3).txt',   #1

                '88Sat6FBlindSearchCommand.txt',            #2直连
                '88Sat6FSuperBlindSearchCommand.txt',       #3直连

                'Z6Sat6FBlindSearchCommand.txt',            #4
                'Z6Sat6FSuperBlindSearchCommand.txt',       #5

                'Y3Sat6FBlindSearchCommand.txt',            #6
                'Y3Sat6FSuperBlindSearchCommand.txt',       #7

                '138Sat6FBlindSearchCommand.txt',           #8
                '138Sat6FSuperBlindSearchCommand.txt',      #9
                '138Sat6FBlindSearchAddCommand.txt',        #10

                'PLPDSat6FBlindSearchCommand.txt',          #11
                'PLPDSat6FSuperBlindSearchCommand.txt',     #12

                'FactoryResetSearchCommand.txt',            #13
                'AddNewSat20SearchCommand.txt',             #14
                'USBUpgradeUser20SatCommand.txt'            #15
                ]

Serial_SER_Dict = {
					"1":"FTDVKA2HA",
					"2":"FTGDWJ64A",
					"3":"FT9SP964A",
					"4":"FTHB6SSTA"
				}

#记录搜索相关信息的列表，对应xlxs_title的各项数据
Search_data = []    #保存写入Excel表的各项数据
TotalNumber = {}    #保存每次搜索的TP数和节目数
TotalNumber["TotalTVChannel"] = []
TotalNumber["TotalRadioChannel"] = []
TotalNumber["TotalTP"] = []
TotalNumber["tv_num"] = 0			#搜索过程中的电视节目数据统计
TotalNumber["radio_num"] = 0		#搜索过程中的广播节目数据统计
TotalNumber["save_tv_num"] = 0		#保存后的电视节目数据统计
TotalNumber["save_radio_num"] = 0	#保存后的广播节目数据统计
for i in range(len(xlxs_title)):
	Search_data.append(0)
Search_data[0] = "Blind"

#获取串口并配置串口信息
CheckPort = {}
CheckPort["PortDescriptor"] = ["USB-SERIAL CH340","USB Serial Port"]
CheckPort["SendPort"] = ''
CheckPort["ReceivePort"] = ''
CheckPort["SER_Value"] = 4   #指定USB转串口线的编号
ports = check_ports()
ser_name1 = CheckPort["SendPort"] #用于发送红外信号
ser_name2 = CheckPort["ReceivePort"] #用于接收串口打印
ser1 = serial.Serial()
ser2 = serial.Serial()
serial_set(ser1,ser_name1,9600)
serial_set(ser2,ser_name2,115200)

State = {}
State['MainLoopState'] = True #主循环程序控制状态
State['SendCommandState'] = True #True时执行ser.write(),False时暂停执行ser.write()

#按键指令相关信息
button = {}
button['data_length'] = 0    #发送指令的总次数
button['data_position'] = 0  #当前发送按键指令的位置
button["Srch_number"] = int(sys.argv[1])    #搜索次数

#创建发送红外指令文件和写数据文件
WriteFileName = {}
send_data = []
WriteFileName["SendCommandFileName"] = ReadFileName[int(sys.argv[2])]  #指定测试项
WriteFileName["ParentOfCurrentPath"] = os.path.abspath(os.path.join(os.getcwd(), "..")) #当前程序路径的上级路径
WriteFileName["SendCommandPath"] = os.path.join(WriteFileName["ParentOfCurrentPath"],"CommandFile",WriteFileName["SendCommandFileName"])
send_data = ReadDataFromFile(WriteFileName["SendCommandPath"],send_data)
#print(send_data)
button['data_length'] = len(send_data)

#保存文件名称处理
WriteFileName["TotalExcel"] = r".\Result\AddNewSatBlindSearchResult.xlsx"   #保存总表Excel的名称
WriteFileName["SatName"] = WriteFileName["SendCommandFileName"].split("Search")[0]
#WriteFileName["SheetName"] = WriteFileName["SendCommandFileName"].split("Sat")[0]
WriteFileName["Excel"] = r".\Result\{}Result.xlsx".format(WriteFileName["SatName"])
WriteFileName["TEXT"] = r".\Printlog\{}PrintLog.txt".format(WriteFileName["SatName"])

#搜索时间相关信息
Search_time = {} 
Search_time["start_time"] = 0  #记录搜索起始时间
Search_time["end_time"] = 0 #记录搜索结束时间
Search_time["Srch_Dur_time"] = 0 #搜索时间差

TP_LIST = [] #存放搜索到的TP信息
DataList = [] #存放盲扫开始的打印信息
PolarInfo = {} #存放极化判断的信息
PolarInfo['GetBlindInfo'] = [] #存放盲扫开始扫描的中频打印
PolarInfo['NumGetBlind'] = []  #存放盲扫中频打印的长度信息
PolarInfo["Countpolar"] = set() #存放过滤盲扫中频打印重复长度信息
PolarInfo["Polar"] = '' #存放H和V信息
PolarInfo["TP"] = ''    #存放搜索过程中的TP信息，会覆盖

ChannelInfo = {} #存放不同TP下的节目信息

#判断保存测试数据和打印保存的目录是否存在，否则创建
if not os.path.exists(os.path.dirname(WriteFileName["Excel"])):
	os.mkdir(os.path.dirname(WriteFileName["Excel"]))

if not os.path.exists(os.path.dirname(WriteFileName["TEXT"])):
	os.mkdir(os.path.dirname(WriteFileName["TEXT"]))

#判断表格是否存在，否则新建，是则打开
alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
if not os.path.exists(WriteFileName["TotalExcel"]):
    wb = Workbook()
    ws = wb.active
    ws.title = WriteFileName["SatName"]
elif os.path.exists(WriteFileName["TotalExcel"]):
    wb = load_workbook(WriteFileName["TotalExcel"])
    sheetnamelist = wb.sheetnames
    print(sheetnamelist)
    if WriteFileName["SatName"] in sheetnamelist:
        ws = wb[WriteFileName["SatName"]]
    elif WriteFileName["SatName"] not in sheetnamelist:
        ws = wb.create_sheet(WriteFileName["SatName"])
ws.column_dimensions['A'].width = 11
for i in range(len(xlxs_title)):
    if i < len(xlxs_title) - 2:
        ws.cell(i+1,1).value = xlxs_title[i]
        ws.cell(i+1,1).alignment = alignment
    elif i == len(xlxs_title) - 2:
        ws.cell(i+1,1).value = list(xlxs_title[i].keys())[0]
        ws.cell(i+1,1).alignment = alignment
    elif i == len(xlxs_title) - 1:
        ws.cell(i+1,1).value = xlxs_title[i]
        ws.cell(i+1,1).alignment = alignment

while State['MainLoopState']:
	data = ser2.readline() #获取serial端口bytes数据
	if data:
		data1 = data.decode('ISO-8859-1')
		data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
		tt = datetime.now()
		data3 = "[{}]    {}\n".format(str(tt),data2)
#		print(data2)
		WriteDataToFile(WriteFileName["TEXT"],data3)

		#监控搜索起始
		if kws_list[0] in data2:
			State['SendCommandState'] = False
			Search_time["start_time"] = datetime.now()
			Search_data[1] += 1
			button["Data_interval"] = 1 + 5 * (Search_data[1] - 1)

		#判断H和V极化方向
		if 'get blind - fre' in data2:
			PolarInfo['GetBlindInfo'].append(data2)

		if len(PolarInfo['GetBlindInfo']) != 0:
			if len(PolarInfo['GetBlindInfo']) not in PolarInfo['NumGetBlind']:
				PolarInfo['NumGetBlind'].append(len(PolarInfo['GetBlindInfo']))
#				print(PolarInfo['NumGetBlind'])
			elif len(PolarInfo['GetBlindInfo']) in PolarInfo['NumGetBlind']:
#				print(PolarInfo['NumGetBlind']) 
				PolarInfo["Countpolar"].add(len(PolarInfo['GetBlindInfo']))
				if (len(PolarInfo["Countpolar"]) % 2) != 0:
					PolarInfo["Polar"] = "H"
				elif (len(PolarInfo["Countpolar"]) % 2) ==0:
					PolarInfo["Polar"] = "V"

		#获取频点信息
		if kws_list[4] in data2:
			fre = data2.split(' ')[5]
			sym = data2.split(' ')[9]
			PolarInfo["TP"] = "{}{}{}".format(fre,PolarInfo["Polar"],sym)
			TP_LIST.append(PolarInfo["TP"])
			ChannelInfo[str(len(TP_LIST))] = []
			ChannelInfo[str(len(TP_LIST))].append([])
			ChannelInfo[str(len(TP_LIST))].append([])
		#获取电视节目个数信息
		if kws_list[1] in data2:
			TotalNumber["tv_num"] = re.split("------|    ",data2)[1]  #提取电视个数信息
#			print(TotalNumber["tv_num"])
			Search_data[3] = "{}/{}".format(TotalNumber["tv_num"], TotalNumber["radio_num"])
			#ChannelInfo[str(len(TP_LIST))].append(re.split("------|    ",data2)[2])  #提取电视节目名称
			ChannelInfo[str(len(TP_LIST))][0].append("[T]{}".format(re.split("------|    ",data2)[2]))
		#获取广播节目个数信息
		if kws_list[2] in data2:
			TotalNumber["radio_num"] = re.split("-----|    ",data2)[1]   #提取广播个数信息
			Search_data[3] = "{}/{}".format(TotalNumber["tv_num"], TotalNumber["radio_num"])
			#ChannelInfo[str(len(TP_LIST))].append(re.split("-----|    ",data2)[2])    #提取广播节目名称
			ChannelInfo[str(len(TP_LIST))][1].append("[R]{}".format(re.split("-----|    ",data2)[2]))
#			print(re.split("-----|    ",data2)[2])
#			print(ChannelInfo[PolarInfo["TP"]])

		#监控搜索结束
		if kws_list[3] in data2:
			State['SendCommandState'] = True
			Search_time["end_time"] = datetime.now()
			Search_time["Srch_Dur_time"] = Search_time["end_time"] - Search_time["start_time"]
			for i in range(len(TP_LIST)):
				print(TP_LIST[i])
			print("本次搜索节目总数为TV/Radio:{},TP总数为:{},盲扫时长:{}".format(Search_data[3],\
																				len(TP_LIST),\
																				Search_time["Srch_Dur_time"]))
		#监控保存的TP个数
		if kws_list[5] in data2:
			Search_data[4] = int(re.split(r'=', data2)[1])

		#监控保存的电视和广播节目个数
		if kws_list[6] in data2:
#			str_pat = re.compile(r'\[(.*?)\]').findall(data2)
			str_pat = re.split(r"]|,", data2)
			for i in range(len(str_pat)):
				if kws_list[7] in str_pat[i]:		#用来检测保存的TV数的监测
					TotalNumber["save_tv_num"] = re.split(r'=',str_pat[i])[1]
				if kws_list[8] in str_pat[i]:		#用来检测保存的TV数的监测
					TotalNumber["save_radio_num"] = re.split(r'=',str_pat[i])[1]
			Search_data[5] = "{}/{}".format(TotalNumber["save_tv_num"], TotalNumber["save_radio_num"])

			Search_data[2] = len(TP_LIST)
			Search_data[6] = str(Search_time["Srch_Dur_time"])[2:10]
			Search_data[8] = TP_LIST
			WriteDataToXlsx()
			TotalNumber["TotalTVChannel"].append(int(TotalNumber["tv_num"]))
			TotalNumber["TotalRadioChannel"].append(int(TotalNumber["radio_num"]))
			TotalNumber["TotalTP"].append(Search_data[2])
			print("本次搜索实际保存TV/Radio:{},保存TP数为:{}".format(Search_data[5],Search_data[4]))
			print("当前轮次:{},累计搜索节目个数:{}/{},TP个数:{}".format(Search_data[1],\
																	sum(TotalNumber["TotalTVChannel"]),\
																	sum(TotalNumber["TotalRadioChannel"]),\
																	sum(TotalNumber["TotalTP"])))
			Search_time.clear()
			TP_LIST.clear()
			PolarInfo['GetBlindInfo'].clear()
			PolarInfo['NumGetBlind'].clear()
			PolarInfo["Countpolar"].clear()
			ChannelInfo.clear()  #不同TP下的节目字典清空
			TotalNumber["tv_num"] = 0  #搜索结束后将节目数清空
			TotalNumber["radio_num"] = 0

		if kws_list[9] in data2:
			print("搜索TP达到上限：{}".format(data2))
			ser1.write(hexStringTobytes("A1 F1 22 DD 15"))
			button["Srch_number"] = 1

		if kws_list[10] in data2:
			print("搜索节目达到上限：{}".format(data2))
			ser1.write(hexStringTobytes("A1 F1 22 DD 15"))
			button["Srch_number"] = 1


	if not data and State['SendCommandState']:
		if button['data_position'] != button['data_length']:
			print('{}:{}'.format(button['data_position'],button['data_length']))
			print(send_data[button['data_position']])
			ser1.write(hexStringTobytes(send_data[button['data_position']]))	
			button['data_position'] += 1
#			time.sleep(1)

		elif button['data_position'] == button['data_length'] and button["Srch_number"] >= 1:
			time.sleep(3)
			button['data_position'] = 0
			button["Srch_number"] -= 1

			if button["Srch_number"] == 0:
				State['MainLoopState'] = False
