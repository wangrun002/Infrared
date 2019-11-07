#!/usr/bin/python
# -*- coding: utf-8 -*-

'''
voltage = { "0":"13V",
            "1":"18V",
            "2":"Off"
            }

22k = { "0":"On",
        "1":"Off"
        }

diseqc 1.0 = {  "0":"Off",
                "1":"Port1",
                "2":"Port2",
                "3":"Port3",
                "4":"Port4"
                }
'''

from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment
from openpyxl.utils import get_column_letter,column_index_from_string
import serial
import serial.tools.list_ports
import re
import time
import os


MAIN_LOOP_STATE = True
UP_KEY_VALUE = "A1 F1 22 DD 11"
DOWN_KEY_VALUE = "A1 F1 22 DD 14"
LEFT_KEY_VALUE = "A1 F1 22 DD 12"
RIGHT_KEY_VALUE = "A1 F1 22 DD 13"
MENU_KEY_VALUE = "A1 F1 22 DD 0C"
OK_KEY_VALUE = "A1 F1 22 DD 15"
PAGEDOWN_KEY_VALUE = "A1 F1 22 DD 18"
EXIT_KEY_VALUE = "A1 F1 22 DD 0D"
ENTER_ANTENNA_SETTING = [MENU_KEY_VALUE,OK_KEY_VALUE]
EXIT_ANTENNA_SETTING = [OK_KEY_VALUE,EXIT_KEY_VALUE,EXIT_KEY_VALUE]
xlsx_title = [
				"搜索模式",
				"搜索次数",
				"搜索TP数",
				"搜索节目数",
				"保存TP数",
				"保存节目数",
				"搜索时间",
				{"数据类别":["TP","All","TV","Radio","CH_Name"]},
				"TP"
			]

sat_list = [
                "6b_blind",         #0
                "6b_superblind",    #1

                "y3_blind",         #2
                "y3_superblind",    #3

                "138_blind",        #4
                "138_superblind",   #5

                "88_blind",         #6
                "88_superblind",    #7
]
choice_search_sat = 2       # 参考sat_list中的选项进行卫星选择
search_time = 2             # 选择搜索轮次


class MyGlobal():
    def __init__(self):
        self.ser_cable_num = 4                          # USB转串口线编号
        self.switch_commd_stage = 0                     # 切换发送命令的阶段
        self.setting_option_numb = 0                    # 设置项位置number
        self.lnb_power = ''                             # 用来获取打印中的电压
        self.switch_lnb_power_state = True              # 用来控制切换本振power选项参数切换的状态变量
        self.next_search_setting_lnb_fre_state = True   # 用来控制第一次搜索完后，控制本振参数不再设置的状态变量
        self.blind_judge_polar = [[],[],set(),'']       # 用于判断极化
        self.all_tp_list = []                           # 用于存放搜索到的TP
        self.channel_info = {}                          # 用于存放各个TP下搜索到的电视和广播节目名称
        self.search_datas = [0,0,0,0,0,0,0,0,0]         # 用于存放xlsx_title中的数据
        self.xlsx_data_interval = 0                     # 用于计算每轮搜索写xlsx时的间隔

        self.sat_param_save = ["", "", "", "", ""]
        self.sat_param_kws =    [
                                    "[PTD]sat_name=",
                                    "[T1]set tp",
                                    "--------set diseqc 1.0",
                                    "--------set diseqc 1.1"
                                ]

        self.search_monitor_kws = [
                                    "[PTD]SearchStart",		#0
                                    "[PTD]TV------",		#1
                                    "[PTD]Radio-----",		#2
                                    "[PTD]SearchFinish",	#3
                                    "[PTD]get :  fre",		#4
                                    "[PTD]TP_save=",		#5
                                    "[PTD]TV_save=",		#6
                                    "TV_save=",				#7
                                    "Radio_save=",			#8
                                    "[PTD]maximum_tp",		#9
                                    "[PTD]maximum_channel",	#10
                                    "get blind - fre",      #11
                                    ]

        self.all_sat_commd =   [
                                    [[OK_KEY_VALUE,PAGEDOWN_KEY_VALUE,PAGEDOWN_KEY_VALUE,
                                     PAGEDOWN_KEY_VALUE,DOWN_KEY_VALUE,DOWN_KEY_VALUE,
                                     DOWN_KEY_VALUE,OK_KEY_VALUE,DOWN_KEY_VALUE],
                                    [RIGHT_KEY_VALUE,OK_KEY_VALUE,OK_KEY_VALUE],
                                    ["Chinas6b_C", "voltage = 0", "sat22k = 1", "2", "0","Blind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                     PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                     DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["Chinas6b_C", "voltage = 0", "sat22k = 1", "2", "0", "SuperBlind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                      PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                      DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["Asiasat 7 C", "voltage = 0", "sat22k = 1", "1", "0", "Blind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                      PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                      DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["Asiasat 7 C", "voltage = 0", "sat22k = 1", "1", "0", "SuperBlind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                      PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                      DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["Telstar 18 K", "voltage = 0", "sat22k = 0", "4", "0", "Blind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                      PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                      DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["Telstar 18 K", "voltage = 0", "sat22k = 0", "4", "0", "SuperBlind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                      PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                      DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["ST 2 K", "voltage = 0", "sat22k = 1", "3", "0", "Blind"]],

                                    [[OK_KEY_VALUE, PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE,
                                      PAGEDOWN_KEY_VALUE, PAGEDOWN_KEY_VALUE, DOWN_KEY_VALUE,
                                      DOWN_KEY_VALUE, OK_KEY_VALUE, DOWN_KEY_VALUE],
                                     [RIGHT_KEY_VALUE, OK_KEY_VALUE, OK_KEY_VALUE],
                                     ["ST 2 K", "voltage = 0", "sat22k = 1", "3", "0", "SuperBlind"]],
                                ]

def check_ports(ser_cable_num):
    serial_ser =    {
                        "1":"FTDVKA2HA",
                        "2":"FTGDWJ64A",
                        "3":"FT9SP964A",
                        "4":"FTHB6SSTA"
                    }
    send_port_desc = "USB-SERIAL CH340"
    receive_port_desc = "USB Serial Port"
    ports = list(serial.tools.list_ports.comports())
    ports_com = []
    if len(ports) <= 0:
        print("无可用端口")
    elif len(ports) == 1:
        print("只有一个可用端口:{}".format(ports[0][0]))
    elif len(ports) >=2:
        for i in range(len(ports)):
            ports_com.append(str(ports[i]))
            if send_port_desc in str(ports[i]):
                send_com = ports[i][0]
            if receive_port_desc in str(ports[i]) and serial_ser[str(ser_cable_num)] in str(ports[i][2]):
                receive_com = ports[i][0]
                print(ports[i][2])
        print("可用端口:{}".format(ports_com))
    return send_com,receive_com

def serial_set(ser,ser_name,ser_baudrate):
    ser.port = ser_name
    ser.baudrate = ser_baudrate
    ser.bytesize = 8
    ser.parity = "N"
    ser.stopbits = 1
    ser.timeout = 1
    ser.open()

def hex_strs_to_bytes(strings):
    strs = strings.replace(" ", "")
    return bytes.fromhex(strs)

def write_data_to_txt(file_path,write_data):
    with open(file_path,"a+",encoding="utf-8") as fo:
        fo.write(write_data)

def judge_write_file_exist():
    if not os.path.exists(write_xlsx_relative_path):
        os.mkdir(write_xlsx_relative_path)
    if not os.path.exists(write_txt_relative_path):
        os.mkdir(write_txt_relative_path)

def judge_and_wirte_data_to_xlsx():
    alignment = Alignment(horizontal="center",vertical="center",wrapText=True)
    if not os.path.exists(write_xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 11
        for i in range(len(xlsx_title)):
            if i < len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = list(xlsx_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 1:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment

    elif os.path.exists(write_xlsx_path):
        wb = load_workbook(write_xlsx_path)
        sheets_name_list = wb.sheetnames
        print(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 11
        for i in range(len(xlsx_title)):
            if i < len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = list(xlsx_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 1:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment

    tp_column_numb = column_index_from_string("A") + GL.xlsx_data_interval
    all_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 1
    tv_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 2
    radio_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 3
    tp_column_char = get_column_letter(tp_column_numb)
    all_column_char = get_column_letter(all_column_numb)
    tv_column_char = get_column_letter(tv_column_numb)
    radio_column_char = get_column_letter(radio_column_numb)
    ws.column_dimensions[tp_column_char].width = 12
    ws.column_dimensions[all_column_char].width = 3
    ws.column_dimensions[tv_column_char].width = 3
    ws.column_dimensions[radio_column_char].width = 3

    for m in range(len(GL.search_datas)):
        if m < len(GL.search_datas) - 2:
            ws.cell((m + 1),(1 + GL.xlsx_data_interval)).value = GL.search_datas[m]
            ws.merge_cells(start_row=(m + 1),start_column=(1 + GL.xlsx_data_interval),\
                           end_row=(m + 1),end_column=(1 + GL.xlsx_data_interval + 4))
            ws.cell((m + 1),(1 + GL.xlsx_data_interval)).alignment = alignment
        elif m == len(GL.search_datas) - 2:
            for n in range(len(xlsx_title[7]["数据类别"])):
                ws.cell((m + 1),(1 + GL.xlsx_data_interval + n)).value = list(xlsx_title[m].values())[0][n]
                ws.cell((m + 1), (1 + GL.xlsx_data_interval + n)).alignment = alignment
                ws.row_dimensions[(m+1)].height = 13.5
        elif m == len(GL.search_datas) - 1:
            for j in range(len(GL.all_tp_list)):
                ws.cell((m+1+j),(1+GL.xlsx_data_interval)).value = GL.search_datas[m][j]
                ws.cell((m+1+j),(1+GL.xlsx_data_interval)+1).value = len(GL.channel_info[str(j+1)][0]) + \
                                                                     len(GL.channel_info[str(j+1)][1])
                ws.cell((m+1+j),(1+GL.xlsx_data_interval)+2).value = len(GL.channel_info[str(j+1)][0])
                ws.cell((m+1+j),(1+GL.xlsx_data_interval)+3).value = len(GL.channel_info[str(j+1)][1])
                ws.cell((m+1+j),(1+GL.xlsx_data_interval)+4).value = ",".join(GL.channel_info[str(j+1)][0] + \
                                                                              GL.channel_info[str(j+1)][1])
                for k in range(len(xlsx_title[7]["数据类别"])):
                    ws.cell((m+1+j),(1+GL.xlsx_data_interval)+k).alignment = alignment
                ws.row_dimensions[(m+1+j)].height = 13.5
    wb.save(write_xlsx_path)

GL = MyGlobal()

sat_name = GL.all_sat_commd[choice_search_sat][2][0]
search_mode = GL.all_sat_commd[choice_search_sat][2][-1]
timestamp = re.sub(r'[-: ]','_',str(datetime.now())[:19])
sheet_name = "{}_{}".format(sat_name,search_mode)

write_xlsx_file_name = "PrestSatSearchResult.xlsx"
write_xlsx_relative_path = r".\Result"
write_xlsx_path = os.path.join(write_xlsx_relative_path,write_xlsx_file_name)

write_txt_file_name = "{}_{}_{}.txt".format(sat_name,search_mode,timestamp)
write_txt_relative_path = r".\PrintLog"
write_txt_path = os.path.join(write_txt_relative_path,write_txt_file_name)

judge_write_file_exist()


send_ser_name,receive_ser_name = check_ports(GL.ser_cable_num)
send_ser = serial.Serial()
receive_ser = serial.Serial()
serial_set(send_ser, send_ser_name, 9600)
serial_set(receive_ser, receive_ser_name, 115200)

while MAIN_LOOP_STATE:
    data = receive_ser.readline()
    if not data:
        print("======================================================================================================")
        if GL.switch_commd_stage == 0:
            for i in range(len(ENTER_ANTENNA_SETTING)):
                send_ser.write(hex_strs_to_bytes(ENTER_ANTENNA_SETTING[i]))
            GL.switch_commd_stage += 1

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 0:
            print("Satellite")
            if GL.sat_param_save[0] == GL.all_sat_commd[choice_search_sat][2][0]:
                send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
                GL.setting_option_numb += 1
            elif GL.sat_param_save[0] != GL.all_sat_commd[choice_search_sat][2][0]:
                send_ser.write(hex_strs_to_bytes(LEFT_KEY_VALUE))

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 1:
            print("LNB POWER")
            power_off = "voltage = 2"
            if GL.sat_param_save[1] != power_off and GL.switch_lnb_power_state:
                send_ser.write(hex_strs_to_bytes(LEFT_KEY_VALUE))
            elif GL.sat_param_save[1] == power_off and GL.switch_lnb_power_state:
                GL.switch_lnb_power_state = False
                send_ser.write(hex_strs_to_bytes(RIGHT_KEY_VALUE))
                send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
                GL.setting_option_numb += 1

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 2:
            print("LBN FREQUENCY")
            if GL.next_search_setting_lnb_fre_state:
                for i in range(len(GL.all_sat_commd[choice_search_sat][0])):
                    send_ser.write(hex_strs_to_bytes(GL.all_sat_commd[choice_search_sat][0][i]))
                    time.sleep(1)
                    if i == len(GL.all_sat_commd[choice_search_sat][0]) - 1:
                        GL.setting_option_numb += 1
                        GL.next_search_setting_lnb_fre_state = False
            elif not GL.next_search_setting_lnb_fre_state:
                send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
                GL.setting_option_numb += 1

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 3:
            print("22k")
            if GL.sat_param_save[2] != GL.all_sat_commd[choice_search_sat][2][2]:
                send_ser.write(hex_strs_to_bytes(LEFT_KEY_VALUE))
            elif GL.sat_param_save[2] == GL.all_sat_commd[choice_search_sat][2][2]:
                send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
                GL.setting_option_numb += 1

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 4:
            print("Diseqc 1.0")
            if GL.sat_param_save[3] != GL.all_sat_commd[choice_search_sat][2][3]:
                send_ser.write(hex_strs_to_bytes(LEFT_KEY_VALUE))
            elif GL.sat_param_save[3] == GL.all_sat_commd[choice_search_sat][2][3]:
                send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
                GL.setting_option_numb += 1

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 5:
            print("Diseqc 1.1")
            if GL.sat_param_save[4] != GL.all_sat_commd[choice_search_sat][2][4]:
                send_ser.write(hex_strs_to_bytes(LEFT_KEY_VALUE))
            elif GL.sat_param_save[4] == GL.all_sat_commd[choice_search_sat][2][4]:
                send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
                GL.setting_option_numb += 1

        elif GL.switch_commd_stage == 1 and GL.setting_option_numb == 6:
            print("TP")
            send_ser.write(hex_strs_to_bytes(DOWN_KEY_VALUE))
            GL.setting_option_numb += 1
            GL.switch_commd_stage += 1

        elif GL.switch_commd_stage == 2 and GL.setting_option_numb == 7:
            print("Start Search")
            for i in range(len(GL.all_sat_commd[choice_search_sat][1])):
                send_ser.write(hex_strs_to_bytes(GL.all_sat_commd[choice_search_sat][1][i]))
                time.sleep(1)
            GL.setting_option_numb += 1

        elif GL.switch_commd_stage == 3:
            print("End Search")
            for i in range(len(EXIT_ANTENNA_SETTING)):
                send_ser.write(hex_strs_to_bytes(EXIT_ANTENNA_SETTING[i]))
                time.sleep(2)
            GL.switch_commd_stage += 1

        elif GL.switch_commd_stage == 4:
            GL.switch_commd_stage = 0
            GL.setting_option_numb = 0
            GL.switch_lnb_power_state = True
            search_time -= 1
            if search_time < 1:
                MAIN_LOOP_STATE = False


    if data:
        tt = datetime.now()
        data1 = data.decode("ISO-8859-1")
        data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
        data3 = "[{}]     {}\n".format(str(tt),data2)
        print(data2)
        write_data_to_txt(write_txt_path,data3)

        if GL.sat_param_kws[0] in data2:
            GL.sat_param_save[0] = re.split("=",data2)[-1]
        if GL.sat_param_kws[1] in data2:  # 判断[T1]set tp
            GL.sat_param_save[1] = re.split(",", data2)[3].strip()  # 获取voltage
            GL.sat_param_save[2] = re.split(",", data2)[4].strip()  # 获取22k
        if GL.sat_param_kws[2] in data2:  # 判断diseqc 1.0
            GL.sat_param_save[3] = list(filter(None, re.split("-{2,}|,", data2)))[-1].strip()
        if GL.sat_param_kws[3] in data2:  # 判断diseqc 1.1
            GL.sat_param_save[4] = list(filter(None, re.split("-{2,}|,", data2)))[-1].strip()

        if GL.search_monitor_kws[0] in data2:       # 监控搜索起始
            start_time = datetime.now()
            GL.search_datas[1] += 1
            GL.xlsx_data_interval = 1 + 5 * (GL.search_datas[1] - 1)

        if GL.search_monitor_kws[11] in data2:      # 监控极化方向
            GL.blind_judge_polar[0].append(data2)

        if len(GL.blind_judge_polar[0]) != 0:
            if len(GL.blind_judge_polar[0]) not in GL.blind_judge_polar[1]:
                GL.blind_judge_polar[1].append(len(GL.blind_judge_polar[0]))
            elif len(GL.blind_judge_polar[0]) in GL.blind_judge_polar[1]:
                GL.blind_judge_polar[2].add(len(GL.blind_judge_polar[1]))
                if (len(GL.blind_judge_polar[2]) % 2) != 0:
                    GL.blind_judge_polar[3] = "H"
                elif (len(GL.blind_judge_polar[2]) % 2) == 0:
                    GL.blind_judge_polar[3] = "V"

        if GL.search_monitor_kws[4] in data2:       # 监控频点信息
            fre = data2.split(" ")[5]
            symb = data2.split(" ")[9]
            tp = "{}{}{}".format(fre,GL.blind_judge_polar[3],symb)
            GL.all_tp_list.append(tp)
            GL.channel_info[str(len(GL.all_tp_list))] = [[],[]]

        if GL.search_monitor_kws[1] in data2:       # 监控搜索过程电视个数和名称信息
            tv_numb = re.split("-{2,}|\s{2,}",data2)[1]     # 提取电视节目数
            tv_name = re.split("-{2,}|\s{2,}",data2)[2]     # 提取电视节目名称
            GL.channel_info[str(len(GL.all_tp_list))][0].append('[T]{}'.format(tv_name))

        if GL.search_monitor_kws[2] in data2:       # 监控搜索过程广播个数和名称信息
            radio_numb = re.split("-{2,}|\s{2,}",data2)[1]      # 提取广播节目数
            radio_name = re.split("-{2,}|\s{2,}",data2)[2]      # 提取电视节目名称
            GL.channel_info[str(len(GL.all_tp_list))][1].append('[R]{}'.format(radio_name))

        if GL.search_monitor_kws[3] in data2:       # 监控搜索结束
            end_time = datetime.now()
            search_dur_time = str(end_time - start_time)[2:10]
            GL.switch_commd_stage += 1
            for i in range(len(GL.all_tp_list)):
                print(GL.all_tp_list[i])
            print("第{}次搜索节目总数为TV/Radio:{}/{},TP总数为:{},盲扫时长:{}".format(GL.search_datas[1],\
                                                                      tv_numb,radio_numb,len(GL.all_tp_list),\
                                                                      search_dur_time))

            # GL.search_datas[5] = '0/0'
            # save_tv_numb, save_radio_numb = 0, 0
            # save_tp_numb = 0


        if GL.search_monitor_kws[5] in data2:       # 监控保存TP的个数
            save_tp_numb = int(re.split("=",data2)[1])
            GL.search_datas[4] = save_tp_numb
            GL.search_datas[0] = GL.all_sat_commd[choice_search_sat][2][-1]
            GL.search_datas[2] = len(GL.all_tp_list)
            GL.search_datas[3] = "{}/{}".format(tv_numb, radio_numb)
            GL.search_datas[6] = search_dur_time
            GL.search_datas[8] = GL.all_tp_list
            judge_and_wirte_data_to_xlsx()

            GL.all_tp_list.clear()
            GL.blind_judge_polar[0].clear()
            GL.blind_judge_polar[1].clear()
            GL.blind_judge_polar[2].clear()
            GL.channel_info.clear()
            tv_numb, radio_numb = 0, 0
            save_tp_numb = 0

        if GL.search_monitor_kws[6] in data2:       # 监控保存TV和Radio的个数
            split_result = re.split(r"]|,",data2)
            save_tv_numb = re.split("=",split_result[1])[1]
            save_radio_numb = re.split("=",split_result[2])[1]
            GL.search_datas[5] = "{}/{}".format(save_tv_numb,save_radio_numb)
            judge_and_wirte_data_to_xlsx()
            GL.search_datas[5] = '0/0'
            save_tv_numb, save_radio_numb = 0, 0
