# !/usr/bin/python3
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, column_index_from_string
import testlink
import os
import logging


def logging_info_setting():
    # 配置logging输出格式
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    date_format = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=log_format, datefmt=date_format)


def get_all_project():
    ret = []
    for i in tlc.getProjects():
        ret.append({'id': i['id'], 'name': i['name']})
    return ret


def get_specify_project_id_by_name(name):
    for i in tlc.getProjects():
        if i['name'] == name:
            return i['id']


def get_all_first_suites_id(project_id):
    all_first_suite_id = {}
    specify_project_first_suites = tlc.getFirstLevelTestSuitesForTestProject(project_id)
    for first_suite in specify_project_first_suites:
        print(first_suite['id'], first_suite['name'])
        all_first_suite_id[first_suite['id']] = first_suite['name']
    return all_first_suite_id


def data_replace(data):
    data = data.replace('<p>', '')
    data = data.replace('</p>', '')
    data = data.replace('&nbsp;', '')
    data = data.replace('&ldquo;', '"')
    data = data.replace('&rdquo;', '"')
    data = data.replace('&quot;', '"')
    data = data.replace('&le;', '<')
    data = data.replace('<span style="color: rgb(255, 0, 0);">', '')
    data = data.replace('</span>', '')
    data = data.replace('<strong>', '')
    data = data.replace('<span style="font-size: larger;">', '')
    data = data.replace('</strong>', '')
    data = data.replace('<br />', '')
    data = data.replace('<span style="font-family: Arial;">', '')
    return data


def write_test_case_to_excel(case_info):
    wb = ''
    ws = ''
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    a_column_numb = column_index_from_string("A")
    module_title = ["用例模块名称", "用例标题", "概要", "前提", "步骤动作", "期望的结果"]
    if not os.path.exists(module_file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = module_name

        # 写module_title、设置列宽、行高
        for i in range(len(module_title)):
            ws.cell(1, i + 1).value = module_title[i]
            ws.cell(1, i + 1).alignment = alignment
            if i in [0, 1, 2, 3]:
                ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 20
            else:
                ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 50
        ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高

    elif os.path.exists(module_file_name):
        wb = load_workbook(module_file_name)
        sheets_name_list = wb.sheetnames
        # logging.info(sheets_name_list)
        if module_name in sheets_name_list:
            ws = wb[module_name]
        elif module_name not in sheets_name_list:
            ws = wb.create_sheet(module_name)
            # 写module_title、设置列宽、行高
            for i in range(len(module_title)):
                ws.cell(1, i + 1).value = module_title[i]
                ws.cell(1, i + 1).alignment = alignment
                if i in [0, 1, 2, 3]:
                    ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 20
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 50
            ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高

    # 获取当前用例修改类型的sheet表的Max_row
    max_row = ws.max_row

    # 写report_data数据
    # if type(case_info) == list and len(case_info) == 5:
    for j in range(len(case_info)):
        if j == 4:
            step_number_action = ''
            step_number_result = ''
            for k in range(len(case_info[j])):
                step_number = list(case_info[j].keys())[k]
                # step_number_action += f'步骤{step_number}、{case_info[j][step_number][0]}\n'
                step_number_action += "步骤{}、{}\n".format(step_number, case_info[j][step_number][0])
                # step_number_result += f'步骤{step_number}、{case_info[j][step_number][1]}\n'
                step_number_result += "步骤{}、{}\n".format(step_number, case_info[j][step_number][1])
            ws.cell(max_row + 1, j + 1).value = step_number_action
            ws.cell(max_row + 1, j + 1).alignment = alignment
            ws.cell(max_row + 1, j + 2).value = step_number_result
            ws.cell(max_row + 1, j + 2).alignment = alignment
        else:
            ws.cell(max_row + 1, j + 1).value = str(case_info[j])
            ws.cell(max_row + 1, j + 1).alignment = alignment

    ws.row_dimensions[(max_row + 1)].height = 70  # 设置每次执行的report预约事件信息的行高
    wb.save(module_file_name)


def get_case_parent_suite_name(case_suite_id, suite_name=None):
    if suite_name is None:
        suite_name = []
    try:
        parent_suite_id = tlc.getTestSuiteByID(case_suite_id)['parent_id']
        if parent_suite_id != get_specify_project_id_by_name(exported_project_name):
            suite_name.insert(0, tlc.getTestSuiteByID(case_suite_id)['name'].strip())
            # print(parent_suite_id, 'parent_suite_id')
            get_case_parent_suite_name(parent_suite_id, suite_name=suite_name)
    except:
        return False
    else:
        return suite_name


def check_suite_and_case_in_suite(suite_id):
    suites = tlc.getTestSuitesForTestSuite(suite_id)    # 单个套件返回｛｝，多个套件返回｛‘id’:{},'id':{}｝
    cases = tlc.getTestCasesForTestSuite(testsuiteid=suite_id, deep=False, details="simple")    # 返回[{case1},{case2}]
    if len(cases) > 0:
        case_info = []
        case_suite_path = get_case_parent_suite_name(suite_id, suite_name=None)
        case_info.append(str(case_suite_path))
        # 分别处理当前套件下的所有的case
        for case in cases:
            test_case = tlc.getTestCase(case['id'])
            # print(f"{case['id']}:{len(test_case)}:{test_case}")
            # 由于每个test_case是一个list，且只有一个元素，所以test_case[0].get('name')就能获取到当前case的名称，其他信息都是如此
            print(test_case[0].get('name'))  # 也可以用print(test_case[0]['name'])
            # 每个case下的摘要、前提、步骤、期望的结果
            case_info.append(data_replace(test_case[0].get('name').strip()))
            case_info.append(data_replace(test_case[0].get('summary')))
            case_info.append(data_replace(test_case[0].get('preconditions')))
            step_action_result = {}
            for m in test_case[0].get("steps"):
                step_action_result[m.get("step_number")] = ['', '']
                # step_action_result[m.get("step_number")][0] += f"{data_replace(m.get('actions'))}"
                # step_action_result[m.get("step_number")][1] += f"{data_replace(m.get('expected_results'))}"
                step_action_result[m.get("step_number")][0] += "{}".format(data_replace(m.get('actions')))
                step_action_result[m.get("step_number")][1] += "{}".format(data_replace(m.get('expected_results')))
            # print(step_action_result)
            case_info.append(step_action_result)
            # print(case_info)
            write_test_case_to_excel(case_info)
            case_info = case_info[:1]
    if len(suites) > 0:
        if 'id' in suites.keys():   # 套件下只有一个套件存在时
            check_suite_and_case_in_suite(suites['id'])

        elif 'id' not in suites.keys():     # 套件下存在多个套件时
            for suite in suites:    # 默认情况下suite是suites的key值（‘id’）
                check_suite_and_case_in_suite(suites[suite]['id'])


if __name__ == '__main__':
    logging_info_setting()

    # url = 'http://git.nationalchip.com/testlink/lib/api/xmlrpc.php'
    url = "http://192.168.190.223/testlink/lib/api/xmlrpc.php"
    key = 'f5df4f1bd2bdd22403ec6b8b118d022c'
    tlc = testlink.TestlinkAPIClient(url, key)

    exported_project_name = 'ABS-S（四代机）'

    projects_info = get_all_project()
    for project in projects_info:
        print(project)

    specify_project_id = get_specify_project_id_by_name(exported_project_name)
    specify_project_first_suite_id = get_all_first_suites_id(specify_project_id)

    print("{}".format(100 * '%'))
    print(specify_project_first_suite_id)

    for id in specify_project_first_suite_id:
        n = 0
        suites_name_list = []
        suites_id_list = []
        module_name = specify_project_first_suite_id[id]
        module_file_name = '{}.xlsx'.format(module_name)
        check_suite_and_case_in_suite(id)
        # if id == '913421':
        #     check_suite_and_case_in_suite(id)
        # elif id == '917453':
        #     check_suite_and_case_in_suite(id)