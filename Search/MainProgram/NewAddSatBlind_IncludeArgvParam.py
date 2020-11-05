#!/usr/bin/python3
# -*- coding: utf-8 -*-


# voltage = { "0":"13V",
#             "1":"18V",
#             "2":"Off"
#             }
#
# 22k = { "0":"On",
#         "1":"Off"
#         }
#
# diseqc 1.0 = {  "0":"Off",
#                 "1":"Port1",
#                 "2":"Port2",
#                 "3":"Port3",
#                 "4":"Port4"
#                 }
#
# all_sat_commd = [
#                     choice_enter_antenna_mode,
#                     search_preparatory_work,
#                     sat_param_list,
#                     choice_search_mode,
#                     choice_save_type,
#                     choice_exit_mode,
#                     other_operate,
#                     normal_cycle_times,
#                     control_upper_limit_cycle_times,
#                 ]


from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from email.mime.text import MIMEText
from email.header import Header
import smtplib
import serial
import serial.tools.list_ports
import re
import time
import os
import sys
import random
import logging
import threading
import platform
import traceback


class CustomException(Exception):
    pass


class FailSendCmdException(CustomException):

    def __init__(self, info):
        self.info = info

    def __str__(self):
        return self.info


class MyGlobal(object):

    def __init__(self):
        self.search_start_state = False
        self.search_end_state = False
        self.all_tp_list = []                           # 用于存放搜索到的TP
        self.channel_info = {}                          # 用于存放各个TP下搜索到的电视和广播节目名称
        # self.search_datas = [0, 0, 0, 0, 0, 0, 0, 0, 0] # 用于存放xlsx_title中的数据
        self.search_datas = ['0', '0', '0', '0', '0', '0', '0', '0', []]  # 用于存放xlsx_title中的数据
        self.searched_time = 0                          # 用于记录搜索的轮次
        # [GL.tv_radio_tp_count[0],GL.tv_radio_tp_count[1],GL.tv_radio_tp_count[2],GL.tv_radio_tp_count[3],GL.tv_radio_tp_count[4]]
        self.tv_radio_tp_count = ['0', '0', '0', '0', '0']
        self.tv_radio_tp_accumulated = [[], [], [], []]     # 用于统计每轮搜索累加的TV、Radio、TP数以及保存TP数的值
        self.xlsx_data_interval = 0                     # 用于计算每轮搜索写xlsx时的间隔
        self.search_dur_time = ''                       # 用于存放搜索花费的时间
        self.send_loop_state = True
        self.receive_loop_state = True
        self.searched_sat_name = []                     # 用于保存搜索过程中搜索过的卫星的名称,便于搜索达到上限后,删除指定的卫星,不能被清空
        self.upper_limit_state = False                  # 用于控制搜索达到上限的时其他操作的状态变量，false不执行，true时执行
        self.random_choice_sat = []                     # 用于存放搜索达到上限后每次随机选择的卫星,然后进行删除其TP
        self.delete_ch_finish_state = False             # 用于删除所有节目成功的状态变量
        # self.save_ch_finish_state = False               # 用于保存节目结束的状态变量
        self.record_maximum_data = []                   # 用于存放达到上限时的打印
        self.start_record_maximum_state = False         # 用于开始记录达到上限时的状态变量
        self.upper_limit_send_ok_commd_state = False    # 用于控制搜索达到上限后是否发送OK命令的状态变量
        self.delete_ch_finish_kws = "[PTD]All programs deleted successfully"  # 删除所有节目成功关键字
        self.infrared_rsv_kws = "[PTD]Infrared_key_values:"     # 获取红外接收关键字
        self.antenna_setting_kws = "[PTD]Antenna_setting:"      # 获取天线与卫星设置界面焦点位置关键字
        self.antenna_setting_focus_pos = ''                     # 天线与卫星设置界面焦点位置
        self.infrared_send_commd = []                           # 所有红外发送命令列表
        self.infrared_rsv_commd = []                            # 所有红外接收命令列表
        self.receive_cmd_list = []                              # 红外接收公版遥控器命令列表
        self.sat_param_save = ["", "", "", "", "", ""]      # [sat_name,LNB_Power,LNB_Fre,22k,diseqc1.0,diseqc1.1]
        self.sat_param_kws = [
                                    "[PTD]sat_name=",
                                    "[PTD]LNB1=",
                                    "[PTD]--[0:ON,1:OFF]---22K",
                                    "[PTD]--set diseqc 1.0",
                                    "[PTD]--set diseqc 1.1",
                                ]

        self.search_monitor_kws = [
                                    "[PTD]SearchStart",		    # 0
                                    "[PTD]TV------",		    # 1
                                    "[PTD]Radio-----",		    # 2
                                    "[PTD]SearchFinish",        # 3
                                    "[PTD]get :  fre",		    # 4
                                    "[PTD]TP_save=",		    # 5
                                    "[PTD]TV_save=",		    # 6
                                    "[PTD]maximum_tp",		    # 7
                                    "[PTD]maximum_channel",     # 8
                                    "[PTD]get polar:",          # 9
                                    ]

        self.all_sat_commd = [
                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["Chinas6b_C", "Polar=0", "5150/0", "22K=1", "2", "0", "Blind"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["Chinas6b_C", "Polar=0", "5150/0", "22K=1", "2", "0", "SuperBlind"],
                                        CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["Asiasat 7 C", "Polar=0", "5150/5750", "22K=1", "1", "0", "Blind"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["Asiasat 7 C", "Polar=0", "5150/5750", "22K=1", "1", "0", "SuperBlind"],
                                        CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "Blind"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "SuperBlind"],
                                        CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["ST 2 K", "Polar=0", "10600/0", "22K=1", "0", "0", "Blind"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["ST 2 K", "Polar=0", "10600/0", "22K=1", "0", "0", "SuperBlind"],
                                        CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["PLPD", "Polar=0", "5150/5750", "22K=0", "1", "0", "Blind"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
                                        ["PLPD", "Polar=0", "5150/5750", "22K=0", "1", "0", "SuperBlind"],
                                        CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "Incremental"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
                                        INCREMENTAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "ChUL"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, DELETE_ALL_CH,
                                        UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_CYCLE_TIMES],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "ChUL_Cont."],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT,
                                        UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "ChUL_DelTp_Cont."],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
                                        EXIT_ANTENNA_SETTING, DELETE_SPECIFY_SAT_ALL_TP,
                                        UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Chinas6b_C", "Polar=0", "5150/0", "22K=1", "2", "0", "TpUL"],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[1],
                                        EXIT_ANTENNA_SETTING, RESET_FACTORY,
                                        UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_CYCLE_TIMES],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Chinas6b_C", "Polar=0", "5150/0", "22K=1", "2", "0", "TpUL_Cont."],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[1],
                                        EXIT_ANTENNA_SETTING, UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT,
                                        UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

                                    [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
                                        ["Chinas6b_C", "Polar=0", "5150/0", "22K=1", "2", "0", "TpUL_DelTp_Cont."],
                                        CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[1],
                                        EXIT_ANTENNA_SETTING, DELETE_SPECIFY_SAT_ALL_TP,
                                        UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

                                    [RESET_FACTORY, SEARCH_PREPARATORY_WORK[0],
                                        ["Reset", "Factory"], ONLY_EXECUTE_ONE_TIME],

                                    [DELETE_ALL_CH, SEARCH_PREPARATORY_WORK[0],
                                        ["Delete", "AllCH"], ONLY_EXECUTE_ONE_TIME],
                                ]


def check_ports():
    # global send_com, receive_com
    send_com, receive_com = '', ''
    send_port_desc, receive_port_desc = '', ''
    serial_board = "FT232R USB UART"
    connection_serial_board_state = False
    ports_info = []
    ports = list(serial.tools.list_ports.comports())
    for i in range(len(ports)):
        logging.info("可用端口:名称:{} + 描述:{} + 硬件id:{}".format(ports[i].device, ports[i].description, ports[i].hwid))
        # print("可用端口:名称:{} + 描述:{} + 硬件id:{}".format(ports[i].device, ports[i].description, ports[i].hwid))
        ports_info.append("{}~{}~{}".format(ports[i].device, ports[i].description, ports[i].hwid))

    ser_cable_num = 7

    if platform.system() == "Windows":
        serial_ser = {
            "1": "FTDVKA2HA",
            "2": "FTGDWJ64A",
            "3": "FT9SP964A",
            "4": "FTHB6SSTA",
            "5": "FTDVKPRSA",
            "6": "FTHI8UIHA",
            "7": "FTHG05TTA",
        }
        send_port_desc = "USB-SERIAL CH340"
        receive_port_desc = serial_ser[str(ser_cable_num)]
    elif platform.system() == "Linux":
        serial_ser = {
            "1": "FTDVKA2H",
            "2": "FTGDWJ64",
            "3": "FT9SP964",
            "4": "FTHB6SST",
            "5": "FTDVKPRS",
            "6": "FTHI8UIH",
            "7": "FTHG05TT",
        }
        send_port_desc = "USB2.0-Serial"
        for i in range(len(ports_info)):
            if serial_board in ports_info[i]:
                connection_serial_board_state = True
        if connection_serial_board_state:
            receive_port_desc = "FT232R USB UART"
        else:
            receive_port_desc = serial_ser[str(ser_cable_num)]
    if len(ports) <= 0:
        logging.info("无可用端口")
    elif len(ports) == 1:
        logging.info("只有一个可用端口:{}".format(ports[0].device))
    elif len(ports) >= 2:
        for i in range(len(ports_info)):
            if send_port_desc in ports_info[i]:
                send_com = ports_info[i].split("~")[0]
            elif receive_port_desc in ports_info[i]:
                receive_com = ports_info[i].split("~")[0]
    return send_com, receive_com


def serial_set(ser, ser_name, ser_baudrate):
    ser.port = ser_name
    ser.baudrate = ser_baudrate
    ser.bytesize = 8
    ser.parity = "N"
    ser.stopbits = 1
    ser.timeout = 1
    ser.write_timeout = 0
    ser.open()


def hex_strs_to_bytes(strings):
    # strs = strings.replace(" ", "")
    return bytes.fromhex(strings)


def send_commd(commd):
    # 红外发送端发送指令
    continuous_transmission_cmd_num = 0
    send_ser.write(hex_strs_to_bytes(commd))
    send_ser.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
    if REVERSE_KEY[commd] != "POWER":
        GL.infrared_send_commd.append(REVERSE_KEY[commd])
    time.sleep(1.0)
    if len(GL.infrared_send_commd) == len(GL.receive_cmd_list):
        pass
    elif len(GL.infrared_send_commd) != len(GL.receive_cmd_list):
        logging.info("检测到发送和接收命令数不一致，等待2秒，查看是否接收端还没有接收到打印")
        time.sleep(2)
        while True:
            if len(GL.infrared_send_commd) == len(GL.receive_cmd_list):
                break
            elif len(GL.infrared_send_commd) != len(GL.receive_cmd_list):
                logging.info(f"此刻补发STB没有接收到的红外命令{GL.infrared_send_commd[-1]}")
                send_ser.write(hex_strs_to_bytes(KEY[GL.infrared_send_commd[-1]]))
                send_ser.flush()
                time.sleep(1.0)
                continuous_transmission_cmd_num += 1
                if continuous_transmission_cmd_num == 10:
                    stb_crash_msg = "STB一直发送指令，疑似死机"
                    # mail(f'{stb_crash_msg}\n\n{msg}')
                    raise FailSendCmdException(stb_crash_msg)


def add_write_data_to_txt(file_path,write_data):    # 追加写文本
    with open(file_path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def build_print_log_and_report_file_path():
    global sat_name, search_mode, sheet_name
    global report_file_path, case_log_txt_path
    # 设计测试数据的目录
    parent_path = os.path.dirname(os.getcwd())
    test_file_folder_name = "test_data"
    test_file_directory = os.path.join(parent_path, test_file_folder_name)
    case_log_folder_name = "print_log"
    case_log_file_directory = os.path.join(parent_path, test_file_folder_name, case_log_folder_name)
    report_folder_name = "report"
    report_file_directory = os.path.join(parent_path, test_file_folder_name, report_folder_name)
    # 检查是否存在测试数据的目录，没有就创建
    if not os.path.exists(test_file_directory):
        os.mkdir(test_file_directory)
    if not os.path.exists(case_log_file_directory):
        os.mkdir(case_log_file_directory)
    if not os.path.exists(report_file_directory):
        os.mkdir(report_file_directory)
    # 设计打印和报告文件的完整路径
    sat_name = GL.all_sat_commd[choice_search_sat][2][0]
    search_mode = GL.all_sat_commd[choice_search_sat][2][-1]
    timestamp = re.sub(r'[-: ]', '_', str(datetime.now())[:19])
    sheet_name = "{}_{}".format(sat_name, search_mode)

    report_file_name = "{}_{}_{}_{}_Result_{}.xlsx".format(
        choice_search_sat, simplify_sat_name[sat_name], sat_name, search_mode, timestamp)
    report_file_path = os.path.join(report_file_directory, report_file_name)

    case_log_file_name = "{}_{}_{}_{}_{}.txt".format(
        choice_search_sat, simplify_sat_name[sat_name], sat_name, search_mode, timestamp)
    case_log_txt_path = os.path.join(case_log_file_directory, case_log_file_name)


def judge_write_file_exist():
    global case_log_file_directory, report_file_directory
    parent_path = os.path.dirname(os.getcwd())
    test_file_folder_name = "test_data"
    test_file_directory = os.path.join(parent_path, test_file_folder_name)
    case_log_folder_name = "print_log"
    case_log_file_directory = os.path.join(parent_path, test_file_folder_name, case_log_folder_name)
    report_folder_name = "report"
    report_file_directory = os.path.join(parent_path, test_file_folder_name, report_folder_name)

    if not os.path.exists(test_file_directory):
        os.mkdir(test_file_directory)
    if not os.path.exists(case_log_file_directory):
        os.mkdir(case_log_file_directory)
    if not os.path.exists(report_file_directory):
        os.mkdir(report_file_directory)


def judge_and_wirte_data_to_xlsx():
    ws = ''
    wb = ''
    GL.xlsx_data_interval = 1 + 5 * (GL.searched_time - 1)
    global xlsx_title
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    if not os.path.exists(report_file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 11
        for i in range(len(xlsx_title)):
            if i < len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = list(xlsx_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 1:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment

    elif os.path.exists(report_file_path):
        wb = load_workbook(report_file_path)
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 11
        for i in range(len(xlsx_title)):
            if i < len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = list(xlsx_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 1:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment

    tp_column_numb = column_index_from_string("A") + GL.xlsx_data_interval
    all_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 1
    tv_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 2
    radio_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 3
    tp_column_char = get_column_letter(tp_column_numb)
    all_column_char = get_column_letter(all_column_numb)
    tv_column_char = get_column_letter(tv_column_numb)
    radio_column_char = get_column_letter(radio_column_numb)
    ws.column_dimensions[tp_column_char].width = 12
    ws.column_dimensions[all_column_char].width = 3
    ws.column_dimensions[tv_column_char].width = 3
    ws.column_dimensions[radio_column_char].width = 3

    for m in range(len(GL.search_datas)):
        if m < len(GL.search_datas) - 2:
            ws.cell((m + 1), (1 + GL.xlsx_data_interval)).value = GL.search_datas[m]
            ws.merge_cells(start_row=(m + 1), start_column=(1 + GL.xlsx_data_interval),
                           end_row=(m + 1), end_column=(1 + GL.xlsx_data_interval + 4))
            ws.cell((m + 1), (1 + GL.xlsx_data_interval)).alignment = alignment
        elif m == len(GL.search_datas) - 2:
            for n in range(len(xlsx_title[7]["数据类别"])):
                ws.cell((m + 1), (1 + GL.xlsx_data_interval + n)).value = list(xlsx_title[m].values())[0][n]
                ws.cell((m + 1), (1 + GL.xlsx_data_interval + n)).alignment = alignment
                ws.row_dimensions[(m+1)].height = 13.5
        elif m == len(GL.search_datas) - 1:
            for j in range(len(GL.all_tp_list)):
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval)).value = GL.search_datas[m][j]
                # ws.cell((m+1+j),(1+GL.xlsx_data_interval)+1).value = len(GL.channel_info[str(j+1)][0]) + \
                #                                                      len(GL.channel_info[str(j+1)][1])
                # ws.cell((m+1+j),(1+GL.xlsx_data_interval)+2).value = len(GL.channel_info[str(j+1)][0])
                # ws.cell((m+1+j),(1+GL.xlsx_data_interval)+3).value = len(GL.channel_info[str(j+1)][1])
                # ws.cell((m+1+j),(1+GL.xlsx_data_interval)+4).value = ",".join(GL.channel_info[str(j+1)][0] + \
                #                                                               GL.channel_info[str(j+1)][1])

                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 1).value = len(
                    GL.channel_info[GL.search_datas[m][j]][0]) + len(GL.channel_info[GL.search_datas[m][j]][1])
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 2).value = len(
                    GL.channel_info[GL.search_datas[m][j]][0])
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 3).value = len(
                    GL.channel_info[GL.search_datas[m][j]][1])
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 4).value = ",".join(
                    GL.channel_info[GL.search_datas[m][j]][0] + GL.channel_info[GL.search_datas[m][j]][1])

                for k in range(len(xlsx_title[7]["数据类别"])):
                    ws.cell((m+1+j), (1+GL.xlsx_data_interval)+k).alignment = alignment
                ws.row_dimensions[(m+1+j)].height = 13.5
    wb.save(report_file_path)


def enter_antenna_setting():
    logging.debug("Enter Antenna Setting")
    GL.sat_param_save[0] = ''
    send_commd(KEY["MENU"])
    send_commd(KEY["OK"])
    time.sleep(1)       # 等待进入天线卫星设置界面，且获取到卫星名称和焦点位置
    while GL.sat_param_save[0] == '':
        logging.info("没有正确进入天线设置界面，重新进入")
        send_commd(KEY["EXIT"])
        send_commd(KEY["EXIT"])
        send_commd(KEY["EXIT"])
        time.sleep(0.5)
        send_commd(KEY["MENU"])
        send_commd(KEY["OK"])
        time.sleep(1)


def judge_preparatory_work():
    if len(GL.all_sat_commd[choice_search_sat][1]) == 0:
        logging.debug("Not Search Preparatory Work")
    elif len(GL.all_sat_commd[choice_search_sat][1]) > 0:
        logging.debug("Search Preparatory Work")
        send_data_1 = GL.all_sat_commd[choice_search_sat][1][0]
        send_data_2 = GL.all_sat_commd[choice_search_sat][1][1]
        for i in range(len(send_data_1)):
            send_commd(send_data_1[i])
        logging.info("等待删除卫星保存结束6秒")
        time.sleep(6)  # 等待删除卫星保存结束
        for j in range(len(send_data_2)):
            send_commd(send_data_2[j])


def check_satellite_param():
    logging.debug("Satellite")
    time.sleep(1)
    while GL.antenna_setting_focus_pos != "Satellite":
        send_commd(KEY["DOWN"])
    else:
        if GL.all_sat_commd[choice_search_sat][1] == SEARCH_PREPARATORY_WORK[0]:  # upper limit or incremental search
            if len(GL.searched_sat_name) == 72:  # 避免程序还没有执行结束，但是搜索的卫星个数满了导致的死循环
                for i in range(len(GL.searched_sat_name) // 2):
                    GL.searched_sat_name.remove(random.choice(GL.searched_sat_name))
            while GL.sat_param_save[0] in GL.searched_sat_name:
                logging.info("sat in list")
                logging.info("{},{}".format(GL.sat_param_save[0], GL.searched_sat_name))
                send_commd(KEY["RIGHT"])
            else:
                logging.info("{},{}".format(GL.sat_param_save[0], GL.searched_sat_name))
                logging.info("sat not in list")
                GL.searched_sat_name.append(GL.sat_param_save[0])
                logging.info("{},{}".format(GL.sat_param_save[0], GL.searched_sat_name))
                send_commd(KEY["DOWN"])

        elif GL.all_sat_commd[choice_search_sat][1] == SEARCH_PREPARATORY_WORK[1]:  # normal sat search
            send_commd(KEY["DOWN"])


def check_lnb_power():
    logging.debug("LNB POWER")
    while GL.antenna_setting_focus_pos != "LNB Power":
        send_commd(KEY["DOWN"])
    else:
        power_off = "Polar=2"
        while GL.sat_param_save[1] != power_off:
            send_commd(KEY["LEFT"])
        else:
            send_commd(KEY["RIGHT"])
            send_commd(KEY["DOWN"])


def check_lnb_fre():
    logging.debug("LBN FREQUENCY")
    while GL.antenna_setting_focus_pos != "LNB Frequency":
        send_commd(KEY["DOWN"])
    else:
        logging.info(GL.sat_param_save)
        while GL.sat_param_save[2] != GL.all_sat_commd[choice_search_sat][2][2]:
            send_commd(KEY["RIGHT"])
        else:
            send_commd(KEY["DOWN"])


def check_22k():
    logging.debug("22k")
    while GL.antenna_setting_focus_pos != "22K":
        send_commd(KEY["DOWN"])
    else:
        while GL.sat_param_save[3] != GL.all_sat_commd[choice_search_sat][2][3]:
            send_commd(KEY["RIGHT"])
        else:
            send_commd(KEY["DOWN"])


def check_diseqc_10():
    logging.debug("Diseqc 1.0")
    while GL.antenna_setting_focus_pos != "DiSEqC 1,0":
        send_commd(KEY["DOWN"])
    else:
        while GL.sat_param_save[4] != GL.all_sat_commd[choice_search_sat][2][4]:
            send_commd(KEY["LEFT"])
        else:
            send_commd(KEY["DOWN"])


def check_diseqc_11():
    logging.debug("Diseqc 1.1")
    while GL.antenna_setting_focus_pos != "DiSEqC 1,1":
        send_commd(KEY["DOWN"])
    else:
        while GL.sat_param_save[5] != GL.all_sat_commd[choice_search_sat][2][5]:
            send_commd(KEY["LEFT"])
        else:
            send_commd(KEY["DOWN"])


def check_tp():
    logging.debug("TP")
    while GL.antenna_setting_focus_pos != "TP":
        send_commd(KEY["DOWN"])
    else:
        send_commd(KEY["DOWN"])


def choice_srh_mode_and_start_srh():
    logging.debug("Choice Search Mode And Start Search")
    while GL.antenna_setting_focus_pos != "Start Search":
        send_commd(KEY["DOWN"])
    else:
        # send_data = GL.all_sat_commd[choice_search_sat][3]
        if GL.all_sat_commd[choice_search_sat][3] == CHOICE_BLIND_MODE:
            send_commd(KEY["RIGHT"])
            send_commd(KEY["OK"])
            send_commd(KEY["OK"])
        elif GL.all_sat_commd[choice_search_sat][3] == CHOICE_SUPERBLIND_MODE:
            send_commd(KEY["BLUE"])
            send_commd(KEY["RIGHT"])
            send_commd(KEY["OK"])
            send_commd(KEY["OK"])

        time.sleep(1)
        while not GL.search_start_state:
            send_commd(KEY["EXIT"])
            if GL.all_sat_commd[choice_search_sat][3] == CHOICE_BLIND_MODE:
                send_commd(KEY["RIGHT"])
                send_commd(KEY["OK"])
                send_commd(KEY["OK"])
            elif GL.all_sat_commd[choice_search_sat][3] == CHOICE_SUPERBLIND_MODE:
                send_commd(KEY["BLUE"])
                send_commd(KEY["RIGHT"])
                send_commd(KEY["OK"])
                send_commd(KEY["OK"])
            time.sleep(1)


def antenna_setting():
    check_satellite_param()
    check_lnb_power()
    check_lnb_fre()
    check_22k()
    check_diseqc_10()
    check_diseqc_11()
    check_tp()
    choice_srh_mode_and_start_srh()


def block_send_thread():
    time.sleep(1)
    # send_ser.send_break(3)


def judge_srh_limit():
    global search_time
    logging.debug("Upper Limit To Save Channel Stage")
    if not GL.upper_limit_state:
        logging.debug("Not Upper Limit")
    elif GL.upper_limit_state:
        if GL.all_sat_commd[choice_search_sat][8] == NOT_UPPER_LIMIT_LATER_SEARCH_TIME:     # 普通搜索
            logging.info("普通搜索，但是达到上限")
        else:   # 上限搜索
            logging.debug("Upper Limit")
            logging.debug("打印搜索达到上限是否有新增节目的记录列表:{}".format(GL.record_maximum_data))
            search_time = 72
            GL.all_sat_commd[choice_search_sat][8] -= 1
            logging.info("搜索到上限剩余次数:{}".format(GL.all_sat_commd[choice_search_sat][8]))
            for i in range(len(GL.record_maximum_data)):
                if GL.search_monitor_kws[6] in GL.record_maximum_data[i]:  # "[PTD]TV_save="
                    GL.upper_limit_send_ok_commd_state = True
            if GL.upper_limit_send_ok_commd_state:
                logging.debug("搜索达到上限但是没有新增节目")
            elif not GL.upper_limit_send_ok_commd_state:
                logging.debug("搜索达到上限但是有新增节目")
                send_commd(KEY["OK"])


def judge_save_ch_mode():
    logging.debug("Whether Or Not Save And End Search")
    logging.info("搜索到上限剩余次数:{}".format(GL.all_sat_commd[choice_search_sat][8]))
    send_data = GL.all_sat_commd[choice_search_sat][4]
    for i in range(len(send_data)):
        send_commd(send_data[i])
    if GL.all_sat_commd[choice_search_sat][4] == CHOICE_SAVE_TYPE[0]:
        logging.info("主动在保存节目时延时3秒")
        time.sleep(3)


def write_data_to_excel():
    logging.debug("Write data to Excel")
    logging.info("保存节目后等待保存TP和保存节目的打印5秒")
    time.sleep(5)
    GL.search_datas[0] = sheet_name
    GL.search_datas[2] = str(len(GL.all_tp_list))
    GL.search_datas[3] = "{}/{}".format(GL.tv_radio_tp_count[0], GL.tv_radio_tp_count[1])
    GL.search_datas[6] = GL.search_dur_time
    GL.search_datas[8] = GL.all_tp_list
    judge_and_wirte_data_to_xlsx()


def clear_variate():
    logging.debug("clear data")
    # 处理循环数据
    GL.record_maximum_data.clear()
    GL.all_tp_list.clear()
    GL.channel_info.clear()
    GL.tv_radio_tp_count = ['0', '0', '0', '0', '0']
    GL.search_datas[2] = '0'
    GL.search_datas[3] = '0/0'
    GL.search_datas[4] = '0'
    GL.search_datas[5] = '0/0'
    GL.search_datas[6] = '0'
    GL.search_datas[8].clear()


def exit_antenna_setting():
    logging.debug("Exit Antenna Setting")
    send_data = GL.all_sat_commd[choice_search_sat][5]
    for i in range(len(send_data)):
        send_commd(send_data[i])


def other_operate_del_all_ch():
    logging.debug("Delete All Channels And Choice Searched First Sat")
    # 执行删除所有节目的命令
    send_data = GL.all_sat_commd[choice_search_sat][6]
    for i in range(len(send_data)):
        send_commd(send_data[i])
    # 等待节目删除完成后返回成功标志
    logging.info("等待所有节目删除完成")
    n = 20
    while True:
        if not GL.delete_ch_finish_state:
            logging.info("还没有删除完成，请等待")
            time.sleep(1)
            n -= 1
            if n == 0:      # 假如20秒后还没有检查到删除成功的标志就重新删除
                send_commd(KEY["EXIT"])
                send_commd(KEY["EXIT"])
                send_commd(KEY["EXIT"])
                send_data = GL.all_sat_commd[choice_search_sat][6]
                for i in range(len(send_data)):
                    send_commd(send_data[i])
                n = 20
        elif GL.delete_ch_finish_state:
            logging.info("删除完成")
            break
    # 进入天线设置界面，并切换到第一个卫星
    send_commd(KEY["EXIT"])
    send_commd(KEY["EXIT"])
    send_commd(KEY["EXIT"])
    # send_commd(KEY["MENU"])
    # send_commd(KEY["OK"])
    enter_antenna_setting()
    first_sat_name = GL.searched_sat_name[0]
    while GL.sat_param_save[0] != first_sat_name:
        send_commd(KEY["LEFT"])
    # 退回大画面
    for i in range(len(EXIT_TO_SCREEN)):
        send_commd(EXIT_TO_SCREEN[i])
    GL.searched_sat_name.clear()


def other_operate_del_specify_sat_all_tp():
    logging.debug("Delete Specify Sat TP And Choice Random Sat")
    # send_commd(KEY["MENU"])
    # send_commd(KEY["OK"])
    enter_antenna_setting()
    GL.random_choice_sat.append(random.choice(GL.searched_sat_name))
    while GL.sat_param_save[0] != GL.random_choice_sat[0]:
        # if PRESET_SAT_NAME.index(GL.sat_param_save[0]) > PRESET_SAT_NAME.index(GL.random_choice_sat[0]):
        #     send_commd(KEY["LEFT"])
        # elif PRESET_SAT_NAME.index(GL.sat_param_save[0]) < PRESET_SAT_NAME.index(GL.random_choice_sat[0]):
        #     send_commd(KEY["RIGHT"])

        logging.info(f'当前卫星为：{GL.sat_param_save[0]}，预期卫星为：{GL.random_choice_sat[0]}')
        cur_sat_pos = PRESET_SAT_NAME.index(GL.sat_param_save[0])
        choice_sat_pos = PRESET_SAT_NAME.index(GL.random_choice_sat[0])
        logging.info(f"当前卫星的位置为：{cur_sat_pos}，预期卫星的位置为：{choice_sat_pos}")
        if cur_sat_pos > choice_sat_pos:
            left_move_steps = cur_sat_pos - choice_sat_pos
            right_move_steps = choice_sat_pos + (len(PRESET_SAT_NAME) - cur_sat_pos)
            logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
            if left_move_steps > right_move_steps:
                logging.info("应该向右移动")
                send_commd(KEY["RIGHT"])
            elif left_move_steps < right_move_steps:
                logging.info("应该向左移动")
                send_commd(KEY["LEFT"])
            elif left_move_steps == right_move_steps:
                logging.info("向左或向右移动距离相等")
                send_commd(KEY["RIGHT"])
        elif cur_sat_pos < choice_sat_pos:
            left_move_steps = cur_sat_pos + (len(PRESET_SAT_NAME) - choice_sat_pos)
            right_move_steps = choice_sat_pos - cur_sat_pos
            logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
            if left_move_steps > right_move_steps:
                logging.info("应该向右移动")
                send_commd(KEY["RIGHT"])
            elif left_move_steps < right_move_steps:
                logging.info("应该向左移动")
                send_commd(KEY["LEFT"])
            elif left_move_steps == right_move_steps:
                logging.info("向左或向右移动距离相等")
                send_commd(KEY["RIGHT"])
    logging.info("{},{},{}".format(GL.sat_param_save[0], GL.random_choice_sat[0], GL.searched_sat_name))
    GL.searched_sat_name.remove(GL.random_choice_sat[0])  # 避免搜索时该卫星在已搜索的卫星列表中，不能进行搜索
    GL.random_choice_sat.clear()
    send_data = DELETE_SPECIFY_SAT_ALL_TP
    for i in range(len(send_data)):
        send_commd(send_data[i])
    logging.info("等待删除指定卫星下的所有TP10秒")
    time.sleep(10)
    send_data = EXIT_TO_SCREEN
    for j in range(len(send_data)):
        send_commd(send_data[j])


def judge_other_operate():
    if len(GL.all_sat_commd[choice_search_sat][6]) == 0:  # 没有额外操作
        logging.debug("Not Other Operate")
    elif len(GL.all_sat_commd[choice_search_sat][6]) > 0:  # 有额外操作
        logging.debug("Exist Other Operate")
        if not GL.upper_limit_state:
            logging.debug("Exist Other Operate But Not Upper Limit")
        elif GL.upper_limit_state:
            if GL.all_sat_commd[choice_search_sat][8] < 0:  # 搜索的次数到最后一次时不再进行额外的操作
                logging.info("搜索的次数到最后一次时不再进行额外的操作")
            else:
                if GL.all_sat_commd[choice_search_sat][6] == RESET_FACTORY:
                    logging.debug("Reset Factory")
                    send_data = GL.all_sat_commd[choice_search_sat][6]
                    for i in range(len(send_data)):
                        send_commd(send_data[i])
                    logging.info("等待恢复出厂设置重启30秒")
                    time.sleep(30)
                    GL.searched_sat_name.clear()
                elif GL.all_sat_commd[choice_search_sat][6] == DELETE_ALL_CH:
                    other_operate_del_all_ch()
                elif GL.all_sat_commd[choice_search_sat][6] == DELETE_SPECIFY_SAT_ALL_TP:
                    other_operate_del_specify_sat_all_tp()
                elif GL.all_sat_commd[choice_search_sat][6] == UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT:
                    logging.debug("Not Delete Specify Sat Tp And Search Continue")
                    send_commd(UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT[0])
                    GL.searched_sat_name.remove(random.choice(GL.searched_sat_name))  # 达到上限后切下一个卫星搜索
                    # GL.searched_sat_name.clear()        # 达到上限后重复搜索最后一个卫星


def cyclic_srh_setting():
    global search_time
    logging.debug("Cyclic Search Setting")
    if GL.all_sat_commd[choice_search_sat][8] == NOT_UPPER_LIMIT_LATER_SEARCH_TIME:
        GL.upper_limit_state = False  # 恢复默认状态
        GL.sat_param_save = ["", "", "", "", "", ""]  # 获取卫星的参数保存数据恢复默认状态
        GL.delete_ch_finish_state = False  # 删除所有节目成功状态恢复默认
        # GL.save_ch_finish_state = False  # 保存节目成功状态恢复默认
        GL.search_end_state = False  # 搜索结束状态恢复默认
        GL.search_start_state = False   # 搜索开始状态恢复默认
        GL.upper_limit_send_ok_commd_state = False  # 搜索达到上限后是否发送OK命令的状态变量恢复默认

        search_time -= 1
        logging.info("进入下一次循环搜索等待5秒")
        time.sleep(5)
        logging.info("剩余搜索次数:{}".format(search_time))
        if search_time < 1:
            logging.info("程序结束")
            GL.send_loop_state = False
            GL.receive_loop_state = False

    elif GL.all_sat_commd[choice_search_sat][8] != NOT_UPPER_LIMIT_LATER_SEARCH_TIME:
        GL.upper_limit_state = False  # 恢复默认状态
        GL.sat_param_save = ["", "", "", "", "", ""]  # 获取卫星的参数保存数据恢复默认状态
        GL.delete_ch_finish_state = False  # 删除所有节目成功状态恢复默认
        # GL.save_ch_finish_state = False  # 保存节目成功状态恢复默认
        GL.search_end_state = False  # 搜索结束状态恢复默认
        GL.search_start_state = False  # 搜索开始状态恢复默认
        GL.upper_limit_send_ok_commd_state = False  # 搜索达到上限后是否发送OK命令的状态变量恢复默认

        logging.info("进入下一次循环搜索等待5秒")
        time.sleep(5)
        logging.info("搜索到上限剩余次数:{}".format(GL.all_sat_commd[choice_search_sat][8]))
        # GL.all_sat_commd[choice_search_sat][8] -= 1
        if GL.all_sat_commd[choice_search_sat][8] < 0:
            logging.info("程序结束")
            GL.send_loop_state = False
            GL.receive_loop_state = False


def data_send_thread():
    global search_time
    # 执行单次运行的场景
    if len(GL.all_sat_commd[choice_search_sat]) < 9:
        send_data = GL.all_sat_commd[choice_search_sat][0]
        for i in range(len(send_data)):
            send_commd(send_data[i])
        search_time -= 1
        if search_time < 1:
            logging.info("单次执行恢复出厂设置等待30秒")
            time.sleep(30)
            GL.send_loop_state = False
            GL.receive_loop_state = False

    # 执行多次运行的场景
    elif len(GL.all_sat_commd[choice_search_sat]) == 9:
        if GL.all_sat_commd[choice_search_sat][-1] == NOT_UPPER_LIMIT_LATER_SEARCH_TIME:    # 普通搜索
            while search_time > 0:
                enter_antenna_setting()
                judge_preparatory_work()
                antenna_setting()
                while True:
                    if GL.search_end_state:
                        judge_srh_limit()
                        judge_save_ch_mode()
                        write_data_to_excel()
                        clear_variate()
                        exit_antenna_setting()
                        judge_other_operate()
                        cyclic_srh_setting()
                        break
                    elif not GL.search_end_state:
                        block_send_thread()

        else:   # 上限搜索
            while GL.all_sat_commd[choice_search_sat][-1] >= 0:
                enter_antenna_setting()
                judge_preparatory_work()
                antenna_setting()
                while True:
                    if GL.search_end_state:
                        judge_srh_limit()
                        judge_save_ch_mode()
                        write_data_to_excel()
                        clear_variate()
                        exit_antenna_setting()
                        judge_other_operate()
                        cyclic_srh_setting()
                        break
                    elif not GL.search_end_state:
                        block_send_thread()


def mail(message):
    my_sender = 'wangrun@nationalchip.com'  # 发件人邮箱账号
    my_pass = 'b8iNRgDiPUfkUVLW'  # 发件人邮箱密码
    my_user = 'wangrun@nationalchip.com'  # 收件人邮箱账号，我这边发送给自己

    return_state = True
    try:
        msg = MIMEText(message, 'plain', 'utf-8')
        # msg['From'] = formataddr(["FromRunoob", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['From'] = Header("Auto_test", 'utf-8')
        msg['To'] = Header("ME", 'utf-8')
        # msg['To'] = formataddr(["FK", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
        msg['Subject'] = "自动化测试终止提醒"  # 邮件的主题，也可以说是标题

        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 关闭连接
    except smtplib.SMTPException:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
        return_state = False
    return return_state


def data_receiver_thread():
    global start_time, end_time
    tp = ''
    while GL.receive_loop_state:
        data = receive_ser.readline()
        if data:
            tt = datetime.now()
            data1 = data.decode("GB18030", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            add_write_data_to_txt(case_log_txt_path, data3)

            if GL.start_record_maximum_state:
                GL.record_maximum_data.append(data2)
                logging.debug(GL.record_maximum_data)

            if GL.infrared_rsv_kws in data2:    # 红外接收打印
                infrared_rsv_commd = re.split(":", data2)[-1]
                GL.infrared_rsv_commd.append(infrared_rsv_commd)
                if infrared_rsv_commd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(infrared_rsv_commd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        GL.infrared_send_commd[-1], reverse_rsv_key[GL.infrared_rsv_commd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(GL.infrared_send_commd), len(GL.infrared_rsv_commd)))
                    GL.receive_cmd_list.append(infrared_rsv_commd)

            if GL.antenna_setting_kws in data2:     # 天线设置界面获取焦点位置
                GL.antenna_setting_focus_pos = re.split(":", data2)[-1]

            if GL.sat_param_kws[0] in data2:  # 判断卫星名称
                GL.sat_param_save[0] = re.split("=", data2)[-1]

            if GL.sat_param_kws[1] in data2:  # 判断LNB Fre
                lnb_split = re.split(r"[],]", data2)
                lnb1 = lnb_split[1].split("=")[-1]
                lnb2 = lnb_split[2].split("=")[-1]
                GL.sat_param_save[2] = "{}/{}".format(lnb1, lnb2)

            if GL.sat_param_kws[2] in data2:  # 判断22k
                GL.sat_param_save[3] = list(filter(None, re.split("-{2,}|,", data2)))[-1].strip()

            if GL.sat_param_kws[3] in data2:  # 判断diseqc 1.0和Polar(LNB Power)
                polar_split = re.split(r"[,\]-]", data2)
                GL.sat_param_save[4] = polar_split[3].split("=")[-1]
                GL.sat_param_save[1] = polar_split[7].strip()

            if GL.sat_param_kws[4] in data2:  # 判断diseqc 1.1
                disqc1_1_info_split = re.split(r"-", data2)
                GL.sat_param_save[5] = disqc1_1_info_split[-1].split("=")[-1]

            if GL.search_monitor_kws[0] in data2:  # 监控搜索起始
                GL.search_start_state = True
                start_time = datetime.now()
                GL.searched_time += 1
                # GL.xlsx_data_interval = 1 + 5 * (GL.searched_time - 1)
                if GL.all_sat_commd[choice_search_sat][6] == NOT_OTHER_OPERATE:
                    GL.search_datas[1] = str(GL.searched_time)
                elif GL.all_sat_commd[choice_search_sat][6] != NOT_OTHER_OPERATE:
                    GL.search_datas[1] = "{}/{}".format(GL.searched_time, GL.all_sat_commd[choice_search_sat][8])

            if GL.search_monitor_kws[4] in data2:  # 监控频点信息
                fre_symb_info_split = re.split(r"[:,]", data2)
                fre = fre_symb_info_split[1].split("=")[-1].strip()
                symb = fre_symb_info_split[2].split("=")[-1].strip()
                polar = fre_symb_info_split[3].split("=")[-1].strip()
                tp = "{}{}{}".format(fre, polar, symb)
                GL.all_tp_list.append(tp)
                # GL.channel_info[str(len(GL.all_tp_list))] = [[], []]
                GL.channel_info[tp] = [[], []]

            if GL.search_monitor_kws[1] in data2:  # 监控搜索过程电视个数和名称信息
                GL.tv_radio_tp_count[0] = re.split(r"-{2,}|\s{2,}", data2)[1]  # 提取电视节目数
                tv_name = re.split(r"-{2,}|\s{2,}", data2)[2]  # 提取电视节目名称
                # GL.channel_info[str(len(GL.all_tp_list))][0].append('[T]{}'.format(tv_name))
                GL.channel_info[tp][0].append('[T]{}'.format(tv_name))

            if GL.search_monitor_kws[2] in data2:  # 监控搜索过程广播个数和名称信息
                GL.tv_radio_tp_count[1] = re.split(r"-{2,}|\s{2,}", data2)[1]  # 提取广播节目数
                radio_name = re.split(r"-{2,}|\s{2,}", data2)[2]  # 提取电视节目名称
                # GL.channel_info[str(len(GL.all_tp_list))][1].append('[R]{}'.format(radio_name))
                GL.channel_info[tp][1].append('[R]{}'.format(radio_name))

            if GL.search_monitor_kws[7] in data2 or GL.search_monitor_kws[8] in data2:  # 监控搜索达到上限
                limit_type = re.split(r"[\s_]", data2)[1]
                logging.debug(limit_type)
                logging.info("搜索{}达到上限:{}".format(limit_type, data2))
                GL.start_record_maximum_state = True
                GL.record_maximum_data.append(data2)
                # search_time = 72
                # GL.all_sat_commd[choice_search_sat][8] -= 1
                # logging.info("搜索到上限剩余次数:{}".format(GL.all_sat_commd[choice_search_sat][8]))
                GL.upper_limit_state = True

            if GL.search_monitor_kws[3] in data2:  # 监控搜索结束
                GL.search_end_state = True
                GL.start_record_maximum_state = False
                end_time = datetime.now()
                GL.search_dur_time = str(end_time - start_time)[2:10]
                for i in range(len(GL.all_tp_list)):
                    print(GL.all_tp_list[i])
                print("第{}次搜索节目总数为TV/Radio:{}/{},TP总数为:{},盲扫时长:{}".format(
                    GL.search_datas[1], GL.tv_radio_tp_count[0], GL.tv_radio_tp_count[1],
                    len(GL.all_tp_list), GL.search_dur_time))

            if GL.search_monitor_kws[5] in data2:  # 监控保存TP的个数
                GL.tv_radio_tp_count[4] = re.split("=", data2)[1]   # 保存TP的个数
                GL.search_datas[4] = GL.tv_radio_tp_count[4]

            if GL.search_monitor_kws[6] in data2:  # 监控保存TV和Radio的个数
                split_result = re.split(r"[,\]]", data2)
                GL.tv_radio_tp_count[2] = re.split("=", split_result[1])[1]     # 保存TV数
                GL.tv_radio_tp_count[3] = re.split("=", split_result[2])[1]     # 保存Radio数
                GL.search_datas[5] = "{}/{}".format(GL.tv_radio_tp_count[2], GL.tv_radio_tp_count[3])

                GL.tv_radio_tp_accumulated[0].append(int(GL.tv_radio_tp_count[0]))
                GL.tv_radio_tp_accumulated[1].append(int(GL.tv_radio_tp_count[1]))
                GL.tv_radio_tp_accumulated[2].append((int(len(GL.all_tp_list))))
                GL.tv_radio_tp_accumulated[3].append(int(GL.tv_radio_tp_count[4]))

                print("本次搜索实际保存TV/Radio:{},保存TP数为:{}".format(GL.search_datas[5], GL.search_datas[4]))
                print("当前轮次:{},累计搜索节目个数:{}/{},累计搜索TP个数:{},累计保存TP个数：{}".format(
                    GL.search_datas[1], sum(GL.tv_radio_tp_accumulated[0]), sum(GL.tv_radio_tp_accumulated[1]),
                    sum(GL.tv_radio_tp_accumulated[2]), sum(GL.tv_radio_tp_accumulated[3])))
                # GL.save_ch_finish_state = True
            if GL.delete_ch_finish_kws in data2:  # 监控删除所有节目成功的关键字
                GL.delete_ch_finish_state = True


if __name__ == "__main__":
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)

    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIMESHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D",
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIMESHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict([val, key] for key, val in rsv_key.items())

    PRESET_SAT_NAME = [
         'Nilesat', 'Hotbird', 'Badr 4/5/6/7 K', 'Thor 5/6/7', 'Turksat 2A/3A', 'BulgariaSat-1', 'Eutelsat 3B C',
         'Eutelsat 4A', 'Eutelsat 9B', 'Amos 5 K', 'Astra 1E/3B', 'Arabsat 5A C', 'Arabsat 5A K', 'Eutelsat 33E K',
         'C_Paksat 1R', 'Intelsat 12', 'Azerspace K', 'Intelsat 10', 'Yamal 202', 'Turksat 4B K', 'Belintersat K',
         'TurkmenAlem', 'Yahsat 1A', 'Express AM6', 'Yamal 402 K', 'NSS 12 C', 'Intelsat 33e C', 'Intelsat 33e K',
         'Intelsat 902 C', 'Intelsat 20 K', 'ABS 2/2A K', 'APSTAR 7 C', 'Thaicom 5/6 C', 'Thaicom 5/8 K',
         'Express MD1 C', 'Insat 4A K', 'ST 2 K', 'Yamal 201 K', 'Measat 3/3A K', 'Measat 3/3A C', 'SES 8',
         'Express AM33 K', 'Koreasat 5 K', 'JCSat 3A K', 'JCSat 3A C', 'Vinasat 1 K', 'Telstar 18 K', 'Express AM5 K',
         'Express AM5 C', 'Optus D1', 'Superbird B2', 'Intelsat 2/8', 'Amos 2/3/7', 'Eutelsat 5 C', 'Eutelsat 5 K',
         'Eutelsat 8 C', 'Express AM44 K', 'Eutelsat 12', 'Telstar 12V', 'ABS-3 K', 'SES 4 K', 'Intelsat 905 C',
         'AlComSat 1', 'Intelsat 907 C', 'Intelsat 907 K', 'Hispasat 4/5/6', 'Intelsat 35e', 'Intelsat 707 K',
         'Intelsat 21 K', 'Amazonas 2/3 K', 'Asiasat 7 C', 'Chinas6b_C'
    ]

    NORMAL_SEARCH_TIMES = 10  # 10 普通盲扫次数
    SUPER_SEARCH_TIMES = 10  # 10 超级盲扫次数
    INCREMENTAL_SEARCH_TIMES = 15  # 15 累加搜索次数
    UPPER_LIMIT_SEARCH_TIMES = 72  # 72 上限搜索初始次数
    UPPER_LIMIT_CYCLE_TIMES = 5  # 5  上限搜索循环次数
    UPPER_LIMIT_LATER_SEARCH_TIMES = 20  # 20 上限搜索后其他情况执行测试
    ONLY_EXECUTE_ONE_TIME = 1  # 单独场景只执行一次
    NOT_UPPER_LIMIT_LATER_SEARCH_TIME = "Normal_search"

    ENTER_ANTENNA_SETTING = [KEY["MENU"], KEY["OK"]]
    DELETE_ALL_SAT = [KEY["RED"], KEY["0"], KEY["RED"], KEY["OK"]]
    ADD_ONE_SAT = [KEY["GREEN"], KEY["UP"], KEY["OK"], KEY["INFO"]]
    SEARCH_PREPARATORY_WORK = [[], [DELETE_ALL_SAT, ADD_ONE_SAT]]
    CHOICE_BLIND_MODE = [KEY["RIGHT"], KEY["OK"], KEY["OK"]]
    CHOICE_SUPERBLIND_MODE = [KEY["BLUE"], KEY["RIGHT"], KEY["OK"], KEY["OK"]]
    CHOICE_NOT_SEARCH = []
    CHOICE_SAVE_TYPE = [[KEY["OK"]], [KEY["LEFT"], KEY["OK"]]]
    EXIT_ANTENNA_SETTING = [KEY["EXIT"], KEY["EXIT"]]
    NOT_OTHER_OPERATE = []
    RESET_FACTORY = [KEY["MENU"], KEY["RIGHT"], KEY["DOWN"], KEY["OK"],
                     KEY["0"], KEY["0"], KEY["0"], KEY["0"],
                     KEY["OK"]]
    DELETE_SPECIFY_SAT_ALL_TP = [KEY["GREEN"], KEY["0"], KEY["RED"], KEY["OK"]]
    DELETE_ALL_CH = [KEY["MENU"], KEY["LEFT"], KEY["LEFT"], KEY["UP"], KEY["OK"], KEY["OK"]]
    UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT = [KEY["EXIT"]]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]

    xlsx_title = [
        "搜索模式",
        "搜索次数",
        "搜索TP数",
        "搜索节目数",
        "保存TP数",
        "保存节目数",
        "搜索时间",
        {"数据类别": ["TP", "All", "TV", "Radio", "CH_Name"]},
        "TP"
    ]
    sat_name, search_mode = '', ''
    GL = MyGlobal()

    sat_search_mode_list = [
                            "6b_blind",                         # 0
                            "6b_super_blind",                   # 1

                            "y3_blind",                         # 2
                            "y3_super_blind",                   # 3

                            "138_blind",                        # 4
                            "138_super_blind",                  # 5

                            "88_blind",                         # 6
                            "88_super_blind",                   # 7

                            "plp_blind",                        # 8
                            "plp_super_blind",                  # 9

                            "138_incremental_blind",            # 10 累加搜索

                            "138_ch_upper_limit_blind",          # 11 搜索节目达到上限,会删除所有节目,重新搜索
                            "138_ch_ul_later_cont_blind",        # 12 搜索节目达到上限后,不删除指定卫星下的tp,继续搜索
                            "138_ch_ul_later_del_tp_blind",      # 13 搜索节目达到上限后,删除指定卫星下的tp,继续搜索

                            "z6_tp_upper_limit_blind",          # 14 搜索tp达到上限,会恢复出厂设置,重新搜索
                            "z6_tp_ul_later_cont_blind",        # 15 搜索tp达到上限后,不删除指定卫星下的tp,继续搜索
                            "z6_tp_ul_later_del_tp_blind",      # 16 搜索tp达到上限后,删除指定卫星下的tp,继续搜索

                            "reset_factory",                    # 17 恢复出厂设置
                            "delete_all_channel",               # 18 删除所有节目
                            ]

    simplify_sat_name = {
        "Chinas6b_C": "Z6",
        "Asiasat 7 C": "Y3",
        "Telstar 18 K": "138",
        "ST 2 K": "88",
        "PLPD": "PLPD",
        "Reset": "Reset",
        "Delete": "Delete",
    }

    msg = "现在开始执行的是:{}_{}".format(sat_name, search_mode)
    logging.critical(format(msg, '*^150'))

    try:
        choice_search_sat = int(sys.argv[1])                        # 参考sat_list中的选项进行卫星选择

        # 选择执行轮次
        if len(GL.all_sat_commd[choice_search_sat]) < 9:
            search_time = GL.all_sat_commd[choice_search_sat][-1]
        elif len(GL.all_sat_commd[choice_search_sat]) == 9:
            search_time = GL.all_sat_commd[choice_search_sat][7]

        # judge_write_file_exist()
        build_print_log_and_report_file_path()

        send_ser_name,receive_ser_name = check_ports()
        send_ser = serial.Serial(send_ser_name, 9600)
        receive_ser = serial.Serial(receive_ser_name, 115200, timeout=1)
        # serial_set(send_ser, send_ser_name, 9600)
        # serial_set(receive_ser, receive_ser_name, 115200)

        thread_send = threading.Thread(target=data_send_thread)
        thread_receive = threading.Thread(target=data_receiver_thread)

        thread_send.start()
        thread_receive.start()

    except Exception as e:
        print(e)
        # cur_py_file_name = sys.argv[0]        # 第0个就是这个python文件本身的路径（全路径）
        cur_py_file_name = os.path.basename(__file__)       # 当前文件名名称
        ret = mail(f"{cur_py_file_name}\n\n"
                   f"{msg}\n\n"
                   f"{traceback.format_exc()}")
        if ret:
            print("邮件发送成功")
        else:
            print("邮件发送失败")

        print("***traceback.format_exc():*** ")
        print(traceback.format_exc())
