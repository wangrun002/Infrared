#!/usr/bin/python
# -*- coding: utf-8 -*-

from datetime import datetime
from multiprocessing import Process, Manager
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import multiprocessing
import serial
import serial.tools.list_ports
import re
import time
import os
import sys
import random
import logging
import platform

class MyGlobal():
    def __init__(self):
        self.send_loop_state = True
        self.receive_loop_state = True
        self.searched_sat_name = []  # 用于保存搜索过程中搜索过的卫星的名称,便于搜索达到上限后,删除指定的卫星,不能被清空
        self.random_choice_sat = []  # 用于存放搜索达到上限后每次随机选择的卫星,然后进行删除其TP
        self.upper_limit_send_ok_commd_state = False  # 用于控制搜索达到上限后是否发送OK命令的状态变量


def check_ports():
    global send_com, receive_com
    ports_info = []
    if platform.system() == "Windows":
        ser_cable_num = 7
        serial_ser = {
            "1": "FTDVKA2HA",
            "2": "FTGDWJ64A",
            "3": "FT9SP964A",
            "4": "FTHB6SSTA",
            "5": "FTDVKPRSA",
            "6": "FTHI8UIHA",
            "7": "FTHG05TTA",
        }
        send_port_desc = "USB-SERIAL CH340"
        receive_port_desc = serial_ser[str(ser_cable_num)]
    elif platform.system() == "Linux":
        ser_cable_num = 7
        serial_ser = {
            "1": "FTDVKA2H",
            "2": "FTGDWJ64",
            "3": "FT9SP964",
            "4": "FTHB6SST",
            "5": "FTDVKPRS",
            "6": "FTHI8UIH",
            "7": "FTHG05TT",
        }
        send_port_desc = "USB2.0-Serial"
        receive_port_desc = serial_ser[str(ser_cable_num)]
    ports = list(serial.tools.list_ports.comports())
    for i in range(len(ports)):
        logging.info("可用端口:名称:{} + 描述:{} + 硬件id:{}".format(ports[i].device, ports[i].description, ports[i].hwid))
        # print("可用端口:名称:{} + 描述:{} + 硬件id:{}".format(ports[i].device, ports[i].description, ports[i].hwid))
        ports_info.append("{}~{}~{}".format(ports[i].device, ports[i].description, ports[i].hwid))
    if len(ports) <= 0:
        logging.info("无可用端口")
    elif len(ports) == 1:
        logging.info("只有一个可用端口:{}".format(ports[0].device))
    elif len(ports) >= 2:
        for i in range(len(ports_info)):
            if send_port_desc in ports_info[i]:
                send_com = ports_info[i].split("~")[0]
            elif receive_port_desc in ports_info[i]:
                receive_com = ports_info[i].split("~")[0]
    return send_com, receive_com


def serial_set(ser, ser_name, ser_baudrate):
    ser.port = ser_name
    ser.baudrate = ser_baudrate
    ser.bytesize = 8
    ser.parity = "N"
    ser.stopbits = 1
    ser.timeout = 1
    ser.write_timeout = 0
    ser.open()


def hex_strs_to_bytes(strings):
    # strs = strings.replace(" ", "")
    return bytes.fromhex(strings)


def send_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    send_ser.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
    infrared_send_commd.append(REVERSE_KEY[commd])
    time.sleep(1.2)


def build_print_log_and_report_file_path():
    global sat_name, search_mode, sheet_name
    global report_file_path, case_log_txt_path

    parent_path = os.path.dirname(os.getcwd())
    test_file_folder_name = "test_data"
    test_file_directory = os.path.join(parent_path, test_file_folder_name)
    case_log_folder_name = "print_log"
    case_log_file_directory = os.path.join(parent_path, test_file_folder_name, case_log_folder_name)
    report_folder_name = "report"
    report_file_directory = os.path.join(parent_path, test_file_folder_name, report_folder_name)

    if not os.path.exists(test_file_directory):
        os.mkdir(test_file_directory)
    if not os.path.exists(case_log_file_directory):
        os.mkdir(case_log_file_directory)
    if not os.path.exists(report_file_directory):
        os.mkdir(report_file_directory)

    sat_name = all_sat_commd[choice_search_sat[0]][2][0]
    search_mode = all_sat_commd[choice_search_sat[0]][2][-1]
    timestamp = re.sub(r'[-: ]', '_', str(datetime.now())[:19])
    sheet_name = "{}_{}".format(sat_name, search_mode)

    report_file_name = "{}_{}_{}_{}_Result_{}.xlsx".format(choice_search_sat[0], simplify_sat_name[sat_name], \
                                                           sat_name, search_mode, timestamp)
    report_file_path = os.path.join(report_file_directory, report_file_name)

    case_log_file_name = "{}_{}_{}_{}_{}.txt".format(choice_search_sat[0], simplify_sat_name[sat_name], \
                                                     sat_name, search_mode, timestamp)
    case_log_txt_path = os.path.join(case_log_file_directory, case_log_file_name)


def enter_antenna_setting():
    logging.debug("Enter Antenna Setting")
    sat_para_list[0] = ''
    send_commd(KEY["MENU"])
    send_commd(KEY["OK"])
    time.sleep(1)       # 等待进入天线卫星设置界面，且获取到卫星名称和焦点位置
    while sat_para_list[0] == '':
        logging.info("没有正确进入天线设置界面，重新进入")
        send_commd(KEY["EXIT"])
        send_commd(KEY["EXIT"])
        send_commd(KEY["EXIT"])
        time.sleep(0.5)
        send_commd(KEY["MENU"])
        send_commd(KEY["OK"])
        time.sleep(1)


def judge_and_wirte_data_to_xlsx():
    xlsx_data_interval = 1 + 5 * (search_time_list[1] - 1)
    xlsx_title = [
        "搜索模式",
        "搜索次数",
        "搜索TP数",
        "搜索节目数",
        "保存TP数",
        "保存节目数",
        "搜索时间",
        {"数据类别": ["TP", "All", "TV", "Radio", "CH_Name"]},
        "TP"
    ]
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    if not os.path.exists(report_file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 11
        for i in range(len(xlsx_title)):
            if i < len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = list(xlsx_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 1:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment

    elif os.path.exists(report_file_path):
        wb = load_workbook(report_file_path)
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 11
        for i in range(len(xlsx_title)):
            if i < len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 2:
                ws.cell(i + 1, 1).value = list(xlsx_title[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(xlsx_title) - 1:
                ws.cell(i + 1, 1).value = xlsx_title[i]
                ws.cell(i + 1, 1).alignment = alignment

    tp_column_numb = column_index_from_string("A") + xlsx_data_interval
    all_column_numb = column_index_from_string("A") + xlsx_data_interval + 1
    tv_column_numb = column_index_from_string("A") + xlsx_data_interval + 2
    radio_column_numb = column_index_from_string("A") + xlsx_data_interval + 3
    tp_column_char = get_column_letter(tp_column_numb)
    all_column_char = get_column_letter(all_column_numb)
    tv_column_char = get_column_letter(tv_column_numb)
    radio_column_char = get_column_letter(radio_column_numb)
    ws.column_dimensions[tp_column_char].width = 12
    ws.column_dimensions[all_column_char].width = 3
    ws.column_dimensions[tv_column_char].width = 3
    ws.column_dimensions[radio_column_char].width = 3

    for m in range(len(search_data)):
        if m < len(search_data) - 2:
            ws.cell((m + 1), (1 + xlsx_data_interval)).value = search_data[m]
            ws.merge_cells(start_row=(m + 1), start_column=(1 + xlsx_data_interval), \
                           end_row=(m + 1), end_column=(1 + xlsx_data_interval + 4))
            ws.cell((m + 1), (1 + xlsx_data_interval)).alignment = alignment
        elif m == len(search_data) - 2:
            for n in range(len(xlsx_title[7]["数据类别"])):
                ws.cell((m + 1), (1 + xlsx_data_interval + n)).value = list(xlsx_title[m].values())[0][n]
                ws.cell((m + 1), (1 + xlsx_data_interval + n)).alignment = alignment
                ws.row_dimensions[(m + 1)].height = 13.5
        elif m == len(search_data) - 1:
            for j in range(len(channel_info.keys())):
                ws.cell((m + 1 + j), (1 + xlsx_data_interval)).value = search_data[m][j]
                # ws.cell((m+1+j),(1+xlsx_data_interval)+1).value = len(channel_info[str(j+1)][0]) + len(channel_info[str(j+1)][1])
                # ws.cell((m+1+j),(1+xlsx_data_interval)+2).value = len(channel_info[str(j+1)][0])
                # ws.cell((m+1+j),(1+xlsx_data_interval)+3).value = len(channel_info[str(j+1)][1])
                # ws.cell((m+1+j),(1+xlsx_data_interval)+4).value = ",".join(channel_info[str(j+1)][0] + channel_info[str(j+1)][1])

                ws.cell((m + 1 + j), (1 + xlsx_data_interval) + 1).value = len(
                    channel_info[search_data[m][j]][0]) + len(channel_info[search_data[m][j]][1])
                ws.cell((m + 1 + j), (1 + xlsx_data_interval) + 2).value = len(channel_info[search_data[m][j]][0])
                ws.cell((m + 1 + j), (1 + xlsx_data_interval) + 3).value = len(channel_info[search_data[m][j]][1])
                ws.cell((m + 1 + j), (1 + xlsx_data_interval) + 4).value = ",".join(
                    channel_info[search_data[m][j]][0] + channel_info[search_data[m][j]][1])

                for k in range(len(xlsx_title[7]["数据类别"])):
                    ws.cell((m + 1 + j), (1 + xlsx_data_interval) + k).alignment = alignment
                ws.row_dimensions[(m + 1 + j)].height = 13.5
    wb.save(report_file_path)


def judge_preparatory_work():
    if len(all_sat_commd[choice_search_sat[0]][1]) == 0:
        logging.debug("Not Search Preparatory Work")
    elif len(all_sat_commd[choice_search_sat[0]][1]) > 0:
        logging.debug("Search Preparatory Work")
        send_data_1 = all_sat_commd[choice_search_sat[0]][1][0]
        send_data_2 = all_sat_commd[choice_search_sat[0]][1][1]
        for i in range(len(send_data_1)):
            send_commd(send_data_1[i])
        logging.info("等待删除卫星保存结束5秒")
        time.sleep(5)  # 等待删除卫星保存结束
        for j in range(len(send_data_2)):
            send_commd(send_data_2[j])


def check_satellite_param():  # 考虑打印出现有延迟的情况
    logging.debug("Satellite")
    # 用于等待卫星名称打印的接收
    while antenna_setting_focus_pos[0] == '':
        time.sleep(2)
    while antenna_setting_focus_pos[0] != "Satellite":
        send_commd(KEY["DOWN"])
    else:
        if all_sat_commd[choice_search_sat[0]][1] == SEARCH_PREPARATORY_WORK[0]:  # upper limit or incremental search
            if len(GL.searched_sat_name) == 72:  # 避免程序还没有执行结束，但是搜索的卫星个数满了导致的死循环
                for i in range(len(GL.searched_sat_name) // 2):
                    GL.searched_sat_name.remove(random.choice(GL.searched_sat_name))
            while sat_para_list[0] in GL.searched_sat_name:
                logging.info("sat in list")
                logging.info("{},{}".format(sat_para_list[0], GL.searched_sat_name))
                send_commd(KEY["RIGHT"])
            else:
                logging.info("{},{}".format(sat_para_list[0], GL.searched_sat_name))
                logging.info("sat not in list")
                GL.searched_sat_name.append(sat_para_list[0])
                logging.info("{},{}".format(sat_para_list[0], GL.searched_sat_name))
                send_commd(KEY["DOWN"])

        elif all_sat_commd[choice_search_sat[0]][1] == SEARCH_PREPARATORY_WORK[1]:  # normal sat search
            send_commd(KEY["DOWN"])


def check_lnb_power():
    logging.debug("LNB POWER")
    while antenna_setting_focus_pos[0] != "LNB Power":
        send_commd(KEY["DOWN"])
    else:
        power_off = "Polar=2"
        while sat_para_list[1] != power_off:
            send_commd(KEY["LEFT"])
        else:
            send_commd(KEY["RIGHT"])
            send_commd(KEY["DOWN"])


def check_lnb_fre():
    logging.debug("LBN FREQUENCY")
    while antenna_setting_focus_pos[0] != "LNB Frequency":
        send_commd(KEY["DOWN"])
    else:
        logging.info(sat_para_list)
        while sat_para_list[2] != all_sat_commd[choice_search_sat[0]][2][2]:
            send_commd(KEY["RIGHT"])
        else:
            send_commd(KEY["DOWN"])


def check_22k():
    logging.debug("22k")
    while antenna_setting_focus_pos[0] != "22K":
        send_commd(KEY["DOWN"])
    else:
        while sat_para_list[3] != all_sat_commd[choice_search_sat[0]][2][3]:
            send_commd(KEY["RIGHT"])
        else:
            send_commd(KEY["DOWN"])


def check_diseqc_10():
    logging.debug("Diseqc 1.0")
    while antenna_setting_focus_pos[0] != "DiSEqC 1,0":
        send_commd(KEY["DOWN"])
    else:
        while sat_para_list[4] != all_sat_commd[choice_search_sat[0]][2][4]:
            send_commd(KEY["LEFT"])
        else:
            send_commd(KEY["DOWN"])


def check_diseqc_11():
    logging.debug("Diseqc 1.1")
    while antenna_setting_focus_pos[0] != "DiSEqC 1,1":
        send_commd(KEY["DOWN"])
    else:
        while sat_para_list[5] != all_sat_commd[choice_search_sat[0]][2][5]:
            send_commd(KEY["LEFT"])
        else:
            send_commd(KEY["DOWN"])


def check_tp():
    logging.debug("TP")
    while antenna_setting_focus_pos[0] != "TP":
        send_commd(KEY["DOWN"])
    else:
        send_commd(KEY["DOWN"])


def choice_srh_mode_and_start_srh():
    logging.debug("Choice Search Mode And Start Search")
    while antenna_setting_focus_pos[0] != "Start Search":
        send_commd(KEY["DOWN"])
    else:
        send_data = all_sat_commd[choice_search_sat[0]][3]
        for i in range(len(send_data)):
            send_commd(send_data[i])
        time.sleep(1)
        while not state["search_start_state"]:
            send_commd(KEY["OK"])
            time.sleep(1)


def antenna_setting():
    check_satellite_param()
    check_lnb_power()
    check_lnb_fre()
    check_22k()
    check_diseqc_10()
    check_diseqc_11()
    check_tp()
    choice_srh_mode_and_start_srh()


def block_send_thread():
    time.sleep(1)
    # send_ser.send_break(3)


def judge_srh_limit():
    logging.debug("Upper Limit To Save Channel Stage")
    if not state["upper_limit_state"]:
        logging.debug("Not Upper Limit")
    elif state["upper_limit_state"]:
        if all_sat_commd[choice_search_sat[0]][-1] != NOT_UPPER_LIMIT_LATER_SEARCH_TIME:  # 上限搜索
            logging.debug("Upper Limit")
            logging.debug("打印搜索达到上限是否有新增节目的记录列表:{}".format(record_maximum_data))
            search_time_list[0] = 72
            # all_sat_commd[choice_search_sat[0]][8] -= 1
            # logging.info("搜索到上限剩余次数:{}".format(all_sat_commd[choice_search_sat[0]][8]))
            search_time_list[3] -= 1
            logging.info("搜索到上限剩余次数:{}".format(search_time_list[3]))
            for i in range(len(record_maximum_data)):
                if "[PTD]TV_save=" in record_maximum_data[i]:  # "[PTD]TV_save="
                    GL.upper_limit_send_ok_commd_state = True
            if GL.upper_limit_send_ok_commd_state:
                logging.debug("搜索达到上限但是没有新增节目")
            elif not GL.upper_limit_send_ok_commd_state:
                logging.debug("搜索达到上限但是有新增节目")
                send_commd(KEY["OK"])
        else:
            logging.info("普通搜索，但是达到上限")


def judge_save_ch_mode():
    logging.debug("Whether Or Not Save And End Search")
    # logging.info("搜索到上限剩余次数:{}".format(all_sat_commd[choice_search_sat[0]][8]))
    logging.info("搜索到上限剩余次数:{}".format(search_time_list[3]))
    send_data = all_sat_commd[choice_search_sat[0]][4]
    for i in range(len(send_data)):
        send_commd(send_data[i])
    if all_sat_commd[choice_search_sat[0]][4] == CHOICE_SAVE_TYPE[0]:
        logging.info("主动在保存节目时延时3秒")
        time.sleep(3)


def write_data_to_excel():
    logging.debug("Write data to Excel")
    logging.info("保存节目后等待保存TP和保存节目的打印5秒")
    time.sleep(5)
    state["save_ch_finish_state"] = True
    search_data[0] = sheet_name
    search_data[2] = len(all_tp_list)
    search_data[3] = "{}/{}".format(tv_radio_tp_count[0], tv_radio_tp_count[1])
    search_data[6] = search_time_list[2]
    search_data[8] = all_tp_list
    logging.info(all_tp_list)
    logging.info(channel_info)
    judge_and_wirte_data_to_xlsx()


def clear_variate():
    logging.debug("clear data")
    # global record_maximum_data, all_tp_list, channel_info

    # 处理循环数据
    # logging.info(record_maximum_data)
    # logging.info(all_tp_list)
    # logging.info(channel_info)
    record_maximum_data = []

    tv_radio_tp_count[0], tv_radio_tp_count[1] = 0, 0
    tv_radio_tp_count[4] = 0
    search_data[5] = '0/0'
    tv_radio_tp_count[2], tv_radio_tp_count[3] = 0, 0
    logging.info(record_maximum_data)
    logging.info(all_tp_list)
    logging.info(channel_info)
    state["clear_variate_state"] = True


def exit_antenna_setting():
    logging.debug("Exit Antenna Setting")
    send_data = all_sat_commd[choice_search_sat[0]][5]
    for i in range(len(send_data)):
        send_commd(send_data[i])


def other_operate_del_all_ch():
    logging.debug("Delete All Channels And Choice Searched First Sat")
    # 执行删除所有节目的命令
    send_data = all_sat_commd[choice_search_sat[0]][6]
    for i in range(len(send_data)):
        send_commd(send_data[i])
    # 等待节目删除完成后返回成功标志
    n = 20  # 等待20秒还没有返回删除成功标志就继续发送ok命令
    while True:
        if not state["delete_ch_finish_state"]:
            logging.info("还没有删除完成，请等待")
            time.sleep(1)
            n -= 1
            if n == 0:
                send_commd(KEY["EXIT"])
                send_commd(KEY["EXIT"])
                send_commd(KEY["EXIT"])
                send_data = all_sat_commd[choice_search_sat[0]][6]
                for i in range(len(send_data)):
                    send_commd(send_data[i])
                n = 20
        elif state["delete_ch_finish_state"]:
            logging.info("删除完成")
            break
    # 进入天线设置界面，并切换到第一个卫星
    send_commd(KEY["EXIT"])
    send_commd(KEY["EXIT"])
    send_commd(KEY["EXIT"])
    # send_commd(KEY["MENU"])
    # send_commd(KEY["OK"])
    enter_antenna_setting()
    first_sat_name = GL.searched_sat_name[0]
    while sat_para_list[0] != first_sat_name:
        send_commd(KEY["LEFT"])
    # 退回大画面
    for i in range(len(EXIT_TO_SCREEN)):
        send_commd(EXIT_TO_SCREEN[i])
    GL.searched_sat_name.clear()


def other_operate_del_specify_sat_all_tp():
    logging.debug("Delete Specify Sat TP And Choice Random Sat")
    # send_commd(KEY["MENU"])
    # send_commd(KEY["OK"])
    enter_antenna_setting()
    GL.random_choice_sat.append(random.choice(GL.searched_sat_name))
    while sat_para_list[0] != GL.random_choice_sat[0]:
        if PRESET_SAT_NAME.index(sat_para_list[0]) > PRESET_SAT_NAME.index(GL.random_choice_sat[0]):
            send_commd(KEY["LEFT"])
        elif PRESET_SAT_NAME.index(sat_para_list[0]) < PRESET_SAT_NAME.index(GL.random_choice_sat[0]):
            send_commd(KEY["RIGHT"])
    logging.info("{},{},{}".format(sat_para_list[0], GL.random_choice_sat[0], GL.searched_sat_name))
    GL.searched_sat_name.remove(GL.random_choice_sat[0])  # 避免搜索时该卫星在已搜索的卫星列表中，不能进行搜索
    GL.random_choice_sat.clear()
    send_data = DELETE_SPECIFY_SAT_ALL_TP
    for i in range(len(send_data)):
        send_commd(send_data[i])
    logging.info("等待删除指定卫星下的所有TP10秒")
    time.sleep(10)
    send_data = EXIT_TO_SCREEN
    for j in range(len(send_data)):
        send_commd(send_data[j])


def judge_other_operate():
    if len(all_sat_commd[choice_search_sat[0]][6]) == 0:  # 没有额外操作
        logging.debug("Not Other Operate")
    elif len(all_sat_commd[choice_search_sat[0]][6]) > 0:  # 有额外操作
        logging.debug("Exist Other Operate")
        if not state["upper_limit_state"]:
            logging.debug("Exist Other Operate But Not Upper Limit")
        elif state["upper_limit_state"]:
            # if all_sat_commd[choice_search_sat[0]][8] < 0:  # 搜索的次数到最后一次时不再进行额外的操作
            if search_time_list[3] < 0:
                logging.info("搜索的次数到最后一次时不再进行额外的操作")
            else:
                if all_sat_commd[choice_search_sat[0]][6] == RESET_FACTORY:
                    logging.debug("Reset Factory")
                    send_data = all_sat_commd[choice_search_sat[0]][6]
                    for i in range(len(send_data)):
                        send_commd(send_data[i])
                    logging.info("等待恢复出厂设置重启30秒")
                    time.sleep(30)
                    GL.searched_sat_name.clear()
                elif all_sat_commd[choice_search_sat[0]][6] == DELETE_ALL_CH:
                    other_operate_del_all_ch()
                elif all_sat_commd[choice_search_sat[0]][6] == DELETE_SPECIFY_SAT_ALL_TP:
                    other_operate_del_specify_sat_all_tp()
                elif all_sat_commd[choice_search_sat[0]][6] == UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT:
                    logging.debug("Not Delete Specify Sat Tp And Search Continue")
                    send_commd(UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT[0])
                    GL.searched_sat_name.remove(random.choice(GL.searched_sat_name))  # 达到上限后切下一个卫星搜索
                    # GL.searched_sat_name.clear()        # 达到上限后重复搜索最后一个卫星


def cyclic_srh_setting():
    global search_time_list
    logging.debug("Cyclic Search Setting")
    if all_sat_commd[choice_search_sat[0]][8] == NOT_UPPER_LIMIT_LATER_SEARCH_TIME:
        state["upper_limit_state"] = False  # 恢复默认状态
        sat_para_list = ["", "", "", "", "", ""]  # 获取卫星的参数保存数据恢复默认状态
        state["delete_ch_finish_state"] = False  # 删除所有节目成功状态恢复默认
        # state["save_ch_finish_state"] = False  # 保存节目成功状态恢复默认
        state["search_end_state"] = False  # 搜索结束状态恢复默认
        state["search_start_state"] = False  # 搜索开始状态恢复默认
        GL.upper_limit_send_ok_commd_state = False  # 搜索达到上限后是否发送OK命令的状态变量恢复默认

        search_time_list[0] -= 1
        logging.info("进入下一次循环搜索等待5秒")
        time.sleep(5)
        logging.info("剩余搜索次数:{}".format(search_time_list[0]))
        if search_time_list[0] < 1:
            logging.info("程序结束")
            # GL.send_loop_state = False
            state["receive_loop_state"] = True  # 触发结束接收进程的状态

    elif all_sat_commd[choice_search_sat[0]][8] != NOT_UPPER_LIMIT_LATER_SEARCH_TIME:
        state["upper_limit_state"] = False  # 恢复默认状态
        sat_para_list = ["", "", "", "", "", ""]  # 获取卫星的参数保存数据恢复默认状态
        state["delete_ch_finish_state"] = False  # 删除所有节目成功状态恢复默认
        # state["save_ch_finish_state"] = False  # 保存节目成功状态恢复默认
        state["search_end_state"] = False  # 搜索结束状态恢复默认
        state["search_start_state"] = False  # 搜索开始状态恢复默认
        GL.upper_limit_send_ok_commd_state = False  # 搜索达到上限后是否发送OK命令的状态变量恢复默认

        logging.info("进入下一次循环搜索等待5秒")
        time.sleep(5)
        # logging.info("搜索到上限剩余次数:{}".format(all_sat_commd[choice_search_sat[0]][8]))
        logging.info("搜索到上限剩余次数:{}".format(search_time_list[3]))
        # all_sat_commd[choice_search_sat[0]][8] -= 1
        # if all_sat_commd[choice_search_sat[0]][8] < 0:
        if search_time_list[3] < 0:
            logging.info("程序结束")
            # GL.send_loop_state = False
            state["receive_loop_state"] = True  # 触发结束接收进程的状态


def receive_ser_process(ser_name_list, sat_para_list, search_time_list, all_sat_commd, choice_search_sat, search_data, \
                        all_tp_list, channel_info, tv_radio_tp_count, record_maximum_data, \
                        infrared_send_commd, infrared_rsv_commd, antenna_setting_focus_pos, state, case_log_path):
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)

    def add_write_data_to_txt(file_path, write_data):  # 追加写文本
        with open(file_path, "a+", encoding="utf-8") as fo:
            fo.write(write_data)

    NOT_OTHER_OPERATE = []
    all_tp_list_process = []
    channel_info_process = {}
    tv_radio_tp_accumulated = [[], [], [], []]  # 用于统计每轮搜索累加的TV、Radio、TP数以及保存TP数的值

    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIMESHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict([val, key] for key, val in rsv_key.items())

    sat_param_kws = [
        "[PTD]sat_name=",
        "[PTD]LNB1=",
        "[PTD]--[0:ON,1:OFF]---22K",
        "[PTD]--set diseqc 1.0",
        "[PTD]--set diseqc 1.1",
    ]
    search_monitor_kws = [
        "[PTD]SearchStart",  # 0
        "[PTD]TV------",  # 1
        "[PTD]Radio-----",  # 2
        "[PTD]SearchFinish",  # 3
        "[PTD]get :  fre",  # 4
        "[PTD]TP_save=",  # 5
        "[PTD]TV_save=",  # 6
        "[PTD]maximum_tp",  # 7
        "[PTD]maximum_channel",  # 8
        "[PTD]get polar:",  # 9
    ]
    other_kws = [
        "[PTD]All programs deleted successfully",  # 删除所有节目成功关键字
        "[PTD]Infrared_key_values:",  # 获取红外接收关键字
        "[PTD]Antenna_setting:",  # 获取天线与卫星设置界面焦点位置关键字
    ]

    rsv_ser_name = ser_name_list[1]
    receive_ser = serial.Serial(rsv_ser_name, 115200, timeout=1)
    while True:
        data = receive_ser.readline()
        if data:
            tt = datetime.now()
            data1 = data.decode("GB18030", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            add_write_data_to_txt(case_log_path[0], data3)

            if state["save_ch_finish_state"]:
                print("本次搜索实际保存TV/Radio:{},保存TP数为:{}".format(search_data[5], search_data[4]))
                print("当前轮次:{},累计搜索节目个数:{}/{},累计搜索TP个数:{},累计保存TP个数：{}".format(search_data[1], \
                                                                              sum(tv_radio_tp_accumulated[0]), \
                                                                              sum(tv_radio_tp_accumulated[1]), \
                                                                              sum(tv_radio_tp_accumulated[2]), \
                                                                              sum(tv_radio_tp_accumulated[3])))
                state["save_ch_finish_state"] = False

            if state["clear_variate_state"]:  # 清除可变变量的值
                all_tp_list_process.clear()
                channel_info_process.clear()
                del all_tp_list[:]
                del record_maximum_data[:]
                channel_info.clear()
                antenna_setting_focus_pos[0] = ''
                state["clear_variate_state"] = False

            if state["start_record_maximum_state"]:
                record_maximum_data.append(data2)
                logging.debug(record_maximum_data)

            if sat_param_kws[0] in data2:  # 判断卫星名称
                sat_para_list[0] = re.split("=", data2)[-1]
                print("receive", sat_para_list[0])

            if sat_param_kws[1] in data2:  # 判断LNB Fre
                lnb_split = re.split("[,\]]", data2)
                lnb1 = lnb_split[1].split("=")[-1]
                lnb2 = lnb_split[2].split("=")[-1]
                sat_para_list[2] = "{}/{}".format(lnb1, lnb2)

            if sat_param_kws[2] in data2:  # 判断22k
                sat_para_list[3] = list(filter(None, re.split("-{2,}|,", data2)))[-1].strip()

            if sat_param_kws[3] in data2:  # 判断diseqc 1.0和Polar(LNB Power)
                polar_split = re.split("[,\]-]", data2)
                sat_para_list[4] = polar_split[3].split("=")[-1]
                sat_para_list[1] = polar_split[7].strip()

            if sat_param_kws[4] in data2:  # 判断diseqc 1.1
                disqc1_1_info_split = re.split(r"-", data2)
                sat_para_list[5] = disqc1_1_info_split[-1].split("=")[-1]

            if search_monitor_kws[0] in data2:  # 监控搜索起始
                start_time = datetime.now()
                state["search_start_state"] = True
                search_time_list[1] += 1  # [expect_search_time,searched_time,search_dur_time,expect_limit_search_time]
                if all_sat_commd[choice_search_sat[0]][6] == NOT_OTHER_OPERATE:
                    search_data[1] = search_time_list[1]
                elif all_sat_commd[choice_search_sat[0]][6] != NOT_OTHER_OPERATE:
                    # search_data[1] = "{}/{}".format(search_time_list[1], all_sat_commd[choice_search_sat[0]][8])
                    search_data[1] = "{}/{}".format(search_time_list[1], search_time_list[3])

            if search_monitor_kws[4] in data2:  # 监控频点信息
                fre_symb_info_split = re.split(r"[:,]", data2)
                fre = fre_symb_info_split[1].split("=")[-1].strip()
                symb = fre_symb_info_split[2].split("=")[-1].strip()
                polar = fre_symb_info_split[3].split("=")[-1].strip()
                tp = "{}{}{}".format(fre, polar, symb)
                all_tp_list_process.append(tp)
                # channel_info[str(len(all_tp_list))] = [[], []]
                # channel_info_process[str(len(all_tp_list_process))] = [[], []]
                channel_info_process[tp] = [[], []]

            if search_monitor_kws[1] in data2:  # 监控搜索过程电视个数和名称信息
                tv_radio_tp_count[0] = re.split("-{2,}|\s{2,}", data2)[1]  # 提取电视节目数
                tv_name = re.split("-{2,}|\s{2,}", data2)[2]  # 提取电视节目名称
                # channel_info_process[str(len(all_tp_list_process))][0].append('[T]{}'.format(tv_name))
                channel_info_process[tp][0].append('[T]{}'.format(tv_name))

            if search_monitor_kws[2] in data2:  # 监控搜索过程广播个数和名称信息
                tv_radio_tp_count[1] = re.split("-{2,}|\s{2,}", data2)[1]  # 提取广播节目数
                radio_name = re.split("-{2,}|\s{2,}", data2)[2]  # 提取电视节目名称
                # channel_info_process[str(len(all_tp_list_process))][1].append('[R]{}'.format(radio_name))
                channel_info_process[tp][1].append('[R]{}'.format(radio_name))

            if search_monitor_kws[7] in data2 or search_monitor_kws[8] in data2:  # 监控搜索达到上限
                limit_type = re.split(r"[\s_]", data2)[1]
                logging.debug(limit_type)
                logging.info("搜索{}达到上限:{}".format(limit_type, data2))
                state["start_record_maximum_state"] = True
                record_maximum_data.append(data2)
                # search_time_list[0] = 72
                # all_sat_commd[choice_search_sat[0]][8] -= 1
                # logging.info("搜索到上限剩余次数:{}".format(all_sat_commd[choice_search_sat[0]][8]))
                state["upper_limit_state"] = True

            if search_monitor_kws[3] in data2:  # 监控搜索结束
                state["search_end_state"] = True
                state["start_record_maximum_state"] = False
                end_time = datetime.now()
                search_time_list[2] = str(end_time - start_time)[2:10]
                channel_info.update(channel_info_process)
                all_tp_list.extend(all_tp_list_process)
                tv_radio_tp_accumulated[2].append((int(len(all_tp_list_process))))
                for i in range(len(all_tp_list_process)):
                    print(all_tp_list_process[i])
                print("第{}次搜索节目总数为TV/Radio:{}/{},TP总数为:{},盲扫时长:{}".format(search_data[1], \
                                                                          tv_radio_tp_count[0], tv_radio_tp_count[1],
                                                                          len(all_tp_list_process), \
                                                                          search_time_list[2]))

            if search_monitor_kws[5] in data2:  # 监控保存TP的个数
                tv_radio_tp_count[4] = re.split("=", data2)[1]
                search_data[4] = tv_radio_tp_count[4]
                tv_radio_tp_accumulated[3].append(int(tv_radio_tp_count[4]))

            if search_monitor_kws[6] in data2:  # 监控保存TV和Radio的个数
                split_result = re.split(r"[,\]]", data2)
                tv_radio_tp_count[2] = re.split("=", split_result[1])[1]
                tv_radio_tp_count[3] = re.split("=", split_result[2])[1]
                search_data[5] = "{}/{}".format(tv_radio_tp_count[2], tv_radio_tp_count[3])

                tv_radio_tp_accumulated[0].append(int(tv_radio_tp_count[0]))
                tv_radio_tp_accumulated[1].append(int(tv_radio_tp_count[1]))

            if other_kws[0] in data2:  # 监控删除所有节目成功的关键字
                state["delete_ch_finish_state"] = True

            if other_kws[1] in data2:  # 红外接收打印
                rsv_commd = re.split(":", data2)[-1]
                infrared_rsv_commd.append(rsv_commd)
                if rsv_commd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_commd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(infrared_send_commd[-1],
                                                                 reverse_rsv_key[infrared_rsv_commd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(len(infrared_send_commd), len(infrared_rsv_commd)))

            if other_kws[2] in data2:  # 天线设置界面获取焦点位置
                antenna_setting_focus_pos[0] = re.split(":", data2)[-1]


if __name__ == "__main__":
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)

    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIMESHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D",
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())

    PRESET_SAT_NAME = ['Nilesat', 'Hotbird', 'Badr 4/5/6/7 K', 'Thor 5/6/7', 'Turksat 2A/3A', 'BulgariaSat-1',
                       'Eutelsat 3B C', 'Eutelsat 4A', 'Eutelsat 9B', 'Amos 5 K', 'Astra 1E/3B', 'Arabsat 5A C',
                       'Arabsat 5A K', 'Eutelsat 33E K', 'C_Paksat 1R', 'Intelsat 12', 'Azerspace K', 'Intelsat 10',
                       'Yamal 202', 'Turksat 4B K', 'Belintersat K', 'TurkmenAlem', 'Yahsat 1A', 'Intelsat 707 C',
                       'Yamal 402 K', 'NSS 12 C', 'Intelsat 33e C', 'Intelsat 33e K', 'Intelsat 902 C', 'Intelsat 20 K',
                       'ABS 2/2A K', 'APSTAR 7 C', 'Thaicom 5/6 C', 'Thaicom 5/8 K', 'Express MD1 C', 'Insat 4A K',
                       'ST 2 K', 'Yamal 201 K', 'Measat 3/3A K', 'Measat 3/3A C', 'NSS 6', 'Express AM33 K',
                       'Koreasat 5 K', 'JCSat 3A K', 'JCSat 3A C', 'Vinasat 1 K', 'Telstar 18 K', 'Express AM5 K',
                       'Express AM5 C', 'Optus D1', 'Superbird B2', 'Intelsat 2/8', 'Amos 2/3/7', 'Eutelsat 5 C',
                       'Eutelsat 5 K', 'Eutelsat 8 C', 'Express AM44 K', 'Eutelsat 12', 'Telstar 12V', 'ABS-3 K',
                       'SES 4 K', 'Intelsat 905 C', 'AlComSat 1', 'Intelsat 907 C', 'Intelsat 907 K', 'Hispasat 4/5/6',
                       'Intelsat 35e', 'Intelsat 707 K', 'Intelsat 21 K', 'Amazonas 2/3 K', 'Asiasat 7 C', 'Chinas6b_C']

    NORMAL_SEARCH_TIMES = 10  # 10 普通盲扫次数
    SUPER_SEARCH_TIMES = 10  # 10 超级盲扫次数
    INCREMENTAL_SEARCH_TIMES = 15  # 15 累加搜索次数
    UPPER_LIMIT_SEARCH_TIMES = 72  # 72 上限搜索初始次数
    UPPER_LIMIT_CYCLE_TIMES = 5  # 5  上限搜索循环次数
    UPPER_LIMIT_LATER_SEARCH_TIMES = 20  # 20 上限搜索后其他情况执行测试
    ONLY_EXECUTE_ONE_TIME = 1  # 单独场景只执行一次
    NOT_UPPER_LIMIT_LATER_SEARCH_TIME = 0

    # choice_search_sat[0] = 4

    ENTER_ANTENNA_SETTING = [KEY["MENU"], KEY["OK"]]
    DELETE_ALL_SAT = [KEY["RED"], KEY["0"], KEY["RED"], KEY["OK"]]
    ADD_ONE_SAT = [KEY["GREEN"], KEY["UP"], KEY["OK"], KEY["INFO"]]
    SEARCH_PREPARATORY_WORK = [[], [DELETE_ALL_SAT, ADD_ONE_SAT]]
    CHOICE_BLIND_MODE = [KEY["RIGHT"], KEY["OK"], KEY["OK"]]
    CHOICE_SUPERBLIND_MODE = [KEY["BLUE"], KEY["RIGHT"], KEY["OK"], KEY["OK"]]
    CHOICE_NOT_SEARCH = []
    CHOICE_SAVE_TYPE = [[KEY["OK"]], [KEY["LEFT"], KEY["OK"]]]
    EXIT_ANTENNA_SETTING = [KEY["EXIT"], KEY["EXIT"]]
    NOT_OTHER_OPERATE = []
    RESET_FACTORY = [KEY["MENU"], KEY["RIGHT"], KEY["DOWN"], KEY["OK"],
                     KEY["0"], KEY["0"], KEY["0"], KEY["0"],
                     KEY["OK"]]
    DELETE_SPECIFY_SAT_ALL_TP = [KEY["GREEN"], KEY["0"], KEY["RED"], KEY["OK"]]
    DELETE_ALL_CH = [KEY["MENU"], KEY["LEFT"], KEY["LEFT"], KEY["UP"], KEY["OK"], KEY["OK"]]
    UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT = [KEY["EXIT"]]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]

    simplify_sat_name = {
        "Chinas6b_C": "Z6",
        "Asiasat 7 C": "Y3",
        "Telstar 18 K": "138",
        "ST 2 K": "88",
        "PLPD": "PLPD",
        "Reset": "Reset",
        "Delete": "Delete",
    }
    all_commd = [
        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["Chinas6b_C", "Polar=0", "5150/5750", "22K=1", "2", "0", "Blind"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["Chinas6b_C", "Polar=0", "5150/5750", "22K=1", "2", "0", "SuperBlind"],
         CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["Asiasat 7 C", "Polar=0", "5150/5750", "22K=1", "1", "0", "Blind"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["Asiasat 7 C", "Polar=0", "5150/5750", "22K=1", "1", "0", "SuperBlind"],
         CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "Blind"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "SuperBlind"],
         CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["ST 2 K", "Polar=0", "10600/0", "22K=1", "0", "0", "Blind"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["ST 2 K", "Polar=0", "10600/0", "22K=1", "0", "0", "SuperBlind"],
         CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["PLPD", "Polar=0", "5150/5750", "22K=0", "1", "0", "Blind"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         NORMAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[1],
         ["PLPD", "Polar=0", "5150/5750", "22K=0", "1", "0", "SuperBlind"],
         CHOICE_SUPERBLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         SUPER_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "Incremental"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, NOT_OTHER_OPERATE,
         INCREMENTAL_SEARCH_TIMES, NOT_UPPER_LIMIT_LATER_SEARCH_TIME],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "ChUL"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, DELETE_ALL_CH,
         UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_CYCLE_TIMES],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "ChUL_Cont."],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT,
         UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Telstar 18 K", "Polar=0", "10600/0", "22K=0", "4", "0", "ChUL_DelTp_Cont."],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[0],
         EXIT_ANTENNA_SETTING, DELETE_SPECIFY_SAT_ALL_TP,
         UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Chinas6b_C", "Polar=0", "5150/5750", "22K=1", "2", "0", "TpUL"],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[1],
         EXIT_ANTENNA_SETTING, RESET_FACTORY,
         UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_CYCLE_TIMES],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Chinas6b_C", "Polar=0", "5150/5750", "22K=1", "2", "0", "TpUL_Cont."],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[1],
         EXIT_ANTENNA_SETTING, UPPER_LIMIT_LATER_NOT_DEL_SAT_TP_SEARCH_CONT,
         UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

        [ENTER_ANTENNA_SETTING, SEARCH_PREPARATORY_WORK[0],
         ["Chinas6b_C", "Polar=0", "5150/5750", "22K=1", "2", "0", "TpUL_DelTp_Cont."],
         CHOICE_BLIND_MODE, CHOICE_SAVE_TYPE[1],
         EXIT_ANTENNA_SETTING, DELETE_SPECIFY_SAT_ALL_TP,
         UPPER_LIMIT_SEARCH_TIMES, UPPER_LIMIT_LATER_SEARCH_TIMES],

        [RESET_FACTORY, SEARCH_PREPARATORY_WORK[0],
         ["Reset", "Factory"], ONLY_EXECUTE_ONE_TIME],

        [DELETE_ALL_CH, SEARCH_PREPARATORY_WORK[0],
         ["Delete", "AllCH"], ONLY_EXECUTE_ONE_TIME],
    ]
    choice_test_case = int(sys.argv[1])

    GL = MyGlobal()
    ser_name = list(check_ports())  # send_ser_name, receive_ser_name
    send_ser = serial.Serial(ser_name[0], 9600)

    all_sat_commd = Manager().list(all_commd)
    choice_search_sat = Manager().list([choice_test_case])
    build_print_log_and_report_file_path()
    ser_name_list = Manager().list(ser_name)
    sat_para_list = Manager().list(["", "", "", "", "", ""])  # [sat_name,LNB_Power,LNB_Fre,22k,diseqc1.0,diseqc1.1]
    search_time_list = Manager().list(
        [0, 0, '', 0])  # [expect_search_time,searched_time,search_dur_time,expect_limit_search_time]
    search_data = Manager().list([0, 0, 0, 0, 0, 0, 0, 0, 0])  # 用于存放xlsx_title中的数据
    all_tp_list = Manager().list([])  # 用于存放搜索到的TP
    channel_info = Manager().dict()  # 用于存放各个TP下搜索到的电视和广播节目名称
    tv_radio_tp_count = Manager().list(["", "", "", "", ""])  # [电视节目个数，广播节目个数,保存电视节目数,保存广播节目数,保存TP个数]
    # tv_radio_tp_accumulated = Manager().list([[],[],[],[]])      # 用于统计每轮搜索累加的TV、Radio、TP数以及保存TP数的值
    record_maximum_data = Manager().list([])  # 用于存放达到上限时的打印
    infrared_send_commd = Manager().list([])  # 所有红外发送命令列表
    infrared_rsv_commd = Manager().list([])  # 所有红外接收命令列表
    antenna_setting_focus_pos = Manager().list([''])  # 天线与卫星设置界面焦点位置
    case_log_path = Manager().list([case_log_txt_path])  # 打印保存路径
    state = Manager().dict(
        {"start_record_maximum_state": False, "upper_limit_state": False, "save_ch_finish_state": False, \
         "delete_ch_finish_state": False, "search_end_state": False, "receive_loop_state": False, \
         "clear_variate_state": False, "search_start_state": False})

    rsc_p = Process(target=receive_ser_process, args=(ser_name_list, sat_para_list, search_time_list, all_sat_commd, \
                                                      choice_search_sat, search_data, all_tp_list, channel_info,
                                                      tv_radio_tp_count, \
                                                      record_maximum_data, infrared_send_commd, infrared_rsv_commd,
                                                      antenna_setting_focus_pos, state, \
                                                      case_log_path))
    rsc_p.start()

    time.sleep(5)
    # 选择执行轮次
    if len(all_sat_commd[choice_search_sat[0]]) < 9:
        search_time_list[0] = all_sat_commd[choice_search_sat[0]][-1]
    elif len(all_sat_commd[choice_search_sat[0]]) == 9:
        search_time_list[0] = all_sat_commd[choice_search_sat[0]][7]
        search_time_list[3] = all_sat_commd[choice_search_sat[0]][8]

    # 执行单次运行的场景
    if len(all_sat_commd[choice_search_sat[0]]) < 9:
        send_data = all_sat_commd[choice_search_sat[0]][0]
        for i in range(len(send_data)):
            send_commd(send_data[i])
        search_time_list[0] -= 1
        if search_time_list[0] < 1:
            logging.info("单次执行恢复出厂设置等待30秒")
            time.sleep(30)
            GL.send_loop_state = False
            state["receive_loop_state"] = True

    # 执行多次运行的场景
    elif len(all_sat_commd[choice_search_sat[0]]) == 9:
        if all_sat_commd[choice_search_sat[0]][-1] == NOT_UPPER_LIMIT_LATER_SEARCH_TIME:  # 普通搜索
            while search_time_list[0] > 0:
                enter_antenna_setting()
                judge_preparatory_work()
                antenna_setting()
                while True:
                    if state["search_end_state"]:
                        judge_srh_limit()
                        judge_save_ch_mode()
                        write_data_to_excel()
                        exit_antenna_setting()
                        judge_other_operate()
                        cyclic_srh_setting()
                        clear_variate()
                        break
                    elif not state["search_end_state"]:
                        block_send_thread()

        else:  # 上限搜索
            while search_time_list[3] >= 0:
                enter_antenna_setting()
                judge_preparatory_work()
                antenna_setting()
                while True:
                    if state["search_end_state"]:
                        judge_srh_limit()
                        judge_save_ch_mode()
                        write_data_to_excel()
                        exit_antenna_setting()
                        judge_other_operate()
                        cyclic_srh_setting()
                        clear_variate()
                        break
                    elif not state["search_end_state"]:
                        block_send_thread()

    if state["receive_loop_state"]:
        rsc_p.terminate()
        logging.info('stop receive process')
        rsc_p.join()