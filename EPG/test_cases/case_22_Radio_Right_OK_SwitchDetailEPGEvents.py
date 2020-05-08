#!/usr/bin/python3
# -*- coding: utf-8 -*-

from serial_setting import check_ports
from datetime import datetime
from random import sample,uniform
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment
from openpyxl.utils import get_column_letter,column_index_from_string
import serial
import serial.tools.list_ports
import logging
import platform
import threading
import time
import re
import os


class MyGlobal():
    def __init__(self):
        self.send_test_case_commd_numb = 10                     # 发送测试用例指令次数

        self.main_loop_state = True
        self.send_commd_state = True                            # 控制发送指令状态
        self.actual_test_numb = 0                               # 实际执行用例指令次数
        # [频道号,频道名称,tp,lock,scramble,频道类型,组别,epg_info]
        self.channel_info = ['', '', '', '', '', '', '', '']
        self.prog_group_name = ''                               # 组别名称
        self.prog_group_total = ''                              # 组别下的节目总数
        self.epg_info_exist = ''                                # 所切节目是否有EPG信息（0为没有,1为有）
        self.TV_channel_groups = {}                             # 存放电视节目的组别和节目数信息
        self.Radio_channel_groups = {}                          # 存放广播节目的组别和节目数信息
        self.TV_ch_attribute = [[], [], [], []]                 # 用于存放TV节目属性的列表(免费\加密\加锁\所有有EPG的节目)
        self.Radio_ch_attribute = [[], [], [], []]              # 用于存放Radio节目属性的列表(免费\加密\加锁\所有有EPG的节目)
        self.report_test_ch_name = ''                           # 用于输出报告记录测试的节目名称
        self.expect_report_data = ['', '', '', '', '', '']      # 用于期望输出报告的数据管理
        self.actual_report_data = ['', '', '', '', '']          # 用于实际输出报告的数据管理
        self.all_ch_epg_info = {}                               # 所有有EPG信息的节目的事件管理
        self.ch_epg_info = ['', '', '']                         # 单个EPG信息的提取[event_date, event_time, event_name]
        self.judge_switch_epg_info_end = []                     # 用于判断对比当前事件是否已经切换完成一个周期
        self.interval = 0

def hex_strs_to_bytes(strings):
    return bytes.fromhex(strings)

def send_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    time.sleep(1)

def send_numb_key_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    time.sleep(0.5)

def write_logs_to_txt(file_path,logs):
    with open(file_path, "a+", encoding="utf-8") as fo:
        fo.write(logs)

def check_test_data_directory_path_exist():
    global full_log_file_directory, report_file_directory
    parent_path = os.path.dirname(os.getcwd())
    test_file_folder_name = "test_data"
    test_file_directory = os.path.join(parent_path, test_file_folder_name)
    full_log_folder_name = "print_log"
    full_log_file_directory = os.path.join(parent_path, test_file_folder_name, full_log_folder_name)
    report_folder_name = "report"
    report_file_directory = os.path.join(parent_path, test_file_folder_name, report_folder_name)

    if not os.path.exists(test_file_directory):
        os.mkdir(test_file_directory)
    if not os.path.exists(full_log_file_directory):
        os.mkdir(full_log_file_directory)
    if not os.path.exists(report_file_directory):
        os.mkdir(report_file_directory)

def build_print_log_and_report_file_path():
    global full_log_txt_path, report_file_path, fmt_name, sheet_name
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    fmt_name = "{}_{}_{}_{}_{}".format(TEST_CASE_COMMD[0], TEST_CASE[0], TEST_CASE[1], TEST_CASE[2], TEST_CASE[3])

    full_log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    full_log_txt_path = os.path.join(full_log_file_directory, full_log_file_name)
    report_file_name = "{}_{}.xlsx".format(fmt_name, time_info)
    report_file_path = os.path.join(report_file_directory, report_file_name)
    sheet_name = "{}_{}".format(TEST_CASE[1], TEST_CASE[2])

def change_numbs_to_commds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_commds_list = []
    for i in range(len(numbs_list)):
        channel_commds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_commds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_commds_list[i].append(KEY[numbs_list[i][j]])
    return channel_commds_list

def get_group_channel_total_info():
    # 切台前获取case节目类别,分组,分组节目数量,以及获取节目属性前的去除加锁的判断
    # 根据所选case切换到对应类型节目的界面
    while GL.channel_info[5] != TEST_CASE[1]:
        send_commd(KEY["TV/R"])
        if GL.channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_commd(KEY["OK"])
    # 采集所有分组的名称和分组下节目总数信息
    if TEST_CASE[1] == "TV":
        while GL.prog_group_name not in GL.TV_channel_groups.keys():
            print(GL.prog_group_name)
            GL.TV_channel_groups[GL.prog_group_name] = GL.prog_group_total
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE[0] not in GL.TV_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的分组：{}，退出程序".format(TEST_CASE[0]))
            send_commd(KEY["EXIT"])
            GL.main_loop_state = False
    elif TEST_CASE[1] == "Radio":
        while GL.prog_group_name not in GL.Radio_channel_groups.keys():
            GL.Radio_channel_groups[GL.prog_group_name] = GL.prog_group_total
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE[0] not in GL.Radio_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的分组：{}，退出程序".format(TEST_CASE[0]))
            send_commd(KEY["EXIT"])
            GL.main_loop_state = False
    # 根据所选case切换到对应的分组
    if TEST_CASE[1] == "TV":
        while GL.prog_group_name != TEST_CASE[0]:
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    elif TEST_CASE[1] == "Radio":
        while GL.prog_group_name != TEST_CASE[0]:
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    # 退出频道列表,回到大画面界面
    send_commd(KEY["EXIT"])
    logging.debug(GL.channel_info)
    logging.debug(GL.TV_channel_groups)
    logging.debug(GL.Radio_channel_groups)

def check_ch_type():
    # 判断节目类型:TV/Radio
    while GL.channel_info[5] != TEST_CASE[1]:
        send_commd(KEY["TV/R"])
        if GL.channel_info[3] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])

def check_preparatory_work():
    if isinstance(TEST_CASE_COMMD[1], str):
        pass
    elif isinstance(TEST_CASE_COMMD[1], list):
        send_data = TEST_CASE_COMMD[1]
        for i in range(len(send_data)):
            send_commd(send_data[i])
        if GL.channel_info[3] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])

def get_group_all_ch_type(choice_group_ch_total_numb):
    send_commd(KEY["EPG"])
    for i in range(int(choice_group_ch_total_numb)):
        GL.channel_info = ['', '', '', '', '', '', GL.prog_group_name, '']
        send_commd(KEY["DOWN"])
        if GL.channel_info[7] == "1":
            time.sleep(0.5)
        elif GL.channel_info[7] == "0" or GL.channel_info[7] == '':
            time.sleep(2.5)
        # time.sleep(1)
        if GL.channel_info[3] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])
        if TEST_CASE[1] == "TV":
            if GL.channel_info[7] == "1":  # 所有有EPG信息的电视节目
                GL.TV_ch_attribute[3].append(GL.channel_info[0])
            if GL.channel_info[3] == "1":  # 加锁电视节目
                GL.TV_ch_attribute[2].append(GL.channel_info[0])
            elif GL.channel_info[4] == "0":  # 免费电视节目
                GL.TV_ch_attribute[0].append(GL.channel_info[0])
            elif GL.channel_info[4] == "1":  # 加密电视节目
                GL.TV_ch_attribute[1].append(GL.channel_info[0])
        elif TEST_CASE[1] == "Radio":
            if GL.channel_info[7] == "1":  # 所有有EPG信息的广播节目
                GL.Radio_ch_attribute[3].append(GL.channel_info[0])
            if GL.channel_info[3] == "1":  # 加锁广播节目
                GL.Radio_ch_attribute[2].append(GL.channel_info[0])
            elif GL.channel_info[4] == "0":  # 免费广播节目
                GL.Radio_ch_attribute[0].append(GL.channel_info[0])
            elif GL.channel_info[4] == "1":  # 加密广播节目
                GL.Radio_ch_attribute[1].append(GL.channel_info[0])
        logging.info(GL.channel_info)
    logging.info(GL.TV_ch_attribute)
    logging.info(GL.Radio_ch_attribute)

def check_epg_info_already_show(): # 检查EPG信息是否已经显示
    while GL.ch_epg_info[-1] == '':         # 假如还没有获取到当前节目的EPG信息，则需要退出等待5秒再进入
        GL.ch_epg_info = ['', '', '']
        send_commd(KEY["EXIT"])
        time.sleep(5)
        send_commd(KEY["EPG"])
        send_commd(KEY["RIGHT"])
    GL.ch_epg_info = ['', '', '']
    send_commd(KEY["EXIT"])
    send_commd(KEY["EPG"])

def choice_test_channel():
    if TEST_CASE[1] == "TV":
        if len(GL.TV_ch_attribute[3]) == 0:
            logging.info("无有EPG信息的电视节目")
        elif len(GL.TV_ch_attribute[3]) > 0:
            for i in range(len(GL.TV_ch_attribute[3])):
                free_tv_numb = GL.TV_ch_attribute[3][i]
                logging.debug("当前所选有EPG信息的电视节目频道号为:{}".format(free_tv_numb))
                free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                send_commd(KEY["EXIT"])
                for i in range(len(free_tv_commd)):
                    for j in range(len(free_tv_commd[i])):
                        send_numb_key_commd(free_tv_commd[i][j])
                send_commd(KEY["OK"])
                time.sleep(2)
                logging.info("当前所选有EPG信息的电视节目名称为:{}".format(GL.channel_info[1]))
                logging.info(GL.channel_info)

                # 将有EPG信息的节目名称添加到字典
                if GL.channel_info[1] not in GL.all_ch_epg_info.keys():
                    GL.all_ch_epg_info[GL.channel_info[1]] = []
                    logging.info(GL.all_ch_epg_info.keys())

                check_preparatory_work()
                check_epg_info_already_show()
                time.sleep(1)
                send_test_case_commd()
                padding_report_data()
                write_data_to_report()
                send_commd(KEY["EXIT"])
    elif TEST_CASE[1] == "Radio":
        if len(GL.Radio_ch_attribute[3]) == 0:
            logging.info("无有EPG信息的广播节目")
        elif len(GL.Radio_ch_attribute[3]) > 0:
            for i in range(len(GL.Radio_ch_attribute[3])):
                free_radio_numb = GL.Radio_ch_attribute[3][i]
                logging.debug("当前所选有EPG信息的广播节目频道号为:{}".format(free_radio_numb))
                free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                send_commd(KEY["EXIT"])
                for i in range(len(free_radio_commd)):
                    for j in range(len(free_radio_commd[i])):
                        send_numb_key_commd(free_radio_commd[i][j])
                send_commd(KEY["OK"])
                time.sleep(2)
                logging.info("当前所选有EPG信息的广播节目名称为:{}".format(GL.channel_info[1]))
                logging.info(GL.channel_info)

                # 将有EPG信息的节目名称添加到字典
                if GL.channel_info[1] not in GL.all_ch_epg_info.keys():
                    GL.all_ch_epg_info[GL.channel_info[1]] = []
                    logging.info(GL.all_ch_epg_info.keys())

                check_preparatory_work()
                check_epg_info_already_show()
                time.sleep(1)
                send_test_case_commd()
                padding_report_data()
                write_data_to_report()
                send_commd(KEY["EXIT"])

def send_test_case_commd():
    GL.send_commd_state = True
    if TEST_CASE[2] == "Left+OK":
        send_data = TEST_CASE_COMMD[2][0]
    elif TEST_CASE[2] == "Right+OK":
        send_data = TEST_CASE_COMMD[2][1]
    while GL.send_commd_state:
        GL.ch_epg_info = ['', '', '']
        send_commd(send_data)
        sleep_time = uniform(0.75, 1.0)
        logging.info(sleep_time)
        time.sleep(sleep_time)
        if GL.ch_epg_info not in GL.all_ch_epg_info[GL.channel_info[1]] and GL.ch_epg_info != ['', '', '']:
            GL.all_ch_epg_info[GL.channel_info[1]].append(GL.ch_epg_info)
            send_commd(KEY["OK"])
            send_commd(KEY["PAGE_DOWN"])
            send_commd(KEY["EXIT"])
        if len(GL.all_ch_epg_info[GL.channel_info[1]]) >= 20:
            GL.judge_switch_epg_info_end = GL.all_ch_epg_info[GL.channel_info[1]][:10]
            if GL.ch_epg_info in GL.judge_switch_epg_info_end:
                GL.send_commd_state = False
                GL.judge_switch_epg_info_end = []

def write_data_to_report():
    expect_report_title = [
        "报告名称",
        "期望分组名称",
        "期望分组节目总数",
        "期望节目类别",
        "期望指令",
        "期望有EPG节目数",
    ]
    actual_report_title = [
        "实际分组名称",
        "实际分组节目总数",
        "实际节目类别",
        "实际指令",
        "实际有EPG节目数",
    ]
    channel_info_title = ["频道名称", "EPG事件日期", "EPG事件时段", "EPG事件名称"]

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    GL.interval = 1 + (len(GL.all_ch_epg_info.keys()) - 1) * len(channel_info_title)    # 根据抬头和当前已经获取到的节目个数来设置间隔

    if not os.path.exists(report_file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['D'].width = 17
        for i in range(len(expect_report_title)):  # 写期望测试项的title信息
            ws.row_dimensions[(i + 1)].height = 13.5
            ws.cell(i + 1, 1).value = expect_report_title[i]
            ws.cell(i + 1, 1).alignment = alignment
        for j in range(len(actual_report_title)):   # 写实际测试项的title信息
            # ws.row_dimensions[(j + 1)].height = 13.5
            ws.cell(j + 2, 4).value = actual_report_title[j]
            ws.cell(j + 2, 4).alignment = alignment

        for k in range(len(channel_info_title)):    # 根据节目个数循环写EPG信息的title信息，并设置列宽
            all_column_numb = column_index_from_string("A") + k + GL.interval
            all_column_char = get_column_letter(all_column_numb)
            ws.column_dimensions[all_column_char].width = 16  # 设置列宽
            ws.cell(len(expect_report_title) + 1, GL.interval + k).value = channel_info_title[k]
            ws.cell(len(expect_report_title) + 1, GL.interval + k).alignment = alignment

    elif os.path.exists(report_file_path):
        wb = load_workbook(report_file_path)
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['D'].width = 17
        for i in range(len(expect_report_title)):       # 写期望测试项的title信息
            ws.row_dimensions[(i + 1)].height = 13.5
            ws.cell(i + 1, 1).value = expect_report_title[i]
            ws.cell(i + 1, 1).alignment = alignment
        for j in range(len(actual_report_title)):       # 写实际测试项的title信息
            ws.cell(j + 2, 4).value = actual_report_title[j]
            ws.cell(j + 2, 4).alignment = alignment

        for k in range(len(channel_info_title)):        # 根据节目个数循环写EPG信息的title信息，并设置列宽
            all_column_numb = column_index_from_string("A") + k + GL.interval
            all_column_char = get_column_letter(all_column_numb)
            ws.column_dimensions[all_column_char].width = 16  # 设置列宽
            ws.cell(len(expect_report_title) + 1, GL.interval + k).value = channel_info_title[k]
            ws.cell(len(expect_report_title) + 1, GL.interval + k).alignment = alignment

    for m in range(len(GL.expect_report_data)):         # 写期望测试项的期望结果
        ws.cell(m + 1, 2).value = GL.expect_report_data[m]
        ws.cell(m + 1, 2).alignment = alignment
        if m == 0:
            ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=6)
        else:
            ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=3)
    for n in range(len(GL.actual_report_data)):         # 写实际测试项的测试结果
        ws.cell(n + 2, 5).value = GL.actual_report_data[n]
        ws.merge_cells(start_row=(n + 2), start_column=5, end_row=(n + 2), end_column=6)
        ws.cell(n + 2, 5).alignment = alignment
    # for x in range(len(GL.all_ch_epg_info.keys())):
    #     ch_name_key = list(GL.all_ch_epg_info.keys())[x]
    ch_name_key = list(GL.all_ch_epg_info.keys())[len(GL.all_ch_epg_info.keys()) - 1]
    for y in range(len(GL.all_ch_epg_info[ch_name_key])):       # 写每个有EPG信息节目下切换后的获取到的EPG信息
        ws.cell(len(expect_report_title) + 2 + y, GL.interval).value = ch_name_key
        ws.cell(len(expect_report_title) + 2 + y, GL.interval).alignment = alignment
        ws.row_dimensions[(len(expect_report_title) + 2 + y)].height = 13.5
        for z in range(len(GL.all_ch_epg_info[ch_name_key][y])):
            ws.cell(len(expect_report_title) + 2 + y, z + GL.interval + 1).value = GL.all_ch_epg_info[ch_name_key][y][z]
            ws.cell(len(expect_report_title) + 2 + y, z + GL.interval + 1).alignment = alignment
            # ws.cell(len(expect_report_title) + 2 + x, y + 1).value = GL.epg_switch_ch_data_report[x][y]
            # ws.cell(len(expect_report_title) + 2 + x, y + 1).alignment = alignment

    wb.save(report_file_path)

def padding_report_data():
    # 期望的报告数据
    GL.expect_report_data[0] = "{}_{}_{}".format(TEST_CASE[1], TEST_CASE[2], TEST_CASE[3])
    GL.expect_report_data[1] = TEST_CASE[0]
    GL.expect_report_data[2] = "None"
    GL.expect_report_data[3] = TEST_CASE[1]
    GL.expect_report_data[4] = TEST_CASE[2]
    GL.expect_report_data[5] = "None"

    # 实际的测试数据
    GL.actual_report_data[0] = GL.channel_info[6]
    GL.actual_report_data[2] = GL.channel_info[5]
    GL.actual_report_data[3] = "{} + OK".format(REVERSE_KEY[SEND_TEST_CASE_COMMD[1]])
    if TEST_CASE[1] == "TV":
        GL.actual_report_data[1] = GL.TV_channel_groups[GL.channel_info[6]]
        GL.actual_report_data[4] = len(GL.TV_ch_attribute[3])
    elif TEST_CASE[1] == "Radio":
        GL.actual_report_data[1] = GL.Radio_channel_groups[GL.channel_info[6]]
        GL.actual_report_data[4] = len(GL.Radio_ch_attribute[3])

def get_choice_group_ch_type():
    # 采集All分组下的节目属性和是否有EPG信息
    if TEST_CASE[1] == "TV":
        get_group_all_ch_type(GL.TV_channel_groups[TEST_CASE[0]])
    elif TEST_CASE[1] == "Radio":
        get_group_all_ch_type(GL.Radio_channel_groups[TEST_CASE[0]])

def exit_to_screen():
    send_data = TEST_CASE_COMMD[3]
    for i in range(len(send_data)):
        send_commd(send_data[i])

def data_send_thread():
    get_group_channel_total_info()
    get_choice_group_ch_type()
    check_ch_type()
    choice_test_channel()
    exit_to_screen()
    GL.main_loop_state = False

def data_receiver_thread():
    while GL.main_loop_state:
        data = receive_ser.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            data1 = data.decode("ISO-8859-1", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data4 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_logs_to_txt(full_log_txt_path, data4)

            if SWITCH_CHANNEL_KWS[0] in data2:
                ch_info_split = re.split(r"[\],]", data2)
                for i in range(len(ch_info_split)):
                    if CH_INFO_KWS[0] in ch_info_split[i]:  # 提取频道号
                        GL.channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if CH_INFO_KWS[1] in ch_info_split[i]:  # 提取频道名称
                        GL.channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[1] in data2:
                flag_info_split = re.split(r"[\],]", data2)
                for i in range(len(flag_info_split)):
                    if CH_INFO_KWS[2] in flag_info_split[i]:  # 提取频道所属TP
                        GL.channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if CH_INFO_KWS[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        GL.channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        GL.channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        GL.channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[3] in data2:
                group_info_split = re.split(r"[\],]", data2)
                for i in range(len(group_info_split)):
                    if GROUP_INFO_KWS[0] in group_info_split[i]:  # 提取频道所属组别
                        GL.prog_group_name = re.split(r"=", group_info_split[i])[-1]
                        GL.channel_info[6] = GL.prog_group_name
                    if GROUP_INFO_KWS[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        GL.prog_group_total = re.split(r"=", group_info_split[i])[-1]

            if EPG_SC_KWS[0] in data2:
                epg_info_split = re.split(r"]", data2)
                for i in range(len(epg_info_split)):
                    if EPG_INFO_KWS[0] in epg_info_split[i]:
                        GL.epg_info_exist = re.split(r"=", epg_info_split[i])[-1]
                        GL.channel_info[7] = GL.epg_info_exist

            if EPG_SC_KWS[1] in data2:
                epg_event_split = re.split(r"t:|,", data2)
                for i in range(len(epg_event_split)):
                    if EPG_INFO_KWS[1] in epg_event_split[i]:
                        time_info_split = re.split(r"=|--|\s", epg_event_split[i])
                        if time_info_split[1] == time_info_split[3]:
                            event_date = "{}".format(time_info_split[1])
                        elif time_info_split[1] != time_info_split[3]:
                            event_date = "{}-{}".format(time_info_split[1], time_info_split[3])
                        event_time = "{}-{}".format(time_info_split[2][:5], time_info_split[4][:5])
                    if EPG_INFO_KWS[2] in epg_event_split[i]:
                        event_name = re.split(r"=", epg_event_split[i])[-1]
                GL.ch_epg_info[0] = event_date
                GL.ch_epg_info[1] = event_time
                GL.ch_epg_info[2] = event_name


if __name__ == "__main__":
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "ZOOM": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "R/L": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "FIND": "A1 F1 22 DD 46", "PAUSE": "A1 F1 22 DD 45", "SUB": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIMESHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D",

        "CH+": "A1 F1 22 DD 11", "CH-": "A1 F1 22 DD 14", "VOL-": "A1 F1 22 DD 12", "VOL+": "A1 F1 22 DD 13",
        "MULTIFEED": "A1 F1 22 DD 0F", "SAT": "A1 F1 22 DD 17", "AUDIO": "A1 F1 22 DD 19", "TTX": "A1 F1 22 DD 1C",
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())

    SWITCH_CHANNEL_KWS = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    CH_INFO_KWS = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name", ]

    GROUP_INFO_KWS = [
        "Group_name",
        "Prog_total"
    ]

    EPG_SC_KWS = [
        "[PTD]Program_epg_info=",
        "[PTD]EPG_event:event_time="
    ]

    EPG_INFO_KWS = [
        "Program_epg_info",
        "event_time",
        "event_name"
    ]

    PREPARATORY_WORK = [KEY["EPG"]]
    SEND_TEST_CASE_COMMD = [KEY["LEFT"], KEY["RIGHT"]]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]

    TEST_CASE = ["GX", "Radio", "Right+OK", "SwitchDetailEPGEvents"]
    TEST_CASE_COMMD = ["22", PREPARATORY_WORK, SEND_TEST_CASE_COMMD, EXIT_TO_SCREEN]

    GL = MyGlobal()
    # 检查打印日志和报告的目录,以及创建文件名称
    check_test_data_directory_path_exist()
    build_print_log_and_report_file_path()

    # 配置日志信息
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)

    # 获取串口并配置串口信息
    send_com = ''
    receive_com = ''
    send_ser_name, receive_ser_name = check_ports()
    send_ser = serial.Serial(send_ser_name, 9600)
    receive_ser = serial.Serial(receive_ser_name, 115200, timeout=1)

    msg = "[现在开始执行:{}]".format(fmt_name)
    logging.critical(format(msg, '*^150'))

    thread_send = threading.Thread(target=data_send_thread)
    thread_receive = threading.Thread(target=data_receiver_thread)
    thread_receive.start()
    thread_send.start()