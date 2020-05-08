#!/usr/bin/python3
# -*- coding: utf-8 -*-

from serial_setting import check_ports
from datetime import datetime
from random import sample,uniform
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment
from openpyxl.utils import get_column_letter,column_index_from_string
import serial
import serial.tools.list_ports
import logging
import platform
import threading
import time
import re
import os


class MyGlobal():
    def __init__(self):
        self.send_test_case_commd_numb = 10                     # 发送测试用例指令次数

        self.main_loop_state = True
        self.actual_test_numb = 0                               # 实际执行用例指令次数
        # [频道号,频道名称,tp,lock,scramble,频道类型,组别,epg_info]
        self.channel_info = ['', '', '', '', '', '', '', '']
        self.prog_group_name = ''                               # 组别名称
        self.prog_group_total = ''                              # 组别下的节目总数
        self.epg_info_exist = ''                                # 所切节目是否有EPG信息（0为没有,1为有）
        self.TV_channel_groups = {}                             # 存放电视节目的组别和节目数信息
        self.Radio_channel_groups = {}                          # 存放广播节目的组别和节目数信息
        self.TV_ch_attribute = [[], [], [], []]                 # 用于存放TV节目属性的列表(免费\加密\加锁\免费且有EPG的节目)
        self.Radio_ch_attribute = [[], [], [], []]              # 用于存放Radio节目属性的列表(免费\加密\加锁\免费且有EPG的节目)
        self.report_test_ch_name = ''                           # 用于输出报告记录测试的节目名称
        self.expect_report_data = ['', '', '', '', '', '']      # 用于期望输出报告的数据管理
        self.actual_report_data = ['', '', '', '', '']          # 用于实际输出报告的数据管理
        self.epg_switch_ch_data_report = []                     # 用于EPG界面切台时的节目信息记录

def hex_strs_to_bytes(strings):
    return bytes.fromhex(strings)

def send_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    time.sleep(1)

def send_numb_key_commd(commd):
    send_ser.write(hex_strs_to_bytes(commd))
    time.sleep(0.5)

def write_logs_to_txt(file_path,logs):
    with open(file_path, "a+", encoding="utf-8") as fo:
        fo.write(logs)

def check_test_data_directory_path_exist():
    global full_log_file_directory, report_file_directory
    parent_path = os.path.dirname(os.getcwd())
    test_file_folder_name = "test_data"
    test_file_directory = os.path.join(parent_path, test_file_folder_name)
    full_log_folder_name = "print_log"
    full_log_file_directory = os.path.join(parent_path, test_file_folder_name, full_log_folder_name)
    report_folder_name = "report"
    report_file_directory = os.path.join(parent_path, test_file_folder_name, report_folder_name)

    if not os.path.exists(test_file_directory):
        os.mkdir(test_file_directory)
    if not os.path.exists(full_log_file_directory):
        os.mkdir(full_log_file_directory)
    if not os.path.exists(report_file_directory):
        os.mkdir(report_file_directory)

def build_print_log_and_report_file_path():
    global full_log_txt_path, report_file_path, fmt_name, sheet_name
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    fmt_name = "{}_{}_{}_{}_{}".format(TEST_CASE_COMMD[0], TEST_CASE[0], TEST_CASE[1], TEST_CASE[2], TEST_CASE[3])

    full_log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    full_log_txt_path = os.path.join(full_log_file_directory, full_log_file_name)
    report_file_name = "{}_{}.xlsx".format(fmt_name, time_info)
    report_file_path = os.path.join(report_file_directory, report_file_name)
    sheet_name = "{}_{}".format(TEST_CASE[1], TEST_CASE[2])

def get_group_channel_total_info():
    # 切台前获取case节目类别,分组,分组节目数量,以及获取节目属性前的去除加锁的判断
    # 根据所选case切换到对应类型节目的界面
    while GL.channel_info[5] != TEST_CASE[1]:
        send_commd(KEY["TV/R"])
        if GL.channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_commd(KEY["OK"])
    # 采集所有分组的名称和分组下节目总数信息
    if TEST_CASE[1] == "TV":
        while GL.prog_group_name not in GL.TV_channel_groups.keys():
            print(GL.prog_group_name)
            GL.TV_channel_groups[GL.prog_group_name] = GL.prog_group_total
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE[0] not in GL.TV_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的分组：{}，退出程序".format(TEST_CASE[0]))
            send_commd(KEY["EXIT"])
            GL.main_loop_state = False
    elif TEST_CASE[1] == "Radio":
        while GL.prog_group_name not in GL.Radio_channel_groups.keys():
            GL.Radio_channel_groups[GL.prog_group_name] = GL.prog_group_total
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE[0] not in GL.Radio_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的分组：{}，退出程序".format(TEST_CASE[0]))
            send_commd(KEY["EXIT"])
            GL.main_loop_state = False
    # 根据所选case切换到对应的分组
    if TEST_CASE[1] == "TV":
        while GL.prog_group_name != TEST_CASE[0]:
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    elif TEST_CASE[1] == "Radio":
        while GL.prog_group_name != TEST_CASE[0]:
            send_commd(KEY["RIGHT"])
            if GL.channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    # 退出频道列表,回到大画面界面
    send_commd(KEY["EXIT"])
    logging.debug(GL.channel_info)
    logging.debug(GL.TV_channel_groups)
    logging.debug(GL.Radio_channel_groups)

def check_ch_type():
    while GL.channel_info[5] != TEST_CASE[1]:
        send_commd(KEY["TV/R"])
        if GL.channel_info[3] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])

def check_preparatory_work():
    if isinstance(TEST_CASE_COMMD[1], str):
        pass
    elif isinstance(TEST_CASE_COMMD[1], list):
        send_data = TEST_CASE_COMMD[1]
        for i in range(len(send_data)):
            send_commd(send_data[i])
        if GL.channel_info[3] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])

def send_test_case_commd():
    send_data = TEST_CASE_COMMD[2]
    for i in range(GL.send_test_case_commd_numb):
        GL.channel_info = ['', '', '', '', '', '', GL.prog_group_name, '']
        send_numb_key_commd(send_data[0])
        sleep_time = uniform(0.5, 1.5)
        logging.info(sleep_time)
        time.sleep(sleep_time)
        if GL.channel_info[3] == "1":
            for i in range(4):
                send_numb_key_commd(KEY["0"])
        logging.info("当前执行次数:{}".format(i + 1))
        GL.epg_switch_ch_data_report.append(GL.channel_info)
        logging.info(GL.channel_info)
        GL.actual_test_numb = i + 1

def write_data_to_report():
    expect_report_title = [
        "报告名称",
        "期望分组名称",
        "期望分组节目总数",
        "期望节目类别",
        "期望切台模式",
        "期望执行次数",
    ]
    actual_report_title = [
        "实际分组名称",
        "实际分组节目总数",
        "实际节目类别",
        "实际切台模式",
        "实际执行次数",
    ]
    channel_info_title = ["频道号", "频道名称", "TP", "加锁标志", "加密标志", "频道类型", "组别", "epg_info"]

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    if not os.path.exists(report_file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['D'].width = 17
        for i in range(len(expect_report_title)):
            ws.row_dimensions[(i + 1)].height = 13.5
            ws.cell(i + 1, 1).value = expect_report_title[i]
            ws.cell(i + 1, 1).alignment = alignment
        for j in range(len(actual_report_title)):
            # ws.row_dimensions[(j + 1)].height = 13.5
            ws.cell(j + 2, 4).value = actual_report_title[j]
            ws.cell(j + 2, 4).alignment = alignment
        for k in range(len(channel_info_title)):
            all_column_numb = column_index_from_string("A") + k
            all_column_char = get_column_letter(all_column_numb)
            ws.column_dimensions[all_column_char].width = 16        # 设置列宽
            ws.cell(len(expect_report_title) + 1, k + 1).value = channel_info_title[k]
            ws.cell(len(expect_report_title) + 1, k + 1).alignment = alignment

    elif os.path.exists(report_file_path):
        wb = load_workbook(report_file_path)
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['D'].width = 17
        for i in range(len(expect_report_title)):
            ws.row_dimensions[(i + 1)].height = 13.5
            ws.cell(i + 1, 1).value = expect_report_title[i]
            ws.cell(i + 1, 1).alignment = alignment
        for j in range(len(actual_report_title)):
            ws.cell(j + 2, 4).value = actual_report_title[j]
            ws.cell(j + 2, 4).alignment = alignment
        for k in range(len(channel_info_title)):
            ws.cell(len(expect_report_title) + 1, k + 1).value = channel_info_title[k]
            ws.cell(len(expect_report_title) + 1, k + 1).alignment = alignment

    for m in range(len(GL.expect_report_data)):
        ws.cell(m + 1, 2).value = GL.expect_report_data[m]
        ws.cell(m + 1, 2).alignment = alignment
        if m == 0:
            ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=6)
        else:
            ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=3)
    for n in range(len(GL.actual_report_data)):
        ws.cell(n + 2, 5).value = GL.actual_report_data[n]
        ws.merge_cells(start_row=(n + 2), start_column=5, end_row=(n + 2), end_column=6)
        ws.cell(n + 2, 5).alignment = alignment
    for x in range(len(GL.epg_switch_ch_data_report)):
        for y in range(len(GL.epg_switch_ch_data_report[x])):
            ws.cell(len(expect_report_title) + 2 + x, y + 1).value = GL.epg_switch_ch_data_report[x][y]
            ws.cell(len(expect_report_title) + 2 + x, y + 1).alignment = alignment

    wb.save(report_file_path)

def padding_report_data():
    # 期望的报告数据
    GL.expect_report_data[0] = "{}_{}_{}".format(TEST_CASE[1], TEST_CASE[2], TEST_CASE[3])
    GL.expect_report_data[1] = TEST_CASE[0]
    GL.expect_report_data[2] = "None"
    GL.expect_report_data[3] = TEST_CASE[1]
    GL.expect_report_data[4] = TEST_CASE[2]
    GL.expect_report_data[5] = GL.send_test_case_commd_numb

    # 实际的测试数据
    GL.actual_report_data[0] = GL.channel_info[6]
    GL.actual_report_data[2] = GL.channel_info[5]
    GL.actual_report_data[3] = REVERSE_KEY[SEND_TEST_CASE_COMMD[0]]
    GL.actual_report_data[4] = GL.actual_test_numb
    if TEST_CASE[1] == "TV":
        GL.actual_report_data[1] = GL.TV_channel_groups[GL.channel_info[6]]
    elif TEST_CASE[1] == "Radio":
        GL.actual_report_data[1] = GL.Radio_channel_groups[GL.channel_info[6]]

def exit_to_screen():
    send_data = TEST_CASE_COMMD[3]
    for i in range(len(send_data)):
        send_commd(send_data[i])

def data_send_thread():
    get_group_channel_total_info()
    check_ch_type()
    check_preparatory_work()
    # 发送所选用例的操作指令,汇总报告数据，写报告数据
    send_test_case_commd()
    padding_report_data()
    write_data_to_report()
    exit_to_screen()
    GL.main_loop_state = False

def data_receiver_thread():
    while GL.main_loop_state:
        data = receive_ser.readline()
        if data:
            tt = datetime.now()
            data1 = data.decode("GB18030", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data4 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_logs_to_txt(full_log_txt_path, data4)

            if SWITCH_CHANNEL_KWS[0] in data2:
                ch_info_split = re.split(r"[\],]", data2)
                for i in range(len(ch_info_split)):
                    if CH_INFO_KWS[0] in ch_info_split[i]:  # 提取频道号
                        GL.channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if CH_INFO_KWS[1] in ch_info_split[i]:  # 提取频道名称
                        GL.channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[1] in data2:
                flag_info_split = re.split(r"[\],]", data2)
                for i in range(len(flag_info_split)):
                    if CH_INFO_KWS[2] in flag_info_split[i]:  # 提取频道所属TP
                        GL.channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if CH_INFO_KWS[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        GL.channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        GL.channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if CH_INFO_KWS[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        GL.channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if SWITCH_CHANNEL_KWS[3] in data2:
                group_info_split = re.split(r"[\],]", data2)
                for i in range(len(group_info_split)):
                    if GROUP_INFO_KWS[0] in group_info_split[i]:  # 提取频道所属组别
                        GL.prog_group_name = re.split(r"=", group_info_split[i])[-1]
                        GL.channel_info[6] = GL.prog_group_name
                    if GROUP_INFO_KWS[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        GL.prog_group_total = re.split(r"=", group_info_split[i])[-1]

            if EPG_SC_KWS[0] in data2:
                epg_info_split = re.split(r"]", data2)
                for i in range(len(epg_info_split)):
                    if EPG_INFO_KWS[0] in epg_info_split[i]:
                        GL.epg_info_exist = re.split(r"=", epg_info_split[i])[-1]
                        GL.channel_info[7] = GL.epg_info_exist

if __name__ == "__main__":
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "ZOOM": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "R/L": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "FIND": "A1 F1 22 DD 46", "PAUSE": "A1 F1 22 DD 45", "SUB": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIMESHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D",

        "CH+": "A1 F1 22 DD 11", "CH-": "A1 F1 22 DD 14", "VOL-": "A1 F1 22 DD 12", "VOL+": "A1 F1 22 DD 13",
        "MULTIFEED": "A1 F1 22 DD 0F", "SAT": "A1 F1 22 DD 17", "AUDIO": "A1 F1 22 DD 19", "TTX": "A1 F1 22 DD 1C",
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())

    SWITCH_CHANNEL_KWS = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    CH_INFO_KWS = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name", ]

    GROUP_INFO_KWS = [
        "Group_name",
        "Prog_total"
    ]

    EPG_SC_KWS = [
        "[PTD]Program_epg_info=",
        "[PTD]EPG_event:event_time="
    ]

    EPG_INFO_KWS = [
        "Program_epg_info",
        "event_time",
        "event_name"
    ]

    PREPARATORY_WORK = [KEY["EPG"]]
    SEND_TEST_CASE_COMMD = [KEY["DOWN"]]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]

    TEST_CASE = ["GX", "TV", "Down", "EPGSwitchChannel"]
    TEST_CASE_COMMD = ["04", PREPARATORY_WORK, SEND_TEST_CASE_COMMD, EXIT_TO_SCREEN]

    GL = MyGlobal()
    # 检查打印日志和报告的目录,以及创建文件名称
    check_test_data_directory_path_exist()
    build_print_log_and_report_file_path()

    # 配置日志信息
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)

    # 获取串口并配置串口信息
    send_com = ''
    receive_com = ''
    send_ser_name, receive_ser_name = check_ports()
    send_ser = serial.Serial(send_ser_name, 9600)
    receive_ser = serial.Serial(receive_ser_name, 115200, timeout=1)

    msg = "[现在开始执行:{}]".format(fmt_name)
    logging.critical(format(msg, '*^150'))

    thread_send = threading.Thread(target=data_send_thread)
    thread_receive = threading.Thread(target=data_receiver_thread)
    thread_receive.start()
    thread_send.start()