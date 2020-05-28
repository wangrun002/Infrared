#!/usr/bin/python3
# -*- coding: utf-8 -*-

from serial_setting import *
from multiprocessing import Process, Manager
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from random import randint
import platform
import os
import time
import logging
import re


class MyGlobal(object):

    def __init__(self):
        self.add_res_event_numb = 5                     # 预约事件响应次数
        self.choice_res_ch = ''                         # 预约Play或PVR事件时所选预约节目
        self.res_event_mgr = []                         # 预约事件管理
        self.report_data = [[], '', '', '', '', '']     # 报告数据汇总[[预约事件信息]，"触发时间", "是否跳转", "跳转节目", "是否录制", "录制时长"]
        self.title_data = ['', '', '', '', '', '']      # ["报告名称", "预约事件类型", "预约事件模式", "预约节目类型", "预约等待界面", "预约跳转模式"]
        self.start_row = 0                              # 用于每次预约事件响应后，写数据增加行数
        self.pvr_rec_dur_time = ''                      # 用于记录PVR事件录制持续时间


def logging_info_setting():
    # 配置logging输出格式
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.INFO, format=LOG_FORMAT, datefmt=DATE_FORMAT)


def hex_strs_to_bytes(strings):
    # 将红外命令字符串转换为字节串
    return bytes.fromhex(strings)


def write_log_data_to_txt(path, write_data):
    with open(path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def build_send_and_receive_serial():
    # 创建发送和接受串口
    ser_name = list(check_ports())  # send_ser_name, receive_ser_name
    send_ser = serial.Serial(ser_name[0], 9600)
    receive_ser = serial.Serial(ser_name[1], 115200, timeout=1)
    return send_ser, receive_ser


def send_commd(commd):
    # 红外发送端发送指令
    send_serial.write(hex_strs_to_bytes(commd))
    send_serial.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
    infrared_send_cmd.append(REVERSE_KEY[commd])
    time.sleep(1.0)


def send_more_commds(commd_list):
    # 用于发送一连串的指令
    for commd in commd_list:
        send_commd(commd)
    time.sleep(1)   # 增加函数切换时的的等待，避免可能出现send_commd函数中的等待时间没有执行的情况


def change_numbs_to_commds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_commds_list = []
    for i in range(len(numbs_list)):
        channel_commds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_commds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_commds_list[i].append(KEY[numbs_list[i][j]])
    return channel_commds_list


def build_log_and_report_file_path():
    # 用于创建打印和报告文件路径
    # 构建存放数据的总目录，以及构建存放打印和报告的目录
    parent_path = os.path.dirname(os.getcwd())
    test_data_directory_name = "test_data"
    test_data_directory_path = os.path.join(parent_path, test_data_directory_name)
    log_directory_name = "print_log"
    log_directory_path = os.path.join(test_data_directory_path, log_directory_name)
    report_directory_name = "report"
    report_directory_path = os.path.join(test_data_directory_path, report_directory_name)
    # 判断目录是否存在，否则创建目录
    if not os.path.exists(test_data_directory_path):
        os.mkdir(test_data_directory_path)
    if not os.path.exists(log_directory_path):
        os.mkdir(log_directory_path)
    if not os.path.exists(report_directory_path):
        os.mkdir(report_directory_path)
    # 创建打印和报告文件的名称和路径
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    fmt_name = "{}_{}_{}_{}_event_{}_triggered".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
        TEST_CASE_INFO[3], TEST_CASE_INFO[4], TEST_CASE_INFO[5])
    log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    log_file_path = os.path.join(log_directory_path, log_file_name)
    report_file_name = "{}_{}.xlsx".format(fmt_name, time_info)
    report_file_path = os.path.join(report_directory_path, report_file_name)
    sheet_name = "{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[4])
    return log_file_path, report_file_path, sheet_name


def clear_timer_setting_all_events():
    # logging_info_setting()
    # 清除Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    delete_all_res_events = [KEY["BLUE"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    # 进入定时器设置界面
    send_more_commds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        if rsv_kws["res_event_numb"] != '0':
            send_more_commds(delete_all_res_events)
        elif rsv_kws["res_event_numb"] == '0':
            logging.info("没有预约事件存在")
            time.sleep(1)
        else:
            logging.debug("警告：预约事件个数获取错误！！！")
        state["res_event_numb_state"] = False
    # 退回大画面
    send_more_commds(exit_to_screen)


def check_sys_time_mode():
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    # 进入时间设置界面
    send_more_commds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        if rsv_kws["sys_time_mode"] == "Auto":
            send_more_commds(change_sys_time_mode)
        elif rsv_kws["sys_time_mode"] == "Manual":
            logging.info("系统时间模式已经为手动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    # 退回大画面
    send_more_commds(exit_to_screen)


def get_current_system_time():
    # 获取当前系统时间
    time.sleep(1)
    while not state["current_sys_time_state"]:
        logging.info("还没有获取到系统时间信息")
        time.sleep(1)
    else:
        logging.info(f"当前系统时间为:{rsv_kws['current_sys_time']}")


def get_exist_event_info():
    # 获取当前已预约的事件信息
    if len(res_event_list) == 0:
        logging.info("当前没有预约事件")
        logging.info(f"预约事件列表为:{res_event_list}")
    else:
        logging.info("当前所有预约事件如下")
        for event in res_event_list:
            logging.info(event)


def choice_ch_for_res_event_type():
    # 根据预约事件类型来选择节目
    group_dict = {}
    choice_ch_numb = []
    # 根据所选case切换到对应类型节目的界面
    while channel_info[5] != TEST_CASE_INFO[2]:
        send_commd(KEY["TV/R"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_commd(KEY["OK"])
    # 切到指定分组和采集分组下节目总数信息
    while rsv_kws["prog_group_name"] != TEST_CASE_INFO[1]:
        send_commd(KEY["RIGHT"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    else:
        if rsv_kws["prog_group_name"] == '':
            logging.info("警告：没有All分组信息")
        else:
            group_dict[rsv_kws["prog_group_name"]] = rsv_kws["prog_group_total"]
            logging.info(f"分组信息{group_dict}")
    # 退出频道列表,回到大画面界面
    send_commd(KEY["EXIT"])
    # 根据用例指定的事件类型来选择节目
    if TEST_CASE_INFO[4] == "Play":
        choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
        choice_ch_cmd = change_numbs_to_commds_list(choice_ch_numb)
        for i in range(len(choice_ch_cmd)):
            for j in choice_ch_cmd[i]:
                send_commd(j)
        send_commd(KEY["OK"])
        time.sleep(2)
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
        logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
        GL.choice_res_ch = channel_info[1]
        logging.info(channel_info)

    elif TEST_CASE_INFO[4] == "PVR":
        choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
        choice_ch_cmd = change_numbs_to_commds_list(choice_ch_numb)
        for i in range(len(choice_ch_cmd)):
            for j in choice_ch_cmd[i]:
                send_commd(j)
        send_commd(KEY["OK"])
        time.sleep(2)
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])

        while channel_info[3] != '0' and channel_info[4] != '0':
            logging.info(f"查看所切节目信息：{channel_info}")
            logging.info("所选节目不为免费节目，不可以进行PVR预约，继续切台")
            send_commd(KEY["UP"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        else:
            logging.info("所选节目为免费节目，可以进行PVR预约")
            logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
            GL.choice_res_ch = channel_info[1]
            logging.info(channel_info)




def calculate_expected_event_start_time():
    # 计算期望的预约事件的时间
    expected_res_time = ['', '', '', '', '']        # 期望的预约事件时间信息[年，月，日，时，分]
    swap_data = [0, 0, 0, 0, 0]                     # 用于交换处理的时间信息[年，月，日，时，分]
    time_interval = 3
    leap_year_month = 29
    nonleap_year_month = 28
    solar_month = [1, 3, 5, 7, 8, 10, 12]
    lunar_month = [4, 6, 9, 11]
    sys_time = rsv_kws['current_sys_time']
    logging.info(sys_time)
    sys_time_split = re.split(r"[\s:/]", sys_time)
    sys_year = int(sys_time_split[0])
    sys_month = int(sys_time_split[1])
    sys_day = int(sys_time_split[2])
    sys_hour = int(sys_time_split[3])
    sys_minute = int(sys_time_split[4])

    swap_data[4] = sys_minute + time_interval
    # 计算分钟和进位到小时
    if swap_data[4] < 60:
        expected_res_time[4] = "{0:02d}".format(swap_data[4])
        swap_data[3] = sys_hour
    elif swap_data[4] >= 60:
        expected_res_time[4] = "{0:02d}".format(swap_data[4] - 60)
        swap_data[3] = sys_hour + 1
    # 计算小时和进位到天数
    if swap_data[3] < 24:
        expected_res_time[3] = "{0:02d}".format(swap_data[3])
        swap_data[2] = sys_day
    elif swap_data[3] >= 24:
        expected_res_time[3] = "{0:02d}".format(swap_data[3] - 24)
        swap_data[2] = sys_day + 1
    # 按照闰年、平年，月份来计算天数和是否进位月份
    if sys_month == 2:
        logging.info("当前月份为二月")
        if sys_year % 100 == 0 and sys_year % 400 == 0:
            logging.info("当前年份为世纪闰年，二月有29天")
            if swap_data[2] <= leap_year_month:
                expected_res_time[2] = "{0:02d}".format(swap_data[2])
                expected_res_time[1] = "{0:02d}".format(sys_month)
                expected_res_time[0] = "{0:02d}".format(sys_year)
            elif swap_data[2] > leap_year_month:
                expected_res_time[2] = "{0:02d}".format(swap_data[2] - leap_year_month)
                expected_res_time[1] = "{0:02d}".format(sys_month + 1)
                expected_res_time[0] = "{0:02d}".format(sys_year)
        elif sys_year % 100 != 0 and sys_year % 4 == 0:
            logging.info("当前年份为普通闰年，二月有29天")
            if swap_data[2] <= leap_year_month:
                expected_res_time[2] = "{0:02d}".format(swap_data[2])
                expected_res_time[1] = "{0:02d}".format(sys_month)
                expected_res_time[0] = "{0:02d}".format(sys_year)
            elif swap_data[2] > leap_year_month:
                expected_res_time[2] = "{0:02d}".format(swap_data[2] - leap_year_month)
                expected_res_time[1] = "{0:02d}".format(sys_month + 1)
                expected_res_time[0] = "{0:02d}".format(sys_year)
        else:
            logging.info("当前年份为平年，二月有28天")
            if sys_month == 2:
                if swap_data[2] <= nonleap_year_month:
                    expected_res_time[2] = "{0:02d}".format(swap_data[2])
                    expected_res_time[1] = "{0:02d}".format(sys_month)
                    expected_res_time[0] = "{0:02d}".format(sys_year)
                elif swap_data[2] > nonleap_year_month:
                    expected_res_time[2] = "{0:02d}".format(swap_data[2] - nonleap_year_month)
                    expected_res_time[1] = "{0:02d}".format(sys_month + 1)
                    expected_res_time[0] = "{0:02d}".format(sys_year)
    elif sys_month in solar_month:
        logging.info("当前月份为大月，大月有31天")
        if swap_data[2] <= 31:
            expected_res_time[2] = "{0:02d}".format(swap_data[2])
            expected_res_time[1] = "{0:02d}".format(sys_month)
            expected_res_time[0] = "{0:02d}".format(sys_year)
        elif swap_data[2] > 31:
            expected_res_time[2] = "{0:02d}".format(swap_data[2] - 31)
            swap_data[1] = sys_month + 1
            if swap_data[1] <= 12:
                expected_res_time[1] = "{0:02d}".format(swap_data[1])
                expected_res_time[0] = "{0:02d}".format(sys_year)
            elif swap_data[1] > 12:
                expected_res_time[1] = "{0:02d}".format(swap_data[1] - 12)
                expected_res_time[0] = "{0:02d}".format(sys_year + 1)
    elif sys_month in lunar_month:
        logging.info("当前月份为小月，小月有30天")
        if swap_data[2] <= 30:
            expected_res_time[2] = "{0:02d}".format(swap_data[2])
            expected_res_time[1] = "{0:02d}".format(sys_month)
            expected_res_time[0] = "{0:02d}".format(sys_year)
        elif swap_data[2] > 30:
            expected_res_time[2] = "{0:02d}".format(swap_data[2] - 30)
            swap_data[1] = sys_month + 1
            if swap_data[1] <= 12:
                expected_res_time[1] = "{0:02d}".format(swap_data[1])
                expected_res_time[0] = "{0:02d}".format(sys_year)
            elif swap_data[1] > 12:
                expected_res_time[1] = "{0:02d}".format(swap_data[1] - 12)
                expected_res_time[0] = "{0:02d}".format(sys_year + 1)

    str_expected_res_time = ''.join(expected_res_time)
    logging.info(f"期望的完整的预约事件时间为{str_expected_res_time}")
    return str_expected_res_time


def create_expected_event_info():
    # 创建期望的事件信息
    expected_event_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    duration_time = "0001"
    if TEST_CASE_INFO[4] == "Play":
        # choice_ch_for_res_event_type()
        # if TEST_CASE_INFO[3] == "Once":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "PVR":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = duration_time
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power Off":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]
    return expected_event_info


def edit_res_event_info():
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面
    send_commd(KEY["GREEN"])
    # 生成预期的预约事件
    expected_res_event_info = create_expected_event_info()
    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[4]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_commd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[3]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_commd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[3] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[3]}")
    elif TEST_CASE_INFO[3] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[3]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_commd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_commds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_commd(KEY["DOWN"])
    else:
        start_time_list.append(expected_res_event_info[0][8:])
        start_time_cmd = change_numbs_to_commds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_commd(j)
        send_commd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[4] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[4]}")
    elif TEST_CASE_INFO[4] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_commd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_commds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[4] == "Power Off":
        logging.info(f"当前事件类型为Power Off，不需要设置Channel：{TEST_CASE_INFO[4]}")
    elif TEST_CASE_INFO[4] != "Power Off":
        logging.info(f"当前事件类型不为Power Off，需要设置Channel：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_commd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["update_event_list_state"] = True
    send_commd(KEY["EXIT"])
    send_commd(KEY["OK"])
    # 退回大画面
    send_more_commds(exit_to_screen)


def add_new_res_event_to_event_mgr_list():
    # 添加新预约事件到事件管理列表
    GL.res_event_mgr.extend(res_event_list)
    GL.report_data[0] = GL.res_event_mgr[0]
    logging.info(type(GL.res_event_mgr))
    logging.info(GL.res_event_mgr)
    state["update_event_list_state"] = False


def add_res_event():
    # 新增预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_commds(enter_timer_setting_interface)
    # 获取当前系统时间
    get_current_system_time()
    # 进入事件编辑界面，设置预约事件参数
    edit_res_event_info()
    # 添加新预约事件到事件管理列表
    add_new_res_event_to_event_mgr_list()


def goto_specified_interface_wait_for_event_triggered():
    logging.info("goto_specified_interface_wait_for_event_triggered")
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    # 切到指定界面
    if TEST_CASE_INFO[5] == "Screen_diff_ch":
        send_more_commds(exit_to_screen)
        # 且到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 等待事件响应
    while not state["res_event_triggered_state"]:
        logging.info("事件还没有触发，等待响应")
        time.sleep(10)
    else:
        logging.info("事件已经触发，正确跳出预约跳转选择框")
        logging.info(type(current_triggered_event_info))
        logging.info(type(GL.res_event_mgr))
        if list(current_triggered_event_info) in GL.res_event_mgr:
            logging.info("当前触发事件在事件列表中")
            state["res_event_triggered_state"] = False
        elif list(current_triggered_event_info) not in GL.res_event_mgr:
            logging.info(f"警告：当前触发事件不在事件列表中，{GL.res_event_mgr}-{current_triggered_event_info}")


def res_event_triggered_and_choice_jump_type():
    unlock_cmd = [KEY["0"], KEY["0"], KEY["0"], KEY["0"]]
    weekly_event_mode = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
    # 事件触发后选择跳转方式
    if TEST_CASE_INFO[4] == "Play":
        if TEST_CASE_INFO[6] == "Manual_jump":
            time.sleep(5)
            logging.info("选择手动跳转")
            send_commd(KEY["OK"])
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

        elif TEST_CASE_INFO[6] == "Auto_jump":
            logging.info("选择自动跳转")
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

        elif TEST_CASE_INFO[6] == "Cancel_jump":
            time.sleep(5)
            logging.info("选择取消跳转")
            send_commd(KEY["LEFT"])
            send_commd(KEY["OK"])
            while not state["res_event_cancel_jump_state"]:
                logging.info("请注意：没有检测到取消事件标志")
                time.sleep(3)
            else:
                logging.info("事件跳转取消成功")
                time.sleep(3)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"正确取消跳转，当前节目与触发事件的节目不一致:{channel_info[1]}--{current_triggered_event_info[2]}")
                else:
                    logging.info(f"警告：没有取消跳转成功，当前节目与触发事件的节目为:{channel_info[1]}--{current_triggered_event_info[2]}")

    if TEST_CASE_INFO[4] == "PVR":
        if TEST_CASE_INFO[6] == "Manual_jump":
            time.sleep(5)
            logging.info("选择手动跳转")
            send_commd(KEY["OK"])
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(1)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(1)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

            logging.info("还没有正确进入录制，请稍候...")
            while not state["rec_start_state"]:
                if state["no_storage_device_state"]:
                    logging.info("警告：没有插入存储设备")
                    state["receive_loop_state"] = True
                    break
                if state["no_enough_space_state"]:
                    logging.info("警告：存储设备没有足够的空间")
                    state["receive_loop_state"] = True
                    break
            else:
                logging.info("正确进入录制")
                rec_start_time = datetime.now()
                n = 0
                while not state["rec_end_state"]:
                    if n == 0:
                        logging.info("正在录制过程中，请等待录制结束")
                    n += 1
                else:
                    logging.info("录制结束")
                    rec_end_time = datetime.now()
                    GL.pvr_rec_dur_time = (rec_end_time - rec_start_time).seconds
                    logging.info(f"录制时长:{GL.pvr_rec_dur_time}")

        elif TEST_CASE_INFO[6] == "Auto_jump":
            logging.info("选择自动跳转")
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(1)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(1)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

            logging.info("还没有正确进入录制，请稍候...")
            while not state["rec_start_state"]:
                if state["no_storage_device_state"]:
                    logging.info("警告：没有插入存储设备")
                    state["receive_loop_state"] = True
                    break
                if state["no_enough_space_state"]:
                    logging.info("警告：存储设备没有足够的空间")
                    state["receive_loop_state"] = True
                    break
            else:
                logging.info("正确进入录制")
                rec_start_time = datetime.now()
                n = 0
                while not state["rec_end_state"]:
                    if n == 0:
                        logging.info("正在录制过程中，请等待录制结束")
                    n += 1
                else:
                    logging.info("录制结束")
                    rec_end_time = datetime.now()
                    GL.pvr_rec_dur_time = (rec_end_time - rec_start_time).seconds

        elif TEST_CASE_INFO[6] == "Cancel_jump":
            time.sleep(5)
            logging.info("选择取消跳转和录制")
            send_commd(KEY["LEFT"])
            send_commd(KEY["OK"])
            while not state["res_event_cancel_jump_state"]:
                logging.info("请注意：没有检测到取消事件标志")
                time.sleep(3)
            else:
                logging.info("事件跳转取消成功")
                time.sleep(3)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"正确取消跳转，当前节目与触发事件的节目不一致:{channel_info[1]}--{current_triggered_event_info[2]}")
                else:
                    logging.info(f"警告：没有取消跳转成功，当前节目与触发事件的节目为:{channel_info[1]}--{current_triggered_event_info[2]}")

    if TEST_CASE_INFO[4] == "Power Off":
        if TEST_CASE_INFO[6] == "Manual_jump":
            time.sleep(5)
            logging.info("选择手动跳转")
            send_commd(KEY["OK"])
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
            while not state["power_off_state"]:
                logging.info("还没有进入到关机状态")
                time.sleep(1)
            else:
                logging.info("进入关机状态")
                logging.info("等待5秒后，开始唤醒操作")
                time.sleep(5)
                logging.info("开始唤醒操作")
                send_commd(KEY["POWER"])
                state["control_power_on_info_rsv_state"] = True
                power_off_start_time = datetime.now()   # 用于关机计时起始时间
            while not state["stb_already_power_on_state"]:
                logging.info("还没有获取到启动成功标志，请等候")
                time.sleep(5)
                power_off_end_time = datetime.now()  # 用于关机计时结束时间
                if (power_off_end_time - power_off_start_time).seconds >= 30:
                    send_commd(KEY["UP"])
            else:
                logging.info("检测到启动成功标志")
                state["control_power_on_info_rsv_state"] = False

        elif TEST_CASE_INFO[6] == "Auto_jump":
            logging.info("选择自动跳转")
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
            while not state["power_off_state"]:
                logging.info("还没有进入到关机状态")
                time.sleep(1)
            else:
                logging.info("进入关机状态")
                logging.info("等待5秒后，开始唤醒操作")
                time.sleep(5)
                logging.info("开始唤醒操作")
                send_commd(KEY["POWER"])
                state["control_power_on_info_rsv_state"] = True
                power_off_start_time = datetime.now()   # 用于关机计时起始时间
            while not state["stb_already_power_on_state"]:
                logging.info("还没有获取到启动成功标志，请等候")
                time.sleep(5)
                power_off_end_time = datetime.now()  # 用于关机计时结束时间
                if (power_off_end_time - power_off_start_time).seconds >= 30:
                    send_commd(KEY["UP"])
            else:
                logging.info("检测到启动成功标志")
                state["control_power_on_info_rsv_state"] = False

        elif TEST_CASE_INFO[6] == "Cancel_jump":
            time.sleep(5)
            logging.info("选择取消跳转")
            send_commd(KEY["LEFT"])
            send_commd(KEY["OK"])
            while not state["res_event_cancel_jump_state"]:
                logging.info("请注意：没有检测到取消事件标志")
                time.sleep(3)
            else:
                logging.info("事件跳转取消成功")
                time.sleep(3)

    logging.info("预约事件数据处理，----------------------------------------------------")
    if current_triggered_event_info[-1] == "Once":  # Once事件触发后，需要从数据库中移除
        logging.info(f"移除Once类型当前触发事件前的列表：{GL.res_event_mgr}")
        GL.res_event_mgr.remove(list(current_triggered_event_info))
        logging.info(f"移除Once类型当前触发事件后的列表：{GL.res_event_mgr}")

    elif current_triggered_event_info[-1] == "Daily":
        logging.info("Daily事件不需要从数据库中删除")
    elif current_triggered_event_info[-1] in weekly_event_mode:
        logging.info("Daily事件不需要从数据库中删除")
    else:
        pass


def res_triggered_later_check_timer_setting_event_list():
    # 预约事件触发后，事件列表事件检查
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    send_more_commds(enter_timer_setting_interface)
    if rsv_kws["res_event_numb"] == '0':
        if len(GL.res_event_mgr) == int(rsv_kws["res_event_numb"]):
            logging.info("数据库与事件列表中的事件个数匹配，都为空")
        else:
            logging.info(f"警告：数据库与事件列表中的事件数不匹配，{len(GL.res_event_mgr)}--{int(rsv_kws['res_event_numb'])}")
    elif rsv_kws["res_event_numb"] != '0':
        if len(GL.res_event_mgr) == int(rsv_kws["res_event_numb"]):
            logging.info("数据库与事件列表中的事件个数匹配，再检查事件信息是否匹配")
            if GL.res_event_mgr == list(res_event_list):
                logging.info(f"数据库与事件列表中的事件信息一致，{GL.res_event_mgr}--{list(res_event_list)}")
            else:
                logging.info(f"警告：数据库与事件列表中的事件信息不一致，{GL.res_event_mgr}--{list(res_event_list)}")
        else:
            logging.info("数据库与事件列表中的事件个数不匹配，请检查事件信息")
            logging.info(f"警告：数据库与事件列表中的事件个数不一致，{GL.res_event_mgr}--{list(res_event_list)}")
    send_more_commds(exit_to_screen)


def write_data_to_excel():
    excel_title_0 = [
        "报告名称",
        "预约事件类型",
        "预约事件模式",
        "预约节目类型",
        "预约等待界面",
        "预约跳转模式"
    ]
    excel_title_1 = ["预约事件信息", "触发响应信息"]
    excel_title_2 = ["起始时间", "事件类型", "节目名称", "持续时间", "事件模式", "触发时间", "是否跳转", "跳转节目", "是否录制", "录制时长"]

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    if not os.path.exists(file_path[1]):
        wb = Workbook()
        ws = wb.active
        ws.title = file_path[2]
        ws.column_dimensions['A'].width = 17
        for i in range(len(excel_title_0)):
            if i == 0:
                ws.cell(i + 1, 1).value = excel_title_0[i]
                ws.cell(i + 1, 1).alignment = alignment
                ws.row_dimensions[(i + 1)].height = 27
            else:
                ws.cell(i + 1, 1).value = excel_title_0[i]
                ws.cell(i + 1, 1).alignment = alignment
        ws.column_dimensions['A'].width = 17
        ws.cell(len(excel_title_0) + 1, 1).value = excel_title_1[0]
        ws["A" + str(len(excel_title_0) + 1)].alignment = alignment
        ws.merge_cells(start_row=len(excel_title_0) + 1, start_column=1, end_row=len(excel_title_0) + 1, end_column=5)
        ws.cell(len(excel_title_0) + 1, 6).value = excel_title_1[1]
        ws["F" + str(len(excel_title_0) + 1)].alignment = alignment
        ws.merge_cells(start_row=len(excel_title_0) + 1, start_column=6, end_row=len(excel_title_0) + 1, end_column=10)
        for j in range(len(excel_title_2)):
            ws.cell(len(excel_title_0) + 2, j + 1).value = excel_title_2[j]
            ws.cell(len(excel_title_0) + 2, j + 1).alignment = alignment
    elif os.path.exists(file_path[1]):
        wb = load_workbook(file_path[1])
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if file_path[2] in sheets_name_list:
            ws = wb[file_path[2]]
        elif file_path[2] not in sheets_name_list:
            ws = wb.create_sheet(file_path[2])
        for i in range(len(excel_title_0)):
            if i == 0:
                ws.cell(i + 1, 1).value = excel_title_0[i]
                ws.cell(i + 1, 1).alignment = alignment
                ws.row_dimensions[(i + 1)].height = 27
            else:
                ws.cell(i + 1, 1).value = excel_title_0[i]
                ws.cell(i + 1, 1).alignment = alignment
        ws.column_dimensions['A'].width = 17
        ws.cell(len(excel_title_0) + 1, 1).value = excel_title_1[0]
        ws["A" + str(len(excel_title_0) + 1)].alignment = alignment
        ws.merge_cells(start_row=len(excel_title_0) + 1, start_column=1, end_row=len(excel_title_0) + 1, end_column=5)
        ws.cell(len(excel_title_0) + 1, 6).value = excel_title_1[1]
        ws["F" + str(len(excel_title_0) + 1)].alignment = alignment
        ws.merge_cells(start_row=len(excel_title_0) + 1, start_column=6, end_row=len(excel_title_0) + 1, end_column=10)

        for j in range(len(excel_title_2)):
            ws.cell(len(excel_title_0) + 2, j + 1).value = excel_title_2[j]
            ws.cell(len(excel_title_0) + 2, j + 1).alignment = alignment

    # 写Title数据
    for x in range(len(GL.title_data)):
        ws.cell(x + 1, 2).value = GL.title_data[x]
        ws.cell(x + 1, 2).alignment = alignment
        ws.merge_cells(start_row=x + 1, start_column=2, end_row=x + 1, end_column=10)

    # 写预约事件数据
    a_column_numb = column_index_from_string("A")
    interval_row = len(excel_title_0) + 2
    for d in range(len(GL.report_data)):
        if d == 0:
            for dd in range(len(GL.report_data[d])):
                ws.cell(GL.start_row + interval_row + 1, dd + 1).value = GL.report_data[d][dd]
                ws.cell(GL.start_row + interval_row + 1, dd + 1).alignment = alignment
                ws.column_dimensions[get_column_letter(a_column_numb + dd + 1)].width = 17
        else:
            ws.cell(GL.start_row + interval_row + 1, d + len(GL.report_data[0])).value = GL.report_data[d]
            ws.cell(GL.start_row + interval_row + 1, d + len(GL.report_data[0])).alignment = alignment
            ws.column_dimensions[get_column_letter(a_column_numb + d + len(GL.report_data[0]))].width = 17
    GL.start_row += 1

    wb.save(file_path[1])


def manage_report_data_and_write_data():
    # 整理数据以及写数据
    GL.title_data[0] = file_path[2]
    GL.title_data[1] = TEST_CASE_INFO[4]
    GL.title_data[2] = TEST_CASE_INFO[3]
    GL.title_data[3] = TEST_CASE_INFO[2]
    GL.title_data[4] = TEST_CASE_INFO[5]
    GL.title_data[5] = TEST_CASE_INFO[6]

    if TEST_CASE_INFO[4] == "Play":
        GL.report_data[1] = list(res_event_list)[0][0]
        GL.report_data[2] = TEST_CASE_INFO[6]
        GL.report_data[3] = channel_info[1]
        GL.report_data[4] = TEST_CASE_INFO[4]
        GL.report_data[5] = "----"
    elif TEST_CASE_INFO[4] == "PVR":
        GL.report_data[1] = list(res_event_list)[0][0]
        GL.report_data[2] = TEST_CASE_INFO[6]
        GL.report_data[3] = channel_info[1]
        GL.report_data[4] = TEST_CASE_INFO[4]
        GL.report_data[5] = str(GL.pvr_rec_dur_time) + 's'
    elif TEST_CASE_INFO[4] == "Power Off":
        GL.report_data[1] = list(res_event_list)[0][0]
        GL.report_data[2] = TEST_CASE_INFO[6]
        GL.report_data[3] = "----"
        GL.report_data[4] = TEST_CASE_INFO[4]
        GL.report_data[5] = "----"

    logging.info(GL.title_data)
    logging.info(GL.report_data)
    time.sleep(2)


def before_cycle_test_clear_data_and_state():
    # 循环测试前，清理数据和状态变量
    logging.info("before_cycle_test_clear_data_and_state")
    GL.res_event_mgr.clear()
    GL.choice_res_ch = ''
    state["clear_variate_state"] = True
    GL.pvr_rec_dur_time = ''

    GL.add_res_event_numb -= 1
    logging.info("循环测试，延时5秒")
    time.sleep(5)
    logging.info(f"剩余循环次数：{GL.add_res_event_numb}")

    if GL.add_res_event_numb < 1:
        logging.info("程序结束")
        state["receive_loop_state"] = True  # 触发结束接收进程的状态


def receive_serial_process(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info):
    logging_info_setting()
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIME_SHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict([val, key] for key, val in rsv_key.items())

    res_kws = [
        "[PTD]Time_mode=",          # 0     获取系统时间模式
        "[PTD]System_time=",        # 1     系统时间
        "[PTD]Res_event_numb=",     # 2     预约事件数量
        "[PTD]Res_event:",          # 3     预约事件信息
        "[PTD]Res_triggered:",      # 4     预约事件触发和当前响应事件的信息
        "[PTD]Res_confirm_jump",    # 5     预约事件确认跳转
        "[PTD]Res_cancel_jump",     # 6     预约事件取消跳转
        "[PTD]REC_start",           # 7     录制开始
        "[PTD]REC_end",             # 8     录制结束
        "[PTD]No_storage_device",   # 9     没有存储设备
        "[PTD]No_enough_space",     # 10    没有足够的空间
        "[PTD]power_cut",           # 11    进入待机
        "[PTD]:switch totle cost",  # 12    开机解码成功
        "[PTD][HOTPLUG] PLUG_IN",   # 13    存储设备插入成功
    ]

    switch_ch_kws = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    ch_info_kws = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name"]

    group_info_kws = [
        "Group_name",
        "Prog_total"
    ]

    edit_event_kws = [
        "[PTD]Mode=",
        "[PTD]Type=",
        "[PTD]Start Date=",
        "[PTD]Start Time=",
        "[PTD]Duration=",
        "[PTD]Channel="
    ]

    other_kws = [
        "[PTD]Infrared_key_values:",    # 获取红外接收关键字
    ]

    infrared_rsv_cmd = []
    receive_serial = serial.Serial(prs_data["receive_serial_name"], 115200, timeout=1)

    while True:
        data = receive_serial.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            data1 = data.decode("ISO-8859-1", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_log_data_to_txt(prs_data["log_file_path"], data3)

            if state["clear_variate_state"]:
                state["sys_time_mode_state"] = False
                state["current_sys_time_state"] = False
                state["res_event_numb_state"] = False
                state["res_event_triggered_state"] = False
                state["res_event_confirm_jump_state"] = False
                state["res_event_cancel_jump_state"] = False
                state["rec_start_state"] = False
                state["rec_end_state"] = False
                state["no_storage_device_state"] = False
                state["no_enough_space_state"] = False
                state["update_event_list_state"] = False
                state["clear_variate_state"] = False
                state["power_off_state"] = False
                state["stb_already_power_on_state"] = False

                del res_event_list[:]
                del current_triggered_event_info[:]
                # channel_info = ['', '', '', '', '', '', '']

            if other_kws[0] in data2:   # 红外接收打印
                rsv_cmd = re.split(":", data2)[-1]
                infrared_rsv_cmd.append(rsv_cmd)
                if rsv_cmd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_cmd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        infrared_send_cmd[-1], reverse_rsv_key[infrared_rsv_cmd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(infrared_send_cmd), len(infrared_rsv_cmd)))
                if state["control_power_on_info_rsv_state"]:    # 用于Power Off事件触发后，软开机没有检测到开机关键字用来避免死循环
                    state["stb_already_power_on_state"] = True

            if res_kws[0] in data2:     # 获取系统时间模式（自动还是手动）
                state["sys_time_mode_state"] = True
                rsv_kws["sys_time_mode"] = re.split(r"=", data2)[-1]

            if res_kws[1] in data2:     # 获取当前系统时间
                state["current_sys_time_state"] = True
                rsv_kws["current_sys_time"] = re.split(r"=", data2)[-1]

            if res_kws[2] in data2:     # 获取预约事件数量
                state["res_event_numb_state"] = True
                rsv_kws["res_event_numb"] = re.split(r"=", data2)[-1]

            if res_kws[3] in data2:     # 获取预约事件信息
                event_split_info = re.split(r"t:|,", data2)
                event_info = ['', '', '', '', '']
                for info in event_split_info:
                    if "Start_time" in info:
                        event_start_time = re.split(r"=", info)[-1]
                        if len(event_start_time) == 5:
                            event_info[0] = ''.join(re.split(r":", event_start_time))
                        elif len(event_start_time) == 16:
                            event_info[0] = ''.join(re.split(r"[/:\s]", event_start_time))
                    if "Event_type" in info:
                        event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        event_info[4] = re.split(r"=", info)[-1]
                if state["update_event_list_state"]:
                    res_event_list.append(event_info)

            if res_kws[4] in data2:     # 获取预约事件跳转触发信息，以及当前响应事件的信息
                state["res_event_triggered_state"] = True
                current_event_split_info = re.split(r"d:|,", data2)
                current_event_info = ['', '', '', '', '']
                for info in current_event_split_info:
                    if "Start_time" in info:
                        current_event_start_time = re.split(r"=", info)[-1]
                        if len(current_event_start_time) == 5:
                            current_event_info[0] = ''.join(re.split(r":", current_event_start_time))
                        elif len(current_event_start_time) == 16:
                            current_event_info[0] = ''.join(re.split(r"[/:\s]", current_event_start_time))
                    if "Event_type" in info:
                        current_event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        current_event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        current_event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        current_event_info[4] = re.split(r"=", info)[-1]
                current_triggered_event_info.extend(current_event_info)

            if res_kws[5] in data2:     # 获取预约事件确认跳转信息
                state["res_event_confirm_jump_state"] = True

            if res_kws[6] in data2:     # 获取预约事件取消跳转信息
                state["res_event_cancel_jump_state"] = True

            if res_kws[7] in data2:     # 获取PVR预约事件录制开始信息
                state["rec_start_state"] = True

            if res_kws[8] in data2:     # 获取PVR预约事件录制结束信息
                state["rec_end_state"] = True

            if res_kws[9] in data2:     # 存储设备没有插入的打印信息
                state["no_storage_device_state"] = True

            if res_kws[10] in data2:    # 存储设备没有足够空间的打印信息
                state["no_enough_space_state"] = True

            if res_kws[11] in data2:    # 软关机打印信息
                state["power_off_state"] = True

            if res_kws[12] in data2 or res_kws[13] in data2:    # 开机解码成功打印信息,或开机存储设备挂载成功信息
                if state["control_power_on_info_rsv_state"]:
                    state["stb_already_power_on_state"] = True

            if switch_ch_kws[0] in data2:
                ch_info_split = re.split(r"[\],]", data2)
                for i in range(len(ch_info_split)):
                    if ch_info_kws[0] in ch_info_split[i]:  # 提取频道号
                        channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if ch_info_kws[1] in ch_info_split[i]:  # 提取频道名称
                        channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if switch_ch_kws[1] in data2:
                flag_info_split = re.split(r"[\],]", data2)
                for i in range(len(flag_info_split)):
                    if ch_info_kws[2] in flag_info_split[i]:  # 提取频道所属TP
                        channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if ch_info_kws[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if switch_ch_kws[3] in data2:
                group_info_split = re.split(r"[\],]", data2)
                for i in range(len(group_info_split)):
                    if group_info_kws[0] in group_info_split[i]:  # 提取频道所属组别
                        rsv_kws["prog_group_name"] = re.split(r"=", group_info_split[i])[-1]
                        channel_info[6] = rsv_kws["prog_group_name"]
                    if group_info_kws[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        rsv_kws["prog_group_total"] = re.split(r"=", group_info_split[i])[-1]

            if edit_event_kws[0] in data2:          # 提取Mode参数
                rsv_kws["edit_event_focus_pos"] = "Mode"
                rsv_kws["edit_event_mode"] = re.split(r"=", data2)[-1]

            if edit_event_kws[1] in data2:          # 提取Type参数
                rsv_kws["edit_event_focus_pos"] = "Type"
                rsv_kws["edit_event_type"] = re.split(r"=", data2)[-1]

            if edit_event_kws[2] in data2:          # 提取Start Date参数
                rsv_kws["edit_event_focus_pos"] = "Start Date"
                rsv_kws["edit_event_date"] = re.split(r"=", data2)[-1]

            if edit_event_kws[3] in data2:          # 提取Start Time参数
                rsv_kws["edit_event_focus_pos"] = "Start Time"
                rsv_kws["edit_event_time"] = re.split(r"=", data2)[-1]

            if edit_event_kws[4] in data2:          # 提取Duration参数
                rsv_kws["edit_event_focus_pos"] = "Duration"
                rsv_kws["edit_event_duration"] = re.split(r"=", data2)[-1]

            if edit_event_kws[5] in data2:          # 提取Channel参数
                rsv_kws["edit_event_focus_pos"] = "Channel"
                rsv_kws["edit_event_ch"] = re.split(r"=", data2)[-1]


if __name__ == "__main__":

    GL = MyGlobal()
    logging_info_setting()
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIME_SHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D"
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())
    TEST_CASE_INFO = ["23", "All", "TV", "Once", "Power Off", "Screen_diff_ch", "Auto_jump"]

    file_path = build_log_and_report_file_path()
    ser_name = list(check_ports())  # send_ser_name, receive_ser_name
    send_serial = serial.Serial(ser_name[0], 9600)
    receive_ser_name = ser_name[1]

    infrared_send_cmd = Manager().list([])
    res_event_list = Manager().list([])
    current_triggered_event_info = Manager().list([])
    channel_info = Manager().list(['', '', '', '', '', '', ''])     # [频道号,频道名称,tp,lock,scramble,频道类型,组别]
    rsv_kws = Manager().dict({
        "sys_time_mode": '', "current_sys_time": '', "res_event_numb": '', "prog_group_name": '',
        "prog_group_total": '', "edit_event_focus_pos": '', "edit_event_mode": '', "edit_event_type": '',
        "edit_event_date": '', "edit_event_time": '', "edit_event_duration": '', "edit_event_ch": ''
    })

    state = Manager().dict({
        "res_event_numb_state": False, "res_event_triggered_state": False, "res_event_confirm_jump_state": False,
        "res_event_cancel_jump_state": False, "rec_start_state": False, "rec_end_state": False,
        "no_storage_device_state": False, "no_enough_space_state": False, "power_off_state": False,
        "sys_time_mode_state": False, "current_sys_time_state": False, "update_event_list_state": False,
        "clear_variate_state": False, "receive_loop_state": False, "control_power_on_info_rsv_state": False,
        "stb_already_power_on_state": False
    })

    prs_data = Manager().dict({
        "log_file_path": file_path[0], "receive_serial_name": receive_ser_name,
    })

    rsv_p = Process(target=receive_serial_process, args=(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info))
    rsv_p.start()

    if platform.system() == "Windows":
        time.sleep(5)
        logging.info("Windows系统接受端响应慢，等待5秒")
    elif platform.system() == "Linux":
        time.sleep(1)
        logging.info("Linux系统接受端响应快，但是增加一个延时保护，等待1秒")

    while GL.add_res_event_numb > 0:
        clear_timer_setting_all_events()
        check_sys_time_mode()
        choice_ch_for_res_event_type()
        add_res_event()
        goto_specified_interface_wait_for_event_triggered()
        res_event_triggered_and_choice_jump_type()
        manage_report_data_and_write_data()
        write_data_to_excel()
        res_triggered_later_check_timer_setting_event_list()
        before_cycle_test_clear_data_and_state()

    if state["receive_loop_state"]:
        rsv_p.terminate()
        logging.info('stop receive process')
        rsv_p.join()
