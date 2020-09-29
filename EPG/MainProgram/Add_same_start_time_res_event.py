#!/usr/bin/python3
# -*- coding: utf-8 -*-


from serial_setting1 import *
from multiprocessing import Process, Manager
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import RED, BLUE
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta, date
from random import randint, choice
from email.mime.text import MIMEText
from email.header import Header
import smtplib
import platform
import os
import time
import logging
import re
import sys

TEST_CASE_INFO = ''

choice_case_numb = int(sys.argv[1])
# choice_case_numb = 1
test_case_info = same_start_time_case[choice_case_numb]
print(test_case_info)

weekly_mode = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
# test_case_info = ["01", "All", "TV", "Weekly", "Play", "TVScreenDiffCH",
#                   "Manual_jump", "Same(time+type)+Diff(mode)", "Weekly", "Play", "screen_test_numb"]

if test_case_info[7] == "Same[time+type+mode]" \
        or test_case_info[7] == "Same[time+mode]+Diff[type]" \
        or test_case_info[7] == "Same[type+dur+mode]+Diff[time]" \
        or test_case_info[7] == "Same[mode]+Diff[time+type+dur]":
    if test_case_info[3] == "Weekly" and test_case_info[8] == "Weekly":
        new_test_case_info = test_case_info.copy()
        print(f"选择之前的new_test_case_info：{new_test_case_info}")
        new_test_case_info[3] = new_test_case_info[8] = choice(weekly_mode)
        print(f"选择之后的new_test_case_info：{new_test_case_info}")
        TEST_CASE_INFO = new_test_case_info
    else:
        TEST_CASE_INFO = test_case_info
elif test_case_info[7] == "Same[time+type]+Diff[mode]" \
        or test_case_info[7] == "Same[time]+Diff[type+mode]" \
        or test_case_info[7] == "Same[type+dur]+Diff[time+mode]" \
        or test_case_info[7] == "Diff[time+type+dur+mode]"\
        or test_case_info[7] == "Same[time+type]+Diff[mode+date]"\
        or test_case_info[7] == "Same[time]+Diff[type+mode+date]":
    if test_case_info[3] == "Weekly" and test_case_info[8] != "Weekly":
        new_test_case_info = test_case_info.copy()
        print(f"选择之前的new_test_case_info：{new_test_case_info}")
        new_test_case_info[3] = choice(weekly_mode)
        print(f"选择之后的new_test_case_info：{new_test_case_info}")
        TEST_CASE_INFO = new_test_case_info

    elif test_case_info[3] != "Weekly" and test_case_info[8] == "Weekly":
        new_test_case_info = test_case_info.copy()
        print(f"选择之前的new_test_case_info：{new_test_case_info}")
        new_test_case_info[8] = choice(weekly_mode)
        print(f"选择之后的new_test_case_info：{new_test_case_info}")
        TEST_CASE_INFO = new_test_case_info

    else:
        TEST_CASE_INFO = test_case_info

scenes_list = [
    "Same[time+type+mode]",
    "Same[time+mode]+Diff[type]",
    "Same[time+type]+Diff[mode]",
    "Same[time]+Diff[type+mode]",
    "Same[type+dur+mode]+Diff[time]",
    "Same[type+dur]+Diff[time+mode]",
    "Same[mode]+Diff[time+type+dur]",
    "Diff[time+type+dur+mode]",
    "Same[time+type]+Diff[mode+date]",
    "Same[time]+Diff[type+mode+date]"
    ]


class MyGlobal(object):

    def __init__(self):
        if test_case_info[-1] == "screen_test_numb":
            self.res_triggered_numb = 1                 # 大画面预约响应的次数
        elif test_case_info[-1] == "other_interface_test_numb":
            self.res_triggered_numb = 1                 # 其他界面预约响应的次数

        self.choice_res_ch = ''                         # 预约Play或PVR事件时所选预约节目

        # 报告数据汇总[[预期事件1]，[保存事件1], [保存事件2]，[预期事件2], "保存事件个数", "无效事件提示", "case编号", "执行case时间"]
        self.report_data = [[], [], [], [], '', '', '', '']


def logging_info_setting():
    # 配置logging输出格式
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    date_format = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=log_format, datefmt=date_format)


def hex_strs_to_bytes(strings):
    # 将红外命令字符串转换为字节串
    return bytes.fromhex(strings)


def write_log_data_to_txt(path, write_data):
    with open(path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def send_cmd(command):
    global receive_cmd_list, infrared_send_cmd
    continuous_transmission_cmd_num = 0     # 连续发送命令数
    # 红外发送端发送指令
    send_serial.write(hex_strs_to_bytes(command))
    send_serial.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[command]))
    if REVERSE_KEY[command] != "POWER":
        infrared_send_cmd.append(REVERSE_KEY[command])
    time.sleep(1.0)
    if len(infrared_send_cmd) == len(receive_cmd_list):
        pass
    elif len(infrared_send_cmd) != len(receive_cmd_list):
        logging.info("检测到发送和接收命令数不一致，等待2秒，查看是否接收端还没有接收到打印")
        time.sleep(2)
        while True:
            if len(infrared_send_cmd) == len(receive_cmd_list):
                break
            elif len(infrared_send_cmd) != len(receive_cmd_list):
                logging.info(f"此刻补发STB没有接收到的红外命令{infrared_send_cmd[-1]}")
                send_serial.write(hex_strs_to_bytes(KEY[infrared_send_cmd[-1]]))
                send_serial.flush()
                continuous_transmission_cmd_num += 1
                time.sleep(1.0)
                if continuous_transmission_cmd_num == 30:
                    stb_crash_msg = "STB一直发送指令，疑似死机"
                    mail(stb_crash_msg)


def send_more_cmds(command_list):
    # 用于发送一连串的指令
    for command in command_list:
        send_cmd(command)
    time.sleep(1)   # 增加函数切换时的的等待，避免可能出现send_commd函数中的等待时间没有执行的情况


def change_numbs_to_cmds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_cmds_list = []
    for i in range(len(numbs_list)):
        channel_cmds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_cmds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_cmds_list[i].append(KEY[numbs_list[i][j]])
    return channel_cmds_list


def create_log_and_report_file_path():
    # 用于创建打印和报告文件路径
    # 构建存放数据的总目录，以及构建存放打印和报告的目录
    parent_path = os.path.dirname(os.getcwd())
    case_name = "Add_same_start_time_res_event"
    test_data_directory_name = "test_data"
    test_data_directory_path = os.path.join(parent_path, test_data_directory_name)
    log_directory_name = "print_log"
    log_directory_path = os.path.join(test_data_directory_path, log_directory_name)
    report_directory_name = "report"
    report_directory_path = os.path.join(test_data_directory_path, report_directory_name)

    log_case_directory_path = os.path.join(test_data_directory_path, log_directory_name, case_name)
    report_case_directory_path = os.path.join(test_data_directory_path, report_directory_name, case_name)
    # 判断目录是否存在，否则创建目录
    if not os.path.exists(test_data_directory_path):
        os.mkdir(test_data_directory_path)
    if not os.path.exists(log_directory_path):
        os.mkdir(log_directory_path)
    if not os.path.exists(report_directory_path):
        os.mkdir(report_directory_path)
    if not os.path.exists(log_case_directory_path):
        os.mkdir(log_case_directory_path)
    if not os.path.exists(report_case_directory_path):
        os.mkdir(report_case_directory_path)
    # 创建打印和报告文件的名称和路径
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    sheet_name = TEST_CASE_INFO[7].replace('[', '(').replace(']', ')')

    fmt_name = "{}_{}_{}_{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2], TEST_CASE_INFO[4],
        TEST_CASE_INFO[3], sheet_name, TEST_CASE_INFO[9], TEST_CASE_INFO[8])
    log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    log_file_path = os.path.join(log_case_directory_path, log_file_name)
    report_file_name = "Add_same_start_time_res_event_result_report.xlsx"
    report_file_path = os.path.join(report_case_directory_path, report_file_name)
    # sheet_name = "{}_{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[4], TEST_CASE_INFO[3], TEST_CASE_INFO[7])
    return log_file_path, report_file_path, sheet_name


def clear_timer_setting_all_events():
    logging.info("clear_timer_setting_all_events")
    # 清除Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    delete_all_res_events = [KEY["BLUE"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入定时器设置界面
    send_more_cmds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        if rsv_kws["res_event_numb"] != '0':
            send_more_cmds(delete_all_res_events)
        elif rsv_kws["res_event_numb"] == '0':
            logging.info("没有预约事件存在")
            time.sleep(1)
        else:
            logging.debug("警告：预约事件个数获取错误！！！")
        state["res_event_numb_state"] = False
    # 退回大画面
    send_more_cmds(exit_to_screen)


def check_sys_time_mode():
    logging.info("check_sys_time_mode")
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入时间设置界面
    send_more_cmds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        if rsv_kws["sys_time_mode"] == "Auto":
            send_more_cmds(change_sys_time_mode)
        elif rsv_kws["sys_time_mode"] == "Manual":
            logging.info("系统时间模式已经为手动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    # 退回大画面
    send_more_cmds(exit_to_screen)


def get_current_system_time():
    logging.info("get_current_system_time")
    # 获取当前系统时间
    time.sleep(1)
    while not state["current_sys_time_state"]:
        logging.info("还没有获取到系统时间信息")
        time.sleep(1)
    else:
        logging.info(f"当前系统时间为:{rsv_kws['current_sys_time']}")


def choice_ch_for_res_event_type(choice_event):
    logging.info("choice_ch_for_res_event_type")
    # 根据预约事件类型来选择节目
    group_dict = {}
    choice_ch_numb = []

    # 根据所选case切换到对应类型节目的界面
    while channel_info[5] != TEST_CASE_INFO[2]:
        send_cmd(KEY["TV/R"])
        if channel_info[3] == "1":
            send_cmd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_cmd(KEY["OK"])
    # 切到指定分组和采集分组下节目总数信息
    while rsv_kws["prog_group_name"] != TEST_CASE_INFO[1]:
        send_cmd(KEY["RIGHT"])
        if channel_info[3] == "1":
            send_cmd(KEY["EXIT"])
    else:
        if rsv_kws["prog_group_name"] == '':
            logging.info("警告：没有All分组信息")
        else:
            group_dict[rsv_kws["prog_group_name"]] = rsv_kws["prog_group_total"]
            logging.info(f"分组信息{group_dict}")
    # 退出频道列表,回到大画面界面
    send_cmd(KEY["EXIT"])

    if choice_event == "event_1":
        # 根据用例指定的事件类型来选择节目
        if TEST_CASE_INFO[4] == "Play":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])
            logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
            GL.choice_res_ch = channel_info[1]
            logging.info(channel_info)

        elif TEST_CASE_INFO[4] == "PVR":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])

            while channel_info[3] != '0' and channel_info[4] != '0':
                logging.info(f"查看所切节目信息：{channel_info}")
                logging.info("所选节目不为免费节目，不可以进行PVR预约，继续切台")
                send_cmd(KEY["UP"])
                time.sleep(2)
                if channel_info[3] == "1":
                    send_cmd(KEY["EXIT"])
            else:
                logging.info("所选节目为免费节目，可以进行PVR预约")
                logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
                GL.choice_res_ch = channel_info[1]
                logging.info(channel_info)

        elif TEST_CASE_INFO[4] == "Power Off":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

        elif TEST_CASE_INFO[4] == "Power On":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

    elif choice_event == "event_2":
        # 根据用例指定的事件类型来选择节目
        if TEST_CASE_INFO[9] == "Play":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])
            logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
            GL.choice_res_ch = channel_info[1]
            logging.info(channel_info)

        elif TEST_CASE_INFO[9] == "PVR":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])

            while channel_info[3] != '0' and channel_info[4] != '0':
                logging.info(f"查看所切节目信息：{channel_info}")
                logging.info("所选节目不为免费节目，不可以进行PVR预约，继续切台")
                send_cmd(KEY["UP"])
                time.sleep(2)
                if channel_info[3] == "1":
                    send_cmd(KEY["EXIT"])
            else:
                logging.info("所选节目为免费节目，可以进行PVR预约")
                logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
                GL.choice_res_ch = channel_info[1]
                logging.info(channel_info)

        elif TEST_CASE_INFO[9] == "Power Off":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

        elif TEST_CASE_INFO[9] == "Power On":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")


def calculate_res_event_expected_start_time():
    logging.info("calculate_res_event_expected_start_time")
    time_interval = 5
    str_expected_res_time = ''
    weekday_num_dict = {"Mon.": 0, "Tues.": 1, "Wed.": 2, "Thurs.": 3, "Fri.": 4, "Sat.": 5, "Sun.": 6}
    sys_time = rsv_kws['current_sys_time']
    sys_time_split = re.split(r"[\s:/]", sys_time)
    sys_year = int(sys_time_split[0])
    sys_month = int(sys_time_split[1])
    sys_day = int(sys_time_split[2])
    sys_hour = int(sys_time_split[3])
    sys_minute = int(sys_time_split[4])
    dt_time = datetime(sys_year, sys_month, sys_day, sys_hour, sys_minute)

    logging.info(dt_time)
    if TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode]" or TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode]" \
            or TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]" \
            or TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]":     # Once与weekly的date要相同才能是invalid事件
        if TEST_CASE_INFO[3] == "Once" and TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:
            cur_weekday = date(sys_year, sys_month, sys_day).weekday()  # 当前系统时间的星期数
            event_2_weekday = weekday_num_dict[TEST_CASE_INFO[8]]   # event_2对应的星期数
            if cur_weekday == event_2_weekday:  # 当前系统时间对应的星期数与event_2对应的星期数相同时
                expected_res_time = dt_time + timedelta(minutes=time_interval)
                logging.info(expected_res_time)
                expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
                str_expected_res_time = ''.join(expected_res_time_split)[:12]
            elif cur_weekday != event_2_weekday:  # 当前系统时间对应的星期数与event_2对应的星期数不同时
                if cur_weekday > event_2_weekday:
                    interval_day = 7 - cur_weekday + event_2_weekday
                    cur_next_weekday = dt_time + timedelta(days=interval_day)
                    cur_next_weekday_time = cur_next_weekday + timedelta(minutes=time_interval)
                    cur_next_weekday_time_split = re.split(r"[-\s:]", str(cur_next_weekday_time))
                    str_expected_res_time = ''.join(cur_next_weekday_time_split)[:12]
                elif cur_weekday < event_2_weekday:
                    interval_day = event_2_weekday - cur_weekday
                    cur_next_weekday = dt_time + timedelta(days=interval_day)
                    cur_next_weekday_time = cur_next_weekday + timedelta(minutes=time_interval)
                    cur_next_weekday_time_split = re.split(r"[-\s:]", str(cur_next_weekday_time))
                    str_expected_res_time = ''.join(cur_next_weekday_time_split)[:12]
        else:
            expected_res_time = dt_time + timedelta(minutes=time_interval)
            logging.info(expected_res_time)
            expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
            str_expected_res_time = ''.join(expected_res_time_split)[:12]

    elif TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]" \
            or TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":     # Once与weekly的date要不同才能是valid事件
        if TEST_CASE_INFO[3] == "Once" and TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:
            cur_weekday = date(sys_year, sys_month, sys_day).weekday()  # 当前系统时间的星期数
            event_2_weekday = weekday_num_dict[TEST_CASE_INFO[8]]  # event_2对应的星期数
            if cur_weekday != event_2_weekday:  # 当前系统时间对应的星期数与event_2对应的星期数不同时
                expected_res_time = dt_time + timedelta(minutes=time_interval)
                logging.info(expected_res_time)
                expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
                str_expected_res_time = ''.join(expected_res_time_split)[:12]
            elif cur_weekday == event_2_weekday:  # 当前系统时间对应的星期数与event_2对应的星期数相同时
                interval_day = 1
                cur_next_weekday = dt_time + timedelta(days=interval_day)
                cur_next_weekday_time = cur_next_weekday + timedelta(minutes=time_interval)
                cur_next_weekday_time_split = re.split(r"[-\s:]", str(cur_next_weekday_time))
                str_expected_res_time = ''.join(cur_next_weekday_time_split)[:12]
        else:
            expected_res_time = dt_time + timedelta(minutes=time_interval)
            logging.info(expected_res_time)
            expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
            str_expected_res_time = ''.join(expected_res_time_split)[:12]

    else:
        expected_res_time = dt_time + timedelta(minutes=time_interval)
        logging.info(expected_res_time)
        expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
        str_expected_res_time = ''.join(expected_res_time_split)[:12]

    logging.info(f"期望的完整的预约事件时间为{str_expected_res_time}")
    return str_expected_res_time


def create_expected_add_event_info():
    logging.info("create_expected_add_event_info")
    # 创建期望的事件信息
    expected_event_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    if TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]" \
            or TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]" \
            or TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]" \
            or TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]":
        duration_time = "0010"
    else:
        duration_time = "0001"
    if TEST_CASE_INFO[4] == "Play":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "PVR":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = duration_time
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power Off":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power On":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]
    return expected_event_info


def edit_add_new_res_event_info():
    logging.info("edit_add_new_res_event_info")
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面(Add按键)
    send_cmd(KEY["GREEN"])
    # 生成预期的预约事件
    if TEST_CASE_INFO[3] == "Once":
        expected_res_event_info = create_expected_add_event_info()
    else:
        pretreatment_res_event_info = create_expected_add_event_info()    # 预处理预约事件信息
        expected_res_event_info = pretreatment_res_event_info.copy()
        expected_res_event_info[0] = expected_res_event_info[0][8:]
    logging.info(f"创建的事件为{expected_res_event_info}")
    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[4]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_cmd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[3]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_cmd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[3] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[3]}")
    elif TEST_CASE_INFO[3] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[3]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_cmd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_cmds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_cmd(KEY["DOWN"])
    else:
        if TEST_CASE_INFO[3] == "Once":
            start_time_list.append(expected_res_event_info[0][8:])
        else:
            start_time_list.append(expected_res_event_info[0])
        start_time_cmd = change_numbs_to_cmds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_cmd(j)
        send_cmd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[4] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[4]}")
    elif TEST_CASE_INFO[4] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_cmd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_cmds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[4] == "Power Off" or TEST_CASE_INFO[4] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[4]}，不需要设置Channel")
    elif TEST_CASE_INFO[4] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_cmd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["update_event_list_state"] = True
    send_cmd(KEY["EXIT"])
    send_cmd(KEY["OK"])
    # 添加新预约事件到report
    if TEST_CASE_INFO[4] == "PVR":  # 手动指定dur的‘：’间隔
        new_expected_res_event_info = expected_res_event_info
        dur_time = new_expected_res_event_info[3]
        new_expected_res_event_info[3] = dur_time[:2] + ":" + dur_time[2:]
        GL.report_data[0].extend(new_expected_res_event_info)
    else:
        GL.report_data[0].extend(expected_res_event_info)
    # 退回大画面
    send_more_cmds(exit_to_screen)


def new_add_res_event_1():
    logging.info("new_add_res_event_1")
    # 新增预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_cmds(enter_timer_setting_interface)
    # 获取当前系统时间
    get_current_system_time()
    # 进入事件编辑界面，设置预约事件参数
    edit_add_new_res_event_info()


def cale_str_time_for_add_day(str_time, interval_day):
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    if len(str_time) == 12:     # once事件时间计算
        fmt_year = int(str_time[:4])
        fmt_month = int(str_time[4:6])
        fmt_day = int(str_time[6:8])
        fmt_hour = int(str_time[8:10])
        fmt_minute = int(str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(days=interval_day)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
        logging.info(f"----------------------------------------------------------------------{str_new_fmt_date}")
    return str_new_fmt_date


def change_str_time_and_fmt_time(str_time, interval_time):
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    if len(str_time) == 12:     # once事件时间计算
        fmt_year = int(str_time[:4])
        fmt_month = int(str_time[4:6])
        fmt_day = int(str_time[6:8])
        fmt_hour = int(str_time[8:10])
        fmt_minute = int(str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(minutes=interval_time)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
    elif len(str_time) == 4:    # daily和weekly事件时间计算
        old_hour = int(str_time[:2])
        old_minute = int(str_time[2:])
        new_hour = 0
        new_minute = 0
        if old_minute + interval_time < 60:
            new_minute = old_minute + interval_time
            new_hour = old_hour
        elif old_minute + interval_time >= 60:
            new_minute = (old_minute + interval_time) - 60
            if old_hour + 1 < 24:
                new_hour = old_hour + 1
            elif old_hour + 1 >= 24:
                new_hour = (old_hour + 1) - 24
        str_new_fmt_date = "{0:02d}".format(new_hour) + "{0:02d}".format(new_minute)
    elif len(str_time) == 5:    # duration时间计算
        old_hour = int(str_time[:2])
        old_minute = int(str_time[3:])
        new_hour = 0
        new_minute = 0
        if old_minute + interval_time < 60:
            new_minute = old_minute + interval_time
            new_hour = old_hour
        elif old_minute + interval_time >= 60:
            new_minute = (old_minute + interval_time) - 60
            if old_hour + 1 < 24:
                new_hour = old_hour + 1
            elif old_hour + 1 >= 24:
                new_hour = (old_hour + 1) - 24
        str_new_fmt_date = "{0:02d}".format(new_hour) + ":" + "{0:02d}".format(new_minute)
    return str_new_fmt_date


def manage_report_data_and_write_data():
    logging.info("manage_report_data_and_write_data")
    # 整理数据以及写数据
    GL.report_data[6] = TEST_CASE_INFO[0]   # 用例编号
    GL.report_data[7] = str(datetime.now())[:19]    # 写该用例报告的时间

    logging.info(GL.report_data)
    time.sleep(2)


def write_data_to_excel():
    logging.info("write_data_to_excel")
    wb = ''
    ws = ''
    excel_title_1 = ["用例编号", "预期新增事件1", "保存事件1", "保存事件2", "预期新增事件2", "新增相同起始时间事件结果"]
    excel_title_2 = ["用例编号", "预期新增事件1", "保存事件1", "保存事件2",
                     "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
                     "事件列表预约事件个数", "无效事件提示", "用例测试时间"]

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    blue_font = Font(color=BLUE)
    red_font = Font(color=RED)
    dark_cyan = '00008B8B'
    dark_cyan_font = Font(color=dark_cyan, bold=True)
    a_column_numb = column_index_from_string("A")
    if not os.path.exists(file_path[1]):
        wb = Workbook()
        ws = wb.active
        ws.title = file_path[2]
        # 写excel_title_1的内容
        for i in range(len(excel_title_1)):
            if i == 4:
                ws.cell(1, i + 1).value = excel_title_1[i]
                ws.cell(1, i + 1).alignment = alignment
                ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=9)
            elif i == 5:
                ws.cell(1, i + 5).value = excel_title_1[i]
                ws.cell(1, i + 5).alignment = alignment
                ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)
            else:
                ws.cell(1, i + 1).value = excel_title_1[i]
                ws.cell(1, i + 1).alignment = alignment

        # 写excel_title_2的内容
        for j in range(len(excel_title_2)):
            ws.cell(2, j + 1).value = excel_title_2[j]
            ws.cell(2, j + 1).alignment = alignment
            if j == 0:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 8
            elif j in [1, 2, 3]:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 16
            else:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 11
        # 设置Title的行高
        ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
        ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
        # 合并用例编号单元格，以及report前4个数据的单元格
        for column in range(4):
            ws.merge_cells(start_row=1, start_column=column + 1, end_row=2, end_column=column + 1)

    elif os.path.exists(file_path[1]):
        wb = load_workbook(file_path[1])
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if file_path[2] in sheets_name_list:
            ws = wb[file_path[2]]
        elif file_path[2] not in sheets_name_list:
            ws = wb.create_sheet(file_path[2])
            # 写excel_title_1的内容
            for i in range(len(excel_title_1)):
                if i == 4:
                    ws.cell(1, i + 1).value = excel_title_1[i]
                    ws.cell(1, i + 1).alignment = alignment
                    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=9)
                elif i == 5:
                    ws.cell(1, i + 5).value = excel_title_1[i]
                    ws.cell(1, i + 5).alignment = alignment
                    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)
                else:
                    ws.cell(1, i + 1).value = excel_title_1[i]
                    ws.cell(1, i + 1).alignment = alignment

            # 写excel_title_2的内容
            for j in range(len(excel_title_2)):
                ws.cell(2, j + 1).value = excel_title_2[j]
                ws.cell(2, j + 1).alignment = alignment
                if j == 0:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 8
                elif j in [1, 2, 3]:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 16
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 11
            # 设置Title的行高
            ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
            ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
            # 合并用例编号单元格，以及report前4个数据的单元格
            for column in range(4):
                ws.merge_cells(start_row=1, start_column=column + 1, end_row=2, end_column=column + 1)

    # 获取当前用例修改类型的sheet表的Max_row
    max_row = ws.max_row

    # 写新增预约事件数据
    for d in range(len(GL.report_data)):
        if d in [0, 1, 2]:
            ws.cell(max_row + 1, d + 2).value = str(GL.report_data[d])
            ws.cell(max_row + 1, d + 2).alignment = alignment

            if d == 1:  # 保存事件1
                if GL.report_data[d] == GL.report_data[0]:
                    ws.cell(max_row + 1, d + 2).font = blue_font
                else:
                    ws.cell(max_row + 1, d + 2).font = red_font

            elif d == 2:
                if TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]" \
                        or TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]" \
                        or TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]" \
                        or TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":
                    if GL.report_data[d] == GL.report_data[3]:
                        ws.cell(max_row + 1, d + 2).font = dark_cyan_font
                    else:
                        ws.cell(max_row + 1, d + 2).font = red_font
                else:
                    if GL.report_data[d] == ["----", "----", "----", "----", "----"]:
                        ws.cell(max_row + 1, d + 2).font = dark_cyan_font
                    else:
                        ws.cell(max_row + 1, d + 2).font = red_font

        elif d == 3:    # 新增预期事件2
            for edit_data in range(len(GL.report_data[d])):
                ws.cell(max_row + 1, (d + 1) + edit_data + 1).value = GL.report_data[d][edit_data]
                ws.cell(max_row + 1, (d + 1) + edit_data + 1).alignment = alignment
                if edit_data == 0:  # 起始时间
                    if TEST_CASE_INFO[7] == "Same[time+type+mode]" or \
                            TEST_CASE_INFO[7] == "Same[time+mode]+Diff[type]":
                        if GL.report_data[d][edit_data] == GL.report_data[0][0]:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    elif TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode]" or \
                            TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode]" or \
                            TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]" or \
                            TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":
                        if TEST_CASE_INFO[8] == "Once" and TEST_CASE_INFO[3] == "Once":
                            if GL.report_data[d][edit_data] == GL.report_data[0][0] and \
                                    len(GL.report_data[d][edit_data]) == 12:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                        elif TEST_CASE_INFO[8] == "Once" and TEST_CASE_INFO[3] != "Once":
                            if GL.report_data[d][edit_data][8:] == GL.report_data[0][0] and \
                                    len(GL.report_data[d][edit_data]) == 12:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                        elif TEST_CASE_INFO[8] != "Once" and TEST_CASE_INFO[3] == "Once":
                            if GL.report_data[d][edit_data] == GL.report_data[0][0][8:] and \
                                    len(GL.report_data[d][edit_data]) == 4:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                        elif TEST_CASE_INFO[8] != "Once" and TEST_CASE_INFO[3] != "Once":
                            if GL.report_data[d][edit_data] == GL.report_data[0][0] and \
                                    len(GL.report_data[d][edit_data]) == 4:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    elif TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]" \
                            or TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]":   # 起始时间需要+5
                        if GL.report_data[d][edit_data] == change_str_time_and_fmt_time(GL.report_data[0][0], 5):
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font

                    elif TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]" \
                            or TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]":     # 起始时间需要+5
                        if TEST_CASE_INFO[8] == "Once" and TEST_CASE_INFO[3] != "Once":
                            if GL.report_data[0][0] == \
                                    change_str_time_and_fmt_time(GL.report_data[d][edit_data], -5)[8:] \
                                    and len(GL.report_data[d][edit_data]) == 12:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                        elif TEST_CASE_INFO[8] != "Once" and TEST_CASE_INFO[3] == "Once":
                            if GL.report_data[d][edit_data] == \
                                    change_str_time_and_fmt_time(GL.report_data[0][0], 5)[8:] \
                                    and len(GL.report_data[d][edit_data]) == 4:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                        elif TEST_CASE_INFO[8] != "Once" and TEST_CASE_INFO[3] != "Once":
                            if GL.report_data[d][edit_data] == \
                                    change_str_time_and_fmt_time(GL.report_data[0][0], 5) \
                                    and len(GL.report_data[d][edit_data]) == 4:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    else:
                        ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font

                elif edit_data == 1:  # 事件type
                    if TEST_CASE_INFO[7] == "Same[time+type+mode]" \
                            or TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode]" \
                            or TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]" \
                            or TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]"\
                            or TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]":
                        if GL.report_data[d][edit_data] == GL.report_data[0][1] and \
                                GL.report_data[d][edit_data] == TEST_CASE_INFO[9]:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    elif TEST_CASE_INFO[7] == "Same[time+mode]+Diff[type]" or \
                            TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode]" or \
                            TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]" or \
                            TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]" or \
                            TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":
                        if GL.report_data[d][edit_data] != GL.report_data[0][1] and \
                                GL.report_data[d][edit_data] == TEST_CASE_INFO[9]:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    else:
                        ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font

                elif edit_data == 3:  # duration
                    if TEST_CASE_INFO[9] == "PVR":
                        if TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]" \
                                or TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]":
                            if GL.report_data[d][edit_data] == "00:10":
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                        else:
                            if GL.report_data[d][edit_data] == "00:01":
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    elif TEST_CASE_INFO[9] != "PVR":
                        if GL.report_data[d][edit_data] == "--:--":
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    else:
                        ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font

                elif edit_data == 4:    # 事件Mode
                    if TEST_CASE_INFO[7] == "Same[time+type+mode]" \
                            or TEST_CASE_INFO[7] == "Same[time+mode]+Diff[type]" \
                            or TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]" \
                            or TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]":
                        if GL.report_data[d][edit_data] == GL.report_data[0][4] and \
                                GL.report_data[d][edit_data] == TEST_CASE_INFO[8]:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    elif TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode]" or \
                            TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode]" or \
                            TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]" or \
                            TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]" or \
                            TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]" or \
                            TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":
                        if GL.report_data[d][edit_data] != GL.report_data[0][4] and \
                                GL.report_data[d][edit_data] == TEST_CASE_INFO[8]:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font
                    else:
                        ws.cell(max_row + 1, (d + 1) + edit_data + 1).font = red_font

        elif d == 4:    # 事件列表事件个数
            ws.cell(max_row + 1, d + 5 + 1).value = GL.report_data[d]
            ws.cell(max_row + 1, d + 5 + 1).alignment = alignment
            if TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]" \
                    or TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]"\
                    or TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]"\
                    or TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":
                if GL.report_data[d] == '2':
                    ws.cell(max_row + 1, d + 5 + 1).font = dark_cyan_font
                else:
                    ws.cell(max_row + 1, d + 5 + 1).font = red_font
            else:
                if GL.report_data[d] == '1':
                    ws.cell(max_row + 1, d + 5 + 1).font = blue_font
                else:
                    ws.cell(max_row + 1, d + 5 + 1).font = red_font

        elif d == 5:    # 无效事件提示
            ws.cell(max_row + 1, d + 5 + 1).value = GL.report_data[d]
            ws.cell(max_row + 1, d + 5 + 1).alignment = alignment
            if TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]" \
                    or TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]" \
                    or TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]" \
                    or TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":
                if GL.report_data[d] == 'Add_same_start_time_res_event_success':
                    ws.cell(max_row + 1, d + 5 + 1).font = dark_cyan_font
                else:
                    ws.cell(max_row + 1, d + 5 + 1).font = red_font
            else:
                if GL.report_data[d] == '[PTD]Res_invalid_timer':
                    ws.cell(max_row + 1, d + 5 + 1).font = blue_font
                else:
                    ws.cell(max_row + 1, d + 5 + 1).font = red_font

        elif d == 6:    # 用例编号
            ws.cell(max_row + 1, 1).value = GL.report_data[d]
            ws.cell(max_row + 1, 1).alignment = alignment

        elif d == 7:    # 写报告时间
            ws.cell(max_row + 1, d + 5).value = GL.report_data[d]
            ws.cell(max_row + 1, d + 5).alignment = alignment
    ws.row_dimensions[(max_row + 1)].height = 70    # 设置每次执行的report预约事件信息的行高

    wb.save(file_path[1])


def before_cycle_test_clear_data_and_state():
    # 循环测试前，清理数据和状态变量
    logging.info("before_cycle_test_clear_data_and_state")
    GL.choice_res_ch = ''
    state["clear_variate_state"] = True
    GL.report_data = [[], [], [], [], '', '', '', '']
    GL.res_triggered_numb -= 1
    logging.info("循环测试，延时5秒")
    time.sleep(5)
    logging.info(f"剩余循环次数：{GL.res_triggered_numb}")

    if GL.res_triggered_numb < 1:
        clear_timer_setting_all_events()
        logging.info("程序结束")
        state["receive_loop_state"] = True  # 触发结束接收进程的状态


def calculate_expected_event_2_start_time():
    # 对新增的事件进行计算，修改后的预期起始时间
    logging.info("calculate_expected_event_2_start_time")
    weekday_num_dict = {"Mon.": 0, "Tues.": 1, "Wed.": 2, "Thurs.": 3, "Fri.": 4, "Sat.": 5, "Sun.": 6}
    # time_interval = 5
    str_expected_event_2_start_time = ''
    start_time = GL.report_data[0][0]       # 原新增预约事件的起始时间
    if TEST_CASE_INFO[7] == "Same[time+type+mode]":
        logging.info("当前编辑不涉及修改时间和修改Mode，所以预约事件时间不变")
        str_expected_event_2_start_time = start_time

    elif TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode]":     # Once与weekly的date要相同才能是valid事件
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time[8:]

        elif TEST_CASE_INFO[3] == "Daily":  # 原事件Mode
            if TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                cur_sys_hour_minute_time = int(fmt_sys_time[8:])
                if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                    str_expected_event_2_start_time = sys_time_date + start_time
                elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                    str_expected_event_2_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)

        elif TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:    # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily":    # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":   # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                sys_time_split = re.split(r"[\s:/]", sys_time)
                cur_year = int(sys_time_split[0])
                cur_month = int(sys_time_split[1])
                cur_day = int(sys_time_split[2])
                cur_hour = int(sys_time_split[3])
                cur_minute = int(sys_time_split[4])
                dt_time = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute)
                cur_sys_hour_minute_time = int(sys_time_split[3] + sys_time_split[4])   # 当前系统时间的时分值
                cur_weekday = date(cur_year, cur_month, cur_day).weekday()  # 当前系统时间对应的星期数
                res_event_weekday = weekday_num_dict[TEST_CASE_INFO[3]]     # event_1事件对应的星期数
                if cur_weekday == res_event_weekday:    # 事件1的起始时间星期与当前系统时间日期星期相同时
                    if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                        str_expected_event_2_start_time = dt_time[:8] + start_time
                    elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                        cur_next_weekday = dt_time + timedelta(days=7)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                elif cur_weekday != res_event_weekday:    # 事件1的起始时间星期与当前系统时间日期星期不同时
                    if cur_weekday > res_event_weekday:     # 当前系统时间星期数大于事件1的星期数时
                        interval_day = 7 - cur_weekday + res_event_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                    elif cur_weekday < res_event_weekday:
                        interval_day = res_event_weekday - cur_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time

    elif TEST_CASE_INFO[7] == "Same[time+mode]+Diff[type]":
        logging.info("当前编辑不涉及修改时间和修改Mode，所以预约事件时间不变")
        str_expected_event_2_start_time = start_time

    elif TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode]":     # Once与weekly的date要相同才能是valid事件
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time[8:]
        elif TEST_CASE_INFO[3] == "Daily":  # 原事件Mode
            if TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                cur_sys_hour_minute_time = int(fmt_sys_time[8:])
                if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                    str_expected_event_2_start_time = sys_time_date + start_time
                elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                    str_expected_event_2_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)

        elif TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:    # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily":    # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":   # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                sys_time_split = re.split(r"[\s:/]", sys_time)
                cur_year = int(sys_time_split[0])
                cur_month = int(sys_time_split[1])
                cur_day = int(sys_time_split[2])
                cur_hour = int(sys_time_split[3])
                cur_minute = int(sys_time_split[4])
                dt_time = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute)
                cur_sys_hour_minute_time = int(sys_time_split[3] + sys_time_split[4])   # 当前系统时间的时分值
                cur_weekday = date(cur_year, cur_month, cur_day).weekday()  # 当前系统时间对应的星期数
                res_event_weekday = weekday_num_dict[TEST_CASE_INFO[3]]     # event_1事件对应的星期数
                if cur_weekday == res_event_weekday:    # 事件1的起始时间星期与当前系统时间日期星期相同时
                    if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                        str_expected_event_2_start_time = dt_time[:8] + start_time
                    elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                        cur_next_weekday = dt_time + timedelta(days=7)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                elif cur_weekday != res_event_weekday:    # 事件1的起始时间星期与当前系统时间日期星期不同时
                    if cur_weekday > res_event_weekday:     # 当前系统时间星期数大于事件1的星期数时
                        interval_day = 7 - cur_weekday + res_event_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                    elif cur_weekday < res_event_weekday:
                        interval_day = res_event_weekday - cur_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time

    elif TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]":    # 需要更改起始时间+5
        logging.info("当前编辑涉及修改时间，所以预约事件时间需要变化")
        str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time, 5)

    elif TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]":    # 需要更改起始时间+5
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time[8:], 5)
        elif TEST_CASE_INFO[3] == "Daily":  # 原事件Mode
            if TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time, 5)
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                cur_sys_hour_minute_time = int(fmt_sys_time[8:])
                if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                    str_expected_event_2_start_time = change_str_time_and_fmt_time((sys_time_date + start_time), 5)
                elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                    next_cycle_day_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)
                    str_expected_event_2_start_time = change_str_time_and_fmt_time(next_cycle_day_start_time, 5)

        elif TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time, 5)
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                sys_time_split = re.split(r"[\s:/]", sys_time)
                cur_year = int(sys_time_split[0])
                cur_month = int(sys_time_split[1])
                cur_day = int(sys_time_split[2])
                cur_hour = int(sys_time_split[3])
                cur_minute = int(sys_time_split[4])
                dt_time = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute)
                cur_sys_hour_minute_time = int(sys_time_split[3] + sys_time_split[4])  # 当前系统时间的时分值
                cur_weekday = date(cur_year, cur_month, cur_day).weekday()  # 当前系统时间对应的星期数
                res_event_weekday = weekday_num_dict[TEST_CASE_INFO[3]]  # event_1事件对应的星期数
                if cur_weekday == res_event_weekday:  # 事件1的起始时间星期与当前系统时间日期星期相同时
                    if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                        str_expected_event_2_start_time = change_str_time_and_fmt_time((dt_time[:8] + start_time), 5)
                    elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                        cur_next_weekday = dt_time + timedelta(days=7)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        cur_next_weekday_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                        str_expected_event_2_start_time = change_str_time_and_fmt_time(cur_next_weekday_start_time, 5)
                elif cur_weekday != res_event_weekday:  # 事件1的起始时间星期与当前系统时间日期星期不同时
                    if cur_weekday > res_event_weekday:  # 当前系统时间星期数大于事件1的星期数时
                        interval_day = 7 - cur_weekday + res_event_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        cur_next_weekday_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                        str_expected_event_2_start_time = change_str_time_and_fmt_time(cur_next_weekday_start_time, 5)
                    elif cur_weekday < res_event_weekday:
                        interval_day = res_event_weekday - cur_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        cur_next_weekday_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                        str_expected_event_2_start_time = change_str_time_and_fmt_time(cur_next_weekday_start_time, 5)

    elif TEST_CASE_INFO[7] == "Same[mode]+Diff[time+type+dur]":    # 需要更改起始时间+5
        logging.info("当前编辑涉及修改时间，所以预约事件时间需要变化")
        str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time, 5)

    elif TEST_CASE_INFO[7] == "Diff[time+type+dur+mode]":    # 需要更改起始时间+5
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time[8:], 5)
        elif TEST_CASE_INFO[3] == "Daily":  # 原事件Mode
            if TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time, 5)
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                cur_sys_hour_minute_time = int(fmt_sys_time[8:])
                if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                    str_expected_event_2_start_time = change_str_time_and_fmt_time((sys_time_date + start_time), 5)
                elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                    next_cycle_day_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)
                    str_expected_event_2_start_time = change_str_time_and_fmt_time(next_cycle_day_start_time, 5)

        elif TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = change_str_time_and_fmt_time(start_time, 5)
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                sys_time_split = re.split(r"[\s:/]", sys_time)
                cur_year = int(sys_time_split[0])
                cur_month = int(sys_time_split[1])
                cur_day = int(sys_time_split[2])
                cur_hour = int(sys_time_split[3])
                cur_minute = int(sys_time_split[4])
                dt_time = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute)
                cur_sys_hour_minute_time = int(sys_time_split[3] + sys_time_split[4])  # 当前系统时间的时分值
                cur_weekday = date(cur_year, cur_month, cur_day).weekday()  # 当前系统时间对应的星期数
                res_event_weekday = weekday_num_dict[TEST_CASE_INFO[3]]  # event_1事件对应的星期数
                if cur_weekday == res_event_weekday:  # 事件1的起始时间星期与当前系统时间日期星期相同时
                    if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                        str_expected_event_2_start_time = change_str_time_and_fmt_time((dt_time[:8] + start_time), 5)
                    elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                        cur_next_weekday = dt_time + timedelta(days=7)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        cur_next_weekday_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                        str_expected_event_2_start_time = change_str_time_and_fmt_time(cur_next_weekday_start_time, 5)
                elif cur_weekday != res_event_weekday:  # 事件1的起始时间星期与当前系统时间日期星期不同时
                    if cur_weekday > res_event_weekday:  # 当前系统时间星期数大于事件1的星期数时
                        interval_day = 7 - cur_weekday + res_event_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        cur_next_weekday_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                        str_expected_event_2_start_time = change_str_time_and_fmt_time(cur_next_weekday_start_time, 5)
                    elif cur_weekday < res_event_weekday:
                        interval_day = res_event_weekday - cur_weekday
                        cur_next_weekday = dt_time + timedelta(days=interval_day)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        cur_next_weekday_start_time = ''.join(cur_next_weekday_split)[:8] + start_time
                        str_expected_event_2_start_time = change_str_time_and_fmt_time(cur_next_weekday_start_time, 5)

    elif TEST_CASE_INFO[7] == "Same[time+type]+Diff[mode+date]" \
            or TEST_CASE_INFO[7] == "Same[time]+Diff[type+mode+date]":  # Once与weekly的date要不同才能是valid事件
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time[8:]
        elif TEST_CASE_INFO[3] == "Daily":  # 原事件Mode
            if TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                cur_sys_hour_minute_time = int(fmt_sys_time[8:])
                if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                    str_expected_event_2_start_time = sys_time_date + start_time
                elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                    str_expected_event_2_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)

        elif TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:    # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily":    # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":   # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                sys_time_split = re.split(r"[\s:/]", sys_time)
                cur_year = int(sys_time_split[0])
                cur_month = int(sys_time_split[1])
                cur_day = int(sys_time_split[2])
                cur_hour = int(sys_time_split[3])
                cur_minute = int(sys_time_split[4])
                dt_time = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute)
                cur_sys_hour_minute_time = int(sys_time_split[3] + sys_time_split[4])   # 当前系统时间的时分值
                cur_weekday = date(cur_year, cur_month, cur_day).weekday()  # 当前系统时间对应的星期数
                res_event_weekday = weekday_num_dict[TEST_CASE_INFO[3]]     # event_1事件对应的星期数
                if cur_weekday != res_event_weekday:    # 事件1的起始时间星期与当前系统时间日期星期不同时
                    if int(start_time) < cur_sys_hour_minute_time:  # 事件1的起始时间早于当前系统时间
                        str_expected_event_2_start_time = dt_time[:8] + start_time
                    elif int(start_time) >= cur_sys_hour_minute_time:  # 事件1的起始时间等于或晚于当前系统时间，需要增加一个循环
                        cur_next_weekday = dt_time + timedelta(days=7)
                        cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                        str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time

                elif cur_weekday == res_event_weekday:    # 事件1的起始时间星期与当前系统时间日期星期相同时
                    interval_day = 1
                    cur_next_weekday = dt_time + timedelta(days=interval_day)
                    cur_next_weekday_split = re.split(r"[-\s:]", str(cur_next_weekday))
                    str_expected_event_2_start_time = ''.join(cur_next_weekday_split)[:8] + start_time

    logging.info(f"期望的完整的预约事件时间为{str_expected_event_2_start_time}")
    return str_expected_event_2_start_time


def calculate_expected_event_2_duration_time():
    logging.info("calculate_expected_event_2_duration_time")
    str_expected_dur_time = ''
    interval_dur = 1        # 更改录制时长的变量
    specified_dur_time = "0002"   # ModifyType+ModifyDuration时会出现无Dur_time改有Dur_time的情况，直接指定时间

    if TEST_CASE_INFO[7] == "Same[type+dur+mode]+Diff[time]" or \
            TEST_CASE_INFO[7] == "Same[type+dur]+Diff[time+mode]":
        logging.info("当前事件需要更改Duration")
        str_expected_dur_time = "0010"
    else:
        logging.info("当前事件不需要更改Duration，保持默认值")
        str_expected_dur_time = "0001"
    return str_expected_dur_time


def create_expected_event_2_info():
    logging.info("create_expected_event_2_info")
    # 创建修改后的期望的事件信息
    expected_event_2_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    duration_time = calculate_expected_event_2_duration_time()
    if TEST_CASE_INFO[9] == "Play":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = channel_info[1]
        expected_event_2_info[3] = "--:--"
        expected_event_2_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "PVR":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = channel_info[1]
        expected_event_2_info[3] = duration_time
        expected_event_2_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "Power Off":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = "----"
        expected_event_2_info[3] = "--:--"
        expected_event_2_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "Power On":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = "----"
        expected_event_2_info[3] = "--:--"
        expected_event_2_info[4] = TEST_CASE_INFO[8]
    return expected_event_2_info


def edit_add_new_res_event_2_info():
    logging.info("edit_add_new_res_event_2_info")
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面
    send_cmd(KEY["GREEN"])
    # 生成预期的预约事件
    expected_res_event_info = create_expected_event_2_info()

    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[9]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[9]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[9]}')
            send_cmd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[8]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[8]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[8]}')
            send_cmd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[8] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[8]}")
    elif TEST_CASE_INFO[8] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[8]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_cmd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_cmds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_cmd(KEY["DOWN"])
    else:
        if len(expected_res_event_info[0]) == 12:
            start_time_list.append(expected_res_event_info[0][8:])
        elif len(expected_res_event_info[0]) == 4:
            start_time_list.append(expected_res_event_info[0])
        start_time_cmd = change_numbs_to_cmds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_cmd(j)
        send_cmd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[9] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[9]}")
    elif TEST_CASE_INFO[9] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[9]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_cmd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_cmds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[9]}，不需要设置Channel")
    elif TEST_CASE_INFO[9] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[9]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_cmd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["clear_res_event_list_state"] = True
    send_cmd(KEY["EXIT"])
    send_cmd(KEY["OK"])
    # 此处编辑完事件后，不会再次打印事件信息，需要重新进入Timer Setting界面，为update_edit_res_event_to_event_mgr_list准备数据
    time.sleep(2)
    # 更新event_2的信息
    if TEST_CASE_INFO[9] == "PVR":      # Invalid事件时，这里获取不到event_2的事件信息，所以只能手动指定dur的‘：’间隔
        new_expected_res_event_info = expected_res_event_info
        dur_time = new_expected_res_event_info[3]
        new_expected_res_event_info[3] = dur_time[:2] + ":" + dur_time[2:]
        GL.report_data[3] = new_expected_res_event_info
    else:
        GL.report_data[3] = expected_res_event_info

    if rsv_kws["event_invalid_msg"] != '':
        GL.report_data[5] = rsv_kws["event_invalid_msg"]
        send_cmd(KEY["OK"])
    elif rsv_kws["event_invalid_msg"] == '':
        GL.report_data[5] = "Add_same_start_time_res_event_success"
    # 退回大画面
    send_more_cmds(exit_to_screen)


def new_add_res_event_2():
    logging.info("new_add_res_event_2")
    # 编辑修改新增的预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_cmds(enter_timer_setting_interface)
    # 进入事件编辑界面，设置预约事件参数
    edit_add_new_res_event_2_info()


def check_event_numb():
    logging.info("check_event_numb")
    # 检查Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 获取已预约的事件信息，清除获取预约事件的list，并激活获取预约事件状态标志
    state["clear_res_event_list_state"] = True
    state["update_event_list_state"] = True
    # 进入定时器设置界面
    send_more_cmds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        GL.report_data[4] = rsv_kws["res_event_numb"]
        state["res_event_numb_state"] = False
        # 获取预约事件的状态标志关闭
        state["update_event_list_state"] = False
        logging.info(list(res_event_list))
        # 根据获取到的事件个数来决定保存事件1和保存事件2的信息
        if int(rsv_kws["res_event_numb"]) == 1 and len(list(res_event_list)) == 1:
            GL.report_data[1] = list(res_event_list)[0]
            GL.report_data[2] = ["----", "----", "----", "----", "----"]
        elif int(rsv_kws["res_event_numb"]) == 2 and len(list(res_event_list)) == 2:
            for event in list(res_event_list):
                if event == GL.report_data[0]:
                    GL.report_data[1] = event
                else:
                    GL.report_data[2] = event
    # 退回大画面
    send_more_cmds(exit_to_screen)


def mail(message):
    my_sender = 'wangrun@nationalchip.com'  # 发件人邮箱账号
    my_pass = 'Wr@372542098'  # 发件人邮箱密码
    my_user = 'wangrun@nationalchip.com'  # 收件人邮箱账号，我这边发送给自己

    return_state = True
    try:
        msg = MIMEText(message, 'plain', 'utf-8')
        # msg['From'] = formataddr(["FromRunoob", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['From'] = Header("Auto_test", 'utf-8')
        msg['To'] = Header("ME", 'utf-8')
        # msg['To'] = formataddr(["FK", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
        msg['Subject'] = "自动化测试终止提醒"  # 邮件的主题，也可以说是标题

        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 关闭连接
    except smtplib.SMTPException:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
        return_state = False
    return return_state


def receive_serial_process(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info,
        receive_cmd_list):
    logging_info_setting()
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIME_SHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict([val, key] for key, val in rsv_key.items())

    res_kws = [
        "[PTD]Time_mode=",              # 0     获取系统时间模式
        "[PTD]System_time=",            # 1     系统时间
        "[PTD]Res_event_numb=",         # 2     预约事件数量
        "[PTD]Res_event:",              # 3     预约事件信息
        "[PTD]Res_triggered:",          # 4     预约事件触发和当前响应事件的信息
        "[PTD]Res_confirm_jump",        # 5     预约事件确认跳转
        "[PTD]Res_cancel_jump",         # 6     预约事件取消跳转
        "[PTD]REC_start",               # 7     录制开始
        "[PTD]REC_end",                 # 8     录制结束
        "[PTD]No_storage_device",       # 9     没有存储设备
        "[PTD]No_enough_space",         # 10    没有足够的空间
        "[PTD]power_cut",               # 11    进入待机
        "[PTD]:switch totle cost",      # 12    开机解码成功
        "[PTD][HOTPLUG] PLUG_IN",       # 13    存储设备插入成功
        "[PTD]PVR_is_not_supported",    # 14    录制无信号、加锁节目、加密节目，跳出PVR is not supported!提示
        "[PTD]Current_sys_time",        # 15    预约事件触发时，检测触发时间信息
    ]

    switch_ch_kws = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    ch_info_kws = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name"]

    group_info_kws = [
        "Group_name",
        "Prog_total"
    ]

    edit_event_kws = [
        "[PTD]Mode=",
        "[PTD]Type=",
        "[PTD]Start Date=",
        "[PTD]Start Time=",
        "[PTD]Duration=",
        "[PTD]Channel="
    ]

    event_invalid_msg = [
        "[PTD]Res_no_channel",
        "[PTD]Res_invalid_date",
        "[PTD]Res_invalid_timer"
    ]

    other_kws = [
        "[PTD]Infrared_key_values:",    # 获取红外接收关键字
    ]

    infrared_rsv_cmd = []       # 红外接受命令
    receive_serial = serial.Serial(prs_data["receive_serial_name"], 115200, timeout=1)

    while True:
        data = receive_serial.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            data1 = data.decode("ISO-8859-1", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}".format(str(tt), data2)
            data4 = "[{}]     {}\n".format(str(tt), data2)
            print(data3)
            write_log_data_to_txt(prs_data["log_file_path"], data4)

            if state["clear_variate_state"]:
                state["sys_time_mode_state"] = False
                state["current_sys_time_state"] = False
                state["res_event_numb_state"] = False
                state["update_event_list_state"] = False
                state["clear_variate_state"] = False
                state["event_no_channel_msg_state"] = False
                state["event_invalid_date_msg_state"] = False
                state["event_invalid_timer_msg_state"] = False

                rsv_kws["event_invalid_msg"] = ''

                del current_triggered_event_info[:]
                if prs_data["case_res_event_mode"] == "Once":
                    del res_event_list[:]

            if state["clear_res_event_list_state"]:
                del res_event_list[:]
                state["clear_res_event_list_state"] = False

            if other_kws[0] in data2:   # 红外接收打印
                rsv_cmd = re.split(":", data2)[-1]
                infrared_rsv_cmd.append(rsv_cmd)        # 存放可以共享的接受命令的列表
                if rsv_cmd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_cmd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        infrared_send_cmd[-1], reverse_rsv_key[infrared_rsv_cmd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(infrared_send_cmd), len(infrared_rsv_cmd)))
                    receive_cmd_list.append(rsv_cmd)

            if res_kws[0] in data2:     # 获取系统时间模式（自动还是手动）
                state["sys_time_mode_state"] = True
                rsv_kws["sys_time_mode"] = re.split(r"=", data2)[-1]

            if res_kws[1] in data2:     # 获取当前系统时间
                state["current_sys_time_state"] = True
                rsv_kws["current_sys_time"] = re.split(r"=", data2)[-1]

            if res_kws[2] in data2:     # 获取预约事件数量
                state["res_event_numb_state"] = True
                rsv_kws["res_event_numb"] = re.split(r"=", data2)[-1]

            if res_kws[3] in data2:     # 获取预约事件信息
                event_split_info = re.split(r"event:|,", data2)
                event_info = ['', '', '', '', '']
                for info in event_split_info:
                    if "Start_time" in info:
                        event_start_time = re.split(r"=", info)[-1]
                        if len(event_start_time) == 5:
                            event_info[0] = ''.join(re.split(r":", event_start_time))
                        elif len(event_start_time) == 16:
                            event_info[0] = ''.join(re.split(r"[/:\s]", event_start_time))
                    if "Event_type" in info:
                        event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        event_info[4] = re.split(r"=", info)[-1]
                if state["update_event_list_state"]:
                    res_event_list.append(event_info)

            if switch_ch_kws[0] in data2:
                ch_info_split = re.split(r"[],]", data2)
                for i in range(len(ch_info_split)):
                    if ch_info_kws[0] in ch_info_split[i]:  # 提取频道号
                        channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if ch_info_kws[1] in ch_info_split[i]:  # 提取频道名称
                        channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if switch_ch_kws[1] in data2:
                flag_info_split = re.split(r"[],]", data2)
                for i in range(len(flag_info_split)):
                    if ch_info_kws[2] in flag_info_split[i]:  # 提取频道所属TP
                        channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if ch_info_kws[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if switch_ch_kws[3] in data2:
                group_info_split = re.split(r"[],]", data2)
                for i in range(len(group_info_split)):
                    if group_info_kws[0] in group_info_split[i]:  # 提取频道所属组别
                        rsv_kws["prog_group_name"] = re.split(r"=", group_info_split[i])[-1]
                        channel_info[6] = rsv_kws["prog_group_name"]
                    if group_info_kws[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        rsv_kws["prog_group_total"] = re.split(r"=", group_info_split[i])[-1]

            if edit_event_kws[0] in data2:          # 提取Mode参数
                rsv_kws["edit_event_focus_pos"] = "Mode"
                rsv_kws["edit_event_mode"] = re.split(r"=", data2)[-1]

            if edit_event_kws[1] in data2:          # 提取Type参数
                rsv_kws["edit_event_focus_pos"] = "Type"
                rsv_kws["edit_event_type"] = re.split(r"=", data2)[-1]

            if edit_event_kws[2] in data2:          # 提取Start Date参数
                rsv_kws["edit_event_focus_pos"] = "Start Date"
                rsv_kws["edit_event_date"] = re.split(r"=", data2)[-1]

            if edit_event_kws[3] in data2:          # 提取Start Time参数
                rsv_kws["edit_event_focus_pos"] = "Start Time"
                rsv_kws["edit_event_time"] = re.split(r"=", data2)[-1]

            if edit_event_kws[4] in data2:          # 提取Duration参数
                rsv_kws["edit_event_focus_pos"] = "Duration"
                rsv_kws["edit_event_duration"] = re.split(r"=", data2)[-1]

            if edit_event_kws[5] in data2:          # 提取Channel参数
                rsv_kws["edit_event_focus_pos"] = "Channel"
                rsv_kws["edit_event_ch"] = re.split(r"=", data2)[-1]

            if event_invalid_msg[0] in data2:
                state["event_no_channel_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2

            if event_invalid_msg[1] in data2:
                state["event_invalid_date_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2

            if event_invalid_msg[2] in data2:
                state["event_invalid_timer_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2


if __name__ == "__main__":

    GL = MyGlobal()
    logging_info_setting()
    msg = "现在开始执行的是:{}_{}_{}_{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2], TEST_CASE_INFO[4],
        TEST_CASE_INFO[3], TEST_CASE_INFO[7], TEST_CASE_INFO[9], TEST_CASE_INFO[8])
    logging.critical(format(msg, '*^150'))
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIME_SHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D"
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())
    WAIT_INTERFACE = ["TVScreenDiffCH", "RadioScreenDiffCH", "ChannelList", "Menu", "EPG", "ChannelEdit"]
    WEEKLY_EVENT_MODE = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
    # TEST_CASE_INFO = ["23", "All", "TV", "Daily", "Play", "EPG", "Manual_jump"]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]

    file_path = create_log_and_report_file_path()
    ser_name = list(check_ports())  # send_ser_name, receive_ser_name
    send_serial = serial.Serial(ser_name[0], 9600)
    receive_ser_name = ser_name[1]

    infrared_send_cmd = Manager().list([])
    receive_cmd_list = Manager().list([])
    res_event_list = Manager().list([])
    current_triggered_event_info = Manager().list([])
    channel_info = Manager().list(['', '', '', '', '', '', ''])     # [频道号,频道名称,tp,lock,scramble,频道类型,组别]
    rsv_kws = Manager().dict({
        "sys_time_mode": '', "current_sys_time": '', "res_event_numb": '', "prog_group_name": '',
        "prog_group_total": '', "edit_event_focus_pos": '', "edit_event_mode": '', "edit_event_type": '',
        "edit_event_date": '', "edit_event_time": '', "edit_event_duration": '', "edit_event_ch": '',
        "event_invalid_msg": '',
    })

    state = Manager().dict({
        "res_event_numb_state": False, "sys_time_mode_state": False,
        "current_sys_time_state": False, "update_event_list_state": False,
        "clear_variate_state": False, "receive_loop_state": False,
        "clear_res_event_list_state": False, "event_no_channel_msg_state": False,
        "event_invalid_date_msg_state": False, "event_invalid_timer_msg_state": False
    })

    prs_data = Manager().dict({
        "log_file_path": file_path[0], "receive_serial_name": receive_ser_name, "case_res_event_mode": TEST_CASE_INFO[8]
    })

    rsv_p = Process(target=receive_serial_process, args=(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info,
        receive_cmd_list))
    rsv_p.start()

    if platform.system() == "Windows":
        time.sleep(5)
        logging.info("Windows系统接收端响应慢，等待5秒")
    elif platform.system() == "Linux":
        time.sleep(1)
        logging.info("Linux系统接收端响应快，但是增加一个延时保护，等待1秒")

    # 主程序开始部分

    while GL.res_triggered_numb > 0:
        clear_timer_setting_all_events()
        check_sys_time_mode()
        choice_ch_for_res_event_type("event_1")
        new_add_res_event_1()
        choice_ch_for_res_event_type("event_2")
        new_add_res_event_2()
        check_event_numb()
        manage_report_data_and_write_data()
        write_data_to_excel()
        before_cycle_test_clear_data_and_state()

    if state["receive_loop_state"]:
        rsv_p.terminate()
        logging.info('stop receive process')
        rsv_p.join()
