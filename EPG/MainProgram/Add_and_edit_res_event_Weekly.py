#!/usr/bin/python3
# -*- coding: utf-8 -*-

from serial_setting1 import *
from multiprocessing import Process, Manager
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import RED, BLUE
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.comments import Comment
from datetime import datetime, timedelta, date
from random import randint, choice
from email.mime.text import MIMEText
from email.header import Header
import smtplib
import platform
import os
import time
import logging
import re
import sys
import traceback


Modify_list = [
    "ModifyTime",
    "ModifyType",
    "ModifyDuration",
    "ModifyMode",
    "ModifyTime+ModifyType",
    "ModifyTime+ModifyDuration",
    "ModifyTime+ModifyMode",
    "ModifyType+ModifyDuration",
    "ModifyType+ModifyMode",
    "ModifyDuration+ModifyMode",
    "ModifyTime+ModifyType+ModifyDuration",
    "ModifyTime+ModifyType+ModifyMode",
    "ModifyType+ModifyDuration+ModifyMode",
    "ModifyTime+ModifyType+ModifyDuration+ModifyMode"
]


class MyGlobal(object):

    def __init__(self):
        if TEST_CASE_INFO[-1] == "screen_test_numb":
            self.res_triggered_numb = 1                 # 大画面预约响应的次数
        elif TEST_CASE_INFO[-1] == "other_interface_test_numb":
            self.res_triggered_numb = 1                 # 其他界面预约响应的次数

        self.choice_res_ch = ''                         # 预约Play或PVR事件时所选预约节目
        self.res_event_mgr = []                         # 预约事件管理
        self.start_row = 0                              # 用于每次预约事件响应后，写数据增加行数
        self.pvr_rec_dur_time = ''                      # 用于记录PVR事件录制持续时间
        self.event_already_triggered_numb = 0           # 用于控制循环事件第二次前后的运行代码界限

        # 报告数据汇总[[预约事件信息]，[修改后的预约事件信息], "系统时间日期", "触发时间", "等待节目", "跳转节目", "录制时长", "case编号", "执行case时间"]
        self.report_data = [[], [], '', '', '', '', '', '', '']
        # ["报告名称", "预约事件类型", "预约事件模式", "预约节目类型", "预约等待界面", "预约跳转模式", "预约执行次数"]
        self.title_data = ['', '', '', '', '', '', '']


def logging_info_setting():
    # 配置logging输出格式
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)


def hex_strs_to_bytes(strings):
    # 将红外命令字符串转换为字节串
    return bytes.fromhex(strings)


def write_log_data_to_txt(path, write_data):
    with open(path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def send_commd(commd):
    global receive_cmd_list, infrared_send_cmd
    continuous_transmission_cmd_num = 0
    # 红外发送端发送指令
    send_serial.write(hex_strs_to_bytes(commd))
    send_serial.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
    if REVERSE_KEY[commd] != "POWER":
        infrared_send_cmd.append(REVERSE_KEY[commd])
    time.sleep(1.0)
    if len(infrared_send_cmd) == len(receive_cmd_list):
        pass
    elif len(infrared_send_cmd) != len(receive_cmd_list):
        logging.info("检测到发送和接收命令数不一致，等待2秒，查看是否接收端还没有接收到打印")
        time.sleep(2)
        while True:
            if len(infrared_send_cmd) == len(receive_cmd_list):
                break
            # elif len(infrared_send_cmd) - len(receive_cmd_list) == 1:
            elif len(infrared_send_cmd) != len(receive_cmd_list):
                logging.info(f"此刻补发STB没有接收到的红外命令{infrared_send_cmd[-1]}")
                send_serial.write(hex_strs_to_bytes(KEY[infrared_send_cmd[-1]]))
                send_serial.flush()
                time.sleep(1.0)
                continuous_transmission_cmd_num += 1
                if continuous_transmission_cmd_num == 10:
                    stb_crash_msg = "STB一直发送指令，疑似死机"
                    # mail(f'{stb_crash_msg}\n\n{msg}')
                    raise FailSendCmdException(stb_crash_msg)


def send_more_commds(commd_list):
    # 用于发送一连串的指令
    for commd in commd_list:
        send_commd(commd)
    time.sleep(1)   # 增加函数切换时的的等待，避免可能出现send_commd函数中的等待时间没有执行的情况


def change_numbs_to_commds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_commds_list = []
    for i in range(len(numbs_list)):
        channel_commds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_commds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_commds_list[i].append(KEY[numbs_list[i][j]])
    return channel_commds_list


def build_log_and_report_file_path():
    # 用于创建打印和报告文件路径
    # 构建存放数据的总目录，以及构建存放打印和报告的目录
    parent_path = os.path.dirname(os.getcwd())
    case_name = "Add_and_edit_res_event"
    test_data_directory_name = "test_data"
    test_data_directory_path = os.path.join(parent_path, test_data_directory_name)
    log_directory_name = "print_log"
    log_directory_path = os.path.join(test_data_directory_path, log_directory_name)
    report_directory_name = "report"
    report_directory_path = os.path.join(test_data_directory_path, report_directory_name)

    log_case_directory_path = os.path.join(test_data_directory_path, log_directory_name, case_name)
    report_case_directory_path = os.path.join(test_data_directory_path, report_directory_name, case_name)
    # 判断目录是否存在，否则创建目录
    if not os.path.exists(test_data_directory_path):
        os.mkdir(test_data_directory_path)
    if not os.path.exists(log_directory_path):
        os.mkdir(log_directory_path)
    if not os.path.exists(report_directory_path):
        os.mkdir(report_directory_path)
    if not os.path.exists(log_case_directory_path):
        os.mkdir(log_case_directory_path)
    if not os.path.exists(report_case_directory_path):
        os.mkdir(report_case_directory_path)
    # 创建打印和报告文件的名称和路径
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    rename_modify_type = TEST_CASE_INFO[7].replace('Modify', '')        # 用于提取TEST_CASE_INFO[7]中重复的Modify
    sheet_name = f"Modify({rename_modify_type})"

    fmt_name = "{}_{}_{}_{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2], TEST_CASE_INFO[4],
        TEST_CASE_INFO[3], sheet_name, TEST_CASE_INFO[9], TEST_CASE_INFO[8])
    log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    log_file_path = os.path.join(log_case_directory_path, log_file_name)
    report_file_name = "Edit_modify_res_event_result_report.xlsx"
    report_file_path = os.path.join(report_case_directory_path, report_file_name)
    # sheet_name = "{}_{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[4], TEST_CASE_INFO[3], TEST_CASE_INFO[7])
    return log_file_path, report_file_path, sheet_name


def clear_timer_setting_all_events():
    logging.info("clear_timer_setting_all_events")
    # 清除Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    delete_all_res_events = [KEY["BLUE"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入定时器设置界面
    send_more_commds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        if rsv_kws["res_event_numb"] != '0':
            send_more_commds(delete_all_res_events)
        elif rsv_kws["res_event_numb"] == '0':
            logging.info("没有预约事件存在")
            time.sleep(1)
        else:
            logging.debug("警告：预约事件个数获取错误！！！")
        state["res_event_numb_state"] = False
    # 退回大画面
    send_more_commds(exit_to_screen)


def check_sys_time_mode():
    logging.info("check_sys_time_mode")
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入时间设置界面
    send_more_commds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        if rsv_kws["sys_time_mode"] == "Auto":
            send_more_commds(change_sys_time_mode)
        elif rsv_kws["sys_time_mode"] == "Manual":
            logging.info("系统时间模式已经为手动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    # 退回大画面
    send_more_commds(exit_to_screen)


def get_current_system_time():
    logging.info("get_current_system_time")
    # 获取当前系统时间
    time.sleep(1)
    while not state["current_sys_time_state"]:
        logging.info("还没有获取到系统时间信息")
        time.sleep(1)
    else:
        logging.info(f"当前系统时间为:{rsv_kws['current_sys_time']}")


def get_exist_event_info():
    # 获取当前已预约的事件信息
    if len(res_event_list) == 0:
        logging.info("当前没有预约事件")
        logging.info(f"预约事件列表为:{res_event_list}")
    else:
        logging.info("当前所有预约事件如下")
        for event in res_event_list:
            logging.info(event)


def choice_ch_for_res_event_type():
    logging.info("choice_ch_for_res_event_type")
    # 根据预约事件类型来选择节目
    group_dict = {}
    choice_ch_numb = []
    # 根据所选case切换到对应类型节目的界面
    while channel_info[5] != TEST_CASE_INFO[2]:
        send_commd(KEY["TV/R"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_commd(KEY["OK"])
    # 切到指定分组和采集分组下节目总数信息
    while rsv_kws["prog_group_name"] != TEST_CASE_INFO[1]:
        send_commd(KEY["RIGHT"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    else:
        if rsv_kws["prog_group_name"] == '':
            logging.info("警告：没有All分组信息")
        else:
            group_dict[rsv_kws["prog_group_name"]] = rsv_kws["prog_group_total"]
            logging.info(f"分组信息{group_dict}")
    # 退出频道列表,回到大画面界面
    send_commd(KEY["EXIT"])
    # 根据用例指定的事件类型来选择节目
    if TEST_CASE_INFO[9] == "Play":
        choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
        choice_ch_cmd = change_numbs_to_commds_list(choice_ch_numb)
        for i in range(len(choice_ch_cmd)):
            for j in choice_ch_cmd[i]:
                send_commd(j)
        send_commd(KEY["OK"])
        time.sleep(2)
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
        logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
        GL.choice_res_ch = channel_info[1]
        logging.info(channel_info)

    elif TEST_CASE_INFO[9] == "PVR":
        choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
        choice_ch_cmd = change_numbs_to_commds_list(choice_ch_numb)
        for i in range(len(choice_ch_cmd)):
            for j in choice_ch_cmd[i]:
                send_commd(j)
        send_commd(KEY["OK"])
        time.sleep(2)
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])

        while channel_info[3] != '0' and channel_info[4] != '0':
            logging.info(f"查看所切节目信息：{channel_info}")
            logging.info("所选节目不为免费节目，不可以进行PVR预约，继续切台")
            send_commd(KEY["UP"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        else:
            logging.info("所选节目为免费节目，可以进行PVR预约")
            logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
            GL.choice_res_ch = channel_info[1]
            logging.info(channel_info)

    elif TEST_CASE_INFO[9] == "Power Off":
        logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

    elif TEST_CASE_INFO[9] == "Power On":
        logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")


def calculate_expected_event_start_time():
    logging.info("calculate_expected_event_start_time")
    time_interval = 5
    sys_time = rsv_kws['current_sys_time']
    sys_time_split = re.split(r"[\s:/]", sys_time)
    sys_year = int(sys_time_split[0])
    sys_month = int(sys_time_split[1])
    sys_day = int(sys_time_split[2])
    sys_hour = int(sys_time_split[3])
    sys_minute = int(sys_time_split[4])
    dt_time = datetime(sys_year, sys_month, sys_day, sys_hour, sys_minute)
    logging.info(dt_time)
    expected_res_time = dt_time + timedelta(minutes=time_interval)
    logging.info(expected_res_time)
    expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
    str_expected_res_time = ''.join(expected_res_time_split)[:12]

    logging.info(f"期望的完整的预约事件时间为{str_expected_res_time}")
    return str_expected_res_time


def create_expected_add_event_info():
    logging.info("create_expected_add_event_info")
    # 创建期望的事件信息
    expected_event_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    duration_time = "0001"
    if TEST_CASE_INFO[4] == "Play":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "PVR":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = duration_time
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power Off":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power On":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]
    return expected_event_info


def edit_add_new_res_event_info():
    logging.info("edit_add_new_res_event_info")
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面
    send_commd(KEY["GREEN"])
    # 生成预期的预约事件
    expected_res_event_info = create_expected_add_event_info()
    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[4]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_commd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[3]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_commd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[3] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[3]}")
    elif TEST_CASE_INFO[3] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[3]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_commd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_commds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_commd(KEY["DOWN"])
    else:
        start_time_list.append(expected_res_event_info[0][8:])
        start_time_cmd = change_numbs_to_commds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_commd(j)
        send_commd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[4] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[4]}")
    elif TEST_CASE_INFO[4] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_commd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_commds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[4] == "Power Off" or TEST_CASE_INFO[4] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[4]}，不需要设置Channel")
    elif TEST_CASE_INFO[4] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_commd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["update_event_list_state"] = True
    send_commd(KEY["EXIT"])
    send_commd(KEY["OK"])
    # 退回大画面
    send_more_commds(exit_to_screen)


def add_new_res_event_to_event_mgr_list():
    logging.info("add_new_res_event_to_event_mgr_list")
    # 添加新预约事件到事件管理列表
    if res_event_list not in GL.res_event_mgr:
        GL.res_event_mgr.extend(res_event_list)
    GL.report_data[0] = GL.res_event_mgr[0]
    logging.info(type(GL.res_event_mgr))
    logging.info(GL.res_event_mgr)
    logging.info(list(res_event_list))
    state["update_event_list_state"] = False


def new_add_res_event():
    logging.info("new_add_res_event")
    # 新增预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_commds(enter_timer_setting_interface)
    # 获取当前系统时间
    get_current_system_time()
    # 进入事件编辑界面，设置预约事件参数
    edit_add_new_res_event_info()
    # 添加新预约事件到事件管理列表
    add_new_res_event_to_event_mgr_list()


def get_cycle_event_start_time_and_sys_date():
    logging.info("get_cycle_event_start_time_and_sys_date")
    # 获取循环事件的起始时间和系统时间的日期，组成完整的时间
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    full_cycle_event_date_time = ''
    # 进入Time_Setting界面
    send_more_commds(enter_time_setting_interface)
    # 获取当前系统时间的date
    get_current_system_time()
    sys_time = rsv_kws['current_sys_time']
    logging.info(sys_time)
    sys_time_split = re.split(r"[\s:/]", sys_time)
    fmt_sys_time = ''.join(sys_time_split)
    sys_time_date = fmt_sys_time[:8]
    # 查看预约事件列表时间信息
    logging.info(res_event_list)
    if len(list(res_event_list)[0][0]) == 4:
        logging.info("当前循环事件起始时间为4位数")
        cycle_event_start_time = list(res_event_list)[0][0]
        logging.info(f"cycle_event_start_time:{cycle_event_start_time}")
        # 处理系统日期和事件时间，合并为一个完整的12位时间
        full_cycle_event_date_time = sys_time_date + cycle_event_start_time
        logging.info(f"系统日期和事件时间合并后的时间为：{full_cycle_event_date_time}")
    elif len(list(res_event_list)[0][0]) == 12:     # 假如事件的起始时间为12位数，不用与系统时间合并，直接使用该事件起始时间
        logging.info("当前循环事件起始时间为12位数")
        full_cycle_event_date_time = list(res_event_list)[0][0]
        logging.info(f"事件的起始时间为12位数，不用与系统时间合并，直接使用该事件起始时间为：{full_cycle_event_date_time}")
    # 将获取信息的状态变量恢复默认
    state["current_sys_time_state"] = False
    state["res_event_numb_state"] = False
    state["update_event_list_state"] = False
    return full_cycle_event_date_time


def calc_modify_system_time():
    logging.info("calc_modify_system_time")
    # 计算循环事件下次响应的系统时间
    # w = (d + 1 + 2 * m + 3 * (m + 1) / 5 + y + y / 4 - y / 100 + y / 400) % 7
    weekly_event_mode = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
    weekday_num_dict = {"Mon.": 0, "Tues.": 1, "Wed.": 2, "Thurs.": 3, "Fri.": 4, "Sat.": 5, "Sun.": 6}
    ahead_of_time = 1   # 提前系统时间到事件起始时间前1分钟
    full_cyc_event_date_time = get_cycle_event_start_time_and_sys_date()
    fmt_next_sys_date_time = ''
    cur_year = int(full_cyc_event_date_time[:4])
    cur_month = int(full_cyc_event_date_time[4:6])
    cur_day = int(full_cyc_event_date_time[6:8])
    cur_hour = int(full_cyc_event_date_time[8:10])
    cur_minute = int(full_cyc_event_date_time[10:12])
    cur_date = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute)
    if TEST_CASE_INFO[8] == "Once":
        next_triggered_time = cur_date - timedelta(minutes=ahead_of_time)
        next_triggered_time_split = re.split(r"[-\s:]", str(next_triggered_time))
        fmt_next_sys_date_time = ''.join(next_triggered_time_split)[:12]  # 去掉末尾的秒钟信息

    elif TEST_CASE_INFO[8] == "Daily":
        if GL.event_already_triggered_numb == 0:
            next_triggered_time = cur_date - timedelta(minutes=ahead_of_time)
            next_triggered_time_split = re.split(r"[-\s:]", str(next_triggered_time))
            fmt_next_sys_date_time = ''.join(next_triggered_time_split)[:12]  # 去掉末尾的秒钟信息
        else:
            next_date = cur_date + timedelta(days=1)
            next_sys_date_time = next_date - timedelta(minutes=ahead_of_time)
            next_sys_date_time_split = re.split(r"[-\s:]", str(next_sys_date_time))
            fmt_next_sys_date_time = ''.join(next_sys_date_time_split)[:12]     # 去掉末尾的秒钟信息

    elif TEST_CASE_INFO[8] in weekly_event_mode:
        cur_weekday = date(cur_year, cur_month, cur_day).weekday()
        res_event_weekday = weekday_num_dict[TEST_CASE_INFO[8]]
        if cur_weekday == res_event_weekday:
            if GL.event_already_triggered_numb == 0:
                next_triggered_time = cur_date - timedelta(minutes=ahead_of_time)
                next_triggered_time_split = re.split(r"[-\s:]", str(next_triggered_time))
                fmt_next_sys_date_time = ''.join(next_triggered_time_split)[:12]  # 去掉末尾的秒钟信息
            else:
                next_triggered_date = cur_date + timedelta(days=7)
                next_triggered_date_time = next_triggered_date - timedelta(minutes=ahead_of_time)
                next_triggered_date_time_split = re.split(r"[-\s:]", str(next_triggered_date_time))
                fmt_next_sys_date_time = ''.join(next_triggered_date_time_split)[:12]     # 去掉末尾的秒钟信息
        elif cur_weekday != res_event_weekday:
            if cur_weekday > res_event_weekday:
                interval_day = 7 - cur_weekday + res_event_weekday
                next_triggered_date = cur_date + timedelta(days=interval_day)
                next_triggered_date_time = next_triggered_date - timedelta(minutes=ahead_of_time)
                next_triggered_date_time_split = re.split(r"[-\s:]", str(next_triggered_date_time))
                fmt_next_sys_date_time = ''.join(next_triggered_date_time_split)[:12]  # 去掉末尾的秒钟信息
            elif cur_weekday < res_event_weekday:
                interval_day = res_event_weekday - cur_weekday
                next_triggered_date = cur_date + timedelta(days=interval_day)
                next_triggered_date_time = next_triggered_date - timedelta(minutes=ahead_of_time)
                next_triggered_date_time_split = re.split(r"[-\s:]", str(next_triggered_date_time))
                fmt_next_sys_date_time = ''.join(next_triggered_date_time_split)[:12]  # 去掉末尾的秒钟信息
    logging.info(f"fmt_next_sys_date_time:{fmt_next_sys_date_time}")
    return fmt_next_sys_date_time


def set_system_time():
    logging.info("set_system_time")
    # 根据计算出的下次等待预约时间触发的系统时间来设置系统时间
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    sys_date_list = []      # 用于将系统日期由字符串转化为发送指令的列表
    sys_time_list = []      # 用于将系统时间由字符串转化为发送指令的列表
    sys_date_time = calc_modify_system_time()
    # if sys_date_time[8:] == "2359":
    #     GL.report_data[2] = change_str_time_and_fmt_time(sys_date_time, 1)[:8]
    # else:
    #     GL.report_data[2] = sys_date_time[:8]
    GL.report_data[2] = sys_date_time[:8]
    # 进入时间设置界面
    # 在get_cycle_event_start_time_and_sys_date已经执行
    # 移动到Date选项
    send_commd(KEY["DOWN"])
    sys_date_list.append(sys_date_time[:8])
    sys_date_cmd = change_numbs_to_commds_list(sys_date_list)
    for i in range(len(sys_date_cmd)):
        for j in sys_date_cmd[i]:
            send_commd(j)
    # 移动到Time选项
    send_commd(KEY["DOWN"])
    sys_time_list.append(sys_date_time[8:])
    sys_time_cmd = change_numbs_to_commds_list(sys_time_list)
    for i in range(len(sys_time_cmd)):
        for j in sys_time_cmd[i]:
            send_commd(j)
    # 退出保存
    send_commd(KEY["EXIT"])
    send_commd(KEY["OK"])
    # 退回大画面
    send_more_commds(exit_to_screen)


def goto_specified_interface_wait_for_event_triggered():
    logging.info("goto_specified_interface_wait_for_event_triggered")
    # WAIT_INTERFACE = ["TVScreenDiffCH", "RadioScreenDiffCH", "ChannelList", "Menu", "EPG", "ChannelEdit"]
    menu_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["DOWN"], KEY["DOWN"], KEY["OK"], KEY["OK"]]
    channel_edit_interface = [KEY["MENU"], KEY["LEFT"], KEY["LEFT"], KEY["OK"]]
    # 切到指定界面
    if TEST_CASE_INFO[5] == "TVScreenDiffCH" and TEST_CASE_INFO[9] != "Power On":
        if channel_info[5] != "TV":
            send_commd(KEY["TV/R"])
        # 切到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":  # 加锁节目判断
            send_commd(KEY["EXIT"])
    elif TEST_CASE_INFO[5] == "TVScreenDiffCH" and TEST_CASE_INFO[9] == "Power On":
        # 发送关机指令
        logging.info("当前为Power On事件，手动进入待机状态，等待唤醒")
        send_commd(KEY["POWER"])
    elif TEST_CASE_INFO[5] == "RadioScreenDiffCH":
        if channel_info[5] != "Radio":
            send_commd(KEY["TV/R"])
        # 切到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":  # 加锁节目判断
            send_commd(KEY["EXIT"])
    elif TEST_CASE_INFO[5] == "ChannelList":
        # 切到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":  # 加锁节目判断
            send_commd(KEY["EXIT"])
        # 调出频道列表
        send_commd(KEY["OK"])
    elif TEST_CASE_INFO[5] == "Menu":
        # 切到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":  # 加锁节目判断
            send_commd(KEY["EXIT"])
        # 进入Menu指定子菜单界面
        send_more_commds(menu_interface)
    elif TEST_CASE_INFO[5] == "EPG":
        # 切到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":  # 加锁节目判断
            send_commd(KEY["EXIT"])
        # 进入EPG界面
        send_commd(KEY["EPG"])
    elif TEST_CASE_INFO[5] == "ChannelEdit":
        # 切到不同的频道等待
        send_commd(KEY["UP"])
        if channel_info[3] == "1":  # 加锁节目判断
            send_commd(KEY["EXIT"])
        # 进入节目编辑界面
        send_more_commds(channel_edit_interface)
    # 等待节目信息赋值
    time.sleep(2)
    GL.report_data[4] = channel_info[1]
    # 等待事件响应
    if TEST_CASE_INFO[9] == "Power On":
        logging.info("当前为Power On事件，等待唤醒自动唤醒")
    else:
        logging.info("事件还没有触发，等待响应")
        while not state["res_event_triggered_state"]:
            time.sleep(0.1)
        else:
            logging.info("事件已经触发，正确跳出预约跳转选择框")
            logging.info(f"触发事件时，系统时间信息为：{rsv_kws['res_triggered_sys_time']}")
            GL.report_data[3] = rsv_kws['res_triggered_sys_time']
            logging.info(type(current_triggered_event_info))
            logging.info(type(GL.res_event_mgr))
            if list(current_triggered_event_info) in GL.res_event_mgr:
                logging.info("当前触发事件在事件列表中")
                state["res_event_triggered_state"] = False
            elif list(current_triggered_event_info) not in GL.res_event_mgr:
                logging.info(f"警告：当前触发事件不在事件列表中，{GL.res_event_mgr}-{current_triggered_event_info}")


def res_event_triggered_and_choice_jump_type():
    logging.info("res_event_triggered_and_choice_jump_type")
    unlock_cmd = [KEY["0"], KEY["0"], KEY["0"], KEY["0"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    weekly_event_mode = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
    # 事件触发后选择跳转方式
    if TEST_CASE_INFO[9] == "Play":
        if TEST_CASE_INFO[6] == "Manual_jump":
            time.sleep(5)
            logging.info("选择手动跳转")
            send_commd(KEY["OK"])
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

        elif TEST_CASE_INFO[6] == "Auto_jump":
            logging.info("选择自动跳转")
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

        elif TEST_CASE_INFO[6] == "Cancel_jump":
            time.sleep(5)
            logging.info("选择取消跳转")
            send_commd(KEY["LEFT"])
            send_commd(KEY["OK"])
            while not state["res_event_cancel_jump_state"]:
                logging.info("请注意：没有检测到取消事件标志")
                time.sleep(3)
            else:
                logging.info("事件跳转取消成功")
                time.sleep(3)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"正确取消跳转，当前节目与触发事件的节目不一致:{channel_info[1]}--{current_triggered_event_info[2]}")
                else:
                    logging.info(f"警告：没有取消跳转成功，当前节目与触发事件的节目为:{channel_info[1]}--{current_triggered_event_info[2]}")

    elif TEST_CASE_INFO[9] == "PVR":
        if TEST_CASE_INFO[6] == "Manual_jump":
            time.sleep(5)
            logging.info("选择手动跳转")
            send_commd(KEY["OK"])
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(1)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(1)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

            logging.info("还没有正确进入录制，请稍候...")
            while not state["rec_start_state"]:
                if state["no_storage_device_state"]:
                    logging.info("警告：没有插入存储设备")
                    logging.info("出现没有插入存储设备提示后，按退出键退出提示")
                    if TEST_CASE_INFO[5] in WAIT_INTERFACE[:2]:
                        send_commd(KEY["EXIT"])     # 等待界面为大画面时，只用发送一个退出键，退出提示框
                    else:   # 等待界面为其他界面时，除了要退出提示框，还要退回到大画面
                        send_more_commds(EXIT_TO_SCREEN)
                    GL.pvr_rec_dur_time = 0
                    break
                if state["no_enough_space_state"]:
                    logging.info("警告：存储设备没有足够的空间")
                    logging.info("出现存储设备没有足够的空间提示后，按退出键退出提示")
                    if TEST_CASE_INFO[5] in WAIT_INTERFACE[:2]:
                        send_commd(KEY["EXIT"])  # 等待界面为大画面时，只用发送一个退出键，退出提示框
                    else:  # 等待界面为其他界面时，除了要退出提示框，还要退回到大画面
                        send_more_commds(EXIT_TO_SCREEN)
                    GL.pvr_rec_dur_time = 0
                    break
                if state["pvr_not_supported_state"]:
                    logging.info("警告：当前录制节目为加密节目，或加锁节目，或无信号，请检查")
                    logging.info("出现pvr_not_supported提示后，按退出键退出提示")
                    if TEST_CASE_INFO[5] in WAIT_INTERFACE[:2]:
                        send_commd(KEY["EXIT"])  # 等待界面为大画面时，只用发送一个退出键，退出提示框
                    else:  # 等待界面为其他界面时，除了要退出提示框，还要退回到大画面
                        send_more_commds(EXIT_TO_SCREEN)
                    GL.pvr_rec_dur_time = 0
                    break
            else:
                logging.info("正确进入录制")
                rec_start_time = datetime.now()
                n = 0
                while not state["rec_end_state"]:
                    if n == 0:
                        logging.info("正在录制过程中，请等待录制结束")
                    n += 1
                else:
                    logging.info("录制结束")
                    rec_end_time = datetime.now()
                    GL.pvr_rec_dur_time = (rec_end_time - rec_start_time).seconds
                    logging.info(f"录制时长:{GL.pvr_rec_dur_time}")

        elif TEST_CASE_INFO[6] == "Auto_jump":
            logging.info("选择自动跳转")
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(1)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(1)
                if channel_info[3] == 1:
                    send_more_commds(unlock_cmd)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"没有正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")
                elif channel_info[1] == current_triggered_event_info[2]:
                    logging.info(f"正确跳转到触发事件的节目:{channel_info[1]}--{current_triggered_event_info[2]}")

            logging.info("还没有正确进入录制，请稍候...")
            while not state["rec_start_state"]:
                if state["no_storage_device_state"]:
                    logging.info("警告：没有插入存储设备")
                    logging.info("出现没有插入存储设备提示后，按退出键退出提示")
                    if TEST_CASE_INFO[5] in WAIT_INTERFACE[:2]:
                        send_commd(KEY["EXIT"])  # 等待界面为大画面时，只用发送一个退出键，退出提示框
                    else:  # 等待界面为其他界面时，除了要退出提示框，还要退回到大画面
                        send_more_commds(EXIT_TO_SCREEN)
                    GL.pvr_rec_dur_time = 0
                    break
                if state["no_enough_space_state"]:
                    logging.info("警告：存储设备没有足够的空间")
                    logging.info("出现存储设备没有足够的空间提示后，按退出键退出提示")
                    if TEST_CASE_INFO[5] in WAIT_INTERFACE[:2]:
                        send_commd(KEY["EXIT"])  # 等待界面为大画面时，只用发送一个退出键，退出提示框
                    else:  # 等待界面为其他界面时，除了要退出提示框，还要退回到大画面
                        send_more_commds(EXIT_TO_SCREEN)
                    GL.pvr_rec_dur_time = 0
                    break
                if state["pvr_not_supported_state"]:
                    logging.info("警告：当前录制节目为加密节目，或加锁节目，或无信号，请检查")
                    logging.info("出现pvr_not_supported提示后，按退出键退出提示")
                    if TEST_CASE_INFO[5] in WAIT_INTERFACE[:2]:
                        send_commd(KEY["EXIT"])  # 等待界面为大画面时，只用发送一个退出键，退出提示框
                    else:  # 等待界面为其他界面时，除了要退出提示框，还要退回到大画面
                        send_more_commds(EXIT_TO_SCREEN)
                    GL.pvr_rec_dur_time = 0
                    break
            else:
                logging.info("正确进入录制")
                rec_start_time = datetime.now()
                n = 0
                while not state["rec_end_state"]:
                    if n == 0:
                        logging.info("正在录制过程中，请等待录制结束")
                    n += 1
                else:
                    logging.info("录制结束")
                    rec_end_time = datetime.now()
                    GL.pvr_rec_dur_time = (rec_end_time - rec_start_time).seconds
                    logging.info(f"录制时长:{GL.pvr_rec_dur_time}")

        elif TEST_CASE_INFO[6] == "Cancel_jump":
            time.sleep(5)
            logging.info("选择取消跳转和录制")
            send_commd(KEY["LEFT"])
            send_commd(KEY["OK"])
            while not state["res_event_cancel_jump_state"]:
                logging.info("请注意：没有检测到取消事件标志")
                time.sleep(3)
            else:
                logging.info("事件跳转取消成功")
                time.sleep(3)

                if channel_info[1] != current_triggered_event_info[2]:
                    logging.info(f"正确取消跳转，当前节目与触发事件的节目不一致:{channel_info[1]}--{current_triggered_event_info[2]}")
                else:
                    logging.info(f"警告：没有取消跳转成功，当前节目与触发事件的节目为:{channel_info[1]}--{current_triggered_event_info[2]}")
            GL.pvr_rec_dur_time = 0     # 取消跳转时，录制持续时长为0

    elif TEST_CASE_INFO[9] == "Power Off":
        if TEST_CASE_INFO[6] == "Manual_jump":
            time.sleep(5)
            logging.info("选择手动跳转")
            send_commd(KEY["OK"])
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
            while not state["power_off_state"]:
                logging.info("还没有进入到关机状态")
                time.sleep(1)
            else:
                logging.info("进入关机状态")
                logging.info("等待5秒后，开始唤醒操作")
                time.sleep(5)
                logging.info("开始唤醒操作")
                send_commd(KEY["POWER"])
                state["control_power_on_info_rsv_state"] = True
                power_off_start_time = datetime.now()   # 用于关机计时起始时间
            while not state["stb_already_power_on_state"]:
                logging.info("还没有获取到启动成功标志，请等候")
                time.sleep(5)
                power_off_end_time = datetime.now()  # 用于关机计时结束时间
                if (power_off_end_time - power_off_start_time).seconds >= 30:
                    send_commd(KEY["UP"])
            else:
                logging.info("检测到启动成功标志")
                state["control_power_on_info_rsv_state"] = False

        elif TEST_CASE_INFO[6] == "Auto_jump":
            logging.info("选择自动跳转")
            while not state["res_event_confirm_jump_state"]:
                logging.info("请注意：没有检测到事件跳转")
                time.sleep(3)
            else:
                logging.info("确认事件跳转成功")
                time.sleep(3)
            while not state["power_off_state"]:
                logging.info("还没有进入到关机状态")
                time.sleep(1)
            else:
                logging.info("进入关机状态")
                logging.info("等待5秒后，开始唤醒操作")
                time.sleep(5)
                logging.info("开始唤醒操作")
                send_commd(KEY["POWER"])
                state["control_power_on_info_rsv_state"] = True
                power_off_start_time = datetime.now()   # 用于关机计时起始时间
            while not state["stb_already_power_on_state"]:
                logging.info("还没有获取到启动成功标志，请等候")
                time.sleep(5)
                power_off_end_time = datetime.now()  # 用于关机计时结束时间
                if (power_off_end_time - power_off_start_time).seconds >= 30:
                    send_commd(KEY["UP"])
            else:
                logging.info("检测到启动成功标志")
                state["control_power_on_info_rsv_state"] = False

        elif TEST_CASE_INFO[6] == "Cancel_jump":
            time.sleep(5)
            logging.info("选择取消跳转")
            send_commd(KEY["LEFT"])
            send_commd(KEY["OK"])
            while not state["res_event_cancel_jump_state"]:
                logging.info("请注意：没有检测到取消事件标志")
                time.sleep(3)
            else:
                logging.info("事件跳转取消成功")
                time.sleep(3)

    elif TEST_CASE_INFO[9] == "Power On":
        while not state["power_off_state"]:
            logging.info("还没有进入到关机状态")
            time.sleep(1)
        else:
            logging.info("进入关机状态")
            logging.info("等待唤醒操作")
            state["control_power_on_info_rsv_state"] = True
            power_off_start_time = datetime.now()  # 用于关机计时起始时间
        while not state["stb_already_power_on_state"]:
            logging.info("还没有获取到启动成功标志，请等候")
            time.sleep(1)
        else:
            logging.info("检测到启动成功标志")
            state["control_power_on_info_rsv_state"] = False
            power_off_end_time = datetime.now()  # 用于关机计时结束时间
            if GL.res_event_mgr[0][-1] == "Once":
                logging.info(f"移除Once类型当前触发事件前的列表：{GL.res_event_mgr}")
                GL.res_event_mgr.remove(GL.res_event_mgr[0])
                logging.info(f"移除Once类型当前触发事件后的列表：{GL.res_event_mgr}")
            elif GL.res_event_mgr[0][-1] == "Daily" or GL.res_event_mgr[0][-1] in weekly_event_mode:
                GL.event_already_triggered_numb += 1  # 预约事件触发后，次数加1
                logging.info(f"{GL.res_event_mgr[0]}事件不需要从数据库中删除")
            GL.report_data[3] = str(power_off_end_time - power_off_start_time)[:7]

    logging.info("预约事件数据处理，----------------------------------------------------")
    if TEST_CASE_INFO[9] != "Power On":
        if current_triggered_event_info[-1] == "Once":  # Once事件触发后，需要从数据库中移除
            logging.info(f"移除Once类型当前触发事件前的列表：{GL.res_event_mgr}")
            GL.res_event_mgr.remove(list(current_triggered_event_info))
            logging.info(f"移除Once类型当前触发事件后的列表：{GL.res_event_mgr}")

        elif current_triggered_event_info[-1] == "Daily":
            GL.event_already_triggered_numb += 1  # 预约事件触发后，次数加1
            logging.info(f"{current_triggered_event_info[-1]}事件不需要从数据库中删除")
        elif current_triggered_event_info[-1] in weekly_event_mode:
            GL.event_already_triggered_numb += 1    # 预约事件触发后，次数加1
            logging.info(f"{current_triggered_event_info[-1]}事件不需要从数据库中删除")

        if TEST_CASE_INFO[6] == "Cancel_jump":
            send_more_commds(exit_to_screen)
    else:
        logging.info("预约事件为Power On，不能从响应的事件来获取当前触发的事件信息")


def res_triggered_later_check_timer_setting_event_list():
    logging.info("res_triggered_later_check_timer_setting_event_list")
    # 预约事件触发后，事件列表事件检查
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    send_more_commds(enter_timer_setting_interface)
    if rsv_kws["res_event_numb"] == '0':
        if len(GL.res_event_mgr) == int(rsv_kws["res_event_numb"]):
            logging.info("数据库与事件列表中的事件个数匹配，都为空")
        else:
            logging.info(f"警告：数据库与事件列表中的事件数不匹配，{len(GL.res_event_mgr)}--{int(rsv_kws['res_event_numb'])}")
    elif rsv_kws["res_event_numb"] != '0':
        if len(GL.res_event_mgr) == int(rsv_kws["res_event_numb"]):
            logging.info("数据库与事件列表中的事件个数匹配，再检查事件信息是否匹配")
            if GL.res_event_mgr == list(res_event_list):
                logging.info(f"数据库与事件列表中的事件信息一致，{GL.res_event_mgr}--{list(res_event_list)}")
            else:
                logging.info(f"警告：数据库与事件列表中的事件信息不一致，{GL.res_event_mgr}--{list(res_event_list)}")
        else:
            logging.info("数据库与事件列表中的事件个数不匹配，请检查事件信息")
            logging.info(f"警告：数据库与事件列表中的事件个数不一致，{GL.res_event_mgr}--{list(res_event_list)}")
    send_more_commds(exit_to_screen)


def cale_str_time_for_add_day(str_time, interval_day):
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    if len(str_time) == 12:     # once事件时间计算
        fmt_year = int(str_time[:4])
        fmt_month = int(str_time[4:6])
        fmt_day = int(str_time[6:8])
        fmt_hour = int(str_time[8:10])
        fmt_minute = int(str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(days=interval_day)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
        logging.info(f"----------------------------------------------------------------------{str_new_fmt_date}")
    return str_new_fmt_date


def change_str_time_and_fmt_time(str_time, interval_time):
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    if len(str_time) == 12:     # once事件时间计算
        fmt_year = int(str_time[:4])
        fmt_month = int(str_time[4:6])
        fmt_day = int(str_time[6:8])
        fmt_hour = int(str_time[8:10])
        fmt_minute = int(str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(minutes=interval_time)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
    elif len(str_time) == 4:    # daily和weekly事件时间计算
        old_hour = int(str_time[:2])
        old_minute = int(str_time[2:])
        new_hour = 0
        new_minute = 0
        if old_minute + interval_time < 60:
            new_minute = old_minute + interval_time
            new_hour = old_hour
        elif old_minute + interval_time >= 60:
            new_minute = (old_minute + interval_time) - 60
            if old_hour + 1 < 24:
                new_hour = old_hour + 1
            elif old_hour + 1 >= 24:
                new_hour = (old_hour + 1) - 24
        str_new_fmt_date = "{0:02d}".format(new_hour) + "{0:02d}".format(new_minute)
    elif len(str_time) == 5:    # duration时间计算
        old_hour = int(str_time[:2])
        old_minute = int(str_time[3:])
        new_hour = 0
        new_minute = 0
        if old_minute + interval_time < 60:
            new_minute = old_minute + interval_time
            new_hour = old_hour
        elif old_minute + interval_time >= 60:
            new_minute = (old_minute + interval_time) - 60
            if old_hour + 1 < 24:
                new_hour = old_hour + 1
            elif old_hour + 1 >= 24:
                new_hour = (old_hour + 1) - 24
        str_new_fmt_date = "{0:02d}".format(new_hour) + ":" + "{0:02d}".format(new_minute)
    return str_new_fmt_date


def write_data_to_excel():
    logging.info("write_data_to_excel")
    wb = ''
    ws = ''
    excel_title_1 = ["用例编号", "新增的预约事件信息", "编辑修改后的预约事件信息", "触发响应结果"]
    excel_title_2 = ["用例编号", "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
                     "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
                     "系统时间日期", "触发时间/关机等待时间", "等待节目", "跳转节目", "录制时长", "用例测试时间"]
    # if TEST_CASE_INFO[9] == "Power On":
    #     excel_title_2 = ["用例编号", "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
    #                      "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
    #                      "系统时间日期", "关机等待时间", "等待节目", "跳转节目", "录制时长", "用例测试时间"]
    # else:
    #     excel_title_2 = ["用例编号", "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
    #                      "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
    #                      "系统时间日期", "触发时间", "等待节目", "跳转节目", "录制时长", "用例测试时间"]

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    blue_font = Font(color=BLUE)
    red_font = Font(color=RED)
    a_column_numb = column_index_from_string("A")
    if not os.path.exists(file_path[1]):
        wb = Workbook()
        ws = wb.active
        ws.title = file_path[2]
        # 写excel_title_1的内容
        ws.cell(1, 1).value = excel_title_1[0]
        ws["A" + str(1)].alignment = alignment

        ws.cell(1, 2).value = excel_title_1[1]
        ws["B" + str(1)].alignment = alignment
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

        ws.cell(1, 7).value = excel_title_1[2]
        ws["G" + str(1)].alignment = alignment
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)

        ws.cell(1, 12).value = excel_title_1[3]
        ws["L" + str(1)].alignment = alignment
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=17)

        # 写excel_title_2的内容
        for j in range(len(excel_title_2)):
            ws.cell(2, j + 1).value = excel_title_2[j]
            ws.cell(2, j + 1).alignment = alignment
            if j == 0:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
            else:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 9
        # 设置Title的行高
        ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
        ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
        # 合并用例编号单元格
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    elif os.path.exists(file_path[1]):
        wb = load_workbook(file_path[1])
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if file_path[2] in sheets_name_list:
            ws = wb[file_path[2]]
        elif file_path[2] not in sheets_name_list:
            ws = wb.create_sheet(file_path[2])
            # 写excel_title_1的内容
            ws.cell(1, 1).value = excel_title_1[0]
            ws["A" + str(1)].alignment = alignment

            ws.cell(1, 2).value = excel_title_1[1]
            ws["B" + str(1)].alignment = alignment
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

            ws.cell(1, 7).value = excel_title_1[2]
            ws["G" + str(1)].alignment = alignment
            ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)

            ws.cell(1, 12).value = excel_title_1[3]
            ws["L" + str(1)].alignment = alignment
            ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16)

            # 写excel_title_2的内容
            for j in range(len(excel_title_2)):
                ws.cell(2, j + 1).value = excel_title_2[j]
                ws.cell(2, j + 1).alignment = alignment
                if j == 0:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 9
            # 设置Title的行高
            ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
            ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
            # 合并用例编号单元格
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # 获取当前用例修改类型的sheet表的Max_row
    max_row = ws.max_row

    # 写新增和编辑后的预约事件数据
    len_res_1 = len(GL.report_data[0])          # 新增预约事件的长度
    len_res_2 = len(GL.report_data[1])          # 编辑修改后的预约事件的长度
    len_total = len_res_1 + len_res_2           # 新增和编辑修改后的预约事件的总长度
    for d in range(len(GL.report_data)):
        if d == 0:      # 新增预约事件的信息
            for add_data in range(len(GL.report_data[d])):
                ws.cell(max_row + 1, add_data + 2).value = GL.report_data[d][add_data]
                ws.cell(max_row + 1, add_data + 2).alignment = alignment
                # ws.column_dimensions[get_column_letter(a_column_numb + add_data + 1)].width = 15

        elif d == 1:    # 修改后的预约事件的信息
            for edit_data in range(len(GL.report_data[d])):
                ws.cell(max_row + 1, len_res_1 + edit_data + 2).value = GL.report_data[d][edit_data]
                ws.cell(max_row + 1, len_res_1 + edit_data + 2).alignment = alignment
                # ws.column_dimensions[get_column_letter(a_column_numb + len_res_1 + edit_data + 1)].width = 15
                if edit_data == 0:     # start time
                    # 包含ModifyTime，不包含ModifyMode，且时间需要+5的
                    if TEST_CASE_INFO[7] == "ModifyTime" or TEST_CASE_INFO[7] == "ModifyTime+ModifyType" or \
                            TEST_CASE_INFO[7] == "ModifyTime+ModifyDuration" or \
                            TEST_CASE_INFO[7] == "ModifyTime+ModifyType+ModifyDuration":
                        contrast_time = change_str_time_and_fmt_time(GL.report_data[0][0], 5)  # 新增预约事件起始时间加5分钟
                        if GL.report_data[d][edit_data] == contrast_time:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        elif GL.report_data[d][edit_data] != contrast_time:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    # 包含ModifyTime，包含ModifyMode，且时间需要+5的
                    elif TEST_CASE_INFO[7] == "ModifyTime+ModifyMode" or \
                            TEST_CASE_INFO[7] == "ModifyTime+ModifyType+ModifyMode" or \
                            TEST_CASE_INFO[7] == "ModifyTime+ModifyType+ModifyDuration+ModifyMode":
                        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
                            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                                str_time = GL.report_data[0][0][8:]
                                contrast_time = change_str_time_and_fmt_time(str_time, 5)  # 新增预约事件起始时间加5分钟
                                if GL.report_data[d][edit_data] == contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                                elif GL.report_data[d][edit_data] != contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                        elif TEST_CASE_INFO[3] == "Daily" or TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
                            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                                str_time = GL.report_data[0][0]
                                contrast_time = change_str_time_and_fmt_time(str_time, 5)  # 新增预约事件起始时间加5分钟
                                if GL.report_data[d][edit_data] == contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                                elif GL.report_data[d][edit_data] != contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                                str_time = GL.report_data[0][0]
                                contrast_time = change_str_time_and_fmt_time(str_time, 5)  # 新增预约事件起始时间加5分钟
                                if GL.report_data[d][edit_data] == GL.report_data[1][0][:8] + contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                                elif GL.report_data[d][edit_data] != GL.report_data[1][0][:8] + contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    # 不包含ModifyTime，但是包含ModifyMode，时间位数有变化的，但是不需要+5分钟
                    elif TEST_CASE_INFO[7] == "ModifyMode" or TEST_CASE_INFO[7] == "ModifyType+ModifyMode" or \
                            TEST_CASE_INFO[7] == "ModifyDuration+ModifyMode" or \
                            TEST_CASE_INFO[7] == "ModifyType+ModifyDuration+ModifyMode":
                        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
                            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                                contrast_time = GL.report_data[0][0][8:]    # 对比时间，用于验证是否正确
                                if GL.report_data[d][edit_data] == contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                                elif GL.report_data[d][edit_data] != contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                        elif TEST_CASE_INFO[3] == "Daily" or TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
                            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                                contrast_time = GL.report_data[0][0]    # 对比时间，用于验证是否正确
                                if GL.report_data[d][edit_data] == contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                                elif GL.report_data[d][edit_data] != contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                                # 起始时间为["2355"]时，新增+5分钟后，会出现跨日期的现象，GL.report_data[2]需要随之变化，否则出错
                                if GL.report_data[0][0] == "0000":
                                    contrast_time = cale_str_time_for_add_day(
                                        (GL.report_data[2] + GL.report_data[0][0]), 1)
                                else:
                                    contrast_time = GL.report_data[2] + GL.report_data[0][0]    # 对比时间，用于验证是否正确
                                if GL.report_data[d][edit_data] == contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                                elif GL.report_data[d][edit_data] != contrast_time:
                                    ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    # 既不包含ModifyTime，也不包含ModifyMode，时间不变
                    else:
                        if GL.report_data[d][edit_data] == GL.report_data[0][0]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        elif GL.report_data[d][edit_data] != GL.report_data[0][0]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font

                elif edit_data == 1:   # event type
                    if TEST_CASE_INFO[7] == "ModifyType" or "ModifyType" in TEST_CASE_INFO[7]:
                        if GL.report_data[d][edit_data] != GL.report_data[0][1] \
                                and GL.report_data[d][edit_data] == TEST_CASE_INFO[9]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    else:
                        if GL.report_data[d][edit_data] == GL.report_data[0][1] \
                                and GL.report_data[d][edit_data] == TEST_CASE_INFO[9]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font

                elif edit_data == 2:   # ch
                    if (TEST_CASE_INFO[4] != "Power Off" and TEST_CASE_INFO[4] != "Power On") \
                            and (TEST_CASE_INFO[9] != "Power Off" and TEST_CASE_INFO[9] != "Power On"):
                        if GL.report_data[d][edit_data] == GL.report_data[0][2]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        elif GL.report_data[d][edit_data] != GL.report_data[0][2]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    elif (TEST_CASE_INFO[4] == "Power Off" or TEST_CASE_INFO[4] == "Power On") \
                            and (TEST_CASE_INFO[9] != "Power Off" and TEST_CASE_INFO[9] != "Power On"):
                        if GL.report_data[d][edit_data] != GL.report_data[0][2]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    elif (TEST_CASE_INFO[4] != "Power Off" and TEST_CASE_INFO[4] != "Power On") \
                            and (TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On"):
                        if GL.report_data[d][edit_data] != GL.report_data[0][2] \
                                and GL.report_data[d][edit_data] == "----":
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    elif (TEST_CASE_INFO[4] == "Power Off" or TEST_CASE_INFO[4] == "Power On") \
                            and (TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On"):
                        if GL.report_data[d][edit_data] == GL.report_data[0][2] == "----":
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font

                elif edit_data == 3:   # duration
                    if TEST_CASE_INFO[7] == "ModifyDuration" or TEST_CASE_INFO[7] == "ModifyTime+ModifyDuration" or \
                            TEST_CASE_INFO[7] == "ModifyDuration+ModifyMode":
                        contrast_dur_time = change_str_time_and_fmt_time(GL.report_data[0][3], 1)   # PVR事件dur时间加1
                        if GL.report_data[d][edit_data] == contrast_dur_time \
                                and TEST_CASE_INFO[4] == TEST_CASE_INFO[9] == "PVR":
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        elif GL.report_data[d][edit_data] != contrast_dur_time:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    elif "ModifyType" in TEST_CASE_INFO[7] and "ModifyDuration" in TEST_CASE_INFO[7]:
                        if TEST_CASE_INFO[4] != "PVR" and TEST_CASE_INFO[9] == "PVR" and \
                                GL.report_data[0][3] == "--:--" and GL.report_data[d][edit_data] == "00:02":
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    elif "ModifyType" in TEST_CASE_INFO[7] and "ModifyDuration" not in TEST_CASE_INFO[7]:
                        if TEST_CASE_INFO[9] == "PVR":
                            if GL.report_data[d][edit_data] == "00:01":
                                ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                            else:
                                ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                        elif TEST_CASE_INFO[9] != "PVR":
                            if GL.report_data[d][edit_data] == "--:--":
                                ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                            else:
                                ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    else:
                        if GL.report_data[d][edit_data] == GL.report_data[0][3]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font

                elif edit_data == 4:   # event mode
                    if TEST_CASE_INFO[7] == "ModifyMode" or "ModifyMode" in TEST_CASE_INFO[7]:
                        if GL.report_data[d][edit_data] != GL.report_data[0][4] \
                                and GL.report_data[d][edit_data] == TEST_CASE_INFO[8]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font
                    else:
                        if GL.report_data[d][edit_data] == GL.report_data[0][4] \
                                and GL.report_data[d][edit_data] == TEST_CASE_INFO[8]:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = blue_font
                        else:
                            ws.cell(max_row + 1, len_res_1 + edit_data + 2).font = red_font

        elif d == 3:  # 触发时间
            ws.cell(max_row + 1, d + len_total).value = GL.report_data[d]
            ws.cell(max_row + 1, d + len_total).alignment = alignment
            if TEST_CASE_INFO[9] != "Power On":
                if TEST_CASE_INFO[8] == "Once":     # （触发时间+1）后与预约事件起始时间进行比对
                    str_triggered_time = change_str_time_and_fmt_time(GL.report_data[3][:12], 1)  # 加1分钟后的触发时间
                    if str_triggered_time == GL.report_data[1][0]:
                        ws.cell(max_row + 1, d + len_total).font = blue_font
                    elif str_triggered_time != GL.report_data[1][0]:
                        ws.cell(max_row + 1, d + len_total).font = red_font
                elif TEST_CASE_INFO[8] == "Daily":
                    if GL.report_data[3][:12][8:] == "2359":  # （触发时间的时分+1）与（系统时间日期+预约事件起始时间）进行比对
                        str_triggered_time = GL.report_data[3][:8] + \
                                             change_str_time_and_fmt_time(GL.report_data[3][8:12], 1)
                        if str_triggered_time == GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = blue_font
                        elif str_triggered_time != GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = red_font
                    elif GL.report_data[3][:12][8:] != "2359":     # （触发时间+1）与（系统时间日期+预约事件起始时间）进行比对
                        str_triggered_time = change_str_time_and_fmt_time(GL.report_data[3][:12], 1)  # 加1分钟后的触发时间
                        if str_triggered_time == GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = blue_font
                        elif str_triggered_time != GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = red_font
                elif TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:
                    if GL.report_data[3][:12][8:] == "2359":  # （触发时间的时分+1）与（系统时间日期+预约事件起始时间）进行比对
                        str_triggered_time = GL.report_data[3][:8] + \
                                             change_str_time_and_fmt_time(GL.report_data[3][8:12], 1)
                        if str_triggered_time == GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = blue_font
                        elif str_triggered_time != GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = red_font
                    elif GL.report_data[3][:12][8:] != "2359":  # （触发时间+1）与（系统时间日期+预约事件起始时间）进行比对
                        str_triggered_time = change_str_time_and_fmt_time(GL.report_data[3][:12], 1)  # 加1分钟后的触发时间
                        if str_triggered_time == GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = blue_font
                        elif str_triggered_time != GL.report_data[2] + GL.report_data[1][0]:
                            ws.cell(max_row + 1, d + len_total).font = red_font
            elif TEST_CASE_INFO[9] == "Power On":
                standby_dur_time = GL.report_data[3]
                standby_dur_time_split = re.split(r":", standby_dur_time)
                standby_dur_hour = int(standby_dur_time_split[0])
                standby_dur_minute = int(standby_dur_time_split[1])
                standby_dur_second = int(standby_dur_time_split[2])
                standby_dur_time_second = standby_dur_hour * 3600 + standby_dur_minute * 60 + standby_dur_second * 1
                stb_boot_time = 15
                real_standby_dur_time = standby_dur_time_second - stb_boot_time
                if (60 - 10) <= real_standby_dur_time <= (60 + 10):
                    ws.cell(max_row + 1, d + len_total).font = blue_font
                else:
                    ws.cell(max_row + 1, d + len_total).font = red_font

        elif d == 5:  # 跳转节目
            ws.cell(max_row + 1, d + len_total).value = GL.report_data[d]
            ws.cell(max_row + 1, d + len_total).alignment = alignment
            if TEST_CASE_INFO[9] == "Play" or TEST_CASE_INFO[9] == "PVR":
                if TEST_CASE_INFO[6] == "Manual_jump" or TEST_CASE_INFO[6] == "Auto_jump":
                    if GL.report_data[d] == GL.report_data[1][2]:
                        ws.cell(max_row + 1, d + len_total).font = blue_font
                    elif GL.report_data[d] != GL.report_data[1][2]:
                        ws.cell(max_row + 1, d + len_total).font = red_font
                elif TEST_CASE_INFO[6] == "Cancel_jump":
                    if GL.report_data[d] != GL.report_data[1][2]:
                        ws.cell(max_row + 1, d + len_total).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + len_total).font = red_font
            elif TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On":
                if GL.report_data[d] == "----":
                    ws.cell(max_row + 1, d + len_total).font = blue_font
                else:
                    ws.cell(max_row + 1, d + len_total).font = red_font

        elif d == 6:  # 录制时长
            ws.cell(max_row + 1, d + len_total).value = GL.report_data[d]
            ws.cell(max_row + 1, d + len_total).alignment = alignment
            if TEST_CASE_INFO[9] == "PVR":
                if TEST_CASE_INFO[6] == "Manual_jump" or TEST_CASE_INFO[6] == "Auto_jump":
                    res_dur_split = re.split(":", GL.report_data[1][3])
                    # 换算录制时常信息与预约时间的Duration时长信息对比值
                    res_dur_sec_time = int(res_dur_split[0]) * 3600 + int(res_dur_split[1]) * 60
                    actual_rec_time = int(GL.report_data[d][:-1])
                    if (res_dur_sec_time - 5) <= actual_rec_time <= (res_dur_sec_time + 5):
                        ws.cell(max_row + 1, d + len_total).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + len_total).font = red_font
                        error_comment = Comment(f'{rsv_kws["pvr_not_work_info"]}', "wangrun")
                        ws.cell(max_row + 1, d + len_total).comment = error_comment
                elif TEST_CASE_INFO[6] == "Cancel_jump":
                    if GL.report_data[d] == "0s":
                        ws.cell(max_row + 1, d + len_total).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + len_total).font = red_font
            elif TEST_CASE_INFO[9] == "Play" or TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On":
                if GL.report_data[d] == "--:--":
                    ws.cell(max_row + 1, d + len_total).font = blue_font
                else:
                    ws.cell(max_row + 1, d + len_total).font = red_font

        elif d == 7:    # 用例编号
            ws.cell(max_row + 1, 1).value = GL.report_data[d]
            ws.cell(max_row + 1, 1).alignment = alignment

        elif d == 8:    # 写报告时间
            ws.cell(max_row + 1, d + len_total - 1).value = GL.report_data[d]   # 由于d==7的坑填到第一列，所以这里需要列数减一
            ws.cell(max_row + 1, d + len_total - 1).alignment = alignment

        else:
            ws.cell(max_row + 1, d + len_total).value = GL.report_data[d]
            ws.cell(max_row + 1, d + len_total).alignment = alignment
    ws.row_dimensions[(max_row + 1)].height = 27    # 设置每次执行的report预约事件信息的行高

    wb.save(file_path[1])


def manage_report_data_and_write_data():
    logging.info("manage_report_data_and_write_data")
    # 整理数据以及写数据
    if TEST_CASE_INFO[9] == "Play":
        # GL.report_data[2] = "pass"                # 系统时间日期
        # GL.report_data[3] = list(res_event_list)[0][0]      # 事件响应时间（跳出跳转提示框的时间）
        # GL.report_data[4] = TEST_CASE_INFO[6]   # 等待节目
        GL.report_data[5] = channel_info[1]     # 跳转节目
        GL.report_data[6] = "--:--"              # 录制时长
        GL.report_data[7] = TEST_CASE_INFO[0]   # 用例编号
        GL.report_data[8] = str(datetime.now())[:19]    # 写该用例报告的时间
    elif TEST_CASE_INFO[9] == "PVR":
        # GL.report_data[3] = list(res_event_list)[0][0]
        # GL.report_data[4] = TEST_CASE_INFO[6]
        GL.report_data[5] = channel_info[1]
        GL.report_data[6] = str(GL.pvr_rec_dur_time) + 's'
        GL.report_data[7] = TEST_CASE_INFO[0]  # 用例编号
        GL.report_data[8] = str(datetime.now())[:19]  # 写该用例报告的时间
    elif TEST_CASE_INFO[9] == "Power Off":
        # GL.report_data[3] = list(res_event_list)[0][0]
        GL.report_data[4] = "----"
        GL.report_data[5] = "----"
        GL.report_data[6] = "--:--"
        GL.report_data[7] = TEST_CASE_INFO[0]  # 用例编号
        GL.report_data[8] = str(datetime.now())[:19]  # 写该用例报告的时间
    elif TEST_CASE_INFO[9] == "Power On":
        # GL.report_data[3] = list(res_event_list)[0][0]
        GL.report_data[4] = "----"
        GL.report_data[5] = "----"
        GL.report_data[6] = "--:--"
        GL.report_data[7] = TEST_CASE_INFO[0]  # 用例编号
        GL.report_data[8] = str(datetime.now())[:19]  # 写该用例报告的时间

    logging.info(GL.title_data)
    logging.info(GL.report_data)
    time.sleep(2)


def before_cycle_test_clear_data_and_state():
    # 循环测试前，清理数据和状态变量
    logging.info("before_cycle_test_clear_data_and_state")
    # GL.res_event_mgr.clear()
    GL.choice_res_ch = ''
    state["clear_variate_state"] = True
    GL.pvr_rec_dur_time = ''

    GL.res_triggered_numb -= 1
    logging.info("循环测试，延时5秒")
    time.sleep(5)
    logging.info(f"剩余循环次数：{GL.res_triggered_numb}")

    if GL.res_triggered_numb < 1:
        logging.info("程序结束")
        state["receive_loop_state"] = True  # 触发结束接收进程的状态


def calculate_expected_edit_event_start_time():
    # 对新增的事件进行计算，修改后的预期起始时间
    logging.info("calculate_expected_edit_event_start_time")
    time_interval = 5
    str_expected_res_time = ''
    start_time = GL.res_event_mgr[0][0]       # 原新增预约事件的起始时间
    # 包含ModifyTime，但是不包含ModifyMode，且时间需要+5的
    if TEST_CASE_INFO[7] == "ModifyTime" or TEST_CASE_INFO[7] == "ModifyTime+ModifyType" or \
                            TEST_CASE_INFO[7] == "ModifyTime+ModifyDuration" or \
                            TEST_CASE_INFO[7] == "ModifyTime+ModifyType+ModifyDuration":
        str_expected_res_time = change_str_time_and_fmt_time(start_time, time_interval)
    # 包含ModifyTime，且包含ModifyMode，且时间需要+5的
    elif TEST_CASE_INFO[7] == "ModifyTime+ModifyMode" or \
            TEST_CASE_INFO[7] == "ModifyTime+ModifyType+ModifyMode" or \
            TEST_CASE_INFO[7] == "ModifyTime+ModifyType+ModifyDuration+ModifyMode":
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_res_time = change_str_time_and_fmt_time(start_time[8:], time_interval)
        elif TEST_CASE_INFO[3] == "Daily" or TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_res_time = change_str_time_and_fmt_time(start_time, time_interval)
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                str_expected_res_time = change_str_time_and_fmt_time(sys_time_date + start_time, time_interval)
    # 不包含ModifyTime，但涉及到ModifyMode，时间位数有变化的，但是不需要+5
    elif TEST_CASE_INFO[7] == "ModifyMode" or TEST_CASE_INFO[7] == "ModifyType+ModifyMode" or \
            TEST_CASE_INFO[7] == "ModifyDuration+ModifyMode" or \
            TEST_CASE_INFO[7] == "ModifyType+ModifyDuration+ModifyMode":
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_res_time = start_time[8:]
        elif TEST_CASE_INFO[3] == "Daily" or TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_res_time = start_time
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                # 起始时间为["2355", "2356", "2357", "2358", "2359"]时，新增+5分钟后，会出现跨日期的现象，导致程序出错
                if start_time in ["0000", "0001", "0002", "0003", "0004"]:
                    str_expected_res_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)
                else:
                    str_expected_res_time = sys_time_date + start_time
    # 既不包含ModifyTime，也不包含ModifyMode
    else:
        logging.info("当前编辑不涉及修改时间和修改Mode，所以预约事件时间不变")
        str_expected_res_time = start_time

    # GL.report_data[2] = str_expected_res_time[:8]       # 用于Once类型事件的report系统时间日期

    logging.info(f"期望的完整的预约事件时间为{str_expected_res_time}")
    return str_expected_res_time


def calculate_expected_edit_event_duration_time():
    logging.info("calculate_expected_edit_event_duration_time")
    str_expected_dur_time = ''
    interval_dur = 1        # 更改录制时长的变量
    specified_dur_time = "0002"   # ModifyType+ModifyDuration时会出现无Dur_time改有Dur_time的情况，直接指定时间

    if TEST_CASE_INFO[7] == "ModifyDuration" or "ModifyDuration" in TEST_CASE_INFO[7]:
        if TEST_CASE_INFO[4] == "PVR" and TEST_CASE_INFO[9] == "PVR":   # 单项ModifyDuration
            dur_time = GL.res_event_mgr[0][3]  # 原新增预约事件的持续时间
            res_dur_split = re.split(":", dur_time)
            res_dur_hour_info = int(res_dur_split[0])
            res_dur_minute_info = int(res_dur_split[1])
            new_hour = 0
            new_minute = 0
            if res_dur_minute_info + interval_dur < 60:
                new_minute = res_dur_minute_info + interval_dur
                new_hour = res_dur_hour_info
            elif res_dur_minute_info + interval_dur >= 60:
                new_minute = (res_dur_minute_info + interval_dur) - 60
                if res_dur_hour_info + 1 < 24:
                    new_hour = res_dur_hour_info + 1
                elif res_dur_hour_info + 1 >= 24:  # 等于24的情况可能会出现问题，但是目前的用例应该不会遇到
                    new_hour = (res_dur_hour_info + 1) - 24
            new_dur_time = "{0:02d}".format(new_hour) + "{0:02d}".format(new_minute)
            str_expected_dur_time = new_dur_time
        elif TEST_CASE_INFO[4] != "PVR" and TEST_CASE_INFO[9] == "PVR":   # 多项修改且包含ModifyDuration
            dur_time = GL.res_event_mgr[0][3]  # 原新增预约事件的持续时间
            if dur_time == "--:--":
                str_expected_dur_time = specified_dur_time
        else:
            logging.info(f"请注意，当前事件不是PVR事件，而是{TEST_CASE_INFO[4]}事件")

    else:
        logging.info("当前事件不需要更改Duration，保持默认值")
        str_expected_dur_time = "0001"
    return str_expected_dur_time


def create_expected_edit_event_info():
    logging.info("create_expected_edit_event_info")
    # 创建修改后的期望的事件信息
    expected_edit_event_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    duration_time = calculate_expected_edit_event_duration_time()
    if TEST_CASE_INFO[9] == "Play":
        expected_event_full_time = calculate_expected_edit_event_start_time()
        expected_edit_event_info[0] = expected_event_full_time
        expected_edit_event_info[1] = TEST_CASE_INFO[9]
        expected_edit_event_info[2] = channel_info[1]
        expected_edit_event_info[3] = "--:--"
        expected_edit_event_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "PVR":
        expected_event_full_time = calculate_expected_edit_event_start_time()
        expected_edit_event_info[0] = expected_event_full_time
        expected_edit_event_info[1] = TEST_CASE_INFO[9]
        expected_edit_event_info[2] = channel_info[1]
        expected_edit_event_info[3] = duration_time
        expected_edit_event_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "Power Off":
        expected_event_full_time = calculate_expected_edit_event_start_time()
        expected_edit_event_info[0] = expected_event_full_time
        expected_edit_event_info[1] = TEST_CASE_INFO[9]
        expected_edit_event_info[2] = "----"
        expected_edit_event_info[3] = "--:--"
        expected_edit_event_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "Power On":
        expected_event_full_time = calculate_expected_edit_event_start_time()
        expected_edit_event_info[0] = expected_event_full_time
        expected_edit_event_info[1] = TEST_CASE_INFO[9]
        expected_edit_event_info[2] = "----"
        expected_edit_event_info[3] = "--:--"
        expected_edit_event_info[4] = TEST_CASE_INFO[8]
    return expected_edit_event_info


def modify_edit_add_new_res_event_info():
    logging.info("modify_edit_add_new_res_event_info")
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面
    send_commd(KEY["YELLOW"])
    # 生成预期的预约事件
    expected_res_event_info = create_expected_edit_event_info()
    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[9]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[9]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[9]}')
            send_commd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[8]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[8]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[8]}')
            send_commd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[8] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[8]}")
    elif TEST_CASE_INFO[8] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[8]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_commd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_commds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_commd(KEY["DOWN"])
    else:
        if len(expected_res_event_info[0]) == 12:
            start_time_list.append(expected_res_event_info[0][8:])
        elif len(expected_res_event_info[0]) == 4:
            start_time_list.append(expected_res_event_info[0])
        start_time_cmd = change_numbs_to_commds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_commd(j)
        send_commd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[9] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[9]}")
    elif TEST_CASE_INFO[9] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[9]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_commd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_commds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[9]}，不需要设置Channel")
    elif TEST_CASE_INFO[9] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[9]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_commd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["update_event_list_state"] = True
    state["clear_res_event_list_state"] = True
    send_commd(KEY["EXIT"])
    send_commd(KEY["OK"])
    # 此处编辑完事件后，不会再次打印事件信息，需要重新进入Timer Setting界面，为update_edit_res_event_to_event_mgr_list准备数据
    send_commd(KEY["EXIT"])
    send_commd(KEY["OK"])
    # 退回大画面
    send_more_commds(exit_to_screen)


def update_edit_res_event_to_event_mgr_list():
    logging.info("update_edit_res_event_to_event_mgr_list")
    # 先清除之前的新增事件
    GL.res_event_mgr.clear()
    # 添加编辑修改后的预约事件到事件管理列表
    if list(res_event_list) not in GL.res_event_mgr:
        GL.res_event_mgr.extend(list(res_event_list))
    GL.report_data[1] = GL.res_event_mgr[0]
    logging.info(type(GL.res_event_mgr))
    logging.info(GL.res_event_mgr)
    logging.info(list(res_event_list))
    state["update_event_list_state"] = False


def modify_edit_res_event():
    logging.info("edit_res_event")
    # 编辑修改新增的预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_commds(enter_timer_setting_interface)
    # 进入事件编辑界面，设置预约事件参数
    modify_edit_add_new_res_event_info()
    # 添加新预约事件到事件管理列表
    update_edit_res_event_to_event_mgr_list()


def mail(message):
    my_sender = 'wangrun@nationalchip.com'  # 发件人邮箱账号
    my_pass = 'b8iNRgDiPUfkUVLW'  # 发件人邮箱密码
    my_user = 'wangrun@nationalchip.com'  # 收件人邮箱账号，我这边发送给自己

    return_state = True
    try:
        msg = MIMEText(message, 'plain', 'utf-8')
        # msg['From'] = formataddr(["FromRunoob", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['From'] = Header("Auto_test", 'utf-8')
        msg['To'] = Header("ME", 'utf-8')
        # msg['To'] = formataddr(["FK", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
        msg['Subject'] = "自动化测试终止提醒"  # 邮件的主题，也可以说是标题

        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 关闭连接
    except smtplib.SMTPException:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
        return_state = False
    return return_state


def receive_serial_process(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info,
        receive_cmd_list):
    logging_info_setting()
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIME_SHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict([val, key] for key, val in rsv_key.items())

    res_kws = [
        "[PTD]Time_mode=",              # 0     获取系统时间模式
        "[PTD]System_time=",            # 1     系统时间
        "[PTD]Res_event_numb=",         # 2     预约事件数量
        "[PTD]Res_event:",              # 3     预约事件信息
        "[PTD]Res_triggered:",          # 4     预约事件触发和当前响应事件的信息
        "[PTD]Res_confirm_jump",        # 5     预约事件确认跳转
        "[PTD]Res_cancel_jump",         # 6     预约事件取消跳转
        "[PTD]REC_start",               # 7     录制开始
        "[PTD]REC_end",                 # 8     录制结束
        "[PTD]No_storage_device",       # 9     没有存储设备
        "[PTD]No_enough_space",         # 10    没有足够的空间
        "[PTD]power_cut",               # 11    进入待机
        "[PTD]:switch totle cost",      # 12    开机解码成功
        "[PTD][HOTPLUG] PLUG_IN",       # 13    存储设备插入成功
        "[PTD]PVR_is_not_supported",    # 14    录制无信号、加锁节目、加密节目，跳出PVR is not supported!提示
        "[PTD]Current_sys_time",        # 15    预约事件触发时，检测触发时间信息
    ]

    switch_ch_kws = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    ch_info_kws = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name"]

    group_info_kws = [
        "Group_name",
        "Prog_total"
    ]

    edit_event_kws = [
        "[PTD]Mode=",
        "[PTD]Type=",
        "[PTD]Start Date=",
        "[PTD]Start Time=",
        "[PTD]Duration=",
        "[PTD]Channel="
    ]

    other_kws = [
        "[PTD]Infrared_key_values:",    # 获取红外接收关键字
    ]

    infrared_rsv_cmd = []       # 红外接受命令
    receive_serial = serial.Serial(prs_data["receive_serial_name"], 115200, timeout=1)

    while True:
        data = receive_serial.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            data1 = data.decode("ISO-8859-1", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_log_data_to_txt(prs_data["log_file_path"], data3)

            if state["clear_variate_state"]:
                state["sys_time_mode_state"] = False
                state["current_sys_time_state"] = False
                state["res_event_numb_state"] = False
                state["res_event_triggered_state"] = False
                state["res_event_confirm_jump_state"] = False
                state["res_event_cancel_jump_state"] = False
                state["rec_start_state"] = False
                state["rec_end_state"] = False
                state["no_storage_device_state"] = False
                state["no_enough_space_state"] = False
                state["pvr_not_supported_state"] = False
                state["update_event_list_state"] = False
                state["clear_variate_state"] = False
                state["power_off_state"] = False
                state["stb_already_power_on_state"] = False

                # del res_event_list[:]
                del current_triggered_event_info[:]
                if prs_data["case_res_event_mode"] == "Once":
                    del res_event_list[:]
                # channel_info = ['', '', '', '', '', '', '']

            if state["clear_res_event_list_state"]:
                del res_event_list[:]
                state["clear_res_event_list_state"] = False

            if other_kws[0] in data2:   # 红外接收打印
                rsv_cmd = re.split(":", data2)[-1]
                if reverse_rsv_key[rsv_cmd] != "POWER":
                    infrared_rsv_cmd.append(rsv_cmd)        # 存放可以共享的接受命令的列表
                if rsv_cmd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_cmd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        infrared_send_cmd[-1], reverse_rsv_key[infrared_rsv_cmd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(infrared_send_cmd), len(infrared_rsv_cmd)))
                    if reverse_rsv_key[rsv_cmd] != "POWER":
                        receive_cmd_list.append(rsv_cmd)
                if state["control_power_on_info_rsv_state"]:    # 用于Power Off事件触发后，软开机没有检测到开机关键字用来避免死循环
                    state["stb_already_power_on_state"] = True

            if res_kws[0] in data2:     # 获取系统时间模式（自动还是手动）
                state["sys_time_mode_state"] = True
                rsv_kws["sys_time_mode"] = re.split(r"=", data2)[-1]

            if res_kws[1] in data2:     # 获取当前系统时间
                state["current_sys_time_state"] = True
                rsv_kws["current_sys_time"] = re.split(r"=", data2)[-1]

            if res_kws[2] in data2:     # 获取预约事件数量
                state["res_event_numb_state"] = True
                rsv_kws["res_event_numb"] = re.split(r"=", data2)[-1]

            if res_kws[3] in data2:     # 获取预约事件信息
                # state["res_event_info_state"] = True
                event_split_info = re.split(r"event:|,", data2)
                event_info = ['', '', '', '', '']
                for info in event_split_info:
                    if "Start_time" in info:
                        event_start_time = re.split(r"=", info)[-1]
                        if len(event_start_time) == 5:
                            event_info[0] = ''.join(re.split(r":", event_start_time))
                        elif len(event_start_time) == 16:
                            event_info[0] = ''.join(re.split(r"[/:\s]", event_start_time))
                    if "Event_type" in info:
                        event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        event_info[4] = re.split(r"=", info)[-1]
                if state["update_event_list_state"]:
                    res_event_list.append(event_info)

            if res_kws[4] in data2:     # 获取预约事件跳转触发信息，以及当前响应事件的信息
                state["res_event_triggered_state"] = True
                current_event_split_info = re.split(r"triggered:|,", data2)
                current_event_info = ['', '', '', '', '']
                for info in current_event_split_info:
                    if "Start_time" in info:
                        current_event_start_time = re.split(r"=", info)[-1]
                        if len(current_event_start_time) == 5:
                            current_event_info[0] = ''.join(re.split(r":", current_event_start_time))
                        elif len(current_event_start_time) == 16:
                            current_event_info[0] = ''.join(re.split(r"[/:\s]", current_event_start_time))
                    if "Event_type" in info:
                        current_event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        current_event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        current_event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        current_event_info[4] = re.split(r"=", info)[-1]
                current_triggered_event_info.extend(current_event_info)

            if res_kws[5] in data2:     # 获取预约事件确认跳转信息
                state["res_event_confirm_jump_state"] = True

            if res_kws[6] in data2:     # 获取预约事件取消跳转信息
                state["res_event_cancel_jump_state"] = True

            if res_kws[7] in data2:     # 获取PVR预约事件录制开始信息
                state["rec_start_state"] = True

            if res_kws[8] in data2:     # 获取PVR预约事件录制结束信息
                state["rec_end_state"] = True

            if res_kws[9] in data2:     # 存储设备没有插入的打印信息
                state["no_storage_device_state"] = True
                rsv_kws["pvr_not_work_info"] = data2

            if res_kws[10] in data2:    # 存储设备没有足够空间的打印信息
                state["no_enough_space_state"] = True
                rsv_kws["pvr_not_work_info"] = data2

            if res_kws[11] in data2:    # 软关机打印信息
                state["power_off_state"] = True

            if res_kws[12] in data2 or res_kws[13] in data2:    # 开机解码成功打印信息,或开机存储设备挂载成功信息
                if state["control_power_on_info_rsv_state"]:
                    state["stb_already_power_on_state"] = True

            if res_kws[14] in data2:    # 录制无信号、加锁节目、加密节目，跳出PVR is not supported!提示
                state["pvr_not_supported_state"] = True
                rsv_kws["pvr_not_work_info"] = data2

            if res_kws[15] in data2:    # 预约事件触发时系统时间信息(备注：此系统时间为年月日 时分秒信息)
                cur_sys_time = re.split(r"=", data2)[-1]
                cur_sys_time_split = re.split(r"[/\s:]", cur_sys_time)
                rsv_kws["res_triggered_sys_time"] = ''.join(cur_sys_time_split)

            if switch_ch_kws[0] in data2:
                ch_info_split = re.split(r"[],]", data2)
                for i in range(len(ch_info_split)):
                    if ch_info_kws[0] in ch_info_split[i]:  # 提取频道号
                        channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if ch_info_kws[1] in ch_info_split[i]:  # 提取频道名称
                        channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if switch_ch_kws[1] in data2:
                flag_info_split = re.split(r"[],]", data2)
                for i in range(len(flag_info_split)):
                    if ch_info_kws[2] in flag_info_split[i]:  # 提取频道所属TP
                        channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if ch_info_kws[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if switch_ch_kws[3] in data2:
                group_info_split = re.split(r"[],]", data2)
                for i in range(len(group_info_split)):
                    if group_info_kws[0] in group_info_split[i]:  # 提取频道所属组别
                        rsv_kws["prog_group_name"] = re.split(r"=", group_info_split[i])[-1]
                        channel_info[6] = rsv_kws["prog_group_name"]
                    if group_info_kws[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        rsv_kws["prog_group_total"] = re.split(r"=", group_info_split[i])[-1]

            if edit_event_kws[0] in data2:          # 提取Mode参数
                rsv_kws["edit_event_focus_pos"] = "Mode"
                rsv_kws["edit_event_mode"] = re.split(r"=", data2)[-1]

            if edit_event_kws[1] in data2:          # 提取Type参数
                rsv_kws["edit_event_focus_pos"] = "Type"
                rsv_kws["edit_event_type"] = re.split(r"=", data2)[-1]

            if edit_event_kws[2] in data2:          # 提取Start Date参数
                rsv_kws["edit_event_focus_pos"] = "Start Date"
                rsv_kws["edit_event_date"] = re.split(r"=", data2)[-1]

            if edit_event_kws[3] in data2:          # 提取Start Time参数
                rsv_kws["edit_event_focus_pos"] = "Start Time"
                rsv_kws["edit_event_time"] = re.split(r"=", data2)[-1]

            if edit_event_kws[4] in data2:          # 提取Duration参数
                rsv_kws["edit_event_focus_pos"] = "Duration"
                rsv_kws["edit_event_duration"] = re.split(r"=", data2)[-1]

            if edit_event_kws[5] in data2:          # 提取Channel参数
                rsv_kws["edit_event_focus_pos"] = "Channel"
                rsv_kws["edit_event_ch"] = re.split(r"=", data2)[-1]


if __name__ == "__main__":
    logging_info_setting()
    choice_case_numb = int(sys.argv[1])
    # choice_case_numb = 57
    test_case_info = edit_res_case[choice_case_numb]
    print(test_case_info)
    TEST_CASE_INFO = ''
    weekly_mode = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]

    if test_case_info[7] == "ModifyTime" \
            or test_case_info[7] == "ModifyType" \
            or test_case_info[7] == "ModifyDuration" \
            or test_case_info[7] == "ModifyTime+ModifyType" \
            or test_case_info[7] == "ModifyTime+ModifyDuration" \
            or test_case_info[7] == "ModifyType+ModifyDuration" \
            or test_case_info[7] == "ModifyTime+ModifyType+ModifyDuration":
        if test_case_info[3] == "Weekly" and test_case_info[8] == "Weekly":
            new_test_case_info = test_case_info.copy()
            logging.info(f"选择之前的new_test_case_info：{new_test_case_info}")
            new_test_case_info[3] = new_test_case_info[8] = choice(weekly_mode)
            logging.info(f"选择之后的new_test_case_info：{new_test_case_info}")
            TEST_CASE_INFO = new_test_case_info
        else:
            TEST_CASE_INFO = test_case_info
    elif test_case_info[7] == "ModifyMode" \
            or test_case_info[7] == "ModifyTime+ModifyMode" \
            or test_case_info[7] == "ModifyType+ModifyMode" \
            or test_case_info[7] == "ModifyDuration+ModifyMode" \
            or test_case_info[7] == "ModifyTime+ModifyType+ModifyMode" \
            or test_case_info[7] == "ModifyType+ModifyDuration+ModifyMode" \
            or test_case_info[7] == "ModifyTime+ModifyType+ModifyDuration+ModifyMode":
        if test_case_info[3] == "Weekly" and test_case_info[8] != "Weekly":
            new_test_case_info = test_case_info.copy()
            logging.info(f"选择之前的new_test_case_info：{new_test_case_info}")
            new_test_case_info[3] = choice(weekly_mode)
            logging.info(f"选择之后的new_test_case_info：{new_test_case_info}")
            TEST_CASE_INFO = new_test_case_info

        elif test_case_info[3] != "Weekly" and test_case_info[8] == "Weekly":
            new_test_case_info = test_case_info.copy()
            logging.info(f"选择之前的new_test_case_info：{new_test_case_info}")
            new_test_case_info[8] = choice(weekly_mode)
            logging.info(f"选择之后的new_test_case_info：{new_test_case_info}")
            TEST_CASE_INFO = new_test_case_info

        elif test_case_info[3] == "Weekly" and test_case_info[8] == "Weekly":
            new_test_case_info = test_case_info.copy()
            logging.info(f"选择之前的new_test_case_info：{new_test_case_info}")
            logging.info(f"new_test_case_info[3]选择Weekly之前的Weekly_mode:{weekly_mode}")
            new_test_case_info[3] = choice(weekly_mode)
            weekly_mode.remove(new_test_case_info[3])
            logging.info(f"new_test_case_info[3]选择Weekly之后的Weekly_mode:{weekly_mode}")
            new_test_case_info[8] = choice(weekly_mode)
            logging.info(f"选择之后的new_test_case_info：{new_test_case_info}")
            TEST_CASE_INFO = new_test_case_info
        else:
            TEST_CASE_INFO = test_case_info

    GL = MyGlobal()

    msg = "现在开始执行的是:{}_{}_{}_{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2], TEST_CASE_INFO[4],
        TEST_CASE_INFO[3], TEST_CASE_INFO[7], TEST_CASE_INFO[9], TEST_CASE_INFO[8])
    logging.critical(format(msg, '*^150'))
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIME_SHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D"
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())
    WAIT_INTERFACE = ["TVScreenDiffCH", "RadioScreenDiffCH", "ChannelList", "Menu", "EPG", "ChannelEdit"]
    WEEKLY_EVENT_MODE = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
    # TEST_CASE_INFO = ["23", "All", "TV", "Daily", "Play", "EPG", "Manual_jump"]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    try:
        file_path = build_log_and_report_file_path()
        ser_name = list(check_ports())  # send_ser_name, receive_ser_name
        send_serial = serial.Serial(ser_name[0], 9600)
        receive_ser_name = ser_name[1]

        infrared_send_cmd = Manager().list([])
        receive_cmd_list = Manager().list([])
        res_event_list = Manager().list([])
        current_triggered_event_info = Manager().list([])
        channel_info = Manager().list(['', '', '', '', '', '', ''])     # [频道号,频道名称,tp,lock,scramble,频道类型,组别]
        rsv_kws = Manager().dict({
            "sys_time_mode": '', "current_sys_time": '', "res_event_numb": '', "prog_group_name": '',
            "prog_group_total": '', "edit_event_focus_pos": '', "edit_event_mode": '', "edit_event_type": '',
            "edit_event_date": '', "edit_event_time": '', "edit_event_duration": '', "edit_event_ch": '',
            "res_triggered_sys_time": '', "pvr_not_work_info": '',
        })

        state = Manager().dict({
            "res_event_numb_state": False, "res_event_triggered_state": False, "res_event_confirm_jump_state": False,
            "res_event_cancel_jump_state": False, "rec_start_state": False, "rec_end_state": False,
            "no_storage_device_state": False, "no_enough_space_state": False, "power_off_state": False,
            "sys_time_mode_state": False, "current_sys_time_state": False, "update_event_list_state": False,
            "clear_variate_state": False, "receive_loop_state": False, "control_power_on_info_rsv_state": False,
            "stb_already_power_on_state": False, "res_event_info_state": False, "pvr_not_supported_state": False,
            "clear_res_event_list_state": False
        })

        prs_data = Manager().dict({
            "log_file_path": file_path[0], "receive_serial_name": receive_ser_name, "case_res_event_mode": TEST_CASE_INFO[8]
        })

        rsv_p = Process(target=receive_serial_process, args=(
            prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info,
            receive_cmd_list))
        rsv_p.start()

        if platform.system() == "Windows":
            time.sleep(5)
            logging.info("Windows系统接收端响应慢，等待5秒")
        elif platform.system() == "Linux":
            time.sleep(1)
            logging.info("Linux系统接收端响应快，但是增加一个延时保护，等待1秒")

        # 主程序开始部分
        clear_timer_setting_all_events()
        while GL.res_triggered_numb > 0:
            if TEST_CASE_INFO[8] == "Once":
                check_sys_time_mode()
                choice_ch_for_res_event_type()
                new_add_res_event()
                modify_edit_res_event()
                set_system_time()
                goto_specified_interface_wait_for_event_triggered()
                res_event_triggered_and_choice_jump_type()
                manage_report_data_and_write_data()
                write_data_to_excel()
                res_triggered_later_check_timer_setting_event_list()
                before_cycle_test_clear_data_and_state()
            elif TEST_CASE_INFO[8] == "Daily":
                while GL.event_already_triggered_numb < 1 and GL.res_triggered_numb > 0:
                    check_sys_time_mode()
                    choice_ch_for_res_event_type()
                    new_add_res_event()
                    modify_edit_res_event()
                    set_system_time()
                    goto_specified_interface_wait_for_event_triggered()
                    res_event_triggered_and_choice_jump_type()
                    manage_report_data_and_write_data()
                    write_data_to_excel()
                    res_triggered_later_check_timer_setting_event_list()
                    before_cycle_test_clear_data_and_state()
                    break
                while GL.event_already_triggered_numb >= 1 and GL.res_triggered_numb >= 1:
                    set_system_time()
                    goto_specified_interface_wait_for_event_triggered()
                    res_event_triggered_and_choice_jump_type()
                    manage_report_data_and_write_data()
                    write_data_to_excel()
                    res_triggered_later_check_timer_setting_event_list()
                    before_cycle_test_clear_data_and_state()
                    break
            elif TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:
                while GL.event_already_triggered_numb < 1 and GL.res_triggered_numb > 0:
                    check_sys_time_mode()
                    choice_ch_for_res_event_type()
                    new_add_res_event()
                    modify_edit_res_event()
                    set_system_time()
                    goto_specified_interface_wait_for_event_triggered()
                    res_event_triggered_and_choice_jump_type()
                    manage_report_data_and_write_data()
                    write_data_to_excel()
                    res_triggered_later_check_timer_setting_event_list()
                    before_cycle_test_clear_data_and_state()
                    break
                while GL.event_already_triggered_numb >= 1 and GL.res_triggered_numb >= 1:
                    set_system_time()
                    goto_specified_interface_wait_for_event_triggered()
                    res_event_triggered_and_choice_jump_type()
                    manage_report_data_and_write_data()
                    write_data_to_excel()
                    res_triggered_later_check_timer_setting_event_list()
                    before_cycle_test_clear_data_and_state()
                    break
        if state["receive_loop_state"]:
            rsv_p.terminate()
            logging.info('stop receive process')
            rsv_p.join()

    except Exception as e:
        print(e)
        # cur_py_file_name = sys.argv[0]        # 第0个就是这个python文件本身的路径（全路径）
        cur_py_file_name = os.path.basename(__file__)       # 当前文件名名称
        ret = mail(f"{cur_py_file_name}\n\n"
                   f"{msg}\n\n"
                   f"{traceback.format_exc()}")
        if ret:
            print("邮件发送成功")
        else:
            print("邮件发送失败")

        print("***traceback.format_exc():*** ")
        print(traceback.format_exc())
