#!/usr/bin/python3
# -*- coding: utf-8 -*-

from serial_setting1 import *
from multiprocessing import Process, Manager
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import RED, BLUE
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, date, timedelta
from random import randint, choice
from email.mime.text import MIMEText
from email.header import Header
import smtplib
import platform
import os
import time
import logging
import re
import sys
import traceback


class MyGlobal(object):

    def __init__(self):
        if TEST_CASE_INFO[-1] == "epg_test_numb":
            self.case_testing_times = 1

        self.TV_channel_groups = {}                     # 存放电视节目的组别和节目数信息
        self.Radio_channel_groups = {}                  # 存放广播节目的组别和节目数信息
        self.TV_ch_attribute = [[], [], [], []]         # 用于存放TV节目属性的列表(免费\加密\加锁\所有有EPG的节目)
        self.Radio_ch_attribute = [[], [], [], []]      # 用于存放Radio节目属性的列表(免费\加密\加锁\所有有EPG的节目)
        self.all_ch_epg_info = {}                       # 所有有EPG信息的节目的事件管理
        self.choice_timezone = ''                       # 根据不同的Case选择不同的时区
        self.choice_res_ch = ''                         # 预约Play或PVR事件时所选预约节目

        if TEST_CASE_INFO[6] == "EPG":                  # EPG界面预约的report_data
            self.report_data = ['', '', '', '', [], '', '', '', '']
        elif TEST_CASE_INFO[6] == "Timer":              # Timer界面预约的report_data
            self.report_data = ['', '', '', '', [], '', '', '', '', '']


def logging_info_setting():
    # 配置logging输出格式
    # LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    # DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    # logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    date_format = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=log_format, datefmt=date_format)


def hex_strs_to_bytes(strings):
    # 将红外命令字符串转换为字节串
    return bytes.fromhex(strings)


def write_log_data_to_txt(path, write_data):
    with open(path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def send_commd(commd):
    global receive_cmd_list, infrared_send_cmd
    continuous_transmission_cmd_num = 0
    # 红外发送端发送指令
    send_serial.write(hex_strs_to_bytes(commd))
    send_serial.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
    if REVERSE_KEY[commd] != "POWER":
        infrared_send_cmd.append(REVERSE_KEY[commd])
    time.sleep(1.0)
    if len(infrared_send_cmd) == len(receive_cmd_list):
        pass
    elif len(infrared_send_cmd) != len(receive_cmd_list):
        logging.info("检测到发送和接收命令数不一致，等待2秒，查看是否接收端还没有接收到打印")
        time.sleep(2)
        while True:
            if len(infrared_send_cmd) == len(receive_cmd_list):
                break
            elif len(infrared_send_cmd) != len(receive_cmd_list):
                logging.info(f"此刻补发STB没有接收到的红外命令{infrared_send_cmd[-1]}")
                send_serial.write(hex_strs_to_bytes(KEY[infrared_send_cmd[-1]]))
                send_serial.flush()
                time.sleep(1.0)
                continuous_transmission_cmd_num += 1
                if continuous_transmission_cmd_num == 10:
                    stb_crash_msg = "STB一直发送指令，疑似死机"
                    # mail(f'{stb_crash_msg}\n\n{msg}')
                    raise FailSendCmdException(stb_crash_msg)


def send_more_commds(commd_list):
    # 用于发送一连串的指令
    for commd in commd_list:
        send_commd(commd)
    time.sleep(1)   # 增加函数切换时的的等待，避免可能出现send_commd函数中的等待时间没有执行的情况


def build_log_and_report_file_path():
    # 用于创建打印和报告文件路径
    # 构建存放数据的总目录，以及构建存放打印和报告的目录
    parent_path = os.path.dirname(os.getcwd())
    case_name = "Invalid_res_event"
    test_data_directory_name = "test_data"
    test_data_directory_path = os.path.join(parent_path, test_data_directory_name)
    log_directory_name = "print_log"
    log_directory_path = os.path.join(test_data_directory_path, log_directory_name)
    report_directory_name = "report"
    report_directory_path = os.path.join(test_data_directory_path, report_directory_name)

    log_case_directory_path = os.path.join(test_data_directory_path, log_directory_name, case_name)
    report_case_directory_path = os.path.join(test_data_directory_path, report_directory_name, case_name)
    # 判断目录是否存在，否则创建目录
    if not os.path.exists(test_data_directory_path):
        os.mkdir(test_data_directory_path)
    if not os.path.exists(log_directory_path):
        os.mkdir(log_directory_path)
    if not os.path.exists(report_directory_path):
        os.mkdir(report_directory_path)
    if not os.path.exists(log_case_directory_path):
        os.mkdir(log_case_directory_path)
    if not os.path.exists(report_case_directory_path):
        os.mkdir(report_case_directory_path)
    # 创建打印和报告文件的名称和路径
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    fmt_name = ''
    if TEST_CASE_INFO[6] == "EPG":
        fmt_name = "{}_{}_{}_{}_{}_{}_{}".format(
            TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
            TEST_CASE_INFO[3], TEST_CASE_INFO[4], TEST_CASE_INFO[5], TEST_CASE_INFO[6])
    elif TEST_CASE_INFO[6] == "Timer":
        fmt_name = "{}_{}_{}_{}_{}_{}_{}_{}_{}_{}".format(
            TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
            TEST_CASE_INFO[3], TEST_CASE_INFO[4], TEST_CASE_INFO[5],
            TEST_CASE_INFO[6], TEST_CASE_INFO[7], TEST_CASE_INFO[8], TEST_CASE_INFO[9])
    log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    log_file_path = os.path.join(log_case_directory_path, log_file_name)
    report_file_name = "Invalid_res_event_result_report.xlsx"
    report_file_path = os.path.join(report_case_directory_path, report_file_name)
    sheet_name = "{}_{}".format(TEST_CASE_INFO[6], TEST_CASE_INFO[5])
    return log_file_path, report_file_path, sheet_name


def change_numbs_to_commds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_commds_list = []
    for i in range(len(numbs_list)):
        channel_commds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_commds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_commds_list[i].append(KEY[numbs_list[i][j]])
    return channel_commds_list


def exit_to_screen():
    send_data = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    send_more_commds(send_data)


def clear_timer_setting_all_events():
    logging.info("clear_timer_setting_all_events")
    # 清除Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    delete_all_res_events = [KEY["BLUE"], KEY["OK"]]
    # 进入定时器设置界面
    send_more_commds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        if rsv_kws["res_event_numb"] != '0':
            send_more_commds(delete_all_res_events)
        elif rsv_kws["res_event_numb"] == '0':
            logging.info("没有预约事件存在")
            time.sleep(1)
        else:
            logging.debug("警告：预约事件个数获取错误！！！")
        state["res_event_numb_state"] = False
    # 退回大画面
    exit_to_screen()


def check_sys_time_auto_mode():
    logging.debug("check_sys_time_auto_mode")
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    # 进入时间设置界面
    send_more_commds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        logging.info(rsv_kws["sys_time_mode"])
        if rsv_kws["sys_time_mode"] == "Manual":
            send_more_commds(change_sys_time_mode)
        elif rsv_kws["sys_time_mode"] == "Auto":
            logging.info("系统时间模式已经为自动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    # 退回大画面
    exit_to_screen()


def check_sys_time_manual_mode():
    logging.debug("check_sys_time_auto_mode")
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    # 进入时间设置界面
    send_more_commds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        if rsv_kws["sys_time_mode"] == "Auto":
            send_more_commds(change_sys_time_mode)
        elif rsv_kws["sys_time_mode"] == "Manual":
            logging.info("系统时间模式已经为手动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    logging.info(rsv_kws["current_sys_time"])
    # 退回大画面
    exit_to_screen()


def get_sys_time_info():
    logging.info("get_sys_time_info")
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    send_more_commds(enter_time_setting_interface)
    logging.info(rsv_kws["current_sys_time"])
    exit_to_screen()


def set_timezone_and_summertime():
    logging.info("set_timezone_and_summertime")
    state_save_prompt_box_jump = False
    other_timezone = [
        '0.5', '1', '1.5', '2', '2.5', '3', '3.5', '4', '4.5', '5', '5.5', '6', '6.5', '7', '7.5', '8', '8.5', '9',
        '9.5', '10', '10.5', '11', '11.5', '12',
        '-0.5', '-1', '-1.5', '-2', '-2.5', '-3', '-3.5', '-4', '-4.5', '-5', '-5.5', '-6', '-6.5', '-7', '-7.5', '-8',
        '-8.5', '-9', '-9.5', '-10', '-10.5', '-11', '-11.5', '-12'
    ]

    timezone = [
        '-12', '-11.5', '-11', '-10.5', '-10', '-9.5', '-9', '-8.5', '-8', '-7.5', '-7', '-6.5', '-6', '-5.5', '-5',
        '-4.5', '-4', '-3.5', '-3', '-2.5', '-2', '-1.5', '-1', '-0.5', '0',
        '0.5', '1', '1.5', '2', '2.5', '3', '3.5', '4', '4.5', '5', '5.5', '6', '6.5', '7', '7.5', '8', '8.5', '9',
        '9.5', '10', '10.5', '11', '11.5', '12'
    ]
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    send_more_commds(enter_time_setting_interface)
    # 检查是否进入到Time setting界面
    while rsv_kws["sys_time_setting_focus_pos"] == "":
        time.sleep(2)  # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Time Mode")
    while rsv_kws["sys_time_setting_focus_pos"] != "Mode":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["sys_time_mode"] != "Auto":
            logging.info(f'Mode参数与预期不符:{rsv_kws["sys_time_mode"]}--Auto')
            state_save_prompt_box_jump = True
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["sys_time_mode"]}--Auto')
            send_commd(KEY["DOWN"])
    # 设置Timezone参数
    logging.info("Timezone")
    while rsv_kws["sys_time_setting_focus_pos"] != "Timezone":
        send_commd(KEY["DOWN"])
    else:
        if TEST_CASE_INFO[6] == "Timer":
            if TEST_CASE_INFO[8] == "ZeroTimezone":
                GL.choice_timezone = "0"
                while rsv_kws["sys_time_timezone"] != GL.choice_timezone:
                    logging.info(f'Timezone参数与预期不符:{rsv_kws["sys_time_timezone"]}--{GL.choice_timezone}')
                    logging.info(f'当前时区为：{rsv_kws["sys_time_timezone"]}，预期时区为：{GL.choice_timezone}')
                    state_save_prompt_box_jump = True
                    cur_tz_pos = timezone.index(rsv_kws["sys_time_timezone"])
                    expected_tz_pos = timezone.index(GL.choice_timezone)
                    logging.info(f"当前时区的位置为：{cur_tz_pos}，预期时区的位置为：{expected_tz_pos}")
                    if cur_tz_pos > expected_tz_pos:
                        left_move_steps = cur_tz_pos - expected_tz_pos
                        right_move_steps = expected_tz_pos + (len(timezone) - cur_tz_pos)
                        logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
                        if left_move_steps > right_move_steps:
                            logging.info("应该向右移动")
                            send_commd(KEY["RIGHT"])
                        elif left_move_steps < right_move_steps:
                            logging.info("应该向左移动")
                            send_commd(KEY["LEFT"])
                        elif left_move_steps == right_move_steps:
                            logging.info("向左或向右移动距离相等")
                            send_commd(KEY["RIGHT"])
                    elif cur_tz_pos < expected_tz_pos:
                        left_move_steps = cur_tz_pos + (len(timezone) - expected_tz_pos)
                        right_move_steps = expected_tz_pos - cur_tz_pos
                        logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
                        if left_move_steps > right_move_steps:
                            logging.info("应该向右移动")
                            send_commd(KEY["RIGHT"])
                        elif left_move_steps < right_move_steps:
                            logging.info("应该向左移动")
                            send_commd(KEY["LEFT"])
                        elif left_move_steps == right_move_steps:
                            logging.info("向左或向右移动距离相等")
                            send_commd(KEY["RIGHT"])
                else:
                    logging.info(f'Timezone参数与预期相符:{rsv_kws["sys_time_timezone"]}--{GL.choice_timezone}')
                    send_commd(KEY["DOWN"])
            elif TEST_CASE_INFO[8] == "OtherTimezone":
                GL.choice_timezone = choice(other_timezone)
                logging.info(f"所选系统时区为：{GL.choice_timezone}")
                while rsv_kws["sys_time_timezone"] != GL.choice_timezone:
                    logging.info(f'Timezone参数与预期不符:{rsv_kws["sys_time_timezone"]}--{GL.choice_timezone}')
                    logging.info(f'当前时区为：{rsv_kws["sys_time_timezone"]}，预期时区为：{GL.choice_timezone}')
                    state_save_prompt_box_jump =True
                    cur_tz_pos = timezone.index(rsv_kws["sys_time_timezone"])
                    expected_tz_pos = timezone.index(GL.choice_timezone)
                    logging.info(f"当前时区的位置为：{cur_tz_pos}，预期时区的位置为：{expected_tz_pos}")
                    if cur_tz_pos > expected_tz_pos:
                        left_move_steps = cur_tz_pos - expected_tz_pos
                        right_move_steps = expected_tz_pos + (len(timezone) - cur_tz_pos)
                        logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
                        if left_move_steps > right_move_steps:
                            logging.info("应该向右移动")
                            send_commd(KEY["RIGHT"])
                        elif left_move_steps < right_move_steps:
                            logging.info("应该向左移动")
                            send_commd(KEY["LEFT"])
                        elif left_move_steps == right_move_steps:
                            logging.info("向左或向右移动距离相等")
                            send_commd(KEY["RIGHT"])
                    elif cur_tz_pos < expected_tz_pos:
                        left_move_steps = cur_tz_pos + (len(timezone) - expected_tz_pos)
                        right_move_steps = expected_tz_pos - cur_tz_pos
                        logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
                        if left_move_steps > right_move_steps:
                            logging.info("应该向右移动")
                            send_commd(KEY["RIGHT"])
                        elif left_move_steps < right_move_steps:
                            logging.info("应该向左移动")
                            send_commd(KEY["LEFT"])
                        elif left_move_steps == right_move_steps:
                            logging.info("向左或向右移动距离相等")
                            send_commd(KEY["RIGHT"])
                else:
                    logging.info(f'Timezone参数与预期相符:{rsv_kws["sys_time_timezone"]}--{GL.choice_timezone}')
                    send_commd(KEY["DOWN"])

        elif TEST_CASE_INFO[6] == "EPG":
            GL.choice_timezone = "0"
            while rsv_kws["sys_time_timezone"] != GL.choice_timezone:
                logging.info(f'Timezone参数与预期不符:{rsv_kws["sys_time_timezone"]}--{GL.choice_timezone}')
                logging.info(f'当前时区为：{rsv_kws["sys_time_timezone"]}，预期时区为：{GL.choice_timezone}')
                state_save_prompt_box_jump = True
                cur_tz_pos = timezone.index(rsv_kws["sys_time_timezone"])
                expected_tz_pos = timezone.index(GL.choice_timezone)
                logging.info(f"当前时区的位置为：{cur_tz_pos}，预期时区的位置为：{expected_tz_pos}")
                if cur_tz_pos > expected_tz_pos:
                    left_move_steps = cur_tz_pos - expected_tz_pos
                    right_move_steps = expected_tz_pos + (len(timezone) - cur_tz_pos)
                    logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
                    if left_move_steps > right_move_steps:
                        logging.info("应该向右移动")
                        send_commd(KEY["RIGHT"])
                    elif left_move_steps < right_move_steps:
                        logging.info("应该向左移动")
                        send_commd(KEY["LEFT"])
                    elif left_move_steps == right_move_steps:
                        logging.info("向左或向右移动距离相等")
                        send_commd(KEY["RIGHT"])
                elif cur_tz_pos < expected_tz_pos:
                    left_move_steps = cur_tz_pos + (len(timezone) - expected_tz_pos)
                    right_move_steps = expected_tz_pos - cur_tz_pos
                    logging.info(f"向左移动的步数：{left_move_steps}，向右移动的步数：{right_move_steps}")
                    if left_move_steps > right_move_steps:
                        logging.info("应该向右移动")
                        send_commd(KEY["RIGHT"])
                    elif left_move_steps < right_move_steps:
                        logging.info("应该向左移动")
                        send_commd(KEY["LEFT"])
                    elif left_move_steps == right_move_steps:
                        logging.info("向左或向右移动距离相等")
                        send_commd(KEY["RIGHT"])
            else:
                logging.info(f'Timezone参数与预期相符:{rsv_kws["sys_time_timezone"]}--{GL.choice_timezone}')
                send_commd(KEY["DOWN"])
    # 设置Summertime参数
    logging.info("Timezone")
    while rsv_kws["sys_time_setting_focus_pos"] != "Summertime":
        send_commd(KEY["DOWN"])
    else:
        if TEST_CASE_INFO[6] == "Timer":
            if TEST_CASE_INFO[9] == "NoSummertime":
                while rsv_kws["sys_time_summertime"] != "Off":
                    logging.info(f'Summertime参数与预期不符:{rsv_kws["sys_time_summertime"]}--Off')
                    state_save_prompt_box_jump = True
                    send_commd(KEY["RIGHT"])
                else:
                    logging.info(f'Summertime参数与预期相符:{rsv_kws["sys_time_summertime"]}--Off')
            elif TEST_CASE_INFO[9] == "Summertime":
                while rsv_kws["sys_time_summertime"] != "On":
                    logging.info(f'Summertime参数与预期不符:{rsv_kws["sys_time_summertime"]}--On')
                    state_save_prompt_box_jump = True
                    send_commd(KEY["RIGHT"])
                else:
                    logging.info(f'Summertime参数与预期相符:{rsv_kws["sys_time_summertime"]}--On')

        elif TEST_CASE_INFO[6] == "EPG":
            while rsv_kws["sys_time_summertime"] != "Off":
                logging.info(f'Summertime参数与预期不符:{rsv_kws["sys_time_summertime"]}--Off')
                state_save_prompt_box_jump = True
                send_commd(KEY["RIGHT"])
            else:
                logging.info(f'Summertime参数与预期相符:{rsv_kws["sys_time_summertime"]}--Off')

    # 退出保存
    if state_save_prompt_box_jump:  # 假如Mode、Timezone、Summertime有任意一项参数与预期不同，就会跳保存提示框
        logging.info("Mode、Timezone、Summertime有参数与预期不同，会跳保存提示框")
        send_commd(KEY["EXIT"])
        send_commd(KEY["OK"])
    else:       # 假如Mode、Timezone、Summertime所有参数都与预期相同，不会跳保存提示框
        logging.info("Mode、Timezone、Summertime所有参数与预期相同，不会跳保存提示框")
        send_commd(KEY["EXIT"])
    # 退回大画面
    exit_to_screen()


def get_group_channel_total_info():
    logging.debug("get_group_channel_total_info")
    # 切台前获取case节目类别,分组,分组节目数量,以及获取节目属性前的去除加锁的判断
    # 根据所选case切换到对应类型节目的界面
    while channel_info[5] != TEST_CASE_INFO[2]:
        send_commd(KEY["TV/R"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_commd(KEY["OK"])
    # 采集所有分组的名称和分组下节目总数信息
    if TEST_CASE_INFO[2] == "TV":
        while rsv_info["prog_group_name"] not in GL.TV_channel_groups.keys():
            print(rsv_info["prog_group_name"])
            GL.TV_channel_groups[rsv_info["prog_group_name"]] = rsv_info["prog_group_total"]
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE_INFO[1] not in GL.TV_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的电视分组：{}，退出程序".format(TEST_CASE_INFO[1]))
            send_commd(KEY["EXIT"])
            state["receive_loop_state"] = True
    elif TEST_CASE_INFO[2] == "Radio":
        while rsv_info["prog_group_name"] not in GL.Radio_channel_groups.keys():
            GL.Radio_channel_groups[rsv_info["prog_group_name"]] = rsv_info["prog_group_total"]
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE_INFO[1] not in GL.Radio_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的广播分组：{}，退出程序".format(TEST_CASE_INFO[1]))
            send_commd(KEY["EXIT"])
            state["receive_loop_state"] = True
    # 根据所选case切换到对应的分组
    if TEST_CASE_INFO[2] == "TV":
        while rsv_info["prog_group_name"] != TEST_CASE_INFO[1]:
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    elif TEST_CASE_INFO[2] == "Radio":
        while rsv_info["prog_group_name"] != TEST_CASE_INFO[1]:
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    # 退出频道列表,回到大画面界面
    send_commd(KEY["EXIT"])
    logging.debug(channel_info)
    logging.debug(GL.TV_channel_groups)
    logging.debug(GL.Radio_channel_groups)


def get_choice_group_ch_type():
    logging.debug("get_choice_group_ch_type")
    global channel_info
    # 采集All分组下的节目属性和是否有EPG信息
    choice_group_ch_total_numb = ''
    if TEST_CASE_INFO[2] == "TV":
        choice_group_ch_total_numb = GL.TV_channel_groups[TEST_CASE_INFO[1]]
    elif TEST_CASE_INFO[2] == "Radio":
        choice_group_ch_total_numb = GL.Radio_channel_groups[TEST_CASE_INFO[1]]
    # 进入EPG界面，切台获取指定分组下所有节目的属性
    send_commd(KEY["EPG"])
    logging.debug(rsv_info["prog_group_name"])
    logging.debug(channel_info)
    for i in range(int(choice_group_ch_total_numb)):
        # channel_info = ['', '', '', '', '', '', rsv_info["prog_group_name"], '']
        send_commd(KEY["DOWN"])
        if channel_info[7] == "1":
            time.sleep(0.5)
        elif channel_info[7] == "0" or channel_info[7] == '':
            time.sleep(2.5)
        # time.sleep(1)
        if channel_info[3] == "1":
            for j in range(4):
                send_commd(KEY["0"])
        if TEST_CASE_INFO[2] == "TV":
            if channel_info[7] == "1":  # 所有有EPG信息的电视节目
                GL.TV_ch_attribute[3].append(channel_info[0])
            if channel_info[3] == "1":  # 加锁电视节目
                GL.TV_ch_attribute[2].append(channel_info[0])
            elif channel_info[4] == "0":  # 免费电视节目
                GL.TV_ch_attribute[0].append(channel_info[0])
            elif channel_info[4] == "1":  # 加密电视节目
                GL.TV_ch_attribute[1].append(channel_info[0])
        elif TEST_CASE_INFO[2] == "Radio":
            if channel_info[7] == "1":  # 所有有EPG信息的广播节目
                GL.Radio_ch_attribute[3].append(channel_info[0])
            if channel_info[3] == "1":  # 加锁广播节目
                GL.Radio_ch_attribute[2].append(channel_info[0])
            elif channel_info[4] == "0":  # 免费广播节目
                GL.Radio_ch_attribute[0].append(channel_info[0])
            elif channel_info[4] == "1":  # 加密广播节目
                GL.Radio_ch_attribute[1].append(channel_info[0])
        logging.info(channel_info)
    logging.info(GL.TV_ch_attribute)
    logging.info(GL.Radio_ch_attribute)
    # 退回大画面
    exit_to_screen()


def choice_test_channel():
    logging.debug("choice_test_channel")

    if TEST_CASE_INFO[2] == "TV":
        if len(GL.TV_ch_attribute[3]) == 0:
            logging.info("无有EPG信息的电视节目")
        elif len(GL.TV_ch_attribute[3]) > 0:
            free_tv_numb = choice(GL.TV_ch_attribute[3])
            logging.debug("当前所选有EPG信息的电视节目频道号为:{}".format(free_tv_numb))
            free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
            send_commd(KEY["EXIT"])
            for j in range(len(free_tv_commd)):
                for k in range(len(free_tv_commd[j])):
                    send_commd(free_tv_commd[j][k])
            send_commd(KEY["OK"])
            time.sleep(2)
            logging.info("当前所选有EPG信息的电视节目名称为:{}".format(channel_info[1]))
            logging.info(channel_info)
            GL.choice_res_ch = ch_epg_info[1]

            # 将有EPG信息的节目名称添加到report_data
            GL.report_data[2] = channel_info[0]
            GL.report_data[3] = channel_info[1]

    elif TEST_CASE_INFO[2] == "Radio":
        if len(GL.Radio_ch_attribute[3]) == 0:
            logging.info("无有EPG信息的广播节目")
        elif len(GL.Radio_ch_attribute[3]) > 0:
            free_radio_numb = choice(GL.Radio_ch_attribute[3])
            logging.debug("当前所选有EPG信息的广播节目频道号为:{}".format(free_radio_numb))
            free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
            send_commd(KEY["EXIT"])
            for j in range(len(free_radio_commd)):
                for k in range(len(free_radio_commd[j])):
                    send_commd(free_radio_commd[j][k])
            send_commd(KEY["OK"])
            time.sleep(2)
            logging.info("当前所选有EPG信息的广播节目名称为:{}".format(channel_info[1]))
            logging.info(channel_info)
            GL.choice_res_ch = ch_epg_info[1]

            # 将有EPG信息的节目名称添加到report_data
            GL.report_data[2] = channel_info[0]
            GL.report_data[3] = channel_info[1]


def check_preparatory_work():
    logging.debug("check_preparatory_work")
    state["clear_ch_epg_info_state"] = True
    if TEST_CASE_INFO[6] == "EPG":
        send_commd(KEY["EPG"])
        if channel_info[3] == "1":
            for i in range(4):
                send_commd(KEY["0"])
    elif TEST_CASE_INFO[6] == "Timer":
        enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
        send_more_commds(enter_timer_setting_interface)
        logging.info(rsv_kws["current_sys_time"])


def check_epg_info_already_show():      # 检查EPG信息是否已经显示
    logging.debug("check_epg_info_already_show")
    global ch_epg_info
    while ch_epg_info[-1] == '':         # 假如还没有获取到当前节目的EPG信息，则需要退出等待5秒再进入
        # ch_epg_info = ['', '', '']
        send_commd(KEY["EXIT"])
        time.sleep(5)
        send_commd(KEY["EPG"])
        send_commd(KEY["RIGHT"])
    # ch_epg_info = ['', '', '']
    send_commd(KEY["EXIT"])


def str_time_to_datetime_time(str_time):
    logging.info("str_time_to_datetime_time")
    '''
        将字符串时间用datetime.datetime处理成datetime时间
    '''
    datetime_time = ''
    if len(str_time) == 16:
        str_time_split = re.split(r"[/\s:]", str_time)
        logging.info(str_time_split)
        if len(str_time_split) == 5:
            year = int(str_time_split[0])
            month = int(str_time_split[1])
            day = int(str_time_split[2])
            hour = int(str_time_split[3])
            minute = int(str_time_split[4])
            datetime_time = datetime(year, month, day, hour, minute)
    elif len(str_time) == 19:
        str_time_split = re.split(r"[/\s:]", str_time)
        logging.info(str_time_split)
        if len(str_time_split) == 6:
            year = int(str_time_split[0])
            month = int(str_time_split[1])
            day = int(str_time_split[2])
            hour = int(str_time_split[3])
            minute = int(str_time_split[4])
            second = int(str_time_split[5])
            datetime_time = datetime(year, month, day, hour, minute, second)
    return datetime_time


def from_date_to_secs(str_time):
    # start_time = '9999/12/31 23:59:59'
    before_cur_year_num = []
    each_year_days = []
    add_each_year_days = []
    if len(str_time) == 19:
        pass
    elif len(str_time) == 16:
        str_time += ':00'
    logging.info(str_time)
    start_time_split = re.split(r"[\s:/]", str_time)
    logging.info(start_time_split)
    year = int(start_time_split[0])
    month = int(start_time_split[1])
    day = int(start_time_split[2])
    hour = int(start_time_split[3])
    minute = int(start_time_split[4])
    second = int(start_time_split[5])
    if year == 1:
        cur_year_day_num = int(date(year, month, day).strftime("%j"))
        total_secs = (cur_year_day_num - 1) * 24 * 3600 + hour * 3600 + minute * 60 + second
    else:
        for i in range(1, year):
            before_cur_year_num.append(i)

        for j in range(len(before_cur_year_num)):
            year_day = int(date(before_cur_year_num[j], 12, 31).strftime("%j"))
            each_year_days.append(year_day)

        for k in range(len(each_year_days)):
            if k != 0:
                add_each_year_days.append(each_year_days[k] + add_each_year_days[k - 1])
            else:
                add_each_year_days.append(each_year_days[k])

        # logging.info(before_cur_year_num)
        # logging.info(each_year_days)
        # logging.info(add_each_year_days)
        # logging.info(f'{len(before_cur_year_num)}--{len(add_each_year_days)}')
        logging.info(f'当前年份之前的总天数：{sum(each_year_days)}')
        before_cur_year_total_secs = sum(each_year_days) * 24 * 3600
        logging.info(f'当前年份之前的总天数换算成秒：{before_cur_year_total_secs}')
        cur_year_day_num = int(date(year, month, day).strftime("%j"))
        logging.info(cur_year_day_num)
        cur_year_secs = (cur_year_day_num - 1) * 24 * 3600 + hour * 3600 + minute * 60 + second
        logging.info(f'当前年份天数换算成秒：{cur_year_secs}')
        total_secs = before_cur_year_total_secs + cur_year_secs

    logging.info(f'总秒数：{total_secs}')
    return total_secs


def from_secs_to_date(random_sec_num):
    fmt_date = ''
    scence_1 = False
    scence_2 = False
    max_year = 9999
    month_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    common_year_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    add_common_year_month = []
    for i in range(len(common_year_month)):
        add_common_year_month.append(sum(common_year_month[:i + 1]))
    # logging.info(add_common_year_month)
    leap_year_month = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    add_leap_year_month = []
    for i in range(len(leap_year_month)):
        add_leap_year_month.append(sum(leap_year_month[:i + 1]))
    # logging.info(add_leap_year_month)
    year_list = []
    each_year_days_list = []
    add_each_year_days_list = []
    add_each_year_secs_list = []
    for i in range(1, max_year + 1):
        year_list.append(i)

    for j in range(len(year_list)):
        year_day = int(date(year_list[j], 12, 31).strftime("%j"))
        each_year_days_list.append(year_day)

    for k in range(len(each_year_days_list)):
        if k != 0:
            add_each_year_days_list.append(each_year_days_list[k] + add_each_year_days_list[k - 1])
        else:
            add_each_year_days_list.append(each_year_days_list[k])

    for n in range(len(add_each_year_days_list)):
        add_each_year_secs = add_each_year_days_list[n] * 24 * 3600
        add_each_year_secs_list.append(add_each_year_secs)

    # logging.info(year_list)
    # logging.info(each_year_days_list)
    # logging.info(add_each_year_days_list)
    # logging.info(add_each_year_secs_list)
    # random_sec = randint(0, 315537897599)   # 0001/01/01 00:00:00 ~ 9999/12/31 23:59:59换算成秒
    random_sec = random_sec_num
    logging.info(f'random_sec的值为：{random_sec}')
    m = 0
    while True:
        if random_sec < add_each_year_secs_list[m]:
            logging.info(f'm的值为{m},1')
            scence_1 = True
            break
        elif random_sec == add_each_year_secs_list[m]:
            logging.info(f'm的值为{m},2')
            scence_2 = True
            break
        else:
            m += 1

    logging.info(f'm的值为{m},3')

    logging.info(scence_1)
    if scence_1:
        logging.info('1')
        cur_month = ''
        cur_day = ''
        cur_year = year_list[m]
        cur_year_days = int(date(cur_year, 12, 31).strftime("%j"))
        if m == 0:
            cur_year_secs = random_sec
        else:
            cur_year_secs = random_sec - add_each_year_secs_list[m - 1]
        cur_days = cur_year_secs // 3600 // 24
        if cur_year_days == 365:
            logging.info(f"{cur_year}年有365天")
            add_year_month = add_common_year_month
            for num in range(len(add_year_month)):
                if cur_days < add_year_month[num]:
                    logging.info(f"{num}, <<")
                    cur_month = month_list[num]
                    cur_day = common_year_month[num] - ((add_year_month[num] - cur_days) - 1)
                    break
                elif cur_days == add_year_month[num]:
                    logging.info(f"{num}, ==")
                    cur_month_1 = month_list[num]
                    cur_day_1 = common_year_month[num]
                    date1 = datetime(cur_year, cur_month_1, cur_day_1)
                    date2 = date1 + timedelta(days=1)
                    date2_split = re.split(r"[-\s:]", str(date2))
                    cur_month = int(date2_split[1])
                    cur_day = int(date2_split[2])
                    break
        elif cur_year_days == 366:
            logging.info(f"{cur_year}年有366天")
            add_year_month = add_leap_year_month
            for num in range(len(add_year_month)):
                if cur_days < add_year_month[num]:
                    logging.info(f"{num}, <<")
                    cur_month = month_list[num]
                    cur_day = leap_year_month[num] - ((add_year_month[num] - cur_days) - 1)
                    break
                elif cur_days == add_year_month[num]:
                    logging.info(f"{num}, ==")
                    cur_month_1 = month_list[num]
                    cur_day_1 = leap_year_month[num]
                    date1 = datetime(cur_year, cur_month_1, cur_day_1)
                    date2 = date1 + timedelta(days=1)
                    date2_split = re.split(r"[-\s:]", str(date2))
                    cur_month = int(date2_split[1])
                    cur_day = int(date2_split[2])
                    break
        secs_left = cur_year_secs - cur_days * 24 * 3600
        cur_hour = secs_left // 3600
        cur_minute = (secs_left % 3600) // 60
        cur_second = (secs_left % 3600) % 60
        logging.info(datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute, cur_second))
        fmt_date = str(datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute, cur_second))

    logging.info(scence_2)
    if scence_2:
        logging.info('2')
        cur_year = year_list[m]
        cur_month = 12
        cur_day = 31
        cur_hour = 23
        cur_minute = 59
        cur_second = 59
        datetime_time = datetime(cur_year, cur_month, cur_day, cur_hour, cur_minute, cur_second)
        new_datetime_time = datetime_time + timedelta(seconds=1)
        logging.info(new_datetime_time)
        fmt_date = str(new_datetime_time)

    return fmt_date


def get_random_time_between_time_period(start_time, end_time):
    logging.info("get_random_time_between_time_period")
    # 在指定的起始和结束时间范围内随机取一个时间值
    start_time_secs = from_date_to_secs(start_time)
    end_time_secs = from_date_to_secs(end_time)
    get_random_time = randint(int(start_time_secs), int(end_time_secs))
    fmt_random_time = from_secs_to_date(get_random_time)     # 将时间秒转成格式化日期字符串
    logging.info(fmt_random_time)
    fmt_random_time_split = re.split(r"[-:\s]", fmt_random_time)
    logging.info(fmt_random_time_split)
    if len(fmt_random_time_split[0]) < 4:   # 假如年份的位数不满4位，需要补齐
        fmt_random_time_split[0] = "{0:04d}".format(int(fmt_random_time_split[0]))
        fmt_time = "".join(fmt_random_time_split)
    else:
        fmt_time = "".join(fmt_random_time_split)
    return fmt_time


def calculate_str_time_to_fmt_time(str_time, interval_time):
    logging.info("calculate_str_time_to_fmt_time")
    '''
        str_time为2000/01/01 00:00的格式;输出为202009111213的格式
        计算某个时间的与interval_time相加或相减后的时间值
    '''
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    str_time_split = re.split(r"[/:\s]", str_time)
    deal_str_time = ''.join(str_time_split)
    if len(deal_str_time) == 12:     # once事件时间计算
        fmt_year = int(deal_str_time[:4])
        fmt_month = int(deal_str_time[4:6])
        fmt_day = int(deal_str_time[6:8])
        fmt_hour = int(deal_str_time[8:10])
        fmt_minute = int(deal_str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(minutes=interval_time)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
    elif len(deal_str_time) == 14:     # once事件时间计算
        fmt_year = int(deal_str_time[:4])
        fmt_month = int(deal_str_time[4:6])
        fmt_day = int(deal_str_time[6:8])
        fmt_hour = int(deal_str_time[8:10])
        fmt_minute = int(deal_str_time[10:12])
        fmt_second = int(deal_str_time[12:14])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute, fmt_second)
        new_fmt_date = fmt_date + timedelta(minutes=interval_time)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
    else:
        logging.info("str_time的时间格式不对")
    return str_new_fmt_date


def str_time_to_fmt_time(str_time):
    logging.info("fmt_time_to_str_time")
    # 将“2020/09/09 12:11”格式转化为”202009091211”格式
    fmt_time = ''
    if len(str_time) == 16:
        str_time_split = re.split(r"[/\s:]", str_time)
        logging.info(str_time_split)
        fmt_time = ''.join(str_time_split)

    elif len(str_time) == 19:
        str_time_split = re.split(r"[/\s:]", str_time)
        logging.info(str_time_split)
        fmt_time = ''.join(str_time_split)[:12]
    return fmt_time


def fmt_time_to_str_time(fmt_time):
    logging.info("fmt_time_to_str_time")
    # 将”202009091211”格式转化为“2020/09/09 12：11”格式
    str_time = ''
    if len(fmt_time) == 12:
        str_time = f"{fmt_time[:4]}/{fmt_time[4:6]}/{fmt_time[6:8]} {fmt_time[8:10]}:{fmt_time[10:12]}"
    elif len(fmt_time) == 14:
        str_time = f"{fmt_time[:4]}/{fmt_time[4:6]}/{fmt_time[6:8]} {fmt_time[8:10]}:{fmt_time[10:12]}:{fmt_time[12:14]}"
    return str_time


def calculate_other_timezone_save_time(str_time):
    logging.info("calculate_other_timezone_save_time")
    '''
            str_time为2000/01/01 00:00的格式;
            返回的new_str_time也仍为2000/01/01 00:00的格式;
            计算某个时间的与interval_timezone相加或相减后的时间值
        '''
    # 字符串时间和格式化时间之间转换
    new_str_time = ''
    interval_timezone = float(GL.choice_timezone)
    str_time_split = re.split(r"[/:\s]", str_time)
    deal_str_time = ''.join(str_time_split)
    if len(deal_str_time) == 12:  # once事件时间计算
        fmt_year = int(deal_str_time[:4])
        fmt_month = int(deal_str_time[4:6])
        fmt_day = int(deal_str_time[6:8])
        fmt_hour = int(deal_str_time[8:10])
        fmt_minute = int(deal_str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(hours=interval_timezone)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        new_fmt_str_time = ''.join(new_fmt_date_split)[:12]  # 去掉末尾的秒钟信息
        new_str_time = fmt_time_to_str_time(new_fmt_str_time)
    else:
        logging.info("str_time的时间格式不对")
    return new_str_time


def calculate_expected_event_start_time():
    logging.info("calculate_expected_event_start_time")
    str_expected_res_time = ''
    input_range_start_time = "0001/01/01 00:00"
    input_range_end_time = "9999/12/31 23:59"
    default_save_range_start_time = "2000/01/01 00:00"
    default_save_range_end_time = "2037/12/31 23:59"
    if TEST_CASE_INFO[8] == "ZeroTimezone":
        if TEST_CASE_INFO[9] == "NoSummertime":     # 无夏令时
            if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                if TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                    # str_expected_res_time = "199912312359"
                    str_expected_res_time = calculate_str_time_to_fmt_time(default_save_range_start_time, -1)

                elif TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                    # str_expected_res_time = "000101010000"
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_start_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_end_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(default_save_range_end_time, 1)

                elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                    # enter_end_time = "9999/12/31 23:59"
                    enter_end_time = input_range_end_time
                    after_save_time_range_start_time = "2038/01/01 00:00"
                    str_expected_res_time = get_random_time_between_time_period(
                        after_save_time_range_start_time, enter_end_time
                    )
                    logging.info(str_expected_res_time)

                elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                    # enter_start_time = "0001/01/01 00:00"
                    enter_start_time = input_range_start_time
                    before_save_time_range_end_time = "1999/12/31 23:59"
                    str_expected_res_time = get_random_time_between_time_period(
                        enter_start_time, before_save_time_range_end_time
                    )
                    logging.info(str_expected_res_time)

            elif TEST_CASE_INFO[5] == "Expired":
                if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                    # str_expected_res_time = "200001010000"
                    str_expected_res_time = calculate_str_time_to_fmt_time(default_save_range_start_time, 0)
                elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                    current_sys_time = rsv_kws["current_sys_time"]
                    str_expected_res_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                    current_sys_time = rsv_kws["current_sys_time"]
                    current_invalid_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                    before_current_time_range_end_time = fmt_time_to_str_time(current_invalid_time)
                    str_expected_res_time = get_random_time_between_time_period(
                        default_save_range_start_time, before_current_time_range_end_time
                    )
            elif TEST_CASE_INFO[5] == "NowPlaying":
                current_sys_time = rsv_kws["current_sys_time"]
                str_expected_res_time = str_time_to_fmt_time(current_sys_time)
            elif TEST_CASE_INFO[5] == "InvalidDuration":
                current_sys_time = rsv_kws["current_sys_time"]
                str_expected_res_time = calculate_str_time_to_fmt_time(current_sys_time, 5)

        elif TEST_CASE_INFO[9] == "Summertime":     # 有夏令时
            summertime_save_range_start_time = fmt_time_to_str_time(
                calculate_str_time_to_fmt_time(default_save_range_start_time, 60))
            summertime_save_range_end_time = fmt_time_to_str_time(
                calculate_str_time_to_fmt_time(default_save_range_end_time, 60))
            if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                if TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                    # str_expected_res_time = "199912312359"
                    str_expected_res_time = calculate_str_time_to_fmt_time(summertime_save_range_start_time, -1)

                elif TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                    # str_expected_res_time = "000101010000"
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_start_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_end_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(summertime_save_range_end_time, 1)

                elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                    # enter_end_time = "9999/12/31 23:59"
                    enter_end_time = input_range_end_time
                    after_save_time_range_start_time = fmt_time_to_str_time(
                        calculate_str_time_to_fmt_time(summertime_save_range_end_time, 1))
                    str_expected_res_time = get_random_time_between_time_period(
                        after_save_time_range_start_time, enter_end_time
                    )
                    logging.info(str_expected_res_time)

                elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                    # enter_start_time = "0001/01/01 00:00"
                    enter_start_time = input_range_start_time
                    before_save_time_range_end_time = fmt_time_to_str_time(
                        calculate_str_time_to_fmt_time(summertime_save_range_start_time, -1))
                    str_expected_res_time = get_random_time_between_time_period(
                        enter_start_time, before_save_time_range_end_time
                    )
                    logging.info(str_expected_res_time)

            elif TEST_CASE_INFO[5] == "Expired":
                if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                    # str_expected_res_time = "200001010000"
                    str_expected_res_time = calculate_str_time_to_fmt_time(summertime_save_range_start_time, 0)
                elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                    current_sys_time = rsv_kws["current_sys_time"]
                    str_expected_res_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                    current_sys_time = rsv_kws["current_sys_time"]
                    current_invalid_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                    before_current_time_range_end_time = fmt_time_to_str_time(current_invalid_time)
                    str_expected_res_time = get_random_time_between_time_period(
                        summertime_save_range_start_time, before_current_time_range_end_time
                    )
    elif TEST_CASE_INFO[8] == "OtherTimezone":
        if TEST_CASE_INFO[9] == "NoSummertime":
            choice_timezone_save_range_start_time = calculate_other_timezone_save_time(default_save_range_start_time)
            choice_timezone_save_range_end_time = calculate_other_timezone_save_time(default_save_range_end_time)
            logging.info(
                f"计算后的保存范围为：{choice_timezone_save_range_start_time}-------{choice_timezone_save_range_end_time}")
            if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                if TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(choice_timezone_save_range_start_time, -1)

                elif TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_start_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_end_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(choice_timezone_save_range_end_time, 1)

                elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                    enter_end_time = input_range_end_time
                    after_save_time_range_start_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                        choice_timezone_save_range_end_time, 1))
                    str_expected_res_time = get_random_time_between_time_period(
                        after_save_time_range_start_time, enter_end_time
                    )
                    logging.info(str_expected_res_time)

                elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                    enter_start_time = input_range_start_time
                    before_save_time_range_end_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                        choice_timezone_save_range_start_time, -1))
                    str_expected_res_time = get_random_time_between_time_period(
                        enter_start_time, before_save_time_range_end_time
                    )
                    logging.info(str_expected_res_time)

            elif TEST_CASE_INFO[5] == "Expired":
                if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                    # str_expected_res_time = "200001010000"
                    str_expected_res_time = calculate_str_time_to_fmt_time(choice_timezone_save_range_start_time, 0)
                elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                    current_sys_time = rsv_kws["current_sys_time"]
                    str_expected_res_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                    current_sys_time = rsv_kws["current_sys_time"]
                    current_invalid_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                    before_current_time_range_end_time = fmt_time_to_str_time(current_invalid_time)
                    str_expected_res_time = get_random_time_between_time_period(
                        choice_timezone_save_range_start_time, before_current_time_range_end_time
                    )

        elif TEST_CASE_INFO[9] == "Summertime":
            choice_timezone_save_range_start_time = calculate_other_timezone_save_time(
                default_save_range_start_time)
            choice_timezone_save_range_end_time = calculate_other_timezone_save_time(
                default_save_range_end_time)
            summertime_choice_timezone_save_range_start_time = fmt_time_to_str_time(
                calculate_str_time_to_fmt_time(choice_timezone_save_range_start_time, 60))
            summertime_choice_timezone_save_range_end_time = fmt_time_to_str_time(
                calculate_str_time_to_fmt_time(choice_timezone_save_range_end_time, 60))
            logging.info(
                f"计算后的保存范围为：{summertime_choice_timezone_save_range_start_time}-------"
                f"{summertime_choice_timezone_save_range_end_time}")
            if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                if TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(
                        summertime_choice_timezone_save_range_start_time, -1)

                elif TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_start_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(input_range_end_time, 0)

                elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                    str_expected_res_time = calculate_str_time_to_fmt_time(
                        summertime_choice_timezone_save_range_end_time, 1)

                elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                    enter_end_time = input_range_end_time
                    after_save_time_range_start_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                        summertime_choice_timezone_save_range_end_time, 1))
                    str_expected_res_time = get_random_time_between_time_period(
                        after_save_time_range_start_time, enter_end_time
                    )
                    logging.info(str_expected_res_time)

                elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                    enter_start_time = input_range_start_time
                    before_save_time_range_end_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                        summertime_choice_timezone_save_range_start_time, -1))
                    str_expected_res_time = get_random_time_between_time_period(
                        enter_start_time, before_save_time_range_end_time
                    )
                    logging.info(str_expected_res_time)

            elif TEST_CASE_INFO[5] == "Expired":
                if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                    # str_expected_res_time = "200001010000"
                    str_expected_res_time = calculate_str_time_to_fmt_time(
                        summertime_choice_timezone_save_range_start_time, 0)
                elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                    current_sys_time = rsv_kws["current_sys_time"]
                    str_expected_res_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                    current_sys_time = rsv_kws["current_sys_time"]
                    current_invalid_time = calculate_str_time_to_fmt_time(current_sys_time, -1)
                    before_current_time_range_end_time = fmt_time_to_str_time(current_invalid_time)
                    str_expected_res_time = get_random_time_between_time_period(
                        summertime_choice_timezone_save_range_start_time, before_current_time_range_end_time
                    )

    logging.info(f"期望的完整的预约事件时间为{str_expected_res_time}")
    return str_expected_res_time


def create_expected_add_event_info():
    logging.info("create_expected_add_event_info")
    # 创建期望的事件信息
    expected_event_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    if TEST_CASE_INFO[5] == "InvalidDuration":
        duration_time = "0000"
    else:
        duration_time = "0001"
    if TEST_CASE_INFO[3] == "Play":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[3]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[4]

    elif TEST_CASE_INFO[3] == "PVR":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[3]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = duration_time
        expected_event_info[4] = TEST_CASE_INFO[4]

    elif TEST_CASE_INFO[3] == "Power Off":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[3]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[4]

    elif TEST_CASE_INFO[3] == "Power On":
        expected_event_full_time = calculate_expected_event_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[3]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[4]
    return expected_event_info


def edit_add_new_res_event_info():
    logging.info("edit_add_new_res_event_info")
    # 编辑预约事件信息
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面
    send_commd(KEY["GREEN"])
    # 生成预期的预约事件
    if TEST_CASE_INFO[4] == "Once":
        expected_res_event_info = create_expected_add_event_info()
    else:
        pretreatment_res_event_info = create_expected_add_event_info()    # 预处理预约事件信息
        expected_res_event_info = pretreatment_res_event_info.copy()
        expected_res_event_info[0] = expected_res_event_info[0][8:]
    logging.info(f"创建的事件为{expected_res_event_info}")
    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[3]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[3]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[3]}')
            send_commd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_commd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[4]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[4]}')
            send_commd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[4]}')
            send_commd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[4] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[4]}")
    elif TEST_CASE_INFO[4] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_commd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_commds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_commd(KEY["DOWN"])
    else:
        if TEST_CASE_INFO[4] == "Once":
            start_time_list.append(expected_res_event_info[0][8:])
        else:
            start_time_list.append(expected_res_event_info[0])
        start_time_cmd = change_numbs_to_commds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_commd(j)
        send_commd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[3] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[3]}")
    elif TEST_CASE_INFO[3] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[3]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_commd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_commds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_commd(j)
            send_commd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[3] == "Power Off" or TEST_CASE_INFO[3] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[3]}，不需要设置Channel")
    elif TEST_CASE_INFO[3] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[3]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_commd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["update_event_list_state"] = True
    send_commd(KEY["EXIT"])
    send_commd(KEY["OK"])
    GL.report_data[8] = rsv_kws["event_invalid_msg"]
    if rsv_kws["event_invalid_msg"] == "[PTD]Res_invalid_timer":
        send_commd(KEY["OK"])
    # 添加新预约事件到report
    if TEST_CASE_INFO[3] == "PVR":  # 手动指定dur的‘：’间隔
        new_expected_res_event_info = expected_res_event_info
        dur_time = new_expected_res_event_info[3]
        new_expected_res_event_info[3] = dur_time[:2] + ":" + dur_time[2:]
        GL.report_data[4].extend(new_expected_res_event_info)
    else:
        GL.report_data[4].extend(expected_res_event_info)
    # 退回大画面
    # exit_to_screen()


def send_test_case_commd():
    logging.info("send_test_case_commd")
    GL.report_data[1] = rsv_kws["current_sys_time"]
    if TEST_CASE_INFO[6] == "EPG":
        if TEST_CASE_INFO[5] == "Expired":
            logging.info(rsv_kws["current_sys_time"])
            while True:
                logging.info(ch_epg_info)
                if ch_epg_info[-1] != "":
                    event_start_time = str_time_to_datetime_time(ch_epg_info[0])
                    event_end_time = str_time_to_datetime_time(ch_epg_info[1])
                    sys_time = str_time_to_datetime_time(rsv_kws["current_sys_time"])
                    if event_start_time < sys_time and event_end_time < sys_time:
                        logging.info("当前事件为过期事件")
                        event_flag = "Expired"
                        if event_flag == TEST_CASE_INFO[5]:
                            GL.report_data[4] = ch_epg_info
                            break
                    elif event_start_time <= sys_time <= event_end_time:
                        logging.info("当前事件为正在播放事件")
                        event_flag = "NowPlaying"
                        if event_flag != TEST_CASE_INFO[5]:
                            send_commd(KEY["LEFT"])
                    elif event_start_time > sys_time and event_end_time > sys_time:
                        logging.info("当前事件为未播放事件")
                        event_flag = "NoPlay"
                        if event_flag != TEST_CASE_INFO[5]:
                            send_commd(KEY["LEFT"])
                elif ch_epg_info[-1] == "":
                    send_commd(KEY["RIGHT"])
                    time.sleep(1)

        elif TEST_CASE_INFO[5] == "NowPlaying":
            logging.info(rsv_kws["current_sys_time"])
            while True:
                logging.info(ch_epg_info)
                if ch_epg_info[-1] != "":
                    event_start_time = str_time_to_datetime_time(ch_epg_info[0])
                    event_end_time = str_time_to_datetime_time(ch_epg_info[1])
                    sys_time = str_time_to_datetime_time(rsv_kws["current_sys_time"])
                    if event_start_time < sys_time and event_end_time < sys_time:
                        logging.info("当前事件为过期事件")
                        event_flag = "Expired"
                        if event_flag != TEST_CASE_INFO[5]:
                            send_commd(KEY["RIGHT"])
                    elif event_start_time <= sys_time <= event_end_time:
                        logging.info("当前事件为正在播放事件")
                        event_flag = "NowPlaying"
                        if event_flag == TEST_CASE_INFO[5]:
                            GL.report_data[4] = ch_epg_info
                            break
                    elif event_start_time > sys_time and event_end_time > sys_time:
                        logging.info("当前事件为未播放事件")
                        event_flag = "NoPlay"
                        if event_flag != TEST_CASE_INFO[5]:
                            send_commd(KEY["LEFT"])

                elif ch_epg_info[-1] == "":
                    send_commd(KEY["RIGHT"])
                    time.sleep(1)

        # 预约事件
        send_commd(KEY["RED"])
        if TEST_CASE_INFO[3] == "Play":
            send_commd(KEY["EXIT"])
            send_commd(KEY["OK"])
        elif TEST_CASE_INFO[3] == "PVR":
            send_commd(KEY["RIGHT"])
            send_commd(KEY["EXIT"])
            send_commd(KEY["OK"])
        # 判断事件是否Book成功
        if rsv_kws["event_invalid_msg"] != '':
            GL.report_data[7] = rsv_kws["event_invalid_msg"]
            logging.info(f'预约不成功：{rsv_kws["event_invalid_msg"]}')
            send_commd(KEY["OK"])
        elif rsv_kws["event_invalid_msg"] == '':
            GL.report_data[7] = "Book_res_event_success"
            logging.info(f'预约成功：{rsv_kws["event_invalid_msg"]}')

    elif TEST_CASE_INFO[6] == "Timer":
        if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
            edit_add_new_res_event_info()
        elif TEST_CASE_INFO[5] == "Expired":
            edit_add_new_res_event_info()
        elif TEST_CASE_INFO[5] == "NowPlaying":
            edit_add_new_res_event_info()
        elif TEST_CASE_INFO[5] == "InvalidDuration":
            edit_add_new_res_event_info()
    # 退回大画面
    exit_to_screen()


def padding_report_data():
    logging.info("padding_report_data")

    if TEST_CASE_INFO[6] == "EPG":
        GL.report_data[0] = TEST_CASE_INFO[0]           # 用例编号
        GL.report_data[5] = TEST_CASE_INFO[3]           # 预约事件类型
        GL.report_data[8] = str(datetime.now())[:19]    # 写该用例报告的时间

    elif TEST_CASE_INFO[6] == "Timer":
        GL.report_data[0] = TEST_CASE_INFO[0]           # 用例编号
        GL.report_data[5] = TEST_CASE_INFO[7]           # 场景描述
        GL.report_data[6] = GL.choice_timezone          # 时区
        GL.report_data[9] = str(datetime.now())[:19]    # 写该用例报告的时间


def write_data_to_report():
    logging.info("write_data_to_report")
    wb = ''
    ws = ''
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    blue_font = Font(color=BLUE)
    red_font = Font(color=RED)
    dark_cyan = '00008B8B'
    dark_cyan_font = Font(color=dark_cyan, bold=True)
    a_column_numb = column_index_from_string("A")

    if TEST_CASE_INFO[6] == "EPG":
        excel_title_1 = ["用例编号", "系统时间", "所选节目频道号", "所选节目频道名称", "所选事件信息", "预约类型", "预约事件结果"]
        excel_title_2 = ["用例编号", "系统时间", "所选节目频道号", "所选节目频道名称",
                         "起始时间", "结束时间", "事件名称", "预约类型",
                         "事件列表预约事件个数", "无效事件提示", "用例测试时间"]

        if not os.path.exists(file_path[1]):
            wb = Workbook()
            ws = wb.active
            ws.title = file_path[2]
            # 写excel_title_1的内容
            for i in range(len(excel_title_1)):
                if i == 4:
                    ws.cell(1, i + 1).value = excel_title_1[i]
                    ws.cell(1, i + 1).alignment = alignment
                    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
                elif i == 5:
                    ws.cell(1, i + 3).value = excel_title_1[i]
                    ws.cell(1, i + 3).alignment = alignment
                elif i == 6:
                    ws.cell(1, i + 3).value = excel_title_1[i]
                    ws.cell(1, i + 3).alignment = alignment
                    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)
                else:
                    ws.cell(1, i + 1).value = excel_title_1[i]
                    ws.cell(1, i + 1).alignment = alignment
            # 写excel_title_2的内容
            for j in range(len(excel_title_2)):
                ws.cell(2, j + 1).value = excel_title_2[j]
                ws.cell(2, j + 1).alignment = alignment
                if j == 0:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
                elif j in [1, 2, 3]:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12

            # 设置Title的行高
            ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
            ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
            # 合并用例编号单元格，以及report前4个数据的单元格
            for column in range(4):
                ws.merge_cells(start_row=1, start_column=column + 1, end_row=2, end_column=column + 1)
            ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=8)

        elif os.path.exists(file_path[1]):
            wb = load_workbook(file_path[1])
            sheets_name_list = wb.sheetnames
            logging.info(sheets_name_list)
            if file_path[2] in sheets_name_list:
                ws = wb[file_path[2]]
            elif file_path[2] not in sheets_name_list:
                ws = wb.create_sheet(file_path[2])
                # 写excel_title_1的内容
                for i in range(len(excel_title_1)):
                    if i == 4:
                        ws.cell(1, i + 1).value = excel_title_1[i]
                        ws.cell(1, i + 1).alignment = alignment
                        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
                    elif i == 5:
                        ws.cell(1, i + 3).value = excel_title_1[i]
                        ws.cell(1, i + 3).alignment = alignment
                    elif i == 6:
                        ws.cell(1, i + 3).value = excel_title_1[i]
                        ws.cell(1, i + 3).alignment = alignment
                        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)
                    else:
                        ws.cell(1, i + 1).value = excel_title_1[i]
                        ws.cell(1, i + 1).alignment = alignment
                # 写excel_title_2的内容
                for j in range(len(excel_title_2)):
                    ws.cell(2, j + 1).value = excel_title_2[j]
                    ws.cell(2, j + 1).alignment = alignment
                    if j == 0:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
                    elif j in [1, 2, 3]:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12
                    else:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12

                # 设置Title的行高
                ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
                ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
                # 合并用例编号单元格，以及report前4个数据的单元格
                for column in range(4):
                    ws.merge_cells(start_row=1, start_column=column + 1, end_row=2, end_column=column + 1)
                ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=8)

        # 获取当前用例修改类型的sheet表的Max_row
        max_row = ws.max_row

        # 写report_data数据
        event_start_time = str_time_to_datetime_time(GL.report_data[4][0])
        event_end_time = str_time_to_datetime_time(GL.report_data[4][1])
        sys_time = str_time_to_datetime_time(GL.report_data[1])

        for d in range(len(GL.report_data)):
            if d in [0, 1, 2, 3]:
                ws.cell(max_row + 1, d + 1).value = GL.report_data[d]
                ws.cell(max_row + 1, d + 1).alignment = alignment
                if d == 1:
                    if TEST_CASE_INFO[5] == "Expired":
                        if sys_time > event_start_time and sys_time > event_end_time:
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font
                    elif TEST_CASE_INFO[5] == "NowPlaying":
                        if event_start_time <= sys_time <= event_end_time:
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font

            elif d == 4:    # 所选EPG事件信息
                for edit_data in range(len(GL.report_data[d])):
                    ws.cell(max_row + 1, d + edit_data + 1).value = GL.report_data[d][edit_data]
                    ws.cell(max_row + 1, d + edit_data + 1).alignment = alignment

                    if TEST_CASE_INFO[5] == "Expired":
                        if edit_data == 0:
                            if event_start_time < sys_time:
                                ws.cell(max_row + 1, d + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + edit_data + 1).font = red_font
                        elif edit_data == 1:
                            if event_end_time < sys_time:
                                ws.cell(max_row + 1, d + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + edit_data + 1).font = red_font
                    elif TEST_CASE_INFO[5] == "NowPlaying":
                        if edit_data == 0:
                            if event_start_time <= sys_time:
                                ws.cell(max_row + 1, d + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + edit_data + 1).font = red_font
                        elif edit_data == 1:
                            if event_end_time >= sys_time:
                                ws.cell(max_row + 1, d + edit_data + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + edit_data + 1).font = red_font

            else:
                ws.cell(max_row + 1, d + 3).value = GL.report_data[d]
                ws.cell(max_row + 1, d + 3).alignment = alignment
                if d == 5:  # 事件类型
                    if GL.report_data[d] == TEST_CASE_INFO[3]:
                        ws.cell(max_row + 1, d + 3).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + 3).font = red_font
                elif d == 6:  # 预约列表事件个数
                    if GL.report_data[d] == "0":
                        ws.cell(max_row + 1, d + 3).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + 3).font = red_font
                elif d == 7:  # 无效事件的提示
                    if GL.report_data[d] == "[PTD]Res_invalid_timer":
                        ws.cell(max_row + 1, d + 3).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + 3).font = red_font

        ws.row_dimensions[(max_row + 1)].height = 70  # 设置每次执行的report预约事件信息的行高
        wb.save(file_path[1])
    elif TEST_CASE_INFO[6] == "Timer":
        excel_title_1 = ["用例编号", "系统时间", "所选节目频道号", "所选节目频道名称", "预期的无效事件信息", "场景描述",
                         "时区", "预约事件结果"]
        excel_title_2 = ["用例编号", "系统时间", "所选节目频道号", "所选节目频道名称", "预期的无效事件信息", "场景描述",
                         "时区", "事件列表预约事件个数", "无效事件提示", "用例测试时间"]

        if not os.path.exists(file_path[1]):
            wb = Workbook()
            ws = wb.active
            ws.title = file_path[2]
            # 写excel_title_1的内容
            for i in range(len(excel_title_1)):
                if i == 7:
                    ws.cell(1, i + 1).value = excel_title_1[i]
                    ws.cell(1, i + 1).alignment = alignment
                    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
                else:
                    ws.cell(1, i + 1).value = excel_title_1[i]
                    ws.cell(1, i + 1).alignment = alignment
            # 写excel_title_2的内容
            for j in range(len(excel_title_2)):
                ws.cell(2, j + 1).value = excel_title_2[j]
                ws.cell(2, j + 1).alignment = alignment
                if j == 0:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
                elif j in [1, 2, 3]:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12
                elif j == 4 or j == 5:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 26
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12

            # 设置Title的行高
            ws.row_dimensions[1].height = 35  # 设置每次执行的report预约事件信息的行高
            ws.row_dimensions[2].height = 35  # 设置每次执行的report预约事件信息的行高
            # 合并用例编号单元格，以及report前7个数据的单元格
            for column in range(7):
                ws.merge_cells(start_row=1, start_column=column + 1, end_row=2, end_column=column + 1)

        elif os.path.exists(file_path[1]):
            wb = load_workbook(file_path[1])
            sheets_name_list = wb.sheetnames
            logging.info(sheets_name_list)
            if file_path[2] in sheets_name_list:
                ws = wb[file_path[2]]
            elif file_path[2] not in sheets_name_list:
                ws = wb.create_sheet(file_path[2])
                # 写excel_title_1的内容
                for i in range(len(excel_title_1)):
                    if i == 7:
                        ws.cell(1, i + 1).value = excel_title_1[i]
                        ws.cell(1, i + 1).alignment = alignment
                        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
                    else:
                        ws.cell(1, i + 1).value = excel_title_1[i]
                        ws.cell(1, i + 1).alignment = alignment
                # 写excel_title_2的内容
                for j in range(len(excel_title_2)):
                    ws.cell(2, j + 1).value = excel_title_2[j]
                    ws.cell(2, j + 1).alignment = alignment
                    if j == 0:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
                    elif j in [1, 2, 3]:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12
                    elif j == 4 or j == 5:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 26
                    else:
                        ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 12

                # 设置Title的行高
                ws.row_dimensions[1].height = 35  # 设置每次执行的report预约事件信息的行高
                ws.row_dimensions[2].height = 35  # 设置每次执行的report预约事件信息的行高
                # 合并用例编号单元格，以及report前7个数据的单元格
                for column in range(7):
                    ws.merge_cells(start_row=1, start_column=column + 1, end_row=2, end_column=column + 1)

        # 获取当前用例修改类型的sheet表的Max_row
        max_row = ws.max_row

        # 写report_data数据
        default_save_range_start_time = "2000/01/01 00:00"
        default_save_range_end_time = "2037/12/31 23:59"
        choice_timezone_save_range_start_time = calculate_other_timezone_save_time(default_save_range_start_time)
        choice_timezone_save_range_end_time = calculate_other_timezone_save_time(default_save_range_end_time)
        for d in range(len(GL.report_data)):
            if d == 4:  # 预期无效事件信息
                ws.cell(max_row + 1, d + 1).value = str(GL.report_data[d])
                ws.cell(max_row + 1, d + 1).alignment = alignment
                if TEST_CASE_INFO[8] == "ZeroTimezone" and TEST_CASE_INFO[9] == "NoSummertime":
                    if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                        if TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                            if GL.report_data[d][0] == '000101010000' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                            if GL.report_data[d][0] == '199912312359' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                            if GL.report_data[d][0] == '203801010000' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                            if GL.report_data[d][0] == '999912312359' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                            enter_start_time = "0001/01/01 00:00"
                            before_save_time_range_end_time = "1999/12/31 23:59"
                            range_start_time = str_time_to_datetime_time(enter_start_time)
                            range_end_time = str_time_to_datetime_time(before_save_time_range_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                            enter_end_time = "9999/12/31 23:59"
                            after_save_time_range_start_time = "2038/01/01 00:00"
                            range_start_time = str_time_to_datetime_time(after_save_time_range_start_time)
                            range_end_time = str_time_to_datetime_time(enter_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                    elif TEST_CASE_INFO[5] == "Expired":
                        if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                            if GL.report_data[d][0] == '200001010000' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                            str_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            if GL.report_data[d][0] == str_expected_res_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                            default_save_range_start_time = "2000/01/01 00:00"
                            fmt_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            str_expected_res_time = fmt_time_to_str_time(fmt_expected_res_time)
                            range_start_time = str_time_to_datetime_time(default_save_range_start_time)
                            range_end_time = str_time_to_datetime_time(str_expected_res_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                    elif TEST_CASE_INFO[5] == "NowPlaying":
                        fmt_current_sys_time = str_time_to_fmt_time(GL.report_data[1])
                        if GL.report_data[d][0] == fmt_current_sys_time \
                                and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font

                    elif TEST_CASE_INFO[5] == "InvalidDuration":
                        event_start_time = calculate_str_time_to_fmt_time(GL.report_data[1], 5)
                        if TEST_CASE_INFO[4] == "Once":
                            if GL.report_data[d][0] == event_start_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4] \
                                    and GL.report_data[d][3] == "00:00":
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        else:
                            if GL.report_data[d][0] == event_start_time[8:] \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4] \
                                    and GL.report_data[d][3] == "00:00":
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                elif TEST_CASE_INFO[8] == "ZeroTimezone" and TEST_CASE_INFO[9] == "Summertime":
                    summertime_save_range_start_time = fmt_time_to_str_time(
                        calculate_str_time_to_fmt_time(default_save_range_start_time, 60))
                    summertime_save_range_end_time = fmt_time_to_str_time(
                        calculate_str_time_to_fmt_time(default_save_range_end_time, 60))
                    if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                        if TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                            if GL.report_data[d][0] == '000101010000' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                            if GL.report_data[d][0] == calculate_str_time_to_fmt_time(
                                    summertime_save_range_start_time, -1) \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                            if GL.report_data[d][0] == calculate_str_time_to_fmt_time(
                                    summertime_save_range_end_time, 1) \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                            if GL.report_data[d][0] == '999912312359' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                            enter_start_time = "0001/01/01 00:00"
                            before_save_time_range_end_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                                    summertime_save_range_start_time, -1))
                            range_start_time = str_time_to_datetime_time(enter_start_time)
                            range_end_time = str_time_to_datetime_time(before_save_time_range_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                            enter_end_time = "9999/12/31 23:59"
                            after_save_time_range_start_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                                    summertime_save_range_end_time, 1))
                            range_start_time = str_time_to_datetime_time(after_save_time_range_start_time)
                            range_end_time = str_time_to_datetime_time(enter_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                    elif TEST_CASE_INFO[5] == "Expired":
                        if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                            if GL.report_data[d][0] == calculate_str_time_to_fmt_time(
                                    default_save_range_start_time, 60) \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                            str_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            if GL.report_data[d][0] == str_expected_res_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                            # default_save_range_start_time = "2000/01/01 00:00"
                            fmt_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            str_expected_res_time = fmt_time_to_str_time(fmt_expected_res_time)
                            range_start_time = str_time_to_datetime_time(summertime_save_range_start_time)
                            range_end_time = str_time_to_datetime_time(str_expected_res_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                elif TEST_CASE_INFO[8] == "OtherTimezone" and TEST_CASE_INFO[9] == "NoSummertime":
                    if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                        if TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                            if GL.report_data[d][0] == '000101010000' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                            boundary_before_upper_limit_time = calculate_str_time_to_fmt_time(
                                choice_timezone_save_range_start_time, -1)
                            if GL.report_data[d][0] == boundary_before_upper_limit_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                            boundary_after_lower_limit_time = calculate_str_time_to_fmt_time(
                                choice_timezone_save_range_end_time, 1)
                            if GL.report_data[d][0] == boundary_after_lower_limit_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                            if GL.report_data[d][0] == '999912312359' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                            enter_start_time = "0001/01/01 00:00"
                            before_save_time_range_end_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                                choice_timezone_save_range_start_time, -1))
                            range_start_time = str_time_to_datetime_time(enter_start_time)
                            range_end_time = str_time_to_datetime_time(before_save_time_range_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                            enter_end_time = "9999/12/31 23:59"
                            after_save_time_range_start_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                                choice_timezone_save_range_end_time, 1))
                            range_start_time = str_time_to_datetime_time(after_save_time_range_start_time)
                            range_end_time = str_time_to_datetime_time(enter_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                    elif TEST_CASE_INFO[5] == "Expired":
                        if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                            if GL.report_data[d][0] == calculate_str_time_to_fmt_time(
                                    choice_timezone_save_range_start_time, 0) \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                            str_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            if GL.report_data[d][0] == str_expected_res_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                            # default_save_range_start_time = "2000/01/01 00:00"
                            fmt_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            str_expected_res_time = fmt_time_to_str_time(fmt_expected_res_time)
                            range_start_time = str_time_to_datetime_time(choice_timezone_save_range_start_time)
                            range_end_time = str_time_to_datetime_time(str_expected_res_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                elif TEST_CASE_INFO[8] == "OtherTimezone" and TEST_CASE_INFO[9] == "Summertime":
                    summertime_choice_timezone_save_range_start_time = fmt_time_to_str_time(
                        calculate_str_time_to_fmt_time(choice_timezone_save_range_start_time, 60))
                    summertime_choice_timezone_save_range_end_time = fmt_time_to_str_time(
                        calculate_str_time_to_fmt_time(choice_timezone_save_range_end_time, 60))
                    if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                        if TEST_CASE_INFO[7] == "Boundary_before_lower_limit":
                            if GL.report_data[d][0] == '000101010000' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Boundary_before_upper_limit":
                            boundary_before_upper_limit_time = calculate_str_time_to_fmt_time(
                                summertime_choice_timezone_save_range_start_time, -1)
                            if GL.report_data[d][0] == boundary_before_upper_limit_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Boundary_after_lower_limit":
                            boundary_after_lower_limit_time = calculate_str_time_to_fmt_time(
                                summertime_choice_timezone_save_range_end_time, 1)
                            if GL.report_data[d][0] == boundary_after_lower_limit_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Boundary_after_upper_limit":
                            if GL.report_data[d][0] == '999912312359' \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Random_before_save_time_range":
                            enter_start_time = "0001/01/01 00:00"
                            before_save_time_range_end_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                                summertime_choice_timezone_save_range_start_time, -1))
                            range_start_time = str_time_to_datetime_time(enter_start_time)
                            range_end_time = str_time_to_datetime_time(before_save_time_range_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                        elif TEST_CASE_INFO[7] == "Random_after_save_time_range":
                            enter_end_time = "9999/12/31 23:59"
                            after_save_time_range_start_time = fmt_time_to_str_time(calculate_str_time_to_fmt_time(
                                summertime_choice_timezone_save_range_end_time, 1))
                            range_start_time = str_time_to_datetime_time(after_save_time_range_start_time)
                            range_end_time = str_time_to_datetime_time(enter_end_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                    elif TEST_CASE_INFO[5] == "Expired":
                        if TEST_CASE_INFO[7] == "Boundary_lower_limit":
                            if GL.report_data[d][0] == calculate_str_time_to_fmt_time(
                                    summertime_choice_timezone_save_range_start_time, 0) \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Boundary_upper_limit":
                            str_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            if GL.report_data[d][0] == str_expected_res_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        elif TEST_CASE_INFO[7] == "Random_expired_time_range":
                            # default_save_range_start_time = "2000/01/01 00:00"
                            fmt_expected_res_time = calculate_str_time_to_fmt_time(GL.report_data[1], -1)
                            str_expected_res_time = fmt_time_to_str_time(fmt_expected_res_time)
                            range_start_time = str_time_to_datetime_time(
                                summertime_choice_timezone_save_range_start_time)
                            range_end_time = str_time_to_datetime_time(str_expected_res_time)
                            event_start_time = fmt_time_to_str_time(GL.report_data[d][0])
                            datetime_event_start_time = str_time_to_datetime_time(event_start_time)
                            if range_start_time <= datetime_event_start_time <= range_end_time \
                                    and GL.report_data[d][1] == TEST_CASE_INFO[3] \
                                    and GL.report_data[d][4] == TEST_CASE_INFO[4]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

            else:
                ws.cell(max_row + 1, d + 1).value = GL.report_data[d]
                ws.cell(max_row + 1, d + 1).alignment = alignment
                if d == 1:      # 系统时间
                    if TEST_CASE_INFO[5] == "Expired":
                        event_start_time = GL.report_data[4][0]
                        str_event_start_time = fmt_time_to_str_time(event_start_time)
                        datetime_event_start_time = str_time_to_datetime_time(str_event_start_time)
                        datetime_sys_time = str_time_to_datetime_time(GL.report_data[d])
                        logging.info(f"系统时间{datetime_sys_time}--事件时间{datetime_event_start_time}")
                        if datetime_sys_time > datetime_event_start_time:
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font
                    elif TEST_CASE_INFO[5] == "NowPlaying":
                        str_event_start_time = fmt_time_to_str_time(GL.report_data[4][0])
                        if GL.report_data[d][:16] == str_event_start_time:
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font
                    elif TEST_CASE_INFO[5] == "InvalidDuration":
                        fmt_event_start_time = calculate_str_time_to_fmt_time(GL.report_data[d], 5)
                        if TEST_CASE_INFO[4] == "Once":
                            if fmt_event_start_time == GL.report_data[4][0]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font
                        else:
                            if fmt_event_start_time[8:] == GL.report_data[4][0]:
                                ws.cell(max_row + 1, d + 1).font = blue_font
                            else:
                                ws.cell(max_row + 1, d + 1).font = red_font

                elif d == 5:    # 场景描述
                    if GL.report_data[d] == TEST_CASE_INFO[7]:
                        ws.cell(max_row + 1, d + 1).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + 1).font = red_font

                elif d == 6:    # 时区
                    if TEST_CASE_INFO[8] == "ZeroTimezone":
                        if GL.report_data[d] == "0":
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font
                    elif TEST_CASE_INFO[8] == "OtherTimezone":
                        if GL.report_data[d] == GL.choice_timezone:
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font

                elif d == 7:    # 预约事件列表框事件个数
                    if GL.report_data[d] == "0":
                        ws.cell(max_row + 1, d + 1).font = blue_font
                    else:
                        ws.cell(max_row + 1, d + 1).font = red_font

                elif d == 8:    # 无效事件提示信息
                    if TEST_CASE_INFO[5] == "OutOfSaveTimeRange":
                        if GL.report_data[d] == "[PTD]Res_invalid_date":
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font
                    else:
                        if GL.report_data[d] == "[PTD]Res_invalid_timer":
                            ws.cell(max_row + 1, d + 1).font = blue_font
                        else:
                            ws.cell(max_row + 1, d + 1).font = red_font

        ws.row_dimensions[(max_row + 1)].height = 40  # 设置每次执行的report预约事件信息的行高
        wb.save(file_path[1])


def before_cycle_test_clear_data_and_state():
    # 循环测试前，清理数据和状态变量
    logging.info("before_cycle_test_clear_data_and_state")
    state["clear_variate_state"] = True
    if TEST_CASE_INFO[6] == "EPG":
        GL.report_data = ['', '', '', '', [], '', '', '', '']
    elif TEST_CASE_INFO[6] == "Timer":
        GL.report_data = ['', '', '', '', [], '', '', '', '', '']
    GL.case_testing_times -= 1
    logging.info("循环测试，延时5秒")
    time.sleep(5)
    logging.info(f"剩余循环次数：{GL.case_testing_times}")

    if GL.case_testing_times < 1:
        clear_timer_setting_all_events()
        logging.info("程序结束")
        state["receive_loop_state"] = True  # 触发结束接收进程的状态


def check_event_numb():
    logging.info("check_event_numb")
    # 检查Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 获取已预约的事件信息，清除获取预约事件的list，并激活获取预约事件状态标志
    state["clear_res_event_list_state"] = True
    state["update_event_list_state"] = True
    # 进入定时器设置界面
    send_more_commds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        if TEST_CASE_INFO[6] == "EPG":
            GL.report_data[6] = rsv_kws["res_event_numb"]
        elif TEST_CASE_INFO[6] == "Timer":
            GL.report_data[7] = rsv_kws["res_event_numb"]
        state["res_event_numb_state"] = False
        # 获取预约事件的状态标志关闭
        state["update_event_list_state"] = False
    # 退回大画面
    exit_to_screen()


def mail(message):
    my_sender = 'wangrun@nationalchip.com'  # 发件人邮箱账号
    my_pass = 'b8iNRgDiPUfkUVLW'  # 发件人邮箱密码
    my_user = 'wangrun@nationalchip.com'  # 收件人邮箱账号，我这边发送给自己

    return_state = True
    try:
        msg = MIMEText(message, 'plain', 'utf-8')
        # msg['From'] = formataddr(["FromRunoob", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['From'] = Header("Auto_test", 'utf-8')
        msg['To'] = Header("ME", 'utf-8')
        # msg['To'] = formataddr(["FK", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
        msg['Subject'] = "自动化测试终止提醒"  # 邮件的主题，也可以说是标题

        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 关闭连接
    except smtplib.SMTPException:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
        return_state = False
    return return_state


def receive_serial_process(
        prs_data, infrared_send_cmd, rsv_kws, state, channel_info, rsv_info, ch_epg_info, receive_cmd_list):
    logging_info_setting()
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIME_SHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict((val, key) for key, val in rsv_key.items())

    switch_ch_kws = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    ch_info_kws = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name"]

    group_info_kws = [
        "Group_name",
        "Prog_total"
    ]

    check_epg_kws = [
        "[PTD]Program_epg_info=",
        "[PTD]EPG_event:event_time="
    ]

    epg_info_kws = [
        "Program_epg_info",
        "event_time",
        "event_name"
    ]

    sys_time_kws = [
        "[PTD]Time_mode=",
        "[PTD]Timezone=",
        "[PTD]Summertime="
    ]

    other_kws = [
        "[PTD]Infrared_key_values:",    # 获取红外接收关键字
    ]

    res_kws = [
        "[PTD]Time_mode=",              # 0     获取系统时间模式
        "[PTD]System_time=",            # 1     系统时间
        "[PTD]Res_event_numb=",         # 2     预约事件数量
        "[PTD]Res_event:",              # 3     预约事件信息
        "[PTD]Res_triggered:",          # 4     预约事件触发和当前响应事件的信息
        "[PTD]Res_confirm_jump",        # 5     预约事件确认跳转
        "[PTD]Res_cancel_jump",         # 6     预约事件取消跳转
        "[PTD]REC_start",               # 7     录制开始
        "[PTD]REC_end",                 # 8     录制结束
        "[PTD]No_storage_device",       # 9     没有存储设备
        "[PTD]No_enough_space",         # 10    没有足够的空间
        "[PTD]power_cut",               # 11    进入待机
        "[PTD]:switch totle cost",      # 12    开机解码成功
        "[PTD][HOTPLUG] PLUG_IN",       # 13    存储设备插入成功
        "[PTD]PVR_is_not_supported",    # 14    录制无信号、加锁节目、加密节目，跳出PVR is not supported!提示
        "[PTD]Current_sys_time",        # 15    预约事件触发时，检测触发时间信息
    ]

    event_invalid_msg = [
        "[PTD]Res_no_channel",
        "[PTD]Res_invalid_date",
        "[PTD]Res_invalid_timer"
    ]

    edit_event_kws = [
        "[PTD]Mode=",
        "[PTD]Type=",
        "[PTD]Start Date=",
        "[PTD]Start Time=",
        "[PTD]Duration=",
        "[PTD]Channel="
    ]

    infrared_rsv_cmd = []
    receive_serial = serial.Serial(prs_data["receive_serial_name"], 115200, timeout=1)

    while True:
        data = receive_serial.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            # data1 = data.decode("ISO-8859-1", "ignore")
            data1 = data.decode("utf-8", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_log_data_to_txt(prs_data["log_file_path"], data3)

            if state["clear_channel_info_state"]:
                each_ch_info = ['', '', '', '', '', '', '', '']
                del channel_info[:]
                channel_info.extend(each_ch_info)
                state["clear_channel_info_state"] = False

            if state["clear_ch_epg_info_state"]:
                each_epg_info = ['', '', '']
                del ch_epg_info[:]
                ch_epg_info.extend(each_epg_info)
                state["clear_ch_epg_info_state"] = False

            if other_kws[0] in data2:   # 红外接收打印
                rsv_cmd = re.split(":", data2)[-1]
                infrared_rsv_cmd.append(rsv_cmd)
                if rsv_cmd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_cmd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        infrared_send_cmd[-1], reverse_rsv_key[infrared_rsv_cmd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(infrared_send_cmd), len(infrared_rsv_cmd)))
                    receive_cmd_list.append(rsv_cmd)

            if res_kws[0] in data2:     # 获取系统时间模式（自动还是手动）
                state["sys_time_mode_state"] = True
                rsv_kws["sys_time_mode"] = re.split(r"=", data2)[-1]

            if res_kws[1] in data2:     # 获取当前系统时间
                state["current_sys_time_state"] = True
                rsv_kws["current_sys_time"] = re.split(r"=", data2)[-1]

            if res_kws[2] in data2:     # 获取预约事件数量
                state["res_event_numb_state"] = True
                rsv_kws["res_event_numb"] = re.split(r"=", data2)[-1]

            if switch_ch_kws[0] in data2:
                ch_info_split = re.split(r"[],]", data2)
                for i in range(len(ch_info_split)):
                    if ch_info_kws[0] in ch_info_split[i]:  # 提取频道号
                        channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if ch_info_kws[1] in ch_info_split[i]:  # 提取频道名称
                        channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if switch_ch_kws[1] in data2:
                flag_info_split = re.split(r"[],]", data2)
                for i in range(len(flag_info_split)):
                    if ch_info_kws[2] in flag_info_split[i]:  # 提取频道所属TP
                        channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if ch_info_kws[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if switch_ch_kws[3] in data2:
                group_info_split = re.split(r"[],]", data2)
                for i in range(len(group_info_split)):
                    if group_info_kws[0] in group_info_split[i]:  # 提取频道所属组别
                        rsv_info["prog_group_name"] = re.split(r"=", group_info_split[i])[-1]
                        channel_info[6] = rsv_info["prog_group_name"]
                    if group_info_kws[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        rsv_info["prog_group_total"] = re.split(r"=", group_info_split[i])[-1]

            if check_epg_kws[0] in data2:       # 判断节目是否存在EPG信息
                epg_info_split = re.split(r"]", data2)
                for i in range(len(epg_info_split)):
                    if epg_info_kws[0] in epg_info_split[i]:
                        rsv_info["epg_info_exist"] = re.split(r"=", epg_info_split[i])[-1]
                        channel_info[7] = rsv_info["epg_info_exist"]

            if check_epg_kws[1] in data2:
                epg_event_split = re.split(r"event_time=|,event_name=", data2)
                time_info_split = re.split(r"--", epg_event_split[1])
                ch_epg_info[0] = time_info_split[0]
                ch_epg_info[1] = time_info_split[1]
                ch_epg_info[2] = epg_event_split[-1]

            if edit_event_kws[0] in data2:          # 提取Mode参数
                rsv_kws["edit_event_focus_pos"] = "Mode"
                rsv_kws["edit_event_mode"] = re.split(r"=", data2)[-1]

            if edit_event_kws[1] in data2:          # 提取Type参数
                rsv_kws["edit_event_focus_pos"] = "Type"
                rsv_kws["edit_event_type"] = re.split(r"=", data2)[-1]

            if edit_event_kws[2] in data2:          # 提取Start Date参数
                rsv_kws["edit_event_focus_pos"] = "Start Date"
                rsv_kws["edit_event_date"] = re.split(r"=", data2)[-1]

            if edit_event_kws[3] in data2:          # 提取Start Time参数
                rsv_kws["edit_event_focus_pos"] = "Start Time"
                rsv_kws["edit_event_time"] = re.split(r"=", data2)[-1]

            if edit_event_kws[4] in data2:          # 提取Duration参数
                rsv_kws["edit_event_focus_pos"] = "Duration"
                rsv_kws["edit_event_duration"] = re.split(r"=", data2)[-1]

            if edit_event_kws[5] in data2:          # 提取Channel参数
                rsv_kws["edit_event_focus_pos"] = "Channel"
                rsv_kws["edit_event_ch"] = re.split(r"=", data2)[-1]

            if sys_time_kws[0] in data2:             # 提取System_mode参数
                rsv_kws["sys_time_setting_focus_pos"] = "Mode"
                rsv_kws["sys_time_mode"] = re.split(r"=", data2)[-1]

            if sys_time_kws[1] in data2:             # 提取Timezone参数
                rsv_kws["sys_time_setting_focus_pos"] = "Timezone"
                rsv_kws["sys_time_timezone"] = re.split(r"=", data2)[-1]

            if sys_time_kws[2] in data2:             # 提取Summertime参数
                rsv_kws["sys_time_setting_focus_pos"] = "Summertime"
                rsv_kws["sys_time_summertime"] = re.split(r"=", data2)[-1]

            if event_invalid_msg[0] in data2:
                # state["event_no_channel_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2

            if event_invalid_msg[1] in data2:
                # state["event_invalid_date_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2

            if event_invalid_msg[2] in data2:
                # state["event_invalid_timer_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2


if __name__ == "__main__":

    choice_case_numb = int(sys.argv[1])
    # # choice_case_numb = 0
    TEST_CASE_INFO = invalid_res_case[choice_case_numb]
    print(TEST_CASE_INFO)

    msg_info = ''
    if TEST_CASE_INFO[6] == "EPG":
        msg_info = "现在开始执行的是:{}_{}_{}_{}_{}_{}_{}".format(
            TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
            TEST_CASE_INFO[3], TEST_CASE_INFO[4], TEST_CASE_INFO[5], TEST_CASE_INFO[6])
    elif TEST_CASE_INFO[6] == "Timer":
        msg_info = "现在开始执行的是:{}_{}_{}_{}_{}_{}_{}_{}_{}_{}".format(
            TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
            TEST_CASE_INFO[3], TEST_CASE_INFO[4], TEST_CASE_INFO[5],
            TEST_CASE_INFO[6], TEST_CASE_INFO[7], TEST_CASE_INFO[8], TEST_CASE_INFO[9])
    try:
        # TEST_CASE_INFO = ["00", "GX", "TV", "Play", "Once", "Expired", "EPG", "epg_test_numb"]
        # TEST_CASE_INFO = ["01", "GX", "TV", "Play", "Once", "OutOfSaveTimeRange", "Timer",
        # "Boundary_before_upper_limit", "ZeroTimezone", "epg_test_numb"]

        GL = MyGlobal()
        logging_info_setting()

        logging.critical(format(msg_info, '*^150'))
        KEY = {
            "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
            "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
            "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
            "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
            "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
            "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
            "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
            "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
            "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17",
            "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
            "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
            "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
            "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
            "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49",
            "TIME_SHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D"
        }
        REVERSE_KEY = dict((val, key) for key, val in KEY.items())

        file_path = build_log_and_report_file_path()
        ser_name = list(check_ports())  # send_ser_name, receive_ser_name
        send_serial = serial.Serial(ser_name[0], 9600)
        receive_ser_name = ser_name[1]

        infrared_send_cmd = Manager().list([])
        receive_cmd_list = Manager().list([])
        # [频道号,频道名称,tp,lock,scramble,频道类型,组别,epg_info]
        channel_info = Manager().list(['', '', '', '', '', '', '', ''])
        ch_epg_info = Manager().list(['', '', ''])  # 单个EPG信息的提取[event_start_time, event_end_time, event_name]
        rsv_kws = Manager().dict({
            "sys_time_mode": '', "current_sys_time": '', "res_event_numb": '', "prog_group_name": '',
            "prog_group_total": '', "edit_event_focus_pos": '', "edit_event_mode": '', "edit_event_type": '',
            "edit_event_date": '', "edit_event_time": '', "edit_event_duration": '', "edit_event_ch": '',
            "res_triggered_sys_time": '', "event_invalid_msg": '', "sys_time_setting_focus_pos": '',
            "sys_time_timezone": '', "sys_time_summertime": ''
        })

        rsv_info = Manager().dict({
            "prog_group_name": '', "prog_group_total": '', "epg_info_exist": '', "sys_time_mode": '',
        })

        state = Manager().dict({
            "receive_loop_state": False, "sys_time_mode_state": False, "clear_channel_info_state": False,
            "send_commd_state": True, "clear_ch_epg_info_state": False, "send_left_cmd_state": False,
            "send_right_cmd_state": False, "res_event_numb_state": False,
        })
        prs_data = Manager().dict({
            "log_file_path": file_path[0], "receive_serial_name": receive_ser_name,
        })

        rsv_p = Process(target=receive_serial_process, args=(
            prs_data, infrared_send_cmd, rsv_kws, state, channel_info, rsv_info, ch_epg_info, receive_cmd_list))
        rsv_p.start()

        if platform.system() == "Windows":
            time.sleep(5)
            logging.info("Windows系统接收端响应慢，等待5秒")
        elif platform.system() == "Linux":
            time.sleep(1)
            logging.info("Linux系统接收端响应快，但是增加一个延时保护，等待1秒")

        # 主程序开始部分
        if TEST_CASE_INFO[6] == "EPG":
            clear_timer_setting_all_events()
            check_sys_time_auto_mode()
            set_timezone_and_summertime()
            get_group_channel_total_info()
            check_epg_info_already_show()
            check_sys_time_manual_mode()
            get_choice_group_ch_type()

            while GL.case_testing_times > 0:
                choice_test_channel()
                check_epg_info_already_show()
                get_sys_time_info()
                check_preparatory_work()
                send_test_case_commd()
                check_event_numb()
                padding_report_data()
                write_data_to_report()
                before_cycle_test_clear_data_and_state()

        elif TEST_CASE_INFO[6] == "Timer":
            clear_timer_setting_all_events()
            check_sys_time_auto_mode()
            set_timezone_and_summertime()
            get_group_channel_total_info()
            check_epg_info_already_show()
            check_sys_time_manual_mode()
            get_choice_group_ch_type()
            while GL.case_testing_times > 0:
                choice_test_channel()
                check_preparatory_work()
                send_test_case_commd()
                check_event_numb()
                padding_report_data()
                write_data_to_report()
                before_cycle_test_clear_data_and_state()

        if state["receive_loop_state"]:
            rsv_p.terminate()
            logging.info("程序结束")
            logging.info('stop receive process')
            rsv_p.join()

    except Exception as e:
        print(e)
        # cur_py_file_name = sys.argv[0]        # 第0个就是这个python文件本身的路径（全路径）
        cur_py_file_name = os.path.basename(__file__)  # 当前文件名名称
        ret = mail(f"{cur_py_file_name}\n\n"
                   f"{msg_info}\n\n"
                   f"{traceback.format_exc()}")
        if ret:
            print("邮件发送成功")
        else:
            print("邮件发送失败")

        print("***traceback.format_exc():*** ")
        print(traceback.format_exc())
