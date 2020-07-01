#!/usr/bin/python3
# -*- coding: utf-8 -*-

from serial_setting import *
from multiprocessing import Process, Manager
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import RED, BLUE
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from random import sample, uniform, randint
import platform
import os
import time
import logging
import re
import sys

choice_case_numb = int(sys.argv[1])
TEST_CASE_INFO = epg_basic_case[choice_case_numb]
print(TEST_CASE_INFO)
# TEST_CASE_INFO = ["40", "GX", "Radio", "RIGHT+Random+OK", "SwitchAndDetailEPGEvent"]


class MyGlobal(object):

    def __init__(self):
        if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
            self.send_test_case_commd_numb = 10                     # 进入退出EPG界面的次数
        elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
            self.send_test_case_commd_numb = 15                     # EPG界面切台执行次数
        elif TEST_CASE_INFO[4] == "SwitchEPGEvent":
            self.send_test_case_commd_numb = 15                     # EPG界面切换事件次数
            self.each_ch_test_numb = 15                             # 每个节目的测试次数，一般与发送测试case次数保持一致

        self.TV_channel_groups = {}                                 # 存放电视节目的组别和节目数信息
        self.Radio_channel_groups = {}                              # 存放广播节目的组别和节目数信息
        self.TV_ch_attribute = [[], [], [], []]                     # 用于存放TV节目属性的列表(免费\加密\加锁\所有有EPG的节目)
        self.Radio_ch_attribute = [[], [], [], []]                  # 用于存放Radio节目属性的列表(免费\加密\加锁\所有有EPG的节目)
        self.all_ch_epg_info = {}                                   # 所有有EPG信息的节目的事件管理
        self.judge_switch_epg_info_end = []                         # 用于判断对比当前事件是否已经切换完成一个周期
        self.actual_test_numb = 0                                   # 实际执行用例指令次数
        self.epg_switch_ch_data_report = []                         # 用于EPG界面切台时的节目信息记录
        self.send_cmd = ''                                          # 用于记录和报告发送指令
        self.report_interval = 0                                    # 用于计算写Excel时，多节目的event信息的间隔计算
        if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
            self.expect_report_data = ['', '', '', '', '', '', '']  # 用于期望输出报告的数据管理
            self.actual_report_data = ['', '', '', '', '', '']      # 用于实际输出报告的数据管理
        elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
            self.expect_report_data = ['', '', '', '', '', '']      # 用于期望输出报告的数据管理
            self.actual_report_data = ['', '', '', '', '']          # 用于实际输出报告的数据管理
        elif TEST_CASE_INFO[4] == "SwitchEPGEvent" or TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
            self.expect_report_data = ['', '', '', '', '', '']      # 用于期望输出报告的数据管理
            self.actual_report_data = ['', '', '', '', '']          # 用于实际输出报告的数据管理


def logging_info_setting():
    # 配置logging输出格式
    LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)


def hex_strs_to_bytes(strings):
    # 将红外命令字符串转换为字节串
    return bytes.fromhex(strings)


def write_log_data_to_txt(path, write_data):
    with open(path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def send_commd(commd):
    global receive_cmd_list, infrared_send_cmd
    # 红外发送端发送指令
    if len(infrared_send_cmd) == len(receive_cmd_list):
        send_serial.write(hex_strs_to_bytes(commd))
        send_serial.flush()
        logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
        if REVERSE_KEY[commd] != "POWER":
            infrared_send_cmd.append(REVERSE_KEY[commd])
        time.sleep(1.0)
    elif len(infrared_send_cmd) != len(receive_cmd_list):
        if len(infrared_send_cmd) - len(receive_cmd_list) == 1:
            logging.info(f"此刻补发STB没有接收到的红外命令{infrared_send_cmd[-1]}")
            send_serial.write(hex_strs_to_bytes(KEY[infrared_send_cmd[-1]]))
            send_serial.flush()
            time.sleep(1.0)

            logging.info(f"此时再发送本次要发送的命令{REVERSE_KEY[commd]}")
            send_commd(commd)


def send_random_commd(commd):
    # 红外发送端发送指令
    send_serial.write(hex_strs_to_bytes(commd))
    send_serial.flush()
    logging.info("红外发送：{}".format(REVERSE_KEY[commd]))
    infrared_send_cmd.append(REVERSE_KEY[commd])
    time.sleep(0.5)


def send_more_commds(commd_list):
    # 用于发送一连串的指令
    for commd in commd_list:
        send_commd(commd)
    time.sleep(1)   # 增加函数切换时的的等待，避免可能出现send_commd函数中的等待时间没有执行的情况


def build_log_and_report_file_path():
    # 用于创建打印和报告文件路径
    # 构建存放数据的总目录，以及构建存放打印和报告的目录
    parent_path = os.path.dirname(os.getcwd())
    test_data_directory_name = "test_data"
    test_data_directory_path = os.path.join(parent_path, test_data_directory_name)
    log_directory_name = "print_log"
    log_directory_path = os.path.join(test_data_directory_path, log_directory_name)
    report_directory_name = "report"
    report_directory_path = os.path.join(test_data_directory_path, report_directory_name)
    # 判断目录是否存在，否则创建目录
    if not os.path.exists(test_data_directory_path):
        os.mkdir(test_data_directory_path)
    if not os.path.exists(log_directory_path):
        os.mkdir(log_directory_path)
    if not os.path.exists(report_directory_path):
        os.mkdir(report_directory_path)
    # 创建打印和报告文件的名称和路径
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    fmt_name = "{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
        TEST_CASE_INFO[3], TEST_CASE_INFO[4])
    log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    log_file_path = os.path.join(log_directory_path, log_file_name)
    report_file_name = "{}_{}.xlsx".format(fmt_name, time_info)
    report_file_path = os.path.join(report_directory_path, report_file_name)
    sheet_name = "{}".format(TEST_CASE_INFO[4])
    return log_file_path, report_file_path, sheet_name


def change_numbs_to_commds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_commds_list = []
    for i in range(len(numbs_list)):
        channel_commds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_commds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_commds_list[i].append(KEY[numbs_list[i][j]])
    return channel_commds_list


def exit_to_screen():
    send_data = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]
    send_more_commds(send_data)


def check_sys_time_mode():
    logging.debug("check_sys_time_mode")
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    # 进入时间设置界面
    send_more_commds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        if rsv_info["sys_time_mode"] == "Manual":
            send_more_commds(change_sys_time_mode)
        elif rsv_info["sys_time_mode"] == "Auto":
            logging.info("系统时间模式已经为自动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    # 退回大画面
    exit_to_screen()


def get_group_channel_total_info():
    logging.debug("get_group_channel_total_info")
    # 切台前获取case节目类别,分组,分组节目数量,以及获取节目属性前的去除加锁的判断
    # 根据所选case切换到对应类型节目的界面
    while channel_info[5] != TEST_CASE_INFO[2]:
        send_commd(KEY["TV/R"])
        if channel_info[3] == "1":
            send_commd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_commd(KEY["OK"])
    # 采集所有分组的名称和分组下节目总数信息
    if TEST_CASE_INFO[2] == "TV":
        while rsv_info["prog_group_name"] not in GL.TV_channel_groups.keys():
            print(rsv_info["prog_group_name"])
            GL.TV_channel_groups[rsv_info["prog_group_name"]] = rsv_info["prog_group_total"]
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE_INFO[1] not in GL.TV_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的电视分组：{}，退出程序".format(TEST_CASE_INFO[1]))
            send_commd(KEY["EXIT"])
            state["receive_loop_state"] = True
    elif TEST_CASE_INFO[2] == "Radio":
        while rsv_info["prog_group_name"] not in GL.Radio_channel_groups.keys():
            GL.Radio_channel_groups[rsv_info["prog_group_name"]] = rsv_info["prog_group_total"]
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
        if TEST_CASE_INFO[1] not in GL.Radio_channel_groups.keys():  # 用于判断当前测试用例的分组存不存在
            logging.debug("不存在当前case指定的广播分组：{}，退出程序".format(TEST_CASE_INFO[1]))
            send_commd(KEY["EXIT"])
            state["receive_loop_state"] = True
    # 根据所选case切换到对应的分组
    if TEST_CASE_INFO[2] == "TV":
        while rsv_info["prog_group_name"] != TEST_CASE_INFO[1]:
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    elif TEST_CASE_INFO[2] == "Radio":
        while rsv_info["prog_group_name"] != TEST_CASE_INFO[1]:
            send_commd(KEY["RIGHT"])
            if channel_info[3] == "1":
                send_commd(KEY["EXIT"])
    # 退出频道列表,回到大画面界面
    send_commd(KEY["EXIT"])
    logging.debug(channel_info)
    logging.debug(GL.TV_channel_groups)
    logging.debug(GL.Radio_channel_groups)


def get_choice_group_ch_type():
    logging.debug("get_choice_group_ch_type")
    global channel_info
    # 采集All分组下的节目属性和是否有EPG信息
    choice_group_ch_total_numb = ''
    if TEST_CASE_INFO[2] == "TV":
        choice_group_ch_total_numb = GL.TV_channel_groups[TEST_CASE_INFO[1]]
    elif TEST_CASE_INFO[2] == "Radio":
        choice_group_ch_total_numb = GL.Radio_channel_groups[TEST_CASE_INFO[1]]
    # 进入EPG界面，切台获取指定分组下所有节目的属性
    send_commd(KEY["EPG"])
    logging.debug(rsv_info["prog_group_name"])
    logging.debug(channel_info)
    for i in range(int(choice_group_ch_total_numb)):
        # channel_info = ['', '', '', '', '', '', rsv_info["prog_group_name"], '']
        send_commd(KEY["DOWN"])
        if channel_info[7] == "1":
            time.sleep(0.5)
        elif channel_info[7] == "0" or channel_info[7] == '':
            time.sleep(2.5)
        # time.sleep(1)
        if channel_info[3] == "1":
            for j in range(4):
                send_commd(KEY["0"])
        if TEST_CASE_INFO[2] == "TV":
            if channel_info[7] == "1":  # 所有有EPG信息的电视节目
                GL.TV_ch_attribute[3].append(channel_info[0])
            if channel_info[3] == "1":  # 加锁电视节目
                GL.TV_ch_attribute[2].append(channel_info[0])
            elif channel_info[4] == "0":  # 免费电视节目
                GL.TV_ch_attribute[0].append(channel_info[0])
            elif channel_info[4] == "1":  # 加密电视节目
                GL.TV_ch_attribute[1].append(channel_info[0])
        elif TEST_CASE_INFO[2] == "Radio":
            if channel_info[7] == "1":  # 所有有EPG信息的广播节目
                GL.Radio_ch_attribute[3].append(channel_info[0])
            if channel_info[3] == "1":  # 加锁广播节目
                GL.Radio_ch_attribute[2].append(channel_info[0])
            elif channel_info[4] == "0":  # 免费广播节目
                GL.Radio_ch_attribute[0].append(channel_info[0])
            elif channel_info[4] == "1":  # 加密广播节目
                GL.Radio_ch_attribute[1].append(channel_info[0])
        logging.info(channel_info)
    logging.info(GL.TV_ch_attribute)
    logging.info(GL.Radio_ch_attribute)
    # 退回大画面
    exit_to_screen()


def choice_test_channel():
    logging.debug("choice_test_channel")
    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        if TEST_CASE_INFO[2] == "TV":
            if len(GL.TV_ch_attribute[3]) == 0:
                logging.info("没有免费且有EPG信息的电视节目")
                send_commd(KEY["EXIT"])
                state["receive_loop_state"] = True
            elif len(GL.TV_ch_attribute[3]) > 0:
                free_tv_numb = sample(GL.TV_ch_attribute[3], 1)
                logging.debug("所选免费且有EPG信息的电视节目为:{}".format(free_tv_numb))
                free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                send_commd(KEY["EXIT"])
                for i in range(len(free_tv_commd)):
                    for j in range(len(free_tv_commd[i])):
                        send_commd(free_tv_commd[i][j])
                send_commd(KEY["OK"])
                time.sleep(2)
                logging.info("所选免费且有EPG信息的电视节目名称为:{}".format(channel_info[1]))
                logging.info(channel_info)
        elif TEST_CASE_INFO[2] == "Radio":
            if len(GL.Radio_ch_attribute[3]) == 0:
                logging.info("没有免费且有EPG信息的广播节目")
                send_commd(KEY["EXIT"])
                state["receive_loop_state"] = True
            elif len(GL.Radio_ch_attribute[3]) > 0:
                free_radio_numb = sample(GL.Radio_ch_attribute[3], 1)
                logging.debug("所选免费且有EPG信息的广播节目为:{}".format(free_radio_numb))
                free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                send_commd(KEY["EXIT"])
                for i in range(len(free_radio_commd)):
                    for j in range(len(free_radio_commd[i])):
                        send_commd(free_radio_commd[i][j])
                send_commd(KEY["OK"])
                time.sleep(2)
                logging.info("所选免费且有EPG信息的广播节目名称为:{}".format(channel_info[1]))
                logging.info(channel_info)

    elif TEST_CASE_INFO[4] == "SwitchEPGEvent":
        if TEST_CASE_INFO[2] == "TV":
            if len(GL.TV_ch_attribute[3]) == 0:
                logging.info("无有EPG信息的电视节目")
            elif len(GL.TV_ch_attribute[3]) > 0:
                for i in range(len(GL.TV_ch_attribute[3])):
                    free_tv_numb = GL.TV_ch_attribute[3][i]
                    logging.debug("当前所选有EPG信息的电视节目频道号为:{}".format(free_tv_numb))
                    free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                    send_commd(KEY["EXIT"])
                    for j in range(len(free_tv_commd)):
                        for k in range(len(free_tv_commd[j])):
                            send_commd(free_tv_commd[j][k])
                    send_commd(KEY["OK"])
                    time.sleep(2)
                    logging.info("当前所选有EPG信息的电视节目名称为:{}".format(channel_info[1]))
                    logging.info(channel_info)

                    # 将有EPG信息的节目名称添加到字典
                    if channel_info[1] not in GL.all_ch_epg_info.keys():
                        GL.all_ch_epg_info[channel_info[1]] = []
                        logging.info(GL.all_ch_epg_info.keys())

                    check_preparatory_work()
                    check_epg_info_already_show()
                    time.sleep(1)
                    send_test_case_commd()
                    padding_report_data()
                    write_data_to_report()
                    send_commd(KEY["EXIT"])
        elif TEST_CASE_INFO[2] == "Radio":
            if len(GL.Radio_ch_attribute[3]) == 0:
                logging.info("无有EPG信息的广播节目")
            elif len(GL.Radio_ch_attribute[3]) > 0:
                for i in range(len(GL.Radio_ch_attribute[3])):
                    free_radio_numb = GL.Radio_ch_attribute[3][i]
                    logging.debug("当前所选有EPG信息的广播节目频道号为:{}".format(free_radio_numb))
                    free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                    send_commd(KEY["EXIT"])
                    for j in range(len(free_radio_commd)):
                        for k in range(len(free_radio_commd[j])):
                            send_commd(free_radio_commd[j][k])
                    send_commd(KEY["OK"])
                    time.sleep(2)
                    logging.info("当前所选有EPG信息的广播节目名称为:{}".format(channel_info[1]))
                    logging.info(channel_info)

                    # 将有EPG信息的节目名称添加到字典
                    if channel_info[1] not in GL.all_ch_epg_info.keys():
                        GL.all_ch_epg_info[channel_info[1]] = []
                        logging.info(GL.all_ch_epg_info.keys())

                    check_preparatory_work()
                    check_epg_info_already_show()
                    time.sleep(1)
                    send_test_case_commd()
                    padding_report_data()
                    write_data_to_report()
                    send_commd(KEY["EXIT"])

    elif TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
        if TEST_CASE_INFO[2] == "TV":
            if len(GL.TV_ch_attribute[3]) == 0:
                logging.info("无有EPG信息的电视节目")
            elif len(GL.TV_ch_attribute[3]) > 0:
                for i in range(len(GL.TV_ch_attribute[3])):
                    free_tv_numb = GL.TV_ch_attribute[3][i]
                    logging.debug("当前所选有EPG信息的电视节目频道号为:{}".format(free_tv_numb))
                    free_tv_commd = change_numbs_to_commds_list(free_tv_numb)
                    send_commd(KEY["EXIT"])
                    for j in range(len(free_tv_commd)):
                        for k in range(len(free_tv_commd[j])):
                            send_commd(free_tv_commd[j][k])
                    send_commd(KEY["OK"])
                    time.sleep(2)
                    logging.info("当前所选有EPG信息的电视节目名称为:{}".format(channel_info[1]))
                    logging.info(channel_info)

                    # 将有EPG信息的节目名称添加到字典
                    if channel_info[1] not in GL.all_ch_epg_info.keys():
                        GL.all_ch_epg_info[channel_info[1]] = []
                        logging.info(GL.all_ch_epg_info.keys())

                    check_preparatory_work()
                    check_epg_info_already_show()
                    time.sleep(1)
                    send_test_case_commd()
                    padding_report_data()
                    write_data_to_report()
                    send_commd(KEY["EXIT"])
        elif TEST_CASE_INFO[2] == "Radio":
            if len(GL.Radio_ch_attribute[3]) == 0:
                logging.info("无有EPG信息的广播节目")
            elif len(GL.Radio_ch_attribute[3]) > 0:
                for i in range(len(GL.Radio_ch_attribute[3])):
                    free_radio_numb = GL.Radio_ch_attribute[3][i]
                    logging.debug("当前所选有EPG信息的广播节目频道号为:{}".format(free_radio_numb))
                    free_radio_commd = change_numbs_to_commds_list(free_radio_numb)
                    send_commd(KEY["EXIT"])
                    for j in range(len(free_radio_commd)):
                        for k in range(len(free_radio_commd[j])):
                            send_commd(free_radio_commd[j][k])
                    send_commd(KEY["OK"])
                    time.sleep(2)
                    logging.info("当前所选有EPG信息的广播节目名称为:{}".format(channel_info[1]))
                    logging.info(channel_info)

                    # 将有EPG信息的节目名称添加到字典
                    if channel_info[1] not in GL.all_ch_epg_info.keys():
                        GL.all_ch_epg_info[channel_info[1]] = []
                        logging.info(GL.all_ch_epg_info.keys())

                    check_preparatory_work()
                    check_epg_info_already_show()
                    time.sleep(1)
                    send_test_case_commd()
                    padding_report_data()
                    write_data_to_report()
                    send_commd(KEY["EXIT"])


def check_preparatory_work():
    logging.debug("check_preparatory_work")
    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        pass
    else:
        send_commd(KEY["EPG"])
        if channel_info[3] == "1":
            for i in range(4):
                send_commd(KEY["0"])


def check_epg_info_already_show():      # 检查EPG信息是否已经显示
    logging.debug("check_epg_info_already_show")
    global ch_epg_info
    while ch_epg_info[-1] == '':         # 假如还没有获取到当前节目的EPG信息，则需要退出等待5秒再进入
        # ch_epg_info = ['', '', '']
        send_commd(KEY["EXIT"])
        time.sleep(5)
        send_commd(KEY["EPG"])
        send_commd(KEY["RIGHT"])
    # ch_epg_info = ['', '', '']
    send_commd(KEY["EXIT"])
    send_commd(KEY["EPG"])


def send_test_case_commd():
    logging.debug("send_test_case_commd")
    global channel_info, ch_epg_info
    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        send_data = [KEY["EPG"], KEY["EXIT"]]
        for i in range(GL.send_test_case_commd_numb):
            send_commd(send_data[0])
            sleep_time = uniform(0.1, 0.8)
            logging.info(sleep_time)
            time.sleep(sleep_time)
            send_commd(send_data[1])
            logging.info("当前执行次数:{}".format(i + 1))
            GL.actual_test_numb = i + 1

    elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
        cmd_set_list = [KEY["UP"], KEY["DOWN"]]
        if TEST_CASE_INFO[3] == "UP" or TEST_CASE_INFO[3] == "UP+Random":
            GL.send_cmd = cmd_set_list[0]
            for i in range(GL.send_test_case_commd_numb):
                each_ch_info = []
                if TEST_CASE_INFO[3] == "UP":
                    send_commd(GL.send_cmd)
                elif TEST_CASE_INFO[3] == "UP+Random":
                    up_key_random_numb = randint(1, 10)     # 每次执行都需要取随即值，不能提到外面
                    for n in range(up_key_random_numb):
                        send_random_commd(GL.send_cmd)
                sleep_time = uniform(0.5, 1.5)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if channel_info[3] == "1":
                    for j in range(4):
                        send_commd(KEY["0"])
                logging.info("当前执行次数:{}".format(i + 1))
                each_ch_info.extend(channel_info)
                GL.epg_switch_ch_data_report.append(each_ch_info)
                logging.info(channel_info)
                GL.actual_test_numb = i + 1
        elif TEST_CASE_INFO[3] == "DOWN" or TEST_CASE_INFO[3] == "DOWN+Random":
            GL.send_cmd = cmd_set_list[1]
            for i in range(GL.send_test_case_commd_numb):
                each_ch_info = []
                if TEST_CASE_INFO[3] == "DOWN":
                    send_commd(GL.send_cmd)
                elif TEST_CASE_INFO[3] == "DOWN+Random":
                    down_key_random_numb = randint(1, 10)       # 每次执行都需要取随即值，不能提到外面
                    for n in range(down_key_random_numb):
                        send_random_commd(GL.send_cmd)
                sleep_time = uniform(0.5, 1.5)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if channel_info[3] == "1":
                    for j in range(4):
                        send_commd(KEY["0"])
                logging.info("当前执行次数:{}".format(i + 1))
                each_ch_info.extend(channel_info)
                GL.epg_switch_ch_data_report.append(each_ch_info)
                logging.info(channel_info)
                GL.actual_test_numb = i + 1
        elif TEST_CASE_INFO[3] == "UP+DOWN+Random":
            GL.send_cmd = cmd_set_list
            for i in range(GL.send_test_case_commd_numb):
                each_ch_info = []
                up_key_random_numb = randint(1, 10)         # 每次执行都需要取随即值，不能提到外面
                for n in range(up_key_random_numb):
                    send_random_commd(GL.send_cmd[0])
                down_key_random_numb = randint(1, 10)       # 每次执行都需要取随即值，不能提到外面
                for n in range(down_key_random_numb):
                    send_random_commd(GL.send_cmd[1])
                sleep_time = uniform(0.5, 1.5)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if channel_info[3] == "1":
                    for j in range(4):
                        send_commd(KEY["0"])
                logging.info("当前执行次数:{}".format(i + 1))
                each_ch_info.extend(channel_info)
                GL.epg_switch_ch_data_report.append(each_ch_info)
                logging.info(channel_info)
                GL.actual_test_numb = i + 1

    elif TEST_CASE_INFO[4] == "SwitchEPGEvent":
        state["send_commd_state"] = True        # 不同节目完成一轮的事件切换后，状态被关闭，会导致下面的节目不会进行切换事件
        cmd_set_list = [KEY["GREEN"], KEY["YELLOW"], KEY["LEFT"], KEY["RIGHT"]]
        if TEST_CASE_INFO[3] == "LEFT":
            GL.send_cmd = cmd_set_list[2:][0]
            while state["send_commd_state"]:
                send_commd(GL.send_cmd)
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) not in GL.all_ch_epg_info[channel_info[1]] and list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                if len(GL.all_ch_epg_info[channel_info[1]]) >= 20:
                    GL.judge_switch_epg_info_end = GL.all_ch_epg_info[channel_info[1]][:10]
                    if list(ch_epg_info) in GL.judge_switch_epg_info_end:
                        state["send_commd_state"] = False
                        GL.judge_switch_epg_info_end = []
        elif TEST_CASE_INFO[3] == "RIGHT":
            GL.send_cmd = cmd_set_list[2:][1]
            while state["send_commd_state"]:
                send_commd(GL.send_cmd)
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) not in GL.all_ch_epg_info[channel_info[1]] and list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                if len(GL.all_ch_epg_info[channel_info[1]]) >= 20:
                    GL.judge_switch_epg_info_end = GL.all_ch_epg_info[channel_info[1]][:10]
                    if list(ch_epg_info) in GL.judge_switch_epg_info_end:
                        state["send_commd_state"] = False
                        GL.judge_switch_epg_info_end = []

        elif TEST_CASE_INFO[3] == "LEFT+Random":
            GL.send_cmd = cmd_set_list[2:][0]
            while state["send_commd_state"]:
                random_send_numb = randint(1, 20)
                for i in range(random_send_numb):
                    send_random_commd(GL.send_cmd)
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "RIGHT+Random":
            GL.send_cmd = cmd_set_list[2:][1]
            while state["send_commd_state"]:
                random_send_numb = randint(1, 20)
                for i in range(random_send_numb):
                    send_random_commd(GL.send_cmd)
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "RIGHT+LEFT+Random":
            GL.send_cmd = cmd_set_list[2:]
            while state["send_commd_state"]:
                right_random_send_numb = randint(1, 20)
                for i in range(right_random_send_numb):
                    send_random_commd(GL.send_cmd[1])
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, right_random_send_numb))
                left_random_send_numb = randint(1, 20)
                for i in range(left_random_send_numb):
                    send_random_commd(GL.send_cmd[0])
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, left_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day-":
            GL.send_cmd = cmd_set_list[:2][0]
            while state["send_commd_state"]:
                send_commd(GL.send_cmd)
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day+":
            GL.send_cmd = cmd_set_list[:2][1]
            while state["send_commd_state"]:
                send_commd(GL.send_cmd)
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day++LEFT+Random":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[1])     # Day+
            GL.send_cmd.append(cmd_set_list[2])     # LEFT
            while state["send_commd_state"]:
                send_commd(GL.send_cmd[0])              # 发送Day+
                left_random_send_numb = randint(1, 20)
                for i in range(left_random_send_numb):
                    send_random_commd(GL.send_cmd[1])          # 发送LEFT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, left_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day++RIGHT+Random":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[1])     # Day+
            GL.send_cmd.append(cmd_set_list[3])     # RIGHT
            while state["send_commd_state"]:
                send_commd(GL.send_cmd[0])              # 发送Day+
                right_random_send_numb = randint(1, 20)
                for i in range(right_random_send_numb):
                    send_random_commd(GL.send_cmd[1])          # 发送RIGHT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, right_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day++LEFTorRIGHT+Random":
            GL.send_cmd = cmd_set_list[1]
            lest_or_right_random_commd = [KEY["LEFT"], KEY["RIGHT"]]
            while state["send_commd_state"]:
                send_commd(GL.send_cmd)              # 发送Day+
                left_or_right_random_send_numb = randint(1, 20)
                random_send_data = sample(lest_or_right_random_commd, 1)[0]
                for i in range(left_or_right_random_send_numb):
                    send_random_commd(random_send_data)      # 随机发送LEFT或者RIGHT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, left_or_right_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day-+LEFT+Random":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[0])     # Day-
            GL.send_cmd.append(cmd_set_list[2])     # LEFT
            while state["send_commd_state"]:
                send_commd(GL.send_cmd[0])              # 发送Day-
                left_random_send_numb = randint(1, 20)
                for i in range(left_random_send_numb):
                    send_random_commd(GL.send_cmd[1])          # 发送LEFT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, left_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day-+RIGHT+Random":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[0])     # Day-
            GL.send_cmd.append(cmd_set_list[3])     # RIGHT
            while state["send_commd_state"]:
                send_commd(GL.send_cmd[0])              # 发送Day-
                right_random_send_numb = randint(1, 20)
                for i in range(right_random_send_numb):
                    send_random_commd(GL.send_cmd[1])          # 发送RIGHT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, right_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day-+LEFTorRIGHT+Random":
            GL.send_cmd = cmd_set_list[0]
            lest_or_right_random_commd = [KEY["LEFT"], KEY["RIGHT"]]
            while state["send_commd_state"]:
                send_commd(GL.send_cmd)                 # 发送Day-
                left_or_right_random_send_numb = randint(1, 20)
                random_send_data = sample(lest_or_right_random_commd, 1)[0]
                for i in range(left_or_right_random_send_numb):
                    send_random_commd(random_send_data)        # 随机发送LEFT或者RIGHT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, left_or_right_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

        elif TEST_CASE_INFO[3] == "Day+orDay-orLEFTorRIGHT+Random":
            while state["send_commd_state"]:
                GL.send_cmd = sample(cmd_set_list, 1)[0]
                left_or_right_random_send_numb = randint(1, 20)
                for i in range(left_or_right_random_send_numb):
                    send_random_commd(GL.send_cmd)        # 随机发送Day+或Day-或LEFT或RIGHT
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, left_or_right_random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                logging.info(ch_epg_info)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                GL.send_test_case_commd_numb -= 1
                if GL.send_test_case_commd_numb == 0:
                    state["send_commd_state"] = False
                    GL.send_test_case_commd_numb = GL.each_ch_test_numb

    elif TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
        state["send_commd_state"] = True  # 不同节目完成一轮的事件切换后，状态被关闭，会导致下面的节目不会进行切换事件
        cmd_set_list = [KEY["LEFT"], KEY["RIGHT"], KEY["OK"]]
        if TEST_CASE_INFO[3] == "LEFT+OK":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[0])  # LEFT
            GL.send_cmd.append(cmd_set_list[2])  # OK
            while state["send_commd_state"]:
                state["clear_ch_epg_info_state"] = True
                send_commd(GL.send_cmd[0])
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                    send_commd(GL.send_cmd[1])
                    send_commd(KEY["PAGE_DOWN"])
                    send_commd(KEY["EXIT"])
                if len(GL.all_ch_epg_info[channel_info[1]]) >= 20:
                    GL.judge_switch_epg_info_end = GL.all_ch_epg_info[channel_info[1]][:10]
                    if list(ch_epg_info) in GL.judge_switch_epg_info_end:
                        state["send_commd_state"] = False
                        GL.judge_switch_epg_info_end = []
        elif TEST_CASE_INFO[3] == "RIGHT+OK":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[1])  # RIGHT
            GL.send_cmd.append(cmd_set_list[2])  # OK
            while state["send_commd_state"]:
                state["clear_ch_epg_info_state"] = True
                send_commd(GL.send_cmd[0])
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                    send_commd(GL.send_cmd[1])
                    send_commd(KEY["PAGE_DOWN"])
                    send_commd(KEY["EXIT"])
                if len(GL.all_ch_epg_info[channel_info[1]]) >= 20:
                    GL.judge_switch_epg_info_end = GL.all_ch_epg_info[channel_info[1]][:10]
                    if list(ch_epg_info) in GL.judge_switch_epg_info_end:
                        state["send_commd_state"] = False
                        GL.judge_switch_epg_info_end = []
        elif TEST_CASE_INFO[3] == "LEFT+Random+OK":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[0])  # LEFT
            GL.send_cmd.append(cmd_set_list[2])  # OK
            while state["send_commd_state"]:
                random_send_numb = randint(1, 10)
                for i in range(random_send_numb):
                    state["clear_ch_epg_info_state"] = True
                    send_commd(GL.send_cmd[0])
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                    send_commd(GL.send_cmd[1])
                    send_commd(KEY["PAGE_DOWN"])
                    send_commd(KEY["EXIT"])
                elif list(ch_epg_info) == ['', '', '']:
                    logging.debug("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                if len(GL.all_ch_epg_info[channel_info[1]]) >= 20:
                    GL.judge_switch_epg_info_end = GL.all_ch_epg_info[channel_info[1]][:10]
                    if list(ch_epg_info) in GL.judge_switch_epg_info_end:
                        state["send_commd_state"] = False
                        GL.judge_switch_epg_info_end = []
        elif TEST_CASE_INFO[3] == "RIGHT+Random+OK":
            GL.send_cmd = []
            GL.send_cmd.append(cmd_set_list[1])  # RIGHT
            GL.send_cmd.append(cmd_set_list[2])  # OK
            while state["send_commd_state"]:
                random_send_numb = randint(1, 10)
                for i in range(random_send_numb):
                    state["clear_ch_epg_info_state"] = True
                    send_commd(GL.send_cmd[0])
                    logging.info("随机切换焦点次数：{}-{}".format(i + 1, random_send_numb))
                sleep_time = uniform(0.75, 1.0)
                logging.info(sleep_time)
                time.sleep(sleep_time)
                if list(ch_epg_info) != ['', '', '']:
                    GL.all_ch_epg_info[channel_info[1]].append(list(ch_epg_info))
                    send_commd(GL.send_cmd[1])
                    send_commd(KEY["PAGE_DOWN"])
                    send_commd(KEY["EXIT"])
                if len(GL.all_ch_epg_info[channel_info[1]]) >= 20:
                    GL.judge_switch_epg_info_end = GL.all_ch_epg_info[channel_info[1]][:10]
                    if list(ch_epg_info) in GL.judge_switch_epg_info_end:
                        state["send_commd_state"] = False
                        GL.judge_switch_epg_info_end = []


def padding_report_data():
    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        # 期望的报告数据
        GL.expect_report_data[0] = "{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[3], TEST_CASE_INFO[4])
        GL.expect_report_data[1] = TEST_CASE_INFO[1]
        GL.expect_report_data[2] = "None"
        GL.expect_report_data[3] = TEST_CASE_INFO[2]
        GL.expect_report_data[4] = "None"
        GL.expect_report_data[5] = "Yes"
        GL.expect_report_data[6] = str(GL.send_test_case_commd_numb)

        # 实际的测试数据
        GL.actual_report_data[0] = channel_info[6]
        GL.actual_report_data[2] = channel_info[5]
        GL.actual_report_data[3] = channel_info[1]
        GL.actual_report_data[4] = channel_info[7]
        GL.actual_report_data[5] = str(GL.actual_test_numb)
        if TEST_CASE_INFO[2] == "TV":
            GL.actual_report_data[1] = GL.TV_channel_groups[channel_info[6]]
        elif TEST_CASE_INFO[2] == "Radio":
            GL.actual_report_data[1] = GL.Radio_channel_groups[channel_info[6]]

    elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
        # 期望的报告数据
        GL.expect_report_data[0] = "{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[3], TEST_CASE_INFO[4])
        GL.expect_report_data[1] = TEST_CASE_INFO[1]
        GL.expect_report_data[2] = "None"
        GL.expect_report_data[3] = TEST_CASE_INFO[2]
        GL.expect_report_data[4] = TEST_CASE_INFO[3]
        GL.expect_report_data[5] = str(GL.send_test_case_commd_numb)

        # 实际的测试数据
        GL.actual_report_data[0] = channel_info[6]
        GL.actual_report_data[2] = channel_info[5]
        # GL.actual_report_data[3] = REVERSE_KEY[GL.send_cmd]
        GL.actual_report_data[4] = str(GL.actual_test_numb)
        if TEST_CASE_INFO[2] == "TV":
            GL.actual_report_data[1] = GL.TV_channel_groups[channel_info[6]]
        elif TEST_CASE_INFO[2] == "Radio":
            GL.actual_report_data[1] = GL.Radio_channel_groups[channel_info[6]]

        if TEST_CASE_INFO[3] == "UP" or TEST_CASE_INFO[3] == "DOWN":
            GL.actual_report_data[3] = REVERSE_KEY[GL.send_cmd]
        elif TEST_CASE_INFO[3] == "UP+Random" or TEST_CASE_INFO[3] == "DOWN+Random":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd]}+Random'
        elif TEST_CASE_INFO[3] == "UP+DOWN+Random":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd[0]]}+{REVERSE_KEY[GL.send_cmd[1]]}+Random'

    elif TEST_CASE_INFO[4] == "SwitchEPGEvent":
        # 期望的报告数据
        GL.expect_report_data[0] = "{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[3], TEST_CASE_INFO[4])
        GL.expect_report_data[1] = TEST_CASE_INFO[1]
        GL.expect_report_data[2] = "None"
        GL.expect_report_data[3] = TEST_CASE_INFO[2]
        GL.expect_report_data[4] = TEST_CASE_INFO[3]
        GL.expect_report_data[5] = "None"

        # 实际的测试数据
        GL.actual_report_data[0] = channel_info[6]
        GL.actual_report_data[2] = channel_info[5]
        # GL.actual_report_data[3] = REVERSE_KEY[GL.send_cmd]
        if TEST_CASE_INFO[2] == "TV":
            GL.actual_report_data[1] = GL.TV_channel_groups[channel_info[6]]
            GL.actual_report_data[4] = str(len(GL.TV_ch_attribute[3]))
        elif TEST_CASE_INFO[2] == "Radio":
            GL.actual_report_data[1] = GL.Radio_channel_groups[channel_info[6]]
            GL.actual_report_data[4] = str(len(GL.Radio_ch_attribute[3]))

        if TEST_CASE_INFO[3] == "LEFT" or TEST_CASE_INFO[3] == "RIGHT" or \
                TEST_CASE_INFO[3] == "Day+" or TEST_CASE_INFO[3] == "Day-":
            GL.actual_report_data[3] = REVERSE_KEY[GL.send_cmd]
        elif TEST_CASE_INFO[3] == "LEFT+Random" or TEST_CASE_INFO[3] == "RIGHT+Random":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd]}+Random'
        elif TEST_CASE_INFO[3] == "RIGHT+LEFT+Random":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd[1]]}+{REVERSE_KEY[GL.send_cmd[0]]}+Random'
        elif TEST_CASE_INFO[3] == "Day++LEFT+Random" or TEST_CASE_INFO[3] == "Day++RIGHT+Random" or \
                TEST_CASE_INFO[3] == "Day-+LEFT+Random" or TEST_CASE_INFO[3] == "Day-+RIGHT+Random":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd[0]]}+{REVERSE_KEY[GL.send_cmd[1]]}+Random'
        elif TEST_CASE_INFO[3] == "Day++LEFTorRIGHT+Random" or TEST_CASE_INFO[3] == "Day-+LEFTorRIGHT+Random":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd]}+LEFTorRIGHT+Random'
        elif TEST_CASE_INFO[3] == "Day+orDay-orLEFTorRIGHT+Random":
            GL.actual_report_data[3] = "Day+orDay-orLEFTorRIGHT+Random"

    elif TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
        # 期望的报告数据
        GL.expect_report_data[0] = "{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[3], TEST_CASE_INFO[4])
        GL.expect_report_data[1] = TEST_CASE_INFO[1]
        GL.expect_report_data[2] = "None"
        GL.expect_report_data[3] = TEST_CASE_INFO[2]
        GL.expect_report_data[4] = TEST_CASE_INFO[3]
        GL.expect_report_data[5] = "None"

        # 实际的测试数据
        GL.actual_report_data[0] = channel_info[6]
        GL.actual_report_data[2] = channel_info[5]
        # GL.actual_report_data[3] = REVERSE_KEY[GL.send_cmd]
        if TEST_CASE_INFO[2] == "TV":
            GL.actual_report_data[1] = GL.TV_channel_groups[channel_info[6]]
            GL.actual_report_data[4] = str(len(GL.TV_ch_attribute[3]))
        elif TEST_CASE_INFO[2] == "Radio":
            GL.actual_report_data[1] = GL.Radio_channel_groups[channel_info[6]]
            GL.actual_report_data[4] = str(len(GL.Radio_ch_attribute[3]))

        if TEST_CASE_INFO[3] == "LEFT+OK" or TEST_CASE_INFO[3] == "RIGHT+OK":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd[0]]}+{REVERSE_KEY[GL.send_cmd[1]]}'
        elif TEST_CASE_INFO[3] == "LEFT+Random+OK" or TEST_CASE_INFO[3] == "RIGHT+Random+OK":
            GL.actual_report_data[3] = f'{REVERSE_KEY[GL.send_cmd[0]]}+Random+{REVERSE_KEY[GL.send_cmd[1]]}'


def write_data_to_report():
    logging.debug("write_data_to_report")
    wb = ''
    ws = ''
    expect_report_title = ''
    actual_report_title = ''
    channel_report_title = ''
    channel_event_info_title = ''
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    blue_font = Font(color=BLUE)
    red_font = Font(color=RED)

    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        expect_report_title = [
            "报告名称", "期望分组名称", "期望分组节目总数", "期望节目类别", "期望节目名称", "期望节目是否有EPG", "期望执行次数"]
        actual_report_title = [
            "实际分组名称", "实际分组节目总数", "实际节目类别", "实际节目名称", "实际节目是否有EPG", "实际执行次数"]
    elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
        expect_report_title = [
            "报告名称", "期望分组名称", "期望分组节目总数", "期望节目类别", "期望切台模式", "期望执行次数"]
        actual_report_title = [
            "实际分组名称", "实际分组节目总数", "实际节目类别", "实际切台模式", "实际执行次数"]
        channel_report_title = ["频道号", "频道名称", "TP", "加锁标志", "加密标志", "频道类型", "组别", "是否有EPG信息"]
    elif TEST_CASE_INFO[4] == "SwitchEPGEvent" or TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
        expect_report_title = [
            "报告名称", "期望分组名称", "期望分组节目总数", "期望节目类别", "期望指令", "期望有EPG节目数"]
        actual_report_title = [
            "实际分组名称", "实际分组节目总数", "实际节目类别", "实际指令", "实际有EPG节目数"]
        channel_event_info_title = ["频道名称", "EPG事件日期", "EPG事件时段", "EPG事件名称"]
        # 根据抬头和当前已经获取到的节目个数来设置间隔
        GL.report_interval = 1 + (len(GL.all_ch_epg_info.keys()) - 1) * len(channel_event_info_title)
    if not os.path.exists(file_path[1]):
        wb = Workbook()
        ws = wb.active
        ws.title = file_path[2]
        if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['D'].width = 25
            # 写期望项的Title
            for i in range(len(expect_report_title)):
                ws.cell(i + 1, 1).value = expect_report_title[i]
                ws.cell(i + 1, 1).alignment = alignment
                if i == 0:
                    ws.row_dimensions[(i + 1)].height = 30
                else:
                    ws.row_dimensions[(i + 1)].height = 13.5
            # 写实际测试项的Title
            for j in range(len(actual_report_title)):
                ws.cell(j + 2, 4).value = actual_report_title[j]
                ws.cell(j + 2, 4).alignment = alignment
        elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
            # 写期望项的Title
            for i in range(len(expect_report_title)):
                ws.cell(i + 1, 1).value = expect_report_title[i]
                ws.cell(i + 1, 1).alignment = alignment
                if i == 0:
                    ws.row_dimensions[(i + 1)].height = 30
                else:
                    ws.row_dimensions[(i + 1)].height = 13.5
            # 写实际测试项的Title
            for j in range(len(actual_report_title)):
                ws.cell(j + 2, 4).value = actual_report_title[j]
                ws.cell(j + 2, 4).alignment = alignment
            # 写EPG界面切台的节目信息Title
            for k in range(len(channel_report_title)):
                all_column_numb = column_index_from_string("A") + k
                all_column_char = get_column_letter(all_column_numb)
                ws.column_dimensions[all_column_char].width = 16  # 设置列宽
                ws.cell(len(expect_report_title) + 1, k + 1).value = channel_report_title[k]
                ws.cell(len(expect_report_title) + 1, k + 1).alignment = alignment
        elif TEST_CASE_INFO[4] == "SwitchEPGEvent" or TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
            ws.column_dimensions['A'].width = 17
            ws.column_dimensions['D'].width = 17
            # 写期望测试项的title信息
            for i in range(len(expect_report_title)):
                ws.cell(i + 1, 1).value = expect_report_title[i]
                ws.cell(i + 1, 1).alignment = alignment
                if i == 0:
                    ws.row_dimensions[(i + 1)].height = 30
                else:
                    ws.row_dimensions[(i + 1)].height = 13.5
            # 写实际测试项的title信息
            for j in range(len(actual_report_title)):
                ws.cell(j + 2, 4).value = actual_report_title[j]
                ws.cell(j + 2, 4).alignment = alignment
            # 写EPG界面切EPG事件信息Title
            for k in range(len(channel_event_info_title)):  # 根据节目个数循环写EPG信息的title信息，并设置列宽
                all_column_numb = column_index_from_string("A") + k + GL.report_interval
                all_column_char = get_column_letter(all_column_numb)
                ws.column_dimensions[all_column_char].width = 16  # 设置列宽
                ws.cell(len(expect_report_title) + 1, GL.report_interval + k).value = channel_event_info_title[k]
                ws.cell(len(expect_report_title) + 1, GL.report_interval + k).alignment = alignment

    elif os.path.exists(file_path[1]):
        wb = load_workbook(file_path[1])
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if file_path[2] in sheets_name_list:
            ws = wb[file_path[2]]
        elif file_path[2] not in sheets_name_list:
            ws = wb.create_sheet(file_path[2])
        if TEST_CASE_INFO[4] == "SwitchEPGEvent" or TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
            # 写EPG界面切EPG事件信息Title
            for k in range(len(channel_event_info_title)):  # 根据节目个数循环写EPG信息的title信息，并设置列宽
                all_column_numb = column_index_from_string("A") + k + GL.report_interval
                all_column_char = get_column_letter(all_column_numb)
                ws.column_dimensions[all_column_char].width = 16  # 设置列宽
                ws.cell(len(expect_report_title) + 1, GL.report_interval + k).value = channel_event_info_title[k]
                ws.cell(len(expect_report_title) + 1, GL.report_interval + k).alignment = alignment

    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        # 写期望项的数据
        for m in range(len(GL.expect_report_data)):
            ws.cell(m + 1, 2).value = GL.expect_report_data[m]
            ws.cell(m + 1, 2).alignment = alignment
            if m == 0:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=6)
            else:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=3)
        # 写实际测试项的数据
        for n in range(len(GL.actual_report_data)):
            ws.cell(n + 2, 5).value = GL.actual_report_data[n]
            ws.merge_cells(start_row=(n + 2), start_column=5, end_row=(n + 2), end_column=6)
            ws.cell(n + 2, 5).alignment = alignment
            if n == 0:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 2:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 4:
                compare_object = ["No", "Yes"]
                if compare_object[int(GL.actual_report_data[n])] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif compare_object[int(GL.actual_report_data[n])] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 5:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font

    elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
        # 写期望项的数据
        for m in range(len(GL.expect_report_data)):
            ws.cell(m + 1, 2).value = GL.expect_report_data[m]
            ws.cell(m + 1, 2).alignment = alignment
            if m == 0:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=6)
            else:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=3)
        # 写实际测试项的数据
        for n in range(len(GL.actual_report_data)):
            ws.cell(n + 2, 5).value = GL.actual_report_data[n]
            ws.merge_cells(start_row=(n + 2), start_column=5, end_row=(n + 2), end_column=6)
            ws.cell(n + 2, 5).alignment = alignment
            if n == 0:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 2:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 3:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 4:
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1] and \
                        GL.actual_report_data[n] == str(len(GL.epg_switch_ch_data_report)):
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1] or \
                        GL.actual_report_data[n] != str(len(GL.epg_switch_ch_data_report)):
                    ws.cell(n + 2, 5).font = red_font
        # 写实际EPG界面切台的节目信息
        for x in range(len(GL.epg_switch_ch_data_report)):
            for y in range(len(GL.epg_switch_ch_data_report[x])):
                ws.cell(len(expect_report_title) + 2 + x, y + 1).value = GL.epg_switch_ch_data_report[x][y]
                ws.cell(len(expect_report_title) + 2 + x, y + 1).alignment = alignment
                if y == 5:
                    if GL.epg_switch_ch_data_report[x][y] == GL.expect_report_data[3]:
                        ws.cell(len(expect_report_title) + 2 + x, y + 1).font = blue_font
                    elif GL.epg_switch_ch_data_report[x][y] != GL.expect_report_data[3]:
                        ws.cell(len(expect_report_title) + 2 + x, y + 1).font = red_font
                elif y == 6:
                    if GL.epg_switch_ch_data_report[x][y] == GL.expect_report_data[1]:
                        ws.cell(len(expect_report_title) + 2 + x, y + 1).font = blue_font
                    elif GL.epg_switch_ch_data_report[x][y] != GL.expect_report_data[1]:
                        ws.cell(len(expect_report_title) + 2 + x, y + 1).font = red_font

    elif TEST_CASE_INFO[4] == "SwitchEPGEvent":
        # 写期望测试项的期望结果
        for m in range(len(GL.expect_report_data)):
            ws.cell(m + 1, 2).value = GL.expect_report_data[m]
            ws.cell(m + 1, 2).alignment = alignment
            if m == 0:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=6)
            else:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=3)
        # 写实际测试项的测试结果
        for n in range(len(GL.actual_report_data)):
            ws.cell(n + 2, 5).value = GL.actual_report_data[n]
            ws.merge_cells(start_row=(n + 2), start_column=5, end_row=(n + 2), end_column=6)
            ws.cell(n + 2, 5).alignment = alignment
            if n == 0:      # 实际分组名称
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 2:    # 实际节目类别
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 3:    # 实际指令
                if TEST_CASE_INFO[3] == "Day+" or TEST_CASE_INFO[3] == "Day-":
                    key_change = {"GREEN": "Day-", "YELLOW": "Day+"}
                    if key_change[GL.actual_report_data[n]] == GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = blue_font
                    elif key_change[GL.actual_report_data[n]] != GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = red_font
                elif TEST_CASE_INFO[3] == "Day++LEFT+Random" or TEST_CASE_INFO[3] == "Day++RIGHT+Random":
                    if GL.actual_report_data[n].replace("YELLOW", "Day+") == GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = blue_font
                    elif GL.actual_report_data[n].replace("YELLOW", "Day+") != GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = red_font
                elif TEST_CASE_INFO[3] == "Day-+LEFT+Random" or TEST_CASE_INFO[3] == "Day-+RIGHT+Random":
                    if GL.actual_report_data[n].replace("GREEN", "Day-") == GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = blue_font
                    elif GL.actual_report_data[n].replace("GREEN", "Day-") != GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = red_font
                elif TEST_CASE_INFO[3] == "Day++LEFTorRIGHT+Random":
                    if GL.actual_report_data[n].replace("YELLOW", "Day+") == GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = blue_font
                    elif GL.actual_report_data[n].replace("YELLOW", "Day+") != GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = red_font
                elif TEST_CASE_INFO[3] == "Day-+LEFTorRIGHT+Random":
                    if GL.actual_report_data[n].replace("GREEN", "Day-") == GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = blue_font
                    elif GL.actual_report_data[n].replace("GREEN", "Day-") != GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = red_font
                else:
                    if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = blue_font
                    elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                        ws.cell(n + 2, 5).font = red_font
            elif n == 4:    # 实际有EPG节目数
                if GL.actual_report_data[n] == str(len(GL.all_ch_epg_info.keys())):
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != str(len(GL.all_ch_epg_info.keys())):
                    ws.cell(n + 2, 5).font = red_font
        ch_name_key = list(GL.all_ch_epg_info.keys())[len(GL.all_ch_epg_info.keys()) - 1]
        # 写每个有EPG信息节目下切换后的获取到的EPG信息
        for y in range(len(GL.all_ch_epg_info[ch_name_key])):
            ws.cell(len(expect_report_title) + 2 + y, GL.report_interval).value = ch_name_key
            ws.cell(len(expect_report_title) + 2 + y, GL.report_interval).alignment = alignment
            ws.row_dimensions[(len(expect_report_title) + 2 + y)].height = 13.5
            for z in range(len(GL.all_ch_epg_info[ch_name_key][y])):
                ws.cell(len(expect_report_title) + 2 + y, z + GL.report_interval + 1).value = \
                    GL.all_ch_epg_info[ch_name_key][y][z]
                ws.cell(len(expect_report_title) + 2 + y, z + GL.report_interval + 1).alignment = alignment

    elif TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
        # 写期望测试项的期望结果
        for m in range(len(GL.expect_report_data)):
            ws.cell(m + 1, 2).value = GL.expect_report_data[m]
            ws.cell(m + 1, 2).alignment = alignment
            if m == 0:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=6)
            else:
                ws.merge_cells(start_row=(m + 1), start_column=2, end_row=(m + 1), end_column=3)
        # 写实际测试项的测试结果
        for n in range(len(GL.actual_report_data)):
            ws.cell(n + 2, 5).value = GL.actual_report_data[n]
            ws.merge_cells(start_row=(n + 2), start_column=5, end_row=(n + 2), end_column=6)
            ws.cell(n + 2, 5).alignment = alignment
            if n == 0:  # 实际分组名称
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 2:  # 实际节目类别
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 3:  # 实际指令
                if GL.actual_report_data[n] == GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != GL.expect_report_data[n + 1]:
                    ws.cell(n + 2, 5).font = red_font
            elif n == 4:    # 实际有EPG节目数
                if GL.actual_report_data[n] == str(len(GL.all_ch_epg_info.keys())):
                    ws.cell(n + 2, 5).font = blue_font
                elif GL.actual_report_data[n] != str(len(GL.all_ch_epg_info.keys())):
                    ws.cell(n + 2, 5).font = red_font
        ch_name_key = list(GL.all_ch_epg_info.keys())[len(GL.all_ch_epg_info.keys()) - 1]
        # 写每个有EPG信息节目下切换后的获取到的EPG信息
        for y in range(len(GL.all_ch_epg_info[ch_name_key])):
            ws.cell(len(expect_report_title) + 2 + y, GL.report_interval).value = ch_name_key
            ws.cell(len(expect_report_title) + 2 + y, GL.report_interval).alignment = alignment
            ws.row_dimensions[(len(expect_report_title) + 2 + y)].height = 13.5
            for z in range(len(GL.all_ch_epg_info[ch_name_key][y])):
                ws.cell(len(expect_report_title) + 2 + y, z + GL.report_interval + 1).value = \
                    GL.all_ch_epg_info[ch_name_key][y][z]
                ws.cell(len(expect_report_title) + 2 + y, z + GL.report_interval + 1).alignment = alignment
    wb.save(file_path[1])


def receive_serial_process(prs_data, infrared_send_cmd, state, channel_info, rsv_info, ch_epg_info):
    logging_info_setting()
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIME_SHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict((val, key) for key, val in rsv_key.items())

    switch_ch_kws = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    ch_info_kws = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name"]

    group_info_kws = [
        "Group_name",
        "Prog_total"
    ]

    check_epg_kws = [
        "[PTD]Program_epg_info=",
        "[PTD]EPG_event:event_time="
    ]

    epg_info_kws = [
        "Program_epg_info",
        "event_time",
        "event_name"
    ]

    other_kws = [
        "[PTD]Infrared_key_values:",    # 获取红外接收关键字
        "[PTD]Time_mode=",              # 获取系统时间模式
    ]

    infrared_rsv_cmd = []
    receive_serial = serial.Serial(prs_data["receive_serial_name"], 115200, timeout=1)

    while True:
        data = receive_serial.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            data1 = data.decode("ISO-8859-1", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_log_data_to_txt(prs_data["log_file_path"], data3)

            if state["clear_channel_info_state"]:
                each_ch_info = ['', '', '', '', '', '', '', '']
                del channel_info[:]
                channel_info.extend(each_ch_info)
                state["clear_channel_info_state"] = False

            if state["clear_ch_epg_info_state"]:
                each_epg_info = ['', '', '']
                del ch_epg_info[:]
                ch_epg_info.extend(each_epg_info)
                state["clear_ch_epg_info_state"] = False

            if other_kws[0] in data2:   # 红外接收打印
                rsv_cmd = re.split(":", data2)[-1]
                infrared_rsv_cmd.append(rsv_cmd)
                if rsv_cmd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_cmd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        infrared_send_cmd[-1], reverse_rsv_key[infrared_rsv_cmd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(infrared_send_cmd), len(infrared_rsv_cmd)))
                    receive_cmd_list.append(rsv_cmd)

            if other_kws[1] in data2:     # 获取系统时间模式（自动还是手动）
                state["sys_time_mode_state"] = True
                rsv_info["sys_time_mode"] = re.split(r"=", data2)[-1]

            if switch_ch_kws[0] in data2:
                ch_info_split = re.split(r"[\],]", data2)
                for i in range(len(ch_info_split)):
                    if ch_info_kws[0] in ch_info_split[i]:  # 提取频道号
                        channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if ch_info_kws[1] in ch_info_split[i]:  # 提取频道名称
                        channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if switch_ch_kws[1] in data2:
                flag_info_split = re.split(r"[\],]", data2)
                for i in range(len(flag_info_split)):
                    if ch_info_kws[2] in flag_info_split[i]:  # 提取频道所属TP
                        channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if ch_info_kws[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if switch_ch_kws[3] in data2:
                group_info_split = re.split(r"[\],]", data2)
                for i in range(len(group_info_split)):
                    if group_info_kws[0] in group_info_split[i]:  # 提取频道所属组别
                        rsv_info["prog_group_name"] = re.split(r"=", group_info_split[i])[-1]
                        channel_info[6] = rsv_info["prog_group_name"]
                    if group_info_kws[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        rsv_info["prog_group_total"] = re.split(r"=", group_info_split[i])[-1]

            if check_epg_kws[0] in data2:       # 判断节目是否存在EPG信息
                epg_info_split = re.split(r"]", data2)
                for i in range(len(epg_info_split)):
                    if epg_info_kws[0] in epg_info_split[i]:
                        rsv_info["epg_info_exist"] = re.split(r"=", epg_info_split[i])[-1]
                        channel_info[7] = rsv_info["epg_info_exist"]

            if check_epg_kws[1] in data2:
                epg_event_split = re.split(r"t:|,", data2)
                for i in range(len(epg_event_split)):
                    if epg_info_kws[1] in epg_event_split[i]:
                        time_info_split = re.split(r"=|--|\s", epg_event_split[i])
                        if time_info_split[1] == time_info_split[3]:
                            event_date = "{}".format(time_info_split[1])
                            ch_epg_info[0] = event_date
                        elif time_info_split[1] != time_info_split[3]:
                            event_date = "{}-{}".format(time_info_split[1], time_info_split[3])
                            ch_epg_info[0] = event_date
                        event_time = "{}-{}".format(time_info_split[2][:5], time_info_split[4][:5])
                        ch_epg_info[1] = event_time
                    if epg_info_kws[2] in epg_event_split[i]:
                        event_name = re.split(r"=", epg_event_split[i])[-1]
                        ch_epg_info[2] = event_name


if __name__ == "__main__":

    GL = MyGlobal()
    logging_info_setting()
    msg = "现在开始执行的是:{}_{}_{}_{}_{}".format(TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2],
                                           TEST_CASE_INFO[3], TEST_CASE_INFO[4])
    logging.critical(format(msg, '*^150'))
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIME_SHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D"
    }
    REVERSE_KEY = dict((val, key) for key, val in KEY.items())

    file_path = build_log_and_report_file_path()
    ser_name = list(check_ports())  # send_ser_name, receive_ser_name
    send_serial = serial.Serial(ser_name[0], 9600)
    receive_ser_name = ser_name[1]

    infrared_send_cmd = Manager().list([])
    receive_cmd_list = Manager().list([])
    channel_info = Manager().list(['', '', '', '', '', '', '', ''])     # [频道号,频道名称,tp,lock,scramble,频道类型,组别,epg_info]
    ch_epg_info = Manager().list(['', '', ''])                          # 单个EPG信息的提取[event_date, event_time, event_name]
    rsv_info = Manager().dict({
        "prog_group_name": '', "prog_group_total": '', "epg_info_exist": '', "sys_time_mode": '',
    })

    state = Manager().dict({
        "receive_loop_state": False, "sys_time_mode_state": False, "clear_channel_info_state": False,
        "send_commd_state": True, "clear_ch_epg_info_state": False
    })
    prs_data = Manager().dict({
        "log_file_path": file_path[0], "receive_serial_name": receive_ser_name,
    })

    rsv_p = Process(target=receive_serial_process, args=(
        prs_data, infrared_send_cmd, state, channel_info, rsv_info, ch_epg_info))
    rsv_p.start()

    if platform.system() == "Windows":
        time.sleep(5)
        logging.info("Windows系统接收端响应慢，等待5秒")
    elif platform.system() == "Linux":
        time.sleep(1)
        logging.info("Linux系统接收端响应快，但是增加一个延时保护，等待1秒")

    # 主程序开始部分
    check_sys_time_mode()
    if TEST_CASE_INFO[4] == "EnterExitEPGInterface":
        get_group_channel_total_info()
        get_choice_group_ch_type()
        choice_test_channel()
        check_preparatory_work()
        send_test_case_commd()
        padding_report_data()
        write_data_to_report()
        exit_to_screen()
        state["receive_loop_state"] = True
    elif TEST_CASE_INFO[4] == "EPGSwitchChannel":
        get_group_channel_total_info()
        check_preparatory_work()
        send_test_case_commd()
        padding_report_data()
        write_data_to_report()
        exit_to_screen()
        state["receive_loop_state"] = True
    elif TEST_CASE_INFO[4] == "SwitchEPGEvent":
        get_group_channel_total_info()
        get_choice_group_ch_type()
        choice_test_channel()
        exit_to_screen()
        state["receive_loop_state"] = True
    elif TEST_CASE_INFO[4] == "SwitchAndDetailEPGEvent":
        get_group_channel_total_info()
        get_choice_group_ch_type()
        choice_test_channel()
        exit_to_screen()
        state["receive_loop_state"] = True

    if state["receive_loop_state"]:
        rsv_p.terminate()
        logging.info("程序结束")
        logging.info('stop receive process')
        rsv_p.join()
