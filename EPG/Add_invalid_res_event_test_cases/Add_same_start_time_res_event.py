#!/usr/bin/python3
# -*- coding: utf-8 -*-


from serial_setting1 import *
from multiprocessing import Process, Manager
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import RED, BLUE
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta, date
from random import randint
import platform
import os
import time
import logging
import re
import sys

TEST_CASE_INFO = ["01", "All", "TV", "Mon.", "Power On", "TVScreenDiffCH",
                  "Manual_jump", "Same(time+type+mode)", "Mon.", "Power On", "screen_test_numb"]

scenes_list = [
    "Same(time+type+mode)",
    "Same(time+type)+Diff(mode)",
    "Same(time+mode)+Diff(type)",
    "Same(time)+Diff(type+mode)",
    ]


class MyGlobal(object):

    def __init__(self):
        if TEST_CASE_INFO[-1] == "screen_test_numb":
            self.res_triggered_numb = 2                 # 大画面预约响应的次数
        elif TEST_CASE_INFO[-1] == "other_interface_test_numb":
            self.res_triggered_numb = 2                 # 其他界面预约响应的次数

        self.choice_res_ch = ''                         # 预约Play或PVR事件时所选预约节目
        self.res_event_mgr = []                         # 预约事件管理
        self.start_row = 0                              # 用于每次预约事件响应后，写数据增加行数
        self.pvr_rec_dur_time = ''                      # 用于记录PVR事件录制持续时间
        self.event_already_triggered_numb = 0           # 用于控制循环事件第二次前后的运行代码界限
        self.res_event_info_1 = []

        # 报告数据汇总[[预约事件1信息]，[预约事件2信息], "事件列表事件个数", "无效事件提示", "case编号", "执行case时间"]
        self.report_data = [[], [], '', '', '', '', ]
        # ["报告名称", "预约事件类型", "预约事件模式", "预约节目类型", "预约等待界面", "预约跳转模式", "预约执行次数"]
        self.title_data = ['', '', '', '', '', '', '']


def logging_info_setting():
    # 配置logging输出格式
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    date_format = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=log_format, datefmt=date_format)


def hex_strs_to_bytes(strings):
    # 将红外命令字符串转换为字节串
    return bytes.fromhex(strings)


def write_log_data_to_txt(path, write_data):
    with open(path, "a+", encoding="utf-8") as fo:
        fo.write(write_data)


def send_cmd(command):
    global receive_cmd_list, infrared_send_cmd
    # 红外发送端发送指令
    if len(infrared_send_cmd) == len(receive_cmd_list):
        send_serial.write(hex_strs_to_bytes(command))
        send_serial.flush()
        logging.info("红外发送：{}".format(REVERSE_KEY[command]))
        if REVERSE_KEY[command] != "POWER":
            infrared_send_cmd.append(REVERSE_KEY[command])
        time.sleep(1.0)
    elif len(infrared_send_cmd) != len(receive_cmd_list):
        logging.info("检测到发送和接收命令数不一致，等待5秒，查看是否接收端还没有接收到打印")
        time.sleep(5)
        if len(infrared_send_cmd) == len(receive_cmd_list):
            send_cmd(command)
        elif len(infrared_send_cmd) != len(receive_cmd_list):
            logging.info(f"此刻补发STB没有接收到的红外命令{infrared_send_cmd[-1]}")
            send_serial.write(hex_strs_to_bytes(KEY[infrared_send_cmd[-1]]))
            send_serial.flush()
            time.sleep(1.0)

            logging.info(f"此时再发送本次要发送的命令{REVERSE_KEY[command]}")
            send_cmd(command)


def send_more_cmds(command_list):
    # 用于发送一连串的指令
    for command in command_list:
        send_cmd(command)
    time.sleep(1)   # 增加函数切换时的的等待，避免可能出现send_commd函数中的等待时间没有执行的情况


def change_numbs_to_cmds_list(numbs_list):
    # 将数值列表转换为指令集列表
    channel_cmds_list = []
    for i in range(len(numbs_list)):
        channel_cmds_list.append([])
        if len(numbs_list[i]) == 1:
            channel_cmds_list[i].append(KEY[numbs_list[i]])
        elif len(numbs_list[i]) > 1:
            for j in range(len(numbs_list[i])):
                channel_cmds_list[i].append(KEY[numbs_list[i][j]])
    return channel_cmds_list


def create_log_and_report_file_path():
    # 用于创建打印和报告文件路径
    # 构建存放数据的总目录，以及构建存放打印和报告的目录
    parent_path = os.path.dirname(os.getcwd())
    test_data_directory_name = "test_data"
    test_data_directory_path = os.path.join(parent_path, test_data_directory_name)
    log_directory_name = "print_log"
    log_directory_path = os.path.join(test_data_directory_path, log_directory_name)
    report_directory_name = "report"
    report_directory_path = os.path.join(test_data_directory_path, report_directory_name)
    # 判断目录是否存在，否则创建目录
    if not os.path.exists(test_data_directory_path):
        os.mkdir(test_data_directory_path)
    if not os.path.exists(log_directory_path):
        os.mkdir(log_directory_path)
    if not os.path.exists(report_directory_path):
        os.mkdir(report_directory_path)
    # 创建打印和报告文件的名称和路径
    time_info = re.sub(r"[-: ]", "_", str(datetime.now())[:19])
    sheet_name = TEST_CASE_INFO[7]

    fmt_name = "{}_{}_{}_{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2], TEST_CASE_INFO[4],
        TEST_CASE_INFO[3], sheet_name, TEST_CASE_INFO[9], TEST_CASE_INFO[8])
    log_file_name = "Log_{}_{}.txt".format(fmt_name, time_info)
    log_file_path = os.path.join(log_directory_path, log_file_name)
    report_file_name = "Result_report.xlsx"
    report_file_path = os.path.join(report_directory_path, report_file_name)
    # sheet_name = "{}_{}_{}_{}".format(TEST_CASE_INFO[2], TEST_CASE_INFO[4], TEST_CASE_INFO[3], TEST_CASE_INFO[7])
    return log_file_path, report_file_path, sheet_name


def clear_timer_setting_all_events():
    logging.info("clear_timer_setting_all_events")
    # 清除Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    delete_all_res_events = [KEY["BLUE"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入定时器设置界面
    send_more_cmds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        if rsv_kws["res_event_numb"] != '0':
            send_more_cmds(delete_all_res_events)
        elif rsv_kws["res_event_numb"] == '0':
            logging.info("没有预约事件存在")
            time.sleep(1)
        else:
            logging.debug("警告：预约事件个数获取错误！！！")
        state["res_event_numb_state"] = False
    # 退回大画面
    send_more_cmds(exit_to_screen)


def check_sys_time_mode():
    logging.info("check_sys_time_mode")
    # 检测系统时间模式
    enter_time_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["OK"]]
    change_sys_time_mode = [KEY["RIGHT"], KEY["EXIT"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入时间设置界面
    send_more_cmds(enter_time_setting_interface)
    # 对当前系统时间模式进行判断
    while not state["sys_time_mode_state"]:
        logging.info("还没有获取到系统时间模式信息")
        time.sleep(1)
    else:
        if rsv_kws["sys_time_mode"] == "Auto":
            send_more_cmds(change_sys_time_mode)
        elif rsv_kws["sys_time_mode"] == "Manual":
            logging.info("系统时间模式已经为手动模式")
        else:
            logging.debug("警告：系统时间模式获取错误！！！")
        state["sys_time_mode_state"] = False
    # 退回大画面
    send_more_cmds(exit_to_screen)


def get_current_system_time():
    logging.info("get_current_system_time")
    # 获取当前系统时间
    time.sleep(1)
    while not state["current_sys_time_state"]:
        logging.info("还没有获取到系统时间信息")
        time.sleep(1)
    else:
        logging.info(f"当前系统时间为:{rsv_kws['current_sys_time']}")


def get_exist_event_info():
    # 获取当前已预约的事件信息
    if len(res_event_list) == 0:
        logging.info("当前没有预约事件")
        logging.info(f"预约事件列表为:{res_event_list}")
    else:
        logging.info("当前所有预约事件如下")
        for event in res_event_list:
            logging.info(event)


def choice_ch_for_res_event_type(choice_event):
    logging.info("choice_ch_for_res_event_type")
    # 根据预约事件类型来选择节目
    group_dict = {}
    choice_ch_numb = []

    # 根据所选case切换到对应类型节目的界面
    while channel_info[5] != TEST_CASE_INFO[2]:
        send_cmd(KEY["TV/R"])
        if channel_info[3] == "1":
            send_cmd(KEY["EXIT"])
    # 调出频道列表,用于判断组别信息
    send_cmd(KEY["OK"])
    # 切到指定分组和采集分组下节目总数信息
    while rsv_kws["prog_group_name"] != TEST_CASE_INFO[1]:
        send_cmd(KEY["RIGHT"])
        if channel_info[3] == "1":
            send_cmd(KEY["EXIT"])
    else:
        if rsv_kws["prog_group_name"] == '':
            logging.info("警告：没有All分组信息")
        else:
            group_dict[rsv_kws["prog_group_name"]] = rsv_kws["prog_group_total"]
            logging.info(f"分组信息{group_dict}")
    # 退出频道列表,回到大画面界面
    send_cmd(KEY["EXIT"])

    if choice_event == "event_1":
        # 根据用例指定的事件类型来选择节目
        if TEST_CASE_INFO[4] == "Play":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])
            logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
            GL.choice_res_ch = channel_info[1]
            logging.info(channel_info)

        elif TEST_CASE_INFO[4] == "PVR":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])

            while channel_info[3] != '0' and channel_info[4] != '0':
                logging.info(f"查看所切节目信息：{channel_info}")
                logging.info("所选节目不为免费节目，不可以进行PVR预约，继续切台")
                send_cmd(KEY["UP"])
                time.sleep(2)
                if channel_info[3] == "1":
                    send_cmd(KEY["EXIT"])
            else:
                logging.info("所选节目为免费节目，可以进行PVR预约")
                logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
                GL.choice_res_ch = channel_info[1]
                logging.info(channel_info)

        elif TEST_CASE_INFO[4] == "Power Off":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

        elif TEST_CASE_INFO[4] == "Power On":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

    elif choice_event == "event_2":
        # 根据用例指定的事件类型来选择节目
        if TEST_CASE_INFO[9] == "Play":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])
            logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
            GL.choice_res_ch = channel_info[1]
            logging.info(channel_info)

        elif TEST_CASE_INFO[9] == "PVR":
            choice_ch_numb.append(str(randint(1, int(group_dict[TEST_CASE_INFO[1]]))))
            choice_ch_cmd = change_numbs_to_cmds_list(choice_ch_numb)
            for i in range(len(choice_ch_cmd)):
                for j in choice_ch_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["OK"])
            time.sleep(2)
            if channel_info[3] == "1":
                send_cmd(KEY["EXIT"])

            while channel_info[3] != '0' and channel_info[4] != '0':
                logging.info(f"查看所切节目信息：{channel_info}")
                logging.info("所选节目不为免费节目，不可以进行PVR预约，继续切台")
                send_cmd(KEY["UP"])
                time.sleep(2)
                if channel_info[3] == "1":
                    send_cmd(KEY["EXIT"])
            else:
                logging.info("所选节目为免费节目，可以进行PVR预约")
                logging.info(f"所选节目频道号和所切到的节目频道号为:{choice_ch_numb}--{channel_info[0]}")
                GL.choice_res_ch = channel_info[1]
                logging.info(channel_info)

        elif TEST_CASE_INFO[9] == "Power Off":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")

        elif TEST_CASE_INFO[9] == "Power On":
            logging.info(f"当前用例为{TEST_CASE_INFO[4]}，不需要切换节目")


def calculate_res_event_expected_start_time():
    logging.info("calculate_res_event_expected_start_time")
    time_interval = 5
    sys_time = rsv_kws['current_sys_time']
    sys_time_split = re.split(r"[\s:/]", sys_time)
    sys_year = int(sys_time_split[0])
    sys_month = int(sys_time_split[1])
    sys_day = int(sys_time_split[2])
    sys_hour = int(sys_time_split[3])
    sys_minute = int(sys_time_split[4])
    dt_time = datetime(sys_year, sys_month, sys_day, sys_hour, sys_minute)
    logging.info(dt_time)
    expected_res_time = dt_time + timedelta(minutes=time_interval)
    logging.info(expected_res_time)
    expected_res_time_split = re.split(r"[-\s:]", str(expected_res_time))
    str_expected_res_time = ''.join(expected_res_time_split)[:12]

    logging.info(f"期望的完整的预约事件时间为{str_expected_res_time}")
    return str_expected_res_time


def create_expected_add_event_info():
    logging.info("create_expected_add_event_info")
    # 创建期望的事件信息
    expected_event_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    duration_time = "0001"
    if TEST_CASE_INFO[4] == "Play":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "PVR":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = channel_info[1]
        expected_event_info[3] = duration_time
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power Off":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]

    elif TEST_CASE_INFO[4] == "Power On":
        expected_event_full_time = calculate_res_event_expected_start_time()
        expected_event_info[0] = expected_event_full_time
        expected_event_info[1] = TEST_CASE_INFO[4]
        expected_event_info[2] = "----"
        expected_event_info[3] = "--:--"
        expected_event_info[4] = TEST_CASE_INFO[3]
    return expected_event_info


def edit_add_new_res_event_info():
    logging.info("edit_add_new_res_event_info")
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面(Add按键)
    send_cmd(KEY["GREEN"])
    # 生成预期的预约事件
    expected_res_event_info = create_expected_add_event_info()
    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[4]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[4]}')
            send_cmd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[3]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[3]}')
            send_cmd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[3] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[3]}")
    elif TEST_CASE_INFO[3] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[3]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_cmd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_cmds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_cmd(KEY["DOWN"])
    else:
        start_time_list.append(expected_res_event_info[0][8:])
        start_time_cmd = change_numbs_to_cmds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_cmd(j)
        send_cmd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[4] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[4]}")
    elif TEST_CASE_INFO[4] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_cmd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_cmds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[4] == "Power Off" or TEST_CASE_INFO[4] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[4]}，不需要设置Channel")
    elif TEST_CASE_INFO[4] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[4]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_cmd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["update_event_list_state"] = True
    send_cmd(KEY["EXIT"])
    send_cmd(KEY["OK"])
    # 退回大画面
    send_more_cmds(exit_to_screen)


def add_new_res_event_to_event_mgr_list():
    logging.info("add_new_res_event_to_event_mgr_list")
    # 添加新预约事件到事件管理列表
    # if list(res_event_list) not in GL.res_event_mgr:
    #     GL.res_event_mgr.extend(list(res_event_list))
    # GL.res_event_info_1.extend(list(res_event_list))
    GL.report_data[0].extend(list(res_event_list)[0])
    logging.info(type(GL.res_event_mgr))
    logging.info("分割线===============================================================================================")
    logging.info(GL.res_event_mgr)
    logging.info(list(res_event_list))
    state["update_event_list_state"] = False


def new_add_res_event_1():
    logging.info("new_add_res_event")
    # 新增预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_cmds(enter_timer_setting_interface)
    # 获取当前系统时间
    get_current_system_time()
    # 进入事件编辑界面，设置预约事件参数
    edit_add_new_res_event_info()
    # 添加新预约事件到事件管理列表
    add_new_res_event_to_event_mgr_list()


def res_triggered_later_check_timer_setting_event_list():
    logging.info("res_triggered_later_check_timer_setting_event_list")
    # 预约事件触发后，事件列表事件检查
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    send_more_cmds(enter_timer_setting_interface)
    if rsv_kws["res_event_numb"] == '0':
        if len(GL.res_event_mgr) == int(rsv_kws["res_event_numb"]):
            logging.info("数据库与事件列表中的事件个数匹配，都为空")
        else:
            logging.info(f"警告：数据库与事件列表中的事件数不匹配，{len(GL.res_event_mgr)}--{int(rsv_kws['res_event_numb'])}")
    elif rsv_kws["res_event_numb"] != '0':
        if len(GL.res_event_mgr) == int(rsv_kws["res_event_numb"]):
            logging.info("数据库与事件列表中的事件个数匹配，再检查事件信息是否匹配")
            if GL.res_event_mgr == list(res_event_list):
                logging.info(f"数据库与事件列表中的事件信息一致，{GL.res_event_mgr}--{list(res_event_list)}")
            else:
                logging.info(f"警告：数据库与事件列表中的事件信息不一致，{GL.res_event_mgr}--{list(res_event_list)}")
        else:
            logging.info("数据库与事件列表中的事件个数不匹配，请检查事件信息")
            logging.info(f"警告：数据库与事件列表中的事件个数不一致，{GL.res_event_mgr}--{list(res_event_list)}")
    send_more_cmds(exit_to_screen)


def cale_str_time_for_add_day(str_time, interval_day):
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    if len(str_time) == 12:     # once事件时间计算
        fmt_year = int(str_time[:4])
        fmt_month = int(str_time[4:6])
        fmt_day = int(str_time[6:8])
        fmt_hour = int(str_time[8:10])
        fmt_minute = int(str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(days=interval_day)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
        logging.info(f"----------------------------------------------------------------------{str_new_fmt_date}")
    return str_new_fmt_date


def change_str_time_and_fmt_time(str_time, interval_time):
    # 字符串时间和格式化时间之间转换
    str_new_fmt_date = ''
    if len(str_time) == 12:     # once事件时间计算
        fmt_year = int(str_time[:4])
        fmt_month = int(str_time[4:6])
        fmt_day = int(str_time[6:8])
        fmt_hour = int(str_time[8:10])
        fmt_minute = int(str_time[10:12])
        fmt_date = datetime(fmt_year, fmt_month, fmt_day, fmt_hour, fmt_minute)
        new_fmt_date = fmt_date + timedelta(minutes=interval_time)
        new_fmt_date_split = re.split(r"[-\s:]", str(new_fmt_date))
        str_new_fmt_date = ''.join(new_fmt_date_split)[:12]     # 去掉末尾的秒钟信息
    elif len(str_time) == 4:    # daily和weekly事件时间计算
        old_hour = int(str_time[:2])
        old_minute = int(str_time[2:])
        new_hour = 0
        new_minute = 0
        if old_minute + interval_time < 60:
            new_minute = old_minute + interval_time
            new_hour = old_hour
        elif old_minute + interval_time >= 60:
            new_minute = (old_minute + interval_time) - 60
            if old_hour + 1 < 24:
                new_hour = old_hour + 1
            elif old_hour + 1 >= 24:
                new_hour = (old_hour + 1) - 24
        str_new_fmt_date = "{0:02d}".format(new_hour) + "{0:02d}".format(new_minute)
    elif len(str_time) == 5:    # duration时间计算
        old_hour = int(str_time[:2])
        old_minute = int(str_time[3:])
        new_hour = 0
        new_minute = 0
        if old_minute + interval_time < 60:
            new_minute = old_minute + interval_time
            new_hour = old_hour
        elif old_minute + interval_time >= 60:
            new_minute = (old_minute + interval_time) - 60
            if old_hour + 1 < 24:
                new_hour = old_hour + 1
            elif old_hour + 1 >= 24:
                new_hour = (old_hour + 1) - 24
        str_new_fmt_date = "{0:02d}".format(new_hour) + ":" + "{0:02d}".format(new_minute)
    return str_new_fmt_date


def manage_report_data_and_write_data():
    logging.info("manage_report_data_and_write_data")
    # 整理数据以及写数据
    GL.report_data[4] = TEST_CASE_INFO[0]   # 用例编号
    GL.report_data[5] = str(datetime.now())[:19]    # 写该用例报告的时间

    logging.info(GL.report_data)
    time.sleep(2)


def write_data_to_excel():
    logging.info("write_data_to_excel")
    wb = ''
    ws = ''
    excel_title_1 = ["用例编号", "新增预约事件1信息", "新增预约事件2信息", "新增事件结果"]
    excel_title_2 = ["用例编号", "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
                     "起始时间", "事件类型", "节目名称", "持续时间", "事件模式",
                     "事件列表预约事件个数", "无效事件提示", "用例测试时间"]

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    blue_font = Font(color=BLUE)
    red_font = Font(color=RED)
    a_column_numb = column_index_from_string("A")
    if not os.path.exists(file_path[1]):
        wb = Workbook()
        ws = wb.active
        ws.title = file_path[2]
        # 写excel_title_1的内容
        ws.cell(1, 1).value = excel_title_1[0]
        ws["A" + str(1)].alignment = alignment

        ws.cell(1, 2).value = excel_title_1[1]
        ws["B" + str(1)].alignment = alignment
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

        ws.cell(1, 7).value = excel_title_1[2]
        ws["G" + str(1)].alignment = alignment
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)

        ws.cell(1, 12).value = excel_title_1[3]
        ws["L" + str(1)].alignment = alignment
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=14)

        # 写excel_title_2的内容
        for j in range(len(excel_title_2)):
            ws.cell(2, j + 1).value = excel_title_2[j]
            ws.cell(2, j + 1).alignment = alignment
            if j == 0:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
            else:
                ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 10
        # 设置Title的行高
        ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
        ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
        # 合并用例编号单元格
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    elif os.path.exists(file_path[1]):
        wb = load_workbook(file_path[1])
        sheets_name_list = wb.sheetnames
        logging.info(sheets_name_list)
        if file_path[2] in sheets_name_list:
            ws = wb[file_path[2]]
        elif file_path[2] not in sheets_name_list:
            ws = wb.create_sheet(file_path[2])
            # 写excel_title_1的内容
            ws.cell(1, 1).value = excel_title_1[0]
            ws["A" + str(1)].alignment = alignment

            ws.cell(1, 2).value = excel_title_1[1]
            ws["B" + str(1)].alignment = alignment
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

            ws.cell(1, 7).value = excel_title_1[2]
            ws["G" + str(1)].alignment = alignment
            ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)

            ws.cell(1, 12).value = excel_title_1[3]
            ws["L" + str(1)].alignment = alignment
            ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=14)

            # 写excel_title_2的内容
            for j in range(len(excel_title_2)):
                ws.cell(2, j + 1).value = excel_title_2[j]
                ws.cell(2, j + 1).alignment = alignment
                if j == 0:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 6
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + j)].width = 10
            # 设置Title的行高
            ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高
            ws.row_dimensions[2].height = 30  # 设置每次执行的report预约事件信息的行高
            # 合并用例编号单元格
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # 获取当前用例修改类型的sheet表的Max_row
    max_row = ws.max_row

    # 写新增和编辑后的预约事件数据
    len_res_1 = len(GL.report_data[0])          # 新增预约事件的长度
    len_res_2 = len(GL.report_data[1])          # 编辑修改后的预约事件的长度
    len_total = len_res_1 + len_res_2           # 新增和编辑修改后的预约事件的总长度
    for d in range(len(GL.report_data)):
        if d == 0:      # 新增预约事件的信息
            for add_data in range(len(GL.report_data[d])):
                ws.cell(max_row + 1, add_data + 2).value = GL.report_data[d][add_data]
                ws.cell(max_row + 1, add_data + 2).alignment = alignment
                # ws.column_dimensions[get_column_letter(a_column_numb + add_data + 1)].width = 15

        elif d == 1:    # 修改后的预约事件的信息
            for edit_data in range(len(GL.report_data[d])):
                ws.cell(max_row + 1, len_res_1 + edit_data + 2).value = GL.report_data[d][edit_data]
                ws.cell(max_row + 1, len_res_1 + edit_data + 2).alignment = alignment

        elif d == 4:    # 用例编号
            ws.cell(max_row + 1, 1).value = GL.report_data[d]
            ws.cell(max_row + 1, 1).alignment = alignment

        elif d == 5:    # 写报告时间
            ws.cell(max_row + 1, d + len_total - 1).value = GL.report_data[d]   # 由于d==7的坑填到第一列，所以这里需要列数减一
            ws.cell(max_row + 1, d + len_total - 1).alignment = alignment

        else:
            ws.cell(max_row + 1, d + len_total).value = GL.report_data[d]
            ws.cell(max_row + 1, d + len_total).alignment = alignment
    ws.row_dimensions[(max_row + 1)].height = 27    # 设置每次执行的report预约事件信息的行高

    wb.save(file_path[1])


def before_cycle_test_clear_data_and_state():
    # 循环测试前，清理数据和状态变量
    logging.info("before_cycle_test_clear_data_and_state")
    # GL.res_event_mgr.clear()
    GL.choice_res_ch = ''
    state["clear_variate_state"] = True
    GL.pvr_rec_dur_time = ''
    GL.report_data = [[], [], '', '', '', '', ]
    GL.res_event_mgr.clear()
    GL.res_event_info_1.clear()
    GL.res_triggered_numb -= 1
    logging.info("循环测试，延时5秒")
    time.sleep(5)
    logging.info(f"剩余循环次数：{GL.res_triggered_numb}")

    if GL.res_triggered_numb < 1:
        logging.info("程序结束")
        state["receive_loop_state"] = True  # 触发结束接收进程的状态


def calculate_expected_event_2_start_time():
    # 对新增的事件进行计算，修改后的预期起始时间
    logging.info("calculate_expected_event_2_start_time")
    time_interval = 5
    str_expected_event_2_start_time = ''
    start_time = GL.report_data[0][0]       # 原新增预约事件的起始时间
    if TEST_CASE_INFO[7] == "Same(time+type+mode)":
        logging.info("当前编辑不涉及修改时间和修改Mode，所以预约事件时间不变")
        str_expected_event_2_start_time = start_time
    elif TEST_CASE_INFO[7] == "Same(time+type)+Diff(mode)":
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time[8:]
        elif TEST_CASE_INFO[3] == "Daily" or TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                # 起始时间为["2355", "2356", "2357", "2358", "2359"]时，新增+5分钟后，会出现跨日期的现象，导致程序出错
                if start_time in ["0000", "0001", "0002", "0003", "0004"]:
                    str_expected_event_2_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)
                else:
                    str_expected_event_2_start_time = sys_time_date + start_time
    elif TEST_CASE_INFO[7] == "Same(time+mode)+Diff(type)":
        logging.info("当前编辑不涉及修改时间和修改Mode，所以预约事件时间不变")
        str_expected_event_2_start_time = start_time
    elif TEST_CASE_INFO[7] == "Same(time)+Diff(type+mode)":
        logging.info("当前编辑不涉及修改时间和修改Mode，所以预约事件时间不变")
        logging.info("当前编辑涉及修改Mode，所以预约事件时间需要变化")
        if TEST_CASE_INFO[3] == "Once":  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"单次事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time[8:]
        elif TEST_CASE_INFO[3] == "Daily" or TEST_CASE_INFO[3] in WEEKLY_EVENT_MODE:  # 原事件Mode
            if TEST_CASE_INFO[8] == "Daily" or TEST_CASE_INFO[8] in WEEKLY_EVENT_MODE:  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改循环事件{TEST_CASE_INFO[8]}")
                str_expected_event_2_start_time = start_time
            elif TEST_CASE_INFO[8] == "Once":  # 新事件Mode
                logging.info(f"循环事件:{TEST_CASE_INFO[3]}--改单次事件{TEST_CASE_INFO[8]}")
                sys_time = rsv_kws['current_sys_time']
                logging.info(sys_time)
                logging.info(start_time)
                sys_time_split = re.split(r"[\s:/]", sys_time)
                fmt_sys_time = ''.join(sys_time_split)
                sys_time_date = fmt_sys_time[:8]
                # 起始时间为["2355", "2356", "2357", "2358", "2359"]时，新增+5分钟后，会出现跨日期的现象，导致程序出错
                if start_time in ["0000", "0001", "0002", "0003", "0004"]:
                    str_expected_event_2_start_time = cale_str_time_for_add_day((sys_time_date + start_time), 1)
                else:
                    str_expected_event_2_start_time = sys_time_date + start_time

    logging.info(f"期望的完整的预约事件时间为{str_expected_event_2_start_time}")
    return str_expected_event_2_start_time


def calculate_expected_event_2_duration_time():
    logging.info("calculate_expected_edit_event_duration_time")
    str_expected_dur_time = ''
    interval_dur = 1        # 更改录制时长的变量
    specified_dur_time = "0002"   # ModifyType+ModifyDuration时会出现无Dur_time改有Dur_time的情况，直接指定时间

    if TEST_CASE_INFO[7] == "ModifyDuration" or "ModifyDuration" in TEST_CASE_INFO[7]:
        if TEST_CASE_INFO[4] == "PVR" and TEST_CASE_INFO[9] == "PVR":   # 单项ModifyDuration
            dur_time = GL.res_event_mgr[0][3]  # 原新增预约事件的持续时间
            res_dur_split = re.split(":", dur_time)
            res_dur_hour_info = int(res_dur_split[0])
            res_dur_minute_info = int(res_dur_split[1])
            new_hour = 0
            new_minute = 0
            if res_dur_minute_info + interval_dur < 60:
                new_minute = res_dur_minute_info + interval_dur
                new_hour = res_dur_hour_info
            elif res_dur_minute_info + interval_dur >= 60:
                new_minute = (res_dur_minute_info + interval_dur) - 60
                if res_dur_hour_info + 1 < 24:
                    new_hour = res_dur_hour_info + 1
                elif res_dur_hour_info + 1 >= 24:  # 等于24的情况可能会出现问题，但是目前的用例应该不会遇到
                    new_hour = (res_dur_hour_info + 1) - 24
            new_dur_time = "{0:02d}".format(new_hour) + "{0:02d}".format(new_minute)
            str_expected_dur_time = new_dur_time
        elif TEST_CASE_INFO[4] != "PVR" and TEST_CASE_INFO[9] == "PVR":   # 多项修改且包含ModifyDuration
            dur_time = GL.res_event_mgr[0][3]  # 原新增预约事件的持续时间
            if dur_time == "--:--":
                str_expected_dur_time = specified_dur_time
        else:
            logging.info(f"请注意，当前事件不是PVR事件，而是{TEST_CASE_INFO[4]}事件")

    else:
        logging.info("当前事件不需要更改Duration，保持默认值")
        str_expected_dur_time = "0001"
    return str_expected_dur_time


def create_expected_event_2_info():
    logging.info("create_expected_event_2_info")
    # 创建修改后的期望的事件信息
    expected_event_2_info = ['', '', '', '', '']      # [起始时间，事件响应类型，节目名称，持续时间，事件触发模式]
    duration_time = calculate_expected_event_2_duration_time()
    if TEST_CASE_INFO[9] == "Play":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = channel_info[1]
        expected_event_2_info[3] = "--:--"
        expected_event_2_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "PVR":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = channel_info[1]
        expected_event_2_info[3] = duration_time
        expected_event_2_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "Power Off":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = "----"
        expected_event_2_info[3] = "--:--"
        expected_event_2_info[4] = TEST_CASE_INFO[8]

    elif TEST_CASE_INFO[9] == "Power On":
        expected_event_full_time = calculate_expected_event_2_start_time()
        expected_event_2_info[0] = expected_event_full_time
        expected_event_2_info[1] = TEST_CASE_INFO[9]
        expected_event_2_info[2] = "----"
        expected_event_2_info[3] = "--:--"
        expected_event_2_info[4] = TEST_CASE_INFO[8]
    return expected_event_2_info


def edit_add_new_res_event_2_info():
    logging.info("edit_add_new_res_event_2_info")
    # 编辑预约事件信息
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    start_date_list = []        # 用于将开始日期由字符串转化为发送指令的列表
    start_time_list = []        # 用于将开始时间由字符串转化为发送指令的列表
    duration_time_list = []     # 用于将录制持续时间由字符转化为发送指令的列表
    # 进入事件编辑界面
    send_cmd(KEY["GREEN"])
    # 生成预期的预约事件
    expected_res_event_info = create_expected_event_2_info()

    # 根据用例来编辑不同的事件
    # 检查是否进入到Timer Edit界面
    while rsv_kws["edit_event_focus_pos"] == "":
        time.sleep(2)       # 用于还没有进入和接收到焦点关键字时加的延时
    # 设置Mode参数
    logging.info("Edit Mode")
    while rsv_kws["edit_event_focus_pos"] != "Mode":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_mode"] != TEST_CASE_INFO[9]:
            logging.info(f'Mode参数与预期不符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[9]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Mode参数与预期相符:{rsv_kws["edit_event_mode"]}--{TEST_CASE_INFO[9]}')
            send_cmd(KEY["DOWN"])
    # 设置Type参数
    logging.info("Edit Type")
    while rsv_kws["edit_event_focus_pos"] != "Type":
        send_cmd(KEY["DOWN"])
    else:
        while rsv_kws["edit_event_type"] != TEST_CASE_INFO[8]:
            logging.info(f'Type参数与预期不符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[8]}')
            send_cmd(KEY["RIGHT"])
        else:
            logging.info(f'Type参数与预期相符:{rsv_kws["edit_event_type"]}--{TEST_CASE_INFO[8]}')
            send_cmd(KEY["DOWN"])
    # 设置Start_Date参数
    logging.info("Edit Start Date")
    if TEST_CASE_INFO[8] != "Once":
        logging.info(f"当前事件触发模式为循环模式，不需要设置Start Date：{TEST_CASE_INFO[8]}")
    elif TEST_CASE_INFO[8] == "Once":
        logging.info(f"当前事件触发模式为单次模式，需要设置Start Date：{TEST_CASE_INFO[8]}")
        while rsv_kws["edit_event_focus_pos"] != "Start Date":
            send_cmd(KEY["DOWN"])
        else:
            start_date_list.append(expected_res_event_info[0][:8])
            start_date_cmd = change_numbs_to_cmds_list(start_date_list)
            for i in range(len(start_date_cmd)):
                for j in start_date_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Start_Time参数
    logging.info("Edit Start Time")
    while rsv_kws["edit_event_focus_pos"] != "Start Time":
        send_cmd(KEY["DOWN"])
    else:
        if len(expected_res_event_info[0]) == 12:
            start_time_list.append(expected_res_event_info[0][8:])
        elif len(expected_res_event_info[0]) == 4:
            start_time_list.append(expected_res_event_info[0])
        start_time_cmd = change_numbs_to_cmds_list(start_time_list)
        for i in range(len(start_time_cmd)):
            for j in start_time_cmd[i]:
                send_cmd(j)
        send_cmd(KEY["DOWN"])
    # 设置Duration参数
    logging.info("Edit Duration")
    if TEST_CASE_INFO[9] != "PVR":
        logging.info(f"当前事件类型不为PVR，不需要设置Duration：{TEST_CASE_INFO[9]}")
    elif TEST_CASE_INFO[9] == "PVR":
        logging.info(f"当前事件类型为PVR，需要设置Duration：{TEST_CASE_INFO[9]}")
        while rsv_kws["edit_event_focus_pos"] != "Duration":
            send_cmd(KEY["DOWN"])
        else:
            duration_time_list.append(expected_res_event_info[3])
            duration_time_cmd = change_numbs_to_cmds_list(duration_time_list)
            for i in range(len(duration_time_cmd)):
                for j in duration_time_cmd[i]:
                    send_cmd(j)
            send_cmd(KEY["DOWN"])
    # 设置Channel参数
    logging.info("Edit Channel")
    if TEST_CASE_INFO[9] == "Power Off" or TEST_CASE_INFO[9] == "Power On":
        logging.info(f"当前事件类型为：{TEST_CASE_INFO[9]}，不需要设置Channel")
    elif TEST_CASE_INFO[9] != "Power Off":
        logging.info(f"当前事件类型不为Power Off/On，需要设置Channel：{TEST_CASE_INFO[9]}")
        while rsv_kws["edit_event_focus_pos"] != "Channel":
            send_cmd(KEY["DOWN"])
        else:
            if rsv_kws["edit_event_ch"] == GL.choice_res_ch:
                logging.info(f"当前节目与所选节目一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")
            else:
                logging.info(f"警告：当前节目与所选节目不一致：{rsv_kws['edit_event_ch']}--{GL.choice_res_ch}")

    # 退出保存
    state["clear_res_event_list_state"] = True
    send_cmd(KEY["EXIT"])
    send_cmd(KEY["OK"])
    # 此处编辑完事件后，不会再次打印事件信息，需要重新进入Timer Setting界面，为update_edit_res_event_to_event_mgr_list准备数据
    time.sleep(2)
    # 更新event_2的信息
    if TEST_CASE_INFO[9] == "PVR":
        new_expected_res_event_info = expected_res_event_info
        dur_time = new_expected_res_event_info[3]
        new_expected_res_event_info[3] = dur_time[:2] + ":" + dur_time[2:]
        GL.report_data[1] = new_expected_res_event_info
    else:
        GL.report_data[1] = expected_res_event_info
    GL.report_data[3] = rsv_kws["event_invalid_msg"]
    # send_cmd(KEY["EXIT"])
    # send_cmd(KEY["OK"])
    if rsv_kws["event_invalid_msg"] != '':
        send_cmd(KEY["OK"])
    # 退回大画面
    send_more_cmds(exit_to_screen)


def update_edit_res_event_to_event_mgr_list():
    logging.info("update_edit_res_event_to_event_mgr_list")
    # 先清除之前的新增事件
    GL.res_event_mgr.clear()
    # 添加编辑修改后的预约事件到事件管理列表
    if list(res_event_list) not in GL.res_event_mgr:
        GL.res_event_mgr.extend(list(res_event_list))
    GL.report_data[1] = GL.res_event_mgr[0]
    logging.info(type(GL.res_event_mgr))
    logging.info(GL.res_event_mgr)
    logging.info(list(res_event_list))


def new_add_res_event_2():
    logging.info("edit_res_event")
    # 编辑修改新增的预约事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    # 进入Timer_Setting界面
    send_more_cmds(enter_timer_setting_interface)
    # 进入事件编辑界面，设置预约事件参数
    edit_add_new_res_event_2_info()


def check_event_numb():
    # 检查Timer_setting界面所有的事件
    enter_timer_setting_interface = [KEY["MENU"], KEY["LEFT"], KEY["DOWN"], KEY["OK"]]
    exit_to_screen = [KEY["EXIT"], KEY["EXIT"]]
    # 进入定时器设置界面
    send_more_cmds(enter_timer_setting_interface)
    # 对定时器设置界面的事件判断和清除
    time.sleep(1)
    while not state["res_event_numb_state"]:
        logging.info("还没有获取到预约事件个数")
        time.sleep(1)
    else:
        logging.info(rsv_kws["res_event_numb"])
        GL.report_data[2] = rsv_kws["res_event_numb"]
        state["res_event_numb_state"] = False
    # 退回大画面
    send_more_cmds(exit_to_screen)


def receive_serial_process(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info,
        receive_cmd_list):
    logging_info_setting()
    rsv_key = {
        "POWER": "0xbbaf", "TV/R": "0xbbbd", "MUTE": "0xbbf7",
        "1": "0xbb7f", "2": "0xbbbf", "3": "0xbb3f",
        "4": "0xbbdf", "5": "0xbb5f", "6": "0xbb9f",
        "7": "0xbb1f", "8": "0xbbef", "9": "0xbb6f",
        "FAV": "0xbb87", "0": "0xbbff", "SAT": "0xbb97",
        "MENU": "0xbbcf", "EPG": "0xbb8f", "INFO": "0xbb07", "EXIT": "0xbb4f",
        "UP": "0xbb77", "DOWN": "0xbbd7",
        "LEFT": "0xbbb7", "RIGHT": "0xbb37", "OK": "0xbb57",
        "P/N": "0xbb0f", "SLEEP": "0xbb17", "PAGE_UP": "0xbb7d", "PAGE_DOWN": "0xbbe7",
        "RED": "0xbb67", "GREEN": "0xbba7", "YELLOW": "0xbb27", "BLUE": "0xbbc7",
        "F1": "0xbb9d", "F2": "0xbb5d", "F3": "0xbbdd", "RECALL": "0xbb3d",
        "REWIND": "0xbb47", "FF": "0xbb1d", "PLAY": "0xbb2f", "RECORD": "0xbbfd",
        "PREVIOUS": "0xbbad", "NEXT": "0xbb6d", "TIME_SHIFT": "0xbbed", "STOP": "0xbb4d"
    }
    reverse_rsv_key = dict([val, key] for key, val in rsv_key.items())

    res_kws = [
        "[PTD]Time_mode=",              # 0     获取系统时间模式
        "[PTD]System_time=",            # 1     系统时间
        "[PTD]Res_event_numb=",         # 2     预约事件数量
        "[PTD]Res_event:",              # 3     预约事件信息
        "[PTD]Res_triggered:",          # 4     预约事件触发和当前响应事件的信息
        "[PTD]Res_confirm_jump",        # 5     预约事件确认跳转
        "[PTD]Res_cancel_jump",         # 6     预约事件取消跳转
        "[PTD]REC_start",               # 7     录制开始
        "[PTD]REC_end",                 # 8     录制结束
        "[PTD]No_storage_device",       # 9     没有存储设备
        "[PTD]No_enough_space",         # 10    没有足够的空间
        "[PTD]power_cut",               # 11    进入待机
        "[PTD]:switch totle cost",      # 12    开机解码成功
        "[PTD][HOTPLUG] PLUG_IN",       # 13    存储设备插入成功
        "[PTD]PVR_is_not_supported",    # 14    录制无信号、加锁节目、加密节目，跳出PVR is not supported!提示
        "[PTD]Current_sys_time",        # 15    预约事件触发时，检测触发时间信息
    ]

    switch_ch_kws = [
        "[PTD]Prog_numb=",
        "[PTD]TP=",
        "[PTD]video_height=",
        "[PTD]Group_name=",
        "Swtich Video interval"]

    ch_info_kws = [
        "Prog_numb",
        "Prog_name",
        "TP",
        "Lock_flag",
        "Scramble_flag",
        "Prog_type",
        "Group_name"]

    group_info_kws = [
        "Group_name",
        "Prog_total"
    ]

    edit_event_kws = [
        "[PTD]Mode=",
        "[PTD]Type=",
        "[PTD]Start Date=",
        "[PTD]Start Time=",
        "[PTD]Duration=",
        "[PTD]Channel="
    ]

    event_invalid_msg = [
        "[PTD]Res_no_channel",
        "[PTD]Res_invalid_date",
        "[PTD]Res_invalid_timer"
    ]

    other_kws = [
        "[PTD]Infrared_key_values:",    # 获取红外接收关键字
    ]

    infrared_rsv_cmd = []       # 红外接受命令
    receive_serial = serial.Serial(prs_data["receive_serial_name"], 115200, timeout=1)

    while True:
        data = receive_serial.readline()
        if data:
            tt = datetime.now()
            # data1 = data.decode("GB18030", "ignore")
            data1 = data.decode("ISO-8859-1", "ignore")
            data2 = re.compile('[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]').sub('', data1).strip()
            data3 = "[{}]     {}\n".format(str(tt), data2)
            print(data2)
            write_log_data_to_txt(prs_data["log_file_path"], data3)

            if state["clear_variate_state"]:
                state["sys_time_mode_state"] = False
                state["current_sys_time_state"] = False
                state["res_event_numb_state"] = False
                state["res_event_triggered_state"] = False
                state["res_event_confirm_jump_state"] = False
                state["res_event_cancel_jump_state"] = False
                state["rec_start_state"] = False
                state["rec_end_state"] = False
                state["no_storage_device_state"] = False
                state["no_enough_space_state"] = False
                state["pvr_not_supported_state"] = False
                state["update_event_list_state"] = False
                state["clear_variate_state"] = False
                state["power_off_state"] = False
                state["stb_already_power_on_state"] = False
                state["event_no_channel_msg_state"] = False
                state["event_invalid_date_msg_state"] = False
                state["event_invalid_timer_msg_state"] = False

                rsv_kws["event_invalid_msg"] = ''

                # del res_event_list[:]
                del current_triggered_event_info[:]
                if prs_data["case_res_event_mode"] == "Once":
                    del res_event_list[:]
                # channel_info = ['', '', '', '', '', '', '']

            if state["clear_res_event_list_state"]:
                del res_event_list[:]
                state["clear_res_event_list_state"] = False

            if other_kws[0] in data2:   # 红外接收打印
                rsv_cmd = re.split(":", data2)[-1]
                infrared_rsv_cmd.append(rsv_cmd)        # 存放可以共享的接受命令的列表
                if rsv_cmd not in reverse_rsv_key.keys():
                    logging.info("红外键值{}不在当前字典中，被其他遥控影响".format(rsv_cmd))
                else:
                    logging.info("红外键值(发送和接受):({})--({})".format(
                        infrared_send_cmd[-1], reverse_rsv_key[infrared_rsv_cmd[-1]]))
                    logging.info("红外次数统计(发送和接受):{}--{}".format(
                        len(infrared_send_cmd), len(infrared_rsv_cmd)))
                    receive_cmd_list.append(rsv_cmd)
                if state["control_power_on_info_rsv_state"]:    # 用于Power Off事件触发后，软开机没有检测到开机关键字用来避免死循环
                    state["stb_already_power_on_state"] = True

            if res_kws[0] in data2:     # 获取系统时间模式（自动还是手动）
                state["sys_time_mode_state"] = True
                rsv_kws["sys_time_mode"] = re.split(r"=", data2)[-1]

            if res_kws[1] in data2:     # 获取当前系统时间
                state["current_sys_time_state"] = True
                rsv_kws["current_sys_time"] = re.split(r"=", data2)[-1]

            if res_kws[2] in data2:     # 获取预约事件数量
                state["res_event_numb_state"] = True
                rsv_kws["res_event_numb"] = re.split(r"=", data2)[-1]

            if res_kws[3] in data2:     # 获取预约事件信息
                # state["res_event_info_state"] = True
                event_split_info = re.split(r"event:|,", data2)
                event_info = ['', '', '', '', '']
                for info in event_split_info:
                    if "Start_time" in info:
                        event_start_time = re.split(r"=", info)[-1]
                        if len(event_start_time) == 5:
                            event_info[0] = ''.join(re.split(r":", event_start_time))
                        elif len(event_start_time) == 16:
                            event_info[0] = ''.join(re.split(r"[/:\s]", event_start_time))
                    if "Event_type" in info:
                        event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        event_info[4] = re.split(r"=", info)[-1]
                if state["update_event_list_state"]:
                    res_event_list.append(event_info)

            if res_kws[4] in data2:     # 获取预约事件跳转触发信息，以及当前响应事件的信息
                state["res_event_triggered_state"] = True
                current_event_split_info = re.split(r"triggered:|,", data2)
                current_event_info = ['', '', '', '', '']
                for info in current_event_split_info:
                    if "Start_time" in info:
                        current_event_start_time = re.split(r"=", info)[-1]
                        if len(current_event_start_time) == 5:
                            current_event_info[0] = ''.join(re.split(r":", current_event_start_time))
                        elif len(current_event_start_time) == 16:
                            current_event_info[0] = ''.join(re.split(r"[/:\s]", current_event_start_time))
                    if "Event_type" in info:
                        current_event_info[1] = re.split(r"=", info)[-1]
                    if "Ch_name" in info:
                        current_event_info[2] = re.split(r"=", info)[-1]
                    if "Duration" in info:
                        current_event_info[3] = re.split(r"=", info)[-1]
                    if "Event_mode" in info:
                        current_event_info[4] = re.split(r"=", info)[-1]
                current_triggered_event_info.extend(current_event_info)

            if res_kws[5] in data2:     # 获取预约事件确认跳转信息
                state["res_event_confirm_jump_state"] = True

            if res_kws[6] in data2:     # 获取预约事件取消跳转信息
                state["res_event_cancel_jump_state"] = True

            if res_kws[7] in data2:     # 获取PVR预约事件录制开始信息
                state["rec_start_state"] = True

            if res_kws[8] in data2:     # 获取PVR预约事件录制结束信息
                state["rec_end_state"] = True

            if res_kws[9] in data2:     # 存储设备没有插入的打印信息
                state["no_storage_device_state"] = True

            if res_kws[10] in data2:    # 存储设备没有足够空间的打印信息
                state["no_enough_space_state"] = True

            if res_kws[11] in data2:    # 软关机打印信息
                state["power_off_state"] = True

            if res_kws[12] in data2 or res_kws[13] in data2:    # 开机解码成功打印信息,或开机存储设备挂载成功信息
                if state["control_power_on_info_rsv_state"]:
                    state["stb_already_power_on_state"] = True

            if res_kws[14] in data2:    # 录制无信号、加锁节目、加密节目，跳出PVR is not supported!提示
                state["pvr_not_supported_state"] = True

            if res_kws[15] in data2:    # 预约事件触发时系统时间信息(备注：此系统时间为年月日 时分秒信息)
                cur_sys_time = re.split(r"=", data2)[-1]
                cur_sys_time_split = re.split(r"[/\s:]", cur_sys_time)
                rsv_kws["res_triggered_sys_time"] = ''.join(cur_sys_time_split)

            if switch_ch_kws[0] in data2:
                ch_info_split = re.split(r"[\],]", data2)
                for i in range(len(ch_info_split)):
                    if ch_info_kws[0] in ch_info_split[i]:  # 提取频道号
                        channel_info[0] = re.split("=", ch_info_split[i])[-1]
                    if ch_info_kws[1] in ch_info_split[i]:  # 提取频道名称
                        channel_info[1] = re.split("=", ch_info_split[i])[-1]

            if switch_ch_kws[1] in data2:
                flag_info_split = re.split(r"[\],]", data2)
                for i in range(len(flag_info_split)):
                    if ch_info_kws[2] in flag_info_split[i]:  # 提取频道所属TP
                        channel_info[2] = re.split(r"=", flag_info_split[i])[-1].replace(" ", "")
                    if ch_info_kws[3] in flag_info_split[i]:  # 提取频道Lock_flag
                        channel_info[3] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[4] in flag_info_split[i]:  # 提取频道Scramble_flag
                        channel_info[4] = re.split(r"=", flag_info_split[i])[-1]
                    if ch_info_kws[5] in flag_info_split[i]:  # 提取频道类别:TV/Radio
                        channel_info[5] = re.split(r"=", flag_info_split[i])[-1]

            if switch_ch_kws[3] in data2:
                group_info_split = re.split(r"[\],]", data2)
                for i in range(len(group_info_split)):
                    if group_info_kws[0] in group_info_split[i]:  # 提取频道所属组别
                        rsv_kws["prog_group_name"] = re.split(r"=", group_info_split[i])[-1]
                        channel_info[6] = rsv_kws["prog_group_name"]
                    if group_info_kws[1] in group_info_split[i]:  # 提取频道所属组别下的节目总数
                        rsv_kws["prog_group_total"] = re.split(r"=", group_info_split[i])[-1]

            if edit_event_kws[0] in data2:          # 提取Mode参数
                rsv_kws["edit_event_focus_pos"] = "Mode"
                rsv_kws["edit_event_mode"] = re.split(r"=", data2)[-1]

            if edit_event_kws[1] in data2:          # 提取Type参数
                rsv_kws["edit_event_focus_pos"] = "Type"
                rsv_kws["edit_event_type"] = re.split(r"=", data2)[-1]

            if edit_event_kws[2] in data2:          # 提取Start Date参数
                rsv_kws["edit_event_focus_pos"] = "Start Date"
                rsv_kws["edit_event_date"] = re.split(r"=", data2)[-1]

            if edit_event_kws[3] in data2:          # 提取Start Time参数
                rsv_kws["edit_event_focus_pos"] = "Start Time"
                rsv_kws["edit_event_time"] = re.split(r"=", data2)[-1]

            if edit_event_kws[4] in data2:          # 提取Duration参数
                rsv_kws["edit_event_focus_pos"] = "Duration"
                rsv_kws["edit_event_duration"] = re.split(r"=", data2)[-1]

            if edit_event_kws[5] in data2:          # 提取Channel参数
                rsv_kws["edit_event_focus_pos"] = "Channel"
                rsv_kws["edit_event_ch"] = re.split(r"=", data2)[-1]

            if event_invalid_msg[0] in data2:
                state["event_no_channel_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2

            if event_invalid_msg[1] in data2:
                state["event_invalid_date_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2

            if event_invalid_msg[2] in data2:
                state["event_invalid_timer_msg_state"] = True
                rsv_kws["event_invalid_msg"] = data2


if __name__ == "__main__":

    GL = MyGlobal()
    logging_info_setting()
    msg = "现在开始执行的是:{}_{}_{}_{}_{}_{}_{}_{}".format(
        TEST_CASE_INFO[0], TEST_CASE_INFO[1], TEST_CASE_INFO[2], TEST_CASE_INFO[4],
        TEST_CASE_INFO[3], TEST_CASE_INFO[7], TEST_CASE_INFO[9], TEST_CASE_INFO[8])
    logging.critical(format(msg, '*^150'))
    KEY = {
        "POWER": "A1 F1 22 DD 0A", "TV/R": "A1 F1 22 DD 42", "MUTE": "A1 F1 22 DD 10",
        "1": "A1 F1 22 DD 01", "2": "A1 F1 22 DD 02", "3": "A1 F1 22 DD 03",
        "4": "A1 F1 22 DD 04", "5": "A1 F1 22 DD 05", "6": "A1 F1 22 DD 06",
        "7": "A1 F1 22 DD 07", "8": "A1 F1 22 DD 08", "9": "A1 F1 22 DD 09",
        "FAV": "A1 F1 22 DD 1E", "0": "A1 F1 22 DD 00", "SAT": "A1 F1 22 DD 16",
        "MENU": "A1 F1 22 DD 0C", "EPG": "A1 F1 22 DD 0E", "INFO": "A1 F1 22 DD 1F", "EXIT": "A1 F1 22 DD 0D",
        "UP": "A1 F1 22 DD 11", "DOWN": "A1 F1 22 DD 14",
        "LEFT": "A1 F1 22 DD 12", "RIGHT": "A1 F1 22 DD 13", "OK": "A1 F1 22 DD 15",
        "P/N": "A1 F1 22 DD 0F", "SLEEP": "A1 F1 22 DD 17", "PAGE_UP": "A1 F1 22 DD 41", "PAGE_DOWN": "A1 F1 22 DD 18",
        "RED": "A1 F1 22 DD 19", "GREEN": "A1 F1 22 DD 1A", "YELLOW": "A1 F1 22 DD 1B", "BLUE": "A1 F1 22 DD 1C",
        "F1": "A1 F1 22 DD 46", "F2": "A1 F1 22 DD 45", "F3": "A1 F1 22 DD 44", "RECALL": "A1 F1 22 DD 43",
        "REWIND": "A1 F1 22 DD 1D", "FF": "A1 F1 22 DD 47", "PLAY": "A1 F1 22 DD 0B", "RECORD": "A1 F1 22 DD 40",
        "PREVIOUS": "A1 F1 22 DD 4A", "NEXT": "A1 F1 22 DD 49", "TIME_SHIFT": "A1 F1 22 DD 48", "STOP": "A1 F1 22 DD 4D"
    }
    REVERSE_KEY = dict([val, key] for key, val in KEY.items())
    WAIT_INTERFACE = ["TVScreenDiffCH", "RadioScreenDiffCH", "ChannelList", "Menu", "EPG", "ChannelEdit"]
    WEEKLY_EVENT_MODE = ["Mon.", "Tues.", "Wed.", "Thurs.", "Fri.", "Sat.", "Sun."]
    # TEST_CASE_INFO = ["23", "All", "TV", "Daily", "Play", "EPG", "Manual_jump"]
    EXIT_TO_SCREEN = [KEY["EXIT"], KEY["EXIT"], KEY["EXIT"]]

    file_path = create_log_and_report_file_path()
    ser_name = list(check_ports())  # send_ser_name, receive_ser_name
    send_serial = serial.Serial(ser_name[0], 9600)
    receive_ser_name = ser_name[1]

    infrared_send_cmd = Manager().list([])
    receive_cmd_list = Manager().list([])
    res_event_list = Manager().list([])
    current_triggered_event_info = Manager().list([])
    channel_info = Manager().list(['', '', '', '', '', '', ''])     # [频道号,频道名称,tp,lock,scramble,频道类型,组别]
    rsv_kws = Manager().dict({
        "sys_time_mode": '', "current_sys_time": '', "res_event_numb": '', "prog_group_name": '',
        "prog_group_total": '', "edit_event_focus_pos": '', "edit_event_mode": '', "edit_event_type": '',
        "edit_event_date": '', "edit_event_time": '', "edit_event_duration": '', "edit_event_ch": '',
        "res_triggered_sys_time": '', "event_invalid_msg": '',
    })

    state = Manager().dict({
        "res_event_numb_state": False, "res_event_triggered_state": False, "res_event_confirm_jump_state": False,
        "res_event_cancel_jump_state": False, "rec_start_state": False, "rec_end_state": False,
        "no_storage_device_state": False, "no_enough_space_state": False, "power_off_state": False,
        "sys_time_mode_state": False, "current_sys_time_state": False, "update_event_list_state": False,
        "clear_variate_state": False, "receive_loop_state": False, "control_power_on_info_rsv_state": False,
        "stb_already_power_on_state": False, "res_event_info_state": False, "pvr_not_supported_state": False,
        "clear_res_event_list_state": False, "event_no_channel_msg_state": False, "event_invalid_date_msg_state": False,
        "event_invalid_timer_msg_state": False
    })

    prs_data = Manager().dict({
        "log_file_path": file_path[0], "receive_serial_name": receive_ser_name, "case_res_event_mode": TEST_CASE_INFO[8]
    })

    rsv_p = Process(target=receive_serial_process, args=(
        prs_data, infrared_send_cmd, rsv_kws, res_event_list, state, current_triggered_event_info, channel_info,
        receive_cmd_list))
    rsv_p.start()

    if platform.system() == "Windows":
        time.sleep(5)
        logging.info("Windows系统接收端响应慢，等待5秒")
    elif platform.system() == "Linux":
        time.sleep(1)
        logging.info("Linux系统接收端响应快，但是增加一个延时保护，等待1秒")

    # 主程序开始部分

    while GL.res_triggered_numb > 0:
        clear_timer_setting_all_events()
        check_sys_time_mode()
        choice_ch_for_res_event_type("event_1")
        new_add_res_event_1()
        choice_ch_for_res_event_type("event_2")
        new_add_res_event_2()
        check_event_numb()
        manage_report_data_and_write_data()
        write_data_to_excel()
        before_cycle_test_clear_data_and_state()

    if state["receive_loop_state"]:
        rsv_p.terminate()
        logging.info('stop receive process')
        rsv_p.join()
