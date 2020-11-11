# !/usr/bin/python3
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import RED, BLUE
from openpyxl.utils import get_column_letter, column_index_from_string
import testlink
import os
import logging
import time


def logging_info_setting():
    # 配置logging输出格式
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    date_format = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=log_format, datefmt=date_format)


def get_all_project():
    ret = []
    for i in tlc.getProjects():
        ret.append({'id': i['id'], 'name': i['name']})
    return ret


def get_specify_project_id_by_name(name):
    for i in tlc.getProjects():
        if i['name'] == name:
            return i['id']


def get_all_first_suites_id(project_id):
    all_first_suite_id = {}
    specify_project_first_suites = tlc.getFirstLevelTestSuitesForTestProject(project_id)
    for first_suite in specify_project_first_suites:
        print(first_suite['id'], first_suite['name'])
        all_first_suite_id[first_suite['id']] = first_suite['name']
    return all_first_suite_id


def data_replace(data):
    data = data.replace('<p>', '')
    data = data.replace('</p>', '')
    data = data.replace('&nbsp;', '')
    data = data.replace('&ldquo;', '"')
    data = data.replace('&rdquo;', '"')
    data = data.replace('&quot;', '"')
    data = data.replace('&le;', '<')
    data = data.replace('<span style="color: rgb(255, 0, 0);">', '')
    data = data.replace('</span>', '')
    data = data.replace('<strong>', '')
    data = data.replace('<span style="font-size: larger;">', '')
    data = data.replace('</strong>', '')
    data = data.replace('<br />', '')
    return data


def write_test_case_to_excel(case_info):
    wb = ''
    ws = ''
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    a_column_numb = column_index_from_string("A")
    module_title = ["用例模块名称", "用例标题", "概要", "前提", "步骤动作", "期望的结果"]
    if not os.path.exists(module_file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = module_name

        # 写module_title、设置列宽、行高
        for i in range(len(module_title)):
            ws.cell(1, i + 1).value = module_title[i]
            ws.cell(1, i + 1).alignment = alignment
            if i in [0, 1, 2, 3]:
                ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 20
            else:
                ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 50
        ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高

    elif os.path.exists(module_file_name):
        wb = load_workbook(module_file_name)
        sheets_name_list = wb.sheetnames
        # logging.info(sheets_name_list)
        if module_name in sheets_name_list:
            ws = wb[module_name]
        elif module_name not in sheets_name_list:
            ws = wb.create_sheet(module_name)
            # 写module_title、设置列宽、行高
            for i in range(len(module_title)):
                ws.cell(1, i + 1).value = module_title[i]
                ws.cell(1, i + 1).alignment = alignment
                if i in [0, 1, 2, 3]:
                    ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 20
                else:
                    ws.column_dimensions[get_column_letter(a_column_numb + i)].width = 50
            ws.row_dimensions[1].height = 30  # 设置每次执行的report预约事件信息的行高

    # 获取当前用例修改类型的sheet表的Max_row
    max_row = ws.max_row

    # 写report_data数据
    # if type(case_info) == list and len(case_info) == 5:
    for j in range(len(case_info)):
        if j == 4:
            step_number_action = ''
            step_number_result = ''
            for k in range(len(case_info[j])):
                step_number = list(case_info[j].keys())[k]
                step_number_action += f'步骤{step_number}、{case_info[j][step_number][0]}\n'
                step_number_result += f'步骤{step_number}、{case_info[j][step_number][1]}\n'
            ws.cell(max_row + 1, j + 1).value = step_number_action
            ws.cell(max_row + 1, j + 1).alignment = alignment
            ws.cell(max_row + 1, j + 2).value = step_number_result
            ws.cell(max_row + 1, j + 2).alignment = alignment
        else:
            ws.cell(max_row + 1, j + 1).value = str(case_info[j])
            ws.cell(max_row + 1, j + 1).alignment = alignment

    ws.row_dimensions[(max_row + 1)].height = 70  # 设置每次执行的report预约事件信息的行高
    wb.save(module_file_name)


def get_suite_id_name(suites_id):
    global n, suites_name_list, suites_id_list
    space = '-'
    # print(n)
    suites = tlc.getTestSuitesForTestSuite(suites_id)
    # print(suites)
    if len(suites) > 0:
        if 'id' in suites.keys():
            n += 1
            print("(=========================================)")
            print(f"{4 * n * space} {suites['id']} {suites['name']}")
            suites_name_list.append([suites['name']])
            suites_id_list.append(suites['id'])
            get_suite_id_name(suites['id'])
            print(f"{'=' * 50}")
            if suites['parent_id'] in suites_id_list:
                cur_suite_parent_id_index_pos = suites_id_list.index(suites['parent_id'])
                suites_name_list = suites_name_list[:cur_suite_parent_id_index_pos]
                suites_id_list = suites_id_list[:cur_suite_parent_id_index_pos]
            else:
                print(f"{'#' * 50}")
            n = 1

        elif 'id' not in suites.keys():
            n += 1

            for suite in suites:
                print(f"{4 * n * space} {suites[suite]['id']}, {suites[suite]['name']}")
                # print(f"父id为{suites[suite]['parent_id']}")
                suites_name_list.append([suites[suite]['name']])
                suites_id_list.append(suites[suite]['id'])

                get_suite_id_name(suites[suite]['id'])
                if suite == list(suites.keys())[-1]:
                    print(f"{'='*50}")
                    if suites[suite]['parent_id'] in suites_id_list:
                        cur_suite_parent_id_index_pos = suites_id_list.index(suites[suite]['parent_id'])
                        suites_name_list = suites_name_list[:cur_suite_parent_id_index_pos]
                        suites_id_list = suites_id_list[:cur_suite_parent_id_index_pos]
                else:
                    print(f"{'#'*50}")
            n = 1

    elif len(suites) == 0:
        logging.info(f'本套件{suites_id}下没有套件，为空')
        case_info = []
        print(suites_name_list)
        print(suites_id_list)
        case_info.append(suites_name_list)
        # 当套件下没有套件时，说明此时套件下都是用例
        cases = tlc.getTestCasesForTestSuite(testsuiteid=suites_id, deep=True, details="simple")
        print(f'套件下所有case为{cases}')

        # 分别处理当前套件下的所有的case
        for case in range(len(cases)):
            test_case = tlc.getTestCase(cases[case]['id'])
            print(f"{cases[case]['id']}:{len(test_case)}:{test_case}")
            # 由于每个test_case是一个list，且只有一个元素，所以test_case[0].get('name')就能获取到当前case的名称，其他信息都是如此
            print(test_case[0].get('name'))
            # 每个case下的摘要、前提、步骤、期望的结果
            case_info.append(data_replace(test_case[0].get('name')))
            case_info.append(data_replace(test_case[0].get('summary')))
            case_info.append(data_replace(test_case[0].get('preconditions')))
            # print(f"测试用例摘要为：{test_case[0].get('summary')}")
            # print(f"测试用例前提为：{test_case[0].get('preconditions')}")
            step_action_result = {}
            for m in test_case[0].get("steps"):
                # print('序列:', m.get("step_number"), '1111111111')
                # print('执行步骤:', m.get("actions"), '22222222222')
                # print('预期结果:', m.get("expected_results"), '333333333')
                step_action_result[m.get("step_number")] = ['', '']
                step_action_result[m.get("step_number")][0] += f"{data_replace(m.get('actions'))}"
                step_action_result[m.get("step_number")][1] += f"{data_replace(m.get('expected_results'))}"
            # print(step_action_result)
            case_info.append(step_action_result)
            # print(case_info)
            write_test_case_to_excel(case_info)
            case_info = case_info[:1]

            # for i in test_case:
            #     logging.info(f"测试用例名称为：{i.get('name')}")
            #     if i == test_case[0]:
            #         case_info.append(i.get('name'))
            #         case_info.append(i.get('summary'))
            #         case_info.append(i.get('preconditions'))
            #     print("序列", "执行步骤", "预期结果")
            #     print(f"测试用例摘要为：{i.get('summary')}")
            #     print(f"测试用例前提为：{i.get('preconditions')}")
            #     step_action_result = {}
            #     for m in i.get("steps"):
            #         print('序列:', m.get("step_number"), '1111111111')
            #         print('执行步骤:', m.get("actions"), '22222222222')
            #         print('预期结果:', m.get("expected_results"), '333333333')
            #         step_action_result[m.get("step_number")] = [[], []]
            #         step_action_result[m.get("step_number")][0].append(m.get("actions"))
            #         step_action_result[m.get("step_number")][1].append(m.get("expected_results"))
            #         case_info.append(step_action_result)
            # if m == i.get('steps')[-1]:
            # logging.info(case_info)
            # write_test_case_to_excel(case_info)

        # 当前套件的第一个case的父id在suites_id_list中的位置，用于处理最小套件的平行套件的套件路径
        print(cases[0]['parent_id'])
        parent_id_index_pos = suites_id_list.index(cases[0]['parent_id'])
        print(f'{parent_id_index_pos}--{suites_id_list[:parent_id_index_pos]}---{suites_id_list[:(parent_id_index_pos + 1)]}')
        suites_name_list = suites_name_list[:parent_id_index_pos]
        suites_id_list = suites_id_list[:parent_id_index_pos]

if __name__ == '__main__':
    logging_info_setting()
    # get_suite_id_name(547713)

    url = 'http://git.nationalchip.com/testlink/lib/api/xmlrpc.php'
    key = 'f5df4f1bd2bdd22403ec6b8b118d022c'
    tlc = testlink.TestlinkAPIClient(url, key)

    projects_info = get_all_project()
    for project in projects_info:
        print(project)

    # print(tlc.countProjects())
    # projects = tlc.getProjects()
    # for project in projects:
    #     print(project['id'], '------>', project['name'])

    specify_project_id = get_specify_project_id_by_name('DVBS-4.0')
    specify_project_first_suite_id = get_all_first_suites_id(specify_project_id)

    # dvb_4_suites = tlc.getFirstLevelTestSuitesForTestProject(506862)
    # print(dvb_4_suites)
    # print(len(dvb_4_suites))
    # for dvb_4_suite in dvb_4_suites:
    #     print(dvb_4_suite["id"], dvb_4_suite["name"])

    print(f"{100 * '%'}")
    # n = 0
    # suites_name_list = []
    # suites_id_list = []
    # module_name = '播放控制'
    # module_file_name = '播放控制.xlsx'
    print(specify_project_first_suite_id)

    for id in specify_project_first_suite_id:
        n = 0
        suites_name_list = []
        suites_id_list = []
        module_name = specify_project_first_suite_id[id]
        module_file_name = f'{module_name}.xlsx'
        get_suite_id_name(id)
        # if id == '510983':
        #     get_suite_id_name(id)