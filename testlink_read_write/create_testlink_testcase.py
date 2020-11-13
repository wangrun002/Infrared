# !/usr/bin/python3
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import testlink
import logging
import re


def logging_info_setting():
    # 配置logging输出格式
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"  # 配置输出日志的格式
    date_format = "%Y-%m-%d %H:%M:%S %a"  # 配置输出时间的格式
    logging.basicConfig(level=logging.DEBUG, format=log_format, datefmt=date_format)


class TestLinkHandle(object):

    def __init__(self):
        url = 'http://git.nationalchip.com/testlink/lib/api/xmlrpc.php'
        key = 'f5df4f1bd2bdd22403ec6b8b118d022c'
        self.tlc = testlink.TestlinkAPIClient(url, key)

    def get_all_projects(self):
        # 获取所有的项目列表
        r = []
        for i in self.tlc.getProjects():
            r.append({'id': i['id'], 'name': i['name']})
        return r

    def get_project_id_by_name(self, name):
        # 通过项目名称获取项目Ｉd
        for i in self.tlc.getProjects():
            if i['name'] == name:
                return i['id']
        return ''

    def get_suites(self, project_id):
        # 通过项目id获取项目测试用例集
        r = []
        for i in self.tlc.getFirstLevelTestSuitesForTestProject(project_id):
            r.append({'id': i['id'], 'name': i['name']})
            print({'id': i['id'], 'name': i['name'], 'node_table': i['node_table']})
            # logging.info(i)
        return r

    def get_case_info_from_suite(self, suite_id):
        # 通过suite id获取该套件下的case的详细信息
        return self.tlc.getTestCase(suite_id)

    def delete_project(self, prefix):
        # 删除项目
        try:
            return self.tlc.deleteTestProject(prefix=prefix)[0]['status']
        except:
            return False

    def create_project(self, name, prefix):
        # 创建项目
        if self.get_project_id_by_name(name) == '':
            return self.tlc.createTestProject(testprojectname=name, testcaseprefix=prefix, active=True)[0]['id']
        else:
            self.delete_project(prefix)
            return self.tlc.createTestProject(testprojectname=name, testcaseprefix=prefix, active=True)[0]['id']

    def create_suite(self, project_id, suite_name, parent_id=None):
        # 创建用例集（每一个模块都是一个用例集）
        try:
            if parent_id is None:
                first_suite = self.tlc.createTestSuite(testprojectid=project_id, testsuitename=suite_name,
                                                       details=suite_name)
            else:
                first_suite = self.tlc.createTestSuite(testprojectid=project_id, testsuitename=suite_name,
                                                       details=suite_name, parentid=parent_id)
            return first_suite[0]['id']
        except:
            return False

    def creat_cases_steps(self, steps, test_case_name, test_suite_id, test_project_id,
                          author_login, summary, preconditions):
        # 创建用例和创建测试步骤
        try:
            if len(steps) <= 0:
                self.tlc.initStep('', '', 1)
            for step in steps:
                self.tlc.appendStep(step['actions'], step['expected_results'], step['execution_type'])
            return self.tlc.createTestCase(
                testcasename=test_case_name, testsuiteid=test_suite_id,
                testprojectid=test_project_id, authorlogin=author_login, summary=summary, preconditions=preconditions)
        except:
            return False

    def check_module_suite_exist(self, project_id, cur_module_name):
        first_suite_info_dict = {}
        for i in self.tlc.getFirstLevelTestSuitesForTestProject(project_id):
            first_suite_info_dict[i['name']] = i['id']
        if cur_module_name in list(first_suite_info_dict.keys()):
            logging.debug('项目中存在当前模块')
            return first_suite_info_dict[cur_module_name]
        else:
            logging.debug('项目中不存在当前模块')
            return self.create_suite(project_id, cur_module_name)

    def get_speify_suite_sub_suite_info(self, suite_id):
        specify_suite_info_dict = {}
        suites = self.tlc.getTestSuitesForTestSuite(suite_id)
        print(f'#######{suites}')
        if len(suites) > 0:
            if 'id' in suites.keys():
                specify_suite_info_dict[suites['name']] = suites['id']
                return specify_suite_info_dict
            elif 'id' not in suites.keys():
                for id_key in suites:
                    specify_suite_info_dict[suites[id_key]['name']] = suites[id_key]['id']
                return specify_suite_info_dict
        elif len(suites) == 0:
            return specify_suite_info_dict

    def check_case_suite_exist(self, project_id, suite_id, suite_name_list):
        # 判断excel中的套件名称list是否已经存在，返回最后一个套件的id，作为创建case的套件id
        if len(suite_name_list) > 0:
            for i in range(len(suite_name_list)):
                specify_suite_info_dict = self.get_speify_suite_sub_suite_info(suite_id)    #　查询当前套件id下的所有子套件
                print(specify_suite_info_dict)

                if suite_name_list[i] not in specify_suite_info_dict.keys():    # 套件名称不在当前suite_id的子套件中
                    suite_name_list = suite_name_list[i:]   # 需要创建的套件就是当前套件和之后的所有套件名称列表
                    parent_suite_id = suite_id  # 将当前查询的套件id，作为父套件id

                    for name in suite_name_list:
                        parent_suite_id = TLH.create_suite(project_id, name, parent_id=parent_suite_id)
                    return parent_suite_id
                elif suite_name_list[i] in specify_suite_info_dict.keys():
                    suite_id = specify_suite_info_dict[suite_name_list[i]]
                    if i == len(suite_name_list) - 1:   # 假如suite_name_list中的所有套件都存在，返回最后一个套件的id
                        return suite_id
        elif len(suite_name_list) == 0:
            return suite_id

    def check_and_create_suite_case(self, case_steps, case_name, case_parent_suite_id, project_id, author,
                                    case_summary, case_preconditions):

        # 先检查最后一个suite下没有套件（还是存在套件下有套件和case共存的现象，这个到时候遇到了再看）
        suites = self.tlc.getTestSuitesForTestSuite(case_parent_suite_id)
        if len(suites) > 0:
            logging.debug('警告：该套件下仍存在套件')
        elif len(suites) == 0:
            logging.debug('该套件下没有套件了')
            # 检查该套件下是否存在case（这里的case只是简单的描述信息，要查看case详细信息，需要执行tlc.getTestCase(case_id)）
            last_suite_cases = self.tlc.getTestCasesForTestSuite(
                testsuiteid=case_parent_suite_id, deep=False, details="simple")
            # logging.debug(last_suite_cases)
            # logging.debug(len(last_suite_cases))
            if len(last_suite_cases) == 0:
                logging.debug('没有case，能直接创建case')
                return self.creat_cases_steps(case_steps, case_name, case_parent_suite_id, project_id, author,
                                              case_summary, case_preconditions)
            elif len(last_suite_cases) > 0:
                logging.debug('有case存在，需要比对要创建的case是否已经存在')
                last_suite_case_name_list = []
                for case in last_suite_cases:
                    last_suite_case_name_list.append(case['name'])
                if case_name in last_suite_case_name_list:
                    logging.debug('要创建的case已经存在')
                    return False
                elif case_name not in last_suite_case_name_list:
                    logging.debug('要创建的case不存在')
                    return self.creat_cases_steps(case_steps, case_name, case_parent_suite_id, project_id, author,
                                           case_summary, case_preconditions)


class ExcelTestCaseHandle(object):

    def __init__(self):
        pass

    # @staticmethod
    def get_single_testcase(self, testcase_file_path):
        # 从Ｅxcel中获取单条case的信息
        all_cases = []
        wb = load_workbook(testcase_file_path)
        sheets_name_list = wb.sheetnames
        print(sheets_name_list)
        if len(sheets_name_list) == 1:
            ws = wb[sheets_name_list[0]]
            # print(ws.max_row)
            for row in range(2, ws.max_row + 1):
                single_case = []
                for column in range(1, 7):
                    # print(ws.cell(row=i, column=j).value)
                    single_case.append(ws.cell(row=row, column=column).value)
                all_cases.append(single_case)
        return all_cases, sheets_name_list[0]

    def handle_multi_suite_name(self, suites_name):
        # 处理从excel获取到的单条case的所有的套件名称
        # eval_suite_name_list = []
        # suite_name_list = eval(suites_name)
        # for name in suite_name_list:
        #     if len(name) == 1:
        #         eval_suite_name_list.append(name[0])
        # return eval_suite_name_list
        suite_name_list = eval(suites_name)
        return suite_name_list

    def handle_add_tag_to_obj(self, old_object):
        def f(x):
            return f'<p>{x}</p>'

        new_object = []
        # 假如原数据对象是list，用于步骤和期望结果的数据处理
        if type(old_object) is list:
            for obj in old_object:
                if '\n' in obj:
                    obj_split = list(filter(lambda x: x, re.split(r'\n', obj)))
                    # print(obj_split)
                    new_obj_split = list(map(f, obj_split))
                    # print(new_obj_split)
                    new_object.append('\n'.join(new_obj_split))
                else:
                    new_object.append(f(obj))
        # 假如原数据对象是字符串，用于摘要和前提的数据处理
        elif type(old_object) is str:
            if '\n' in old_object:
                obj_split = list(filter(lambda x: x, re.split(r'\n', old_object)))
                # print(obj_split)
                new_obj_split = list(map(f, obj_split))
                # print(new_obj_split)
                new_object = '\n'.join(new_obj_split)
            else:
                new_object = f(old_object)
        return new_object


    def handle_testcase_step(self, step_action, step_result):

        steps = []
        new_list_action_split = []
        new_list_result_split = []
        if step_action is not None:
            # 处理步骤数＋步骤内容
            action_split = re.split(r'步骤\d+、', step_action)
            list_action_split = list(filter(lambda x: x, action_split))
            new_list_action_split = self.handle_add_tag_to_obj(list_action_split)

        if step_result is not None:
            # 处理步骤数＋期望的结果内容
            result_split = re.split(r'步骤\d+、', step_result)
            list_result_split = list(filter(lambda x: x, result_split))
            new_list_result_split = self.handle_add_tag_to_obj(list_result_split)

        if len(new_list_action_split) == len(new_list_result_split):
            if len(new_list_action_split) != 0:
                for numb in range(len(new_list_action_split)):
                    step_info = dict()
                    step_info['step_number'] = numb + 1
                    step_info['actions'] = new_list_action_split[numb]
                    step_info['expected_results'] = new_list_result_split[numb]
                    step_info['execution_type'] = 1
                    steps.append(step_info)
                return steps
            elif len(new_list_action_split) == 0:
                return steps
        else:
            return False


if __name__ == '__main__':
    logging_info_setting()
    TLH = TestLinkHandle()
    ETCH = ExcelTestCaseHandle()
    projects = TLH.get_all_projects()
    for project in projects:
        print(project)

    specify_project_name = 'test_for_using_testlink'
    pj_id = TLH.get_project_id_by_name(specify_project_name)
    print(f'{specify_project_name}的id为{pj_id}')

    pj_first_suites = TLH.get_suites(pj_id)

    # 单条case的所有信息和当前模块名称
    import_project_name = '这是一个测试集1.xlsx'

    all_cases_info, module_name = ETCH.get_single_testcase(import_project_name)
    for single_case_info in all_cases_info:
        print(single_case_info)

        # 单条case中所有的套件名称
        handle_suite_name_list = ETCH.handle_multi_suite_name(single_case_info[0])
        # print(handle_suite_name_list)

        # 单条case中的用例标题名称
        single_case_name = single_case_info[1]
        # print(single_case_name)

        # 单条case中用例的摘要
        single_case_summary = ETCH.handle_add_tag_to_obj(single_case_info[2])

        # 单条case中用例的前提
        single_case_preconditions = ETCH.handle_add_tag_to_obj(single_case_info[3])

        # 处理单个case中的所有步骤和结果
        single_case_steps_info = ETCH.handle_testcase_step(single_case_info[-2], single_case_info[-1])
        # print(single_case_steps_info)

        # 用例作者
        author_name = 'wangrun'

        # 检查模块套件是否存在，返回模块套件id
        module_suite_id = TLH.check_module_suite_exist(pj_id, module_name)
        print(module_suite_id)

        # 检测excel文件中的套件组中的各个套件在模块套件中是否存在，返回最后一个套件的id
        case_parent_suite_id = TLH.check_case_suite_exist(pj_id, module_suite_id, handle_suite_name_list)

        # 创建单个测试用例
        return_testcase_info = TLH.check_and_create_suite_case(
            single_case_steps_info, single_case_name, case_parent_suite_id, pj_id, author_name,
            single_case_summary, single_case_preconditions)
        print(return_testcase_info)
