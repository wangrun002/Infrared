#!/usr/bin/python
# -*- coding: utf-8 -*-

from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment
from openpyxl.utils import get_column_letter,column_index_from_string
import serial
import serial.tools.list_ports
import re
import time
import os

MAIN_LOOP_STATE = True
SEND_COMMD_STATE = True
OK_KEY_VALUE = "A1 F1 22 DD 15"
EXIT_KEY_VALUE = "A1 F1 22 DD 0D"

XLSX_TITLE = [
                "搜索模式",
                "搜索次数",
                "搜索TP数",
                "搜索节目数",
                "保存TP数",
                "保存节目数",
                "搜索时间",
                {"数据类别":["TP","All","TV","Radio","CH_Name"]},
                "TP"
            ]

COMMD_FILE_NAME = [
                "Z6Sat_6F_Blind_SearchCommand.txt",                 #0
                "Z6Sat_6F_SuperBlind_SearchCommand.txt",            #1

                "Y3Sat_6F_Blind_SearchCommand.txt",                 #2
                "Y3Sat_6F_SuperBlind_SearchCommand.txt",            #3

                "88Sat_6F_Blind_SearchCommand.txt",                 #4直连
                "88Sat_6F_SuperBlind_SearchCommand.txt",            #5直连

                "138Sat_6F_Blind_SearchCommand.txt",                #6
                "138Sat_6F_SuperBlind_SearchCommand.txt",           #7
                "138Sat_6F_BlindAdd_SearchCommand.txt",             #8

                "PLPDSat_6F_Blind_SearchCommand.txt",               #9
                "PLPDSat_6F_SuperBlind_SearchCommand.txt",          #10

                "Z6Sat_6F_UpperLimitTP_SearchCommand.txt",          #11
                "Y3Sat_6F_UpperLimitChannel_SearchCommand.txt",     #12

                'Factory_6F_Reset_SearchCommand.txt',               #13
                'Add_6F_20NewSat_SearchCommand.txt',                #14
                'USBUpgradeUser20SatCommand.txt'                    #15
                ]

choice_commd_file_numb = 11              # 选择想要搜索的卫星的指令文件
search_numb = 72                          # 搜索次数

class MyGlobal():
    def __init__(self):
        self.ser_cable_numb = 4                         # USB转串口线编号
        self.search_datas = [0,0,0,0,0,0,0,0,0]         # 用于存放XLSX_TITLE中的数据
        self.blind_judge_polar = [[],[],set(),""]       # 用于判断极化
        self.xlsx_data_interval = 0                     # 用于计算每轮搜索写xlsx时的列间隔数
        self.all_tp_list = []                           # 用于存放搜索到的TP
        self.channel_info = {}                          # 用于存放各个TP下搜索到的电视和广播节目名称
        self.send_commds_pos = 0                        # 用于显示当前发送指令的位置
        self.send_commds_length = 0                     # 用于显示当前指令集的总长度
        self.tv_radio_tp_data = [0,0,0,0,0]             # [tv_numb,radio_numb,save_tv_numb,save_radio_numb,save_tv_numb]
        self.tv_radio_tp_accumulated = [[],[],[]]       # 用于统计每轮搜索累加的TV、Radio、TP的值

        self.search_monitor_kws = [
            "[PTD]SearchStart",         # 0
            "[PTD]TV------",            # 1
            "[PTD]Radio-----",          # 2
            "[PTD]SearchFinish",        # 3
            "[PTD]get :  fre",          # 4
            "[PTD]TP_save=",            # 5
            "[PTD]TV_save=",            # 6
            "TV_save=",                 # 7
            "Radio_save=",              # 8
            "[PTD]maximum_tp",          # 9
            "[PTD]maximum_channel",     # 10
            "get blind - fre",          # 11
        ]

def check_ports(ser_cable_numb):
    serial_ser_value = {
        "1": "FTDVKA2HA",
        "2": "FTGDWJ64A",
        "3": "FT9SP964A",
        "4": "FTHB6SSTA"
    }
    send_port_desc = "USB-SERIAL CH340"
    receive_port_desc = "USB Serial Port"
    ports = list(serial.tools.list_ports.comports())
    ports_com = []
    if len(ports) <= 0:
        print("无可用端口")
    elif len(ports) == 1:
        print("只有一个可用端口:{}".format(ports[0][0]))
    elif len(ports) >= 2:
        for i in range(len(ports)):
            ports_com.append(str(ports[i]))
            if send_port_desc in str(ports[i]):
                send_com = ports[i][0]
            if receive_port_desc in str(ports[i]) and serial_ser_value[str(GL.ser_cable_numb)] in str(ports[i][2]):
                receive_com = ports[i][0]
                print(ports[i][2])
        print("可用端口:{}".format(ports_com))
    return send_com, receive_com

def serial_set(ser, ser_name, ser_baudrate):
    ser.port = ser_name
    ser.baudrate = ser_baudrate
    ser.bytesize = 8
    ser.parity = "N"
    ser.stopbits = 1
    ser.timeout = 1
    ser.open()

def hex_strs_to_bytes(strings):
    return bytes.fromhex(strings)

def rend_commds_from_txt(txt_path):
    with open(txt_path, "r", encoding="utf-8") as fo:
        send_commds = fo.read().split("\n")
    return send_commds

def write_logs_to_txt(txt_path, logs):
    with open(txt_path, "a+", encoding="utf-8") as fo:
        fo.write(logs)

def judge_write_logs_file_exist():
    if not os.path.exists(txt_relative_path):
        os.mkdir(txt_relative_path)
    if not os.path.exists(xlsx_relative_path):
        os.mkdir(xlsx_relative_path)

def judge_and_write_data_to_xlsx():
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    if not os.path.exists(write_xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 11
        for i in range(len(XLSX_TITLE)):
            if i < len(XLSX_TITLE) - 2:
                ws.cell(i + 1, 1).value = XLSX_TITLE[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(XLSX_TITLE) - 2:
                ws.cell(i + 1, 1).value = list(XLSX_TITLE[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(XLSX_TITLE) - 1:
                ws.cell(i + 1, 1).value = XLSX_TITLE[i]
                ws.cell(i + 1, 1).alignment = alignment

    elif os.path.exists(write_xlsx_path):
        wb = load_workbook(write_xlsx_path)
        sheets_name_list = wb.sheetnames
        print(sheets_name_list)
        if sheet_name in sheets_name_list:
            ws = wb[sheet_name]
        elif sheet_name not in sheets_name_list:
            ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 11
        for i in range(len(XLSX_TITLE)):
            if i < len(XLSX_TITLE) - 2:
                ws.cell(i + 1, 1).value = XLSX_TITLE[i]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(XLSX_TITLE) - 2:
                ws.cell(i + 1, 1).value = list(XLSX_TITLE[i].keys())[0]
                ws.cell(i + 1, 1).alignment = alignment
            elif i == len(XLSX_TITLE) - 1:
                ws.cell(i + 1, 1).value = XLSX_TITLE[i]
                ws.cell(i + 1, 1).alignment = alignment

    tp_column_numb = column_index_from_string("A") + GL.xlsx_data_interval
    all_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 1
    tv_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 2
    radio_column_numb = column_index_from_string("A") + GL.xlsx_data_interval + 3
    tp_column_char = get_column_letter(tp_column_numb)
    all_column_char = get_column_letter(all_column_numb)
    tv_column_char = get_column_letter(tv_column_numb)
    radio_column_char = get_column_letter(radio_column_numb)
    ws.column_dimensions[tp_column_char].width = 12
    ws.column_dimensions[all_column_char].width = 3
    ws.column_dimensions[tv_column_char].width = 3
    ws.column_dimensions[radio_column_char].width = 3

    for m in range(len(GL.search_datas)):
        if m < len(GL.search_datas) - 2:
            ws.cell((m + 1), (1 + GL.xlsx_data_interval)).value = GL.search_datas[m]
            ws.merge_cells(start_row=(m + 1), start_column=(1 + GL.xlsx_data_interval), \
                           end_row=(m + 1), end_column=(1 + GL.xlsx_data_interval + 4))
            ws.cell((m + 1), (1 + GL.xlsx_data_interval)).alignment = alignment
        elif m == len(GL.search_datas) - 2:
            for n in range(len(XLSX_TITLE[7]["数据类别"])):
                ws.cell((m + 1), (1 + GL.xlsx_data_interval + n)).value = list(XLSX_TITLE[m].values())[0][n]
                ws.cell((m + 1), (1 + GL.xlsx_data_interval + n)).alignment = alignment
                ws.row_dimensions[(m + 1)].height = 13.5
        elif m == len(GL.search_datas) - 1:
            for j in range(len(GL.all_tp_list)):
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval)).value = GL.search_datas[m][j]
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 1).value = len(GL.channel_info[str(j + 1)][0]) + \
                                                                              len(GL.channel_info[str(j + 1)][1])
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 2).value = len(GL.channel_info[str(j + 1)][0])
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 3).value = len(GL.channel_info[str(j + 1)][1])
                ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + 4).value = ",".join(GL.channel_info[str(j + 1)][0] + \
                                                                                       GL.channel_info[str(j + 1)][1])
                for k in range(len(XLSX_TITLE[7]["数据类别"])):
                    ws.cell((m + 1 + j), (1 + GL.xlsx_data_interval) + k).alignment = alignment
                ws.row_dimensions[(m + 1 + j)].height = 13.5
    wb.save(write_xlsx_path)

GL = MyGlobal()

# 保存xlsx和txt文件的名称和路径处理
sat_name = re.split("_",COMMD_FILE_NAME[choice_commd_file_numb])[0]
search_mode = re.split("_",COMMD_FILE_NAME[choice_commd_file_numb])[2]
timestamp = re.sub(r"[-: ]","_",str(datetime.now()))[:19]
sheet_name = "{}_{}".format(sat_name,search_mode)

xlsx_name = "AddSatSearchResult.xlsx"
xlsx_relative_path = r".\Result"
write_xlsx_path = os.path.join(xlsx_relative_path,xlsx_name)

txt_name = "{}_{}_{}.txt".format(sat_name,search_mode,timestamp)
txt_relative_path = r".\PrintLog"
write_txt_path = os.path.join(txt_relative_path,txt_name)

judge_write_logs_file_exist()

# 获取所选测试项指令文件，并创建指令集
send_commds = []
send_commds_file_name = COMMD_FILE_NAME[choice_commd_file_numb]     # 所选测试项
parent_of_current_path = os.path.abspath(os.path.join(os.getcwd(),".."))        # 当前程序路径的上级路径
send_commds_path = os.path.join(parent_of_current_path,"CommandFile",send_commds_file_name)     # 所选测试项文件所在路径
send_commds = rend_commds_from_txt(send_commds_path)
GL.send_commds_length = len(send_commds)

# 获取串口并配置串口信息
send_ser_name,receive_ser_name = check_ports(GL.ser_cable_numb)
send_ser = serial.Serial()
receive_ser = serial.Serial()
serial_set(send_ser,send_ser_name,9600)
serial_set(receive_ser,receive_ser_name,115200)

while MAIN_LOOP_STATE:
    data = receive_ser.readline()
    if data:
        tt = datetime.now()
        data1 = data.decode("ISO-8859-1")
        data2 = re.compile("[\\x00-\\x08\\x0b-\\x0c\\x0e-\\x1f]").sub("",data1).strip()
        data3 = "[{}]     {}\n".format(str(tt),data2)
#        print(data2)
        write_logs_to_txt(write_txt_path, data3)

        if GL.search_monitor_kws[0] in data2:       # 监控搜索起始
            SEND_COMMD_STATE = False
            start_time = datetime.now()
            GL.search_datas[1] += 1
            GL.xlsx_data_interval = 1 + 5 * (GL.search_datas[1] - 1)

        if GL.search_monitor_kws[11] in data2:      # 监控极化方向
            GL.blind_judge_polar[0].append(data2)

        if len(GL.blind_judge_polar[0]) != 0:
            if len(GL.blind_judge_polar[0]) not in GL.blind_judge_polar[1]:
                GL.blind_judge_polar[1].append(len(GL.blind_judge_polar[0]))
            elif len(GL.blind_judge_polar[0]) in GL.blind_judge_polar[1]:
                GL.blind_judge_polar[2].add(len(GL.blind_judge_polar[1]))
                if (len(GL.blind_judge_polar[2]) % 2) != 0:
                    GL.blind_judge_polar[3] = "H"
                elif (len(GL.blind_judge_polar[2]) % 2) == 0:
                    GL.blind_judge_polar[3] = "V"

        if GL.search_monitor_kws[4] in data2:       # 监控频点信息
            fre = data2.split(" ")[5]
            symb = data2.split(" ")[9]
            tp = "{}{}{}".format(fre,GL.blind_judge_polar[3],symb)
            GL.all_tp_list.append(tp)
            GL.channel_info[str(len(GL.all_tp_list))] = [[],[]]

        if GL.search_monitor_kws[1] in data2:       # 监控搜索过程电视个数和名称信息
            GL.tv_radio_tp_data[0] = re.split("-{2,}|\s{2,}", data2)[1]    # 提取电视节目数
            tv_name = re.split("-{2,}|\s{2,}", data2)[2]    # 提取电视节目名称
            GL.channel_info[str(len(GL.all_tp_list))][0].append('[T]{}'.format(tv_name))

        if GL.search_monitor_kws[2] in data2:       # 监控搜索过程广播个数和名称信息
            GL.tv_radio_tp_data[1] = re.split("-{2,}|\s{2,}", data2)[1]     # 提取广播节目数
            radio_name = re.split("-{2,}|\s{2,}", data2)[2]     # 提取电视节目名称
            GL.channel_info[str(len(GL.all_tp_list))][1].append('[R]{}'.format(radio_name))

        if GL.search_monitor_kws[9] in data2 or GL.search_monitor_kws[10] in data2:     # 监控搜索达到上限
            limit_type = re.split(r"[ _]",data2)[1]
            print(limit_type)
            print("搜索{}达到上限:{}".format(limit_type,data2))
            if int(GL.tv_radio_tp_data[0]) != 0 or int(GL.tv_radio_tp_data[1]) != 0:
                send_ser.write(hex_strs_to_bytes(OK_KEY_VALUE))
            elif int(GL.tv_radio_tp_data[0]) == 0 and int(GL.tv_radio_tp_data[1]) == 0:
                pass
            search_numb = 1

        if GL.search_monitor_kws[3] in data2:       # 监控搜索结束
            end_time = datetime.now()
            SEND_COMMD_STATE = True
            search_dur_time = str(end_time - start_time)[2:10]
            print("第{}次搜索节目总数为TV/Radio:{}/{},TP总数为:{},盲扫时长:{}".format(GL.search_datas[1], \
                                                                      GL.tv_radio_tp_data[0], GL.tv_radio_tp_data[1], \
                                                                      len(GL.all_tp_list), search_dur_time))
            for i in range(len(GL.all_tp_list)):
                print(GL.all_tp_list[i])
            GL.search_datas[0] = search_mode
            GL.search_datas[2] = len(GL.all_tp_list)
            GL.search_datas[3] = "{}/{}".format(GL.tv_radio_tp_data[0], GL.tv_radio_tp_data[1])
            GL.search_datas[6] = search_dur_time
            GL.search_datas[8] = GL.all_tp_list
            judge_and_write_data_to_xlsx()

        if GL.search_monitor_kws[5] in data2:  # 监控保存TP的个数
            GL.tv_radio_tp_data[4] = int(re.split("=", data2)[1])
            GL.search_datas[4] = GL.tv_radio_tp_data[4]

        if GL.search_monitor_kws[6] in data2:  # 监控保存TV和Radio的个数
            split_result = re.split(r"[],]", data2)
            GL.tv_radio_tp_data[2] = re.split("=", split_result[1])[1]
            GL.tv_radio_tp_data[3] = re.split("=", split_result[2])[1]
            GL.search_datas[5] = "{}/{}".format(GL.tv_radio_tp_data[2], GL.tv_radio_tp_data[3])
            judge_and_write_data_to_xlsx()
            GL.tv_radio_tp_accumulated[0].append(int(GL.tv_radio_tp_data[0]))  # 累加TV数
            GL.tv_radio_tp_accumulated[1].append(int(GL.tv_radio_tp_data[1]))  # 累加Radio数
            GL.tv_radio_tp_accumulated[2].append(len(GL.all_tp_list))          # 累加TP数

            print("本次搜索实际保存TV/Radio:{},保存TP数为:{}".format(GL.search_datas[5],GL.search_datas[4]))
            print("当前轮次:{},累计搜索节目个数:{}/{},TP个数:{}".format(GL.search_datas[1],\
                                                          sum(GL.tv_radio_tp_accumulated[0]),\
                                                          sum(GL.tv_radio_tp_accumulated[1]),\
                                                          sum(GL.tv_radio_tp_accumulated[2])))
            GL.all_tp_list.clear()
            GL.blind_judge_polar = [[],[],set(),'']
            GL.channel_info.clear()
            GL.tv_radio_tp_data = [0] * 5
            GL.search_datas[4] = 0
            GL.search_datas[5] = "{}/{}".format(GL.tv_radio_tp_data[2], GL.tv_radio_tp_data[3])

    if not data and SEND_COMMD_STATE:
        if GL.send_commds_pos != GL.send_commds_length:
            print("{}:{}".format(GL.send_commds_pos, GL.send_commds_length - 1))
            print(send_commds[GL.send_commds_pos])
            send_ser.write(hex_strs_to_bytes(send_commds[GL.send_commds_pos]))
            GL.send_commds_pos += 1

        elif GL.send_commds_pos == GL.send_commds_length and search_numb >= 1:
            time.sleep(3)
            GL.send_commds_pos = 0
            search_numb -= 1
            if search_numb == 0:
                MAIN_LOOP_STATE = False
